from __future__ import annotations

import ast
import json
import logging
import os
import re
import time
import urllib.error
import urllib.request
from pathlib import Path
from typing import Any, Iterable

from kmwe.core.config_loader import ConfigError
from openpyxl import load_workbook
from kmwe.core.utils import iso_now
from kmwe.core.run_context import RunContext
from kmwe.utils.jsonio import write_json, write_jsonl_line

LLM_PROMPTS_FILENAME = "llm_prompts.jsonl"
LLM_RESPONSES_FILENAME = "llm_responses.jsonl"
RERANK_FREEZE_SPEC_V1 = {
    "schema_version": "rerank_freeze_v1",
    "index_relpath": "INDEX.json",
    "required_report_keys": [
        "status",
        "llm_mode",
        "input_pred_path",
        "responses_path",
        "prompts_written",
        "responses_rows",
        "decision_none_rows",
        "applied_selected",
    ],
}


def _load_expredict_prompt_meta(
    *,
    dict_xlsx_path: Path,
    expredict_sheet: str,
    logger: logging.Logger,
) -> dict[str, dict[str, str]]:
    meta_by_eid: dict[str, dict[str, str]] = {}
    try:
        if not dict_xlsx_path.exists():
            return meta_by_eid
        wb = load_workbook(dict_xlsx_path, read_only=True, data_only=True)
        if expredict_sheet not in wb.sheetnames:
            wb.close()
            return meta_by_eid
        ws = wb[expredict_sheet]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        header = [str(h).strip() for h in (header_row or [])]
        idx = {name: i for i, name in enumerate(header) if name}
        e_idx = idx.get("e_id")
        if e_idx is None:
            wb.close()
            return meta_by_eid
        canonical_idx = idx.get("canonical_form")
        gloss_idx = idx.get("gloss")
        prag_idx = idx.get("pragmatics")
        disamb_idx = idx.get("disambiguation_hint")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue
            if e_idx >= len(row):
                continue
            eid = row[e_idx]
            eid_s = str(eid).strip() if eid is not None else ""
            if not eid_s:
                continue
            canonical_form = (
                str(row[canonical_idx]).strip()
                if canonical_idx is not None and canonical_idx < len(row) and row[canonical_idx] is not None
                else ""
            )
            gloss = (
                str(row[gloss_idx]).strip()
                if gloss_idx is not None and gloss_idx < len(row) and row[gloss_idx] is not None
                else ""
            )
            pragmatics = (
                str(row[prag_idx]).strip()
                if prag_idx is not None and prag_idx < len(row) and row[prag_idx] is not None
                else ""
            )
            disamb = (
                str(row[disamb_idx]).strip()
                if disamb_idx is not None and disamb_idx < len(row) and row[disamb_idx] is not None
                else ""
            )
            meta_by_eid[eid_s] = {
                "canonical_form": canonical_form,
                "gloss": gloss,
                "pragmatics": pragmatics,
                "disambiguation_hint": disamb,
            }
        wb.close()
    except Exception as exc:
        logger.warning("infer_step2_rerank failed to load expredict prompt meta: %s", exc)
    return meta_by_eid


def _load_llm_examples_by_eid(
    *,
    dict_xlsx_path: Path,
    llm_examples_sheet: str,
    logger: logging.Logger,
) -> dict[str, dict[str, list[dict[str, str]]]]:
    examples_by_eid: dict[str, dict[str, list[dict[str, str]]]] = {}
    allowed_roles = {"pos", "dispos", "neg", "conf"}
    try:
        if not dict_xlsx_path.exists():
            return examples_by_eid
        wb = load_workbook(dict_xlsx_path, read_only=True, data_only=True)
        if llm_examples_sheet not in wb.sheetnames:
            wb.close()
            return examples_by_eid
        ws = wb[llm_examples_sheet]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        header = [str(h).strip() for h in (header_row or [])]
        idx = {name: i for i, name in enumerate(header) if name}

        e_idx = idx.get("e_id")
        role_idx = idx.get("example_role")
        sent_idx = idx.get("target_sentence")
        if sent_idx is None:
            sent_idx = idx.get("raw_sentence")
        span_segments_idx = idx.get("span_segments")
        span_text_idx = idx.get("span_text")
        note_idx = idx.get("note")
        conf_note_idx = idx.get("conf_note")
        if e_idx is None or role_idx is None or sent_idx is None:
            wb.close()
            return examples_by_eid

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue
            if e_idx >= len(row) or role_idx >= len(row) or sent_idx >= len(row):
                continue
            eid = str(row[e_idx]).strip() if row[e_idx] is not None else ""
            role = str(row[role_idx]).strip().lower() if row[role_idx] is not None else ""
            sent = str(row[sent_idx]).strip() if row[sent_idx] is not None else ""
            if not eid or not role or not sent:
                continue
            if role not in allowed_roles:
                continue

            span_segments = (
                str(row[span_segments_idx]).strip()
                if span_segments_idx is not None and span_segments_idx < len(row) and row[span_segments_idx] is not None
                else ""
            )
            span_text = (
                str(row[span_text_idx]).strip()
                if span_text_idx is not None and span_text_idx < len(row) and row[span_text_idx] is not None
                else ""
            )
            note = (
                str(row[note_idx]).strip()
                if note_idx is not None and note_idx < len(row) and row[note_idx] is not None
                else ""
            )
            if not note and conf_note_idx is not None and conf_note_idx < len(row) and row[conf_note_idx] is not None:
                note = str(row[conf_note_idx]).strip()

            role_map = examples_by_eid.setdefault(eid, {})
            bucket = role_map.setdefault(role, [])
            entry = {
                "example_role": role,
                "raw_sentence": sent,
                "span_segments": span_segments,
                "span_text": span_text,
                "note": note,
            }
            if entry not in bucket:
                bucket.append(entry)
        wb.close()
    except Exception as exc:
        logger.warning("infer_step2_rerank failed to load llm_examples: %s", exc)
    return examples_by_eid


def _resolve_llm_runtime_config(
    *,
    rerank_cfg: dict[str, Any],
    llm_cfg: dict[str, Any],
) -> dict[str, Any]:
    legacy_llm_cfg = rerank_cfg.get("llm", {}) or {}
    return {
        "base_url": str(
            llm_cfg.get("base_url")
            or legacy_llm_cfg.get("base_url")
            or os.getenv("OPENAI_BASE_URL")
            or "https://api.openai.com/v1"
        ),
        "api_key_env": str(
            llm_cfg.get("api_key_env")
            or legacy_llm_cfg.get("api_key_env")
            or "OPENAI_API_KEY"
        ),
        "model": str(
            llm_cfg.get("model")
            or legacy_llm_cfg.get("model")
            or os.getenv("OPENAI_MODEL")
            or "gpt-4.1-mini"
        ),
        "temperature": float(
            llm_cfg.get("temperature", legacy_llm_cfg.get("temperature", 0))
        ),
        "max_tokens": int(
            llm_cfg.get("max_tokens", legacy_llm_cfg.get("max_tokens", 64))
        ),
        "batch_size": int(
            llm_cfg.get("batch_size", legacy_llm_cfg.get("batch_size", 20))
        ),
        "retry": int(
            llm_cfg.get("retry", legacy_llm_cfg.get("retry", 2))
        ),
        "backoff_seconds": float(
            llm_cfg.get("backoff_seconds", legacy_llm_cfg.get("backoff_seconds", 1.0))
        ),
        "timeout_seconds": float(
            llm_cfg.get("timeout_seconds", legacy_llm_cfg.get("timeout_seconds", 45))
        ),
    }


def run_infer_step2_rerank(*, cfg: dict[str, Any], run_context: RunContext) -> None:
    logger = logging.getLogger("kmwe")
    outputs_dir = run_context.outputs_dir
    outputs_dir.mkdir(parents=True, exist_ok=True)

    rerank_cfg = cfg.get("infer_step2_rerank", {}) or {}
    input_path, input_path_source, auto_selected = _resolve_input_pred_path(
        cfg=cfg,
        run_context=run_context,
        logger=logger,
        rerank_cfg=rerank_cfg,
    )
    if not input_path.exists():
        raise ConfigError(f"infer_step2_rerank 입력 JSONL이 존재하지 않습니다: {input_path}")

    output_path = outputs_dir / "infer_candidates.reranked.jsonl"
    report_path = outputs_dir / "infer_step2_rerank_report.json"

    llm_mode = str(rerank_cfg.get("llm_mode") or "export_only")
    if llm_mode not in ("export_only", "apply_responses", "call_llm", "full_run"):
        raise ConfigError(f"infer_step2_rerank.llm_mode 값이 올바르지 않습니다: {llm_mode}")
    prompts_path = rerank_cfg.get("prompts_path") or rerank_cfg.get("prompts_jsonl")
    responses_path = rerank_cfg.get("responses_path") or rerank_cfg.get("responses_jsonl")
    llm_prompts_path = Path(prompts_path) if prompts_path else (outputs_dir / LLM_PROMPTS_FILENAME)
    llm_responses_path = (
        Path(responses_path) if responses_path else (outputs_dir / LLM_RESPONSES_FILENAME)
    )
    if not llm_prompts_path.is_absolute():
        parts = llm_prompts_path.parts
        if parts and parts[0] == "outputs":
            llm_prompts_path = outputs_dir / Path(*parts[1:])
        else:
            llm_prompts_path = outputs_dir / llm_prompts_path
    if not llm_responses_path.is_absolute():
        parts = llm_responses_path.parts
        if parts and parts[0] == "outputs":
            llm_responses_path = outputs_dir / Path(*parts[1:])
        else:
            llm_responses_path = outputs_dir / llm_responses_path
    if llm_mode == "full_run":
        llm_responses_path = outputs_dir / LLM_RESPONSES_FILENAME
    llm_prompts_path.parent.mkdir(parents=True, exist_ok=True)
    llm_responses_path.parent.mkdir(parents=True, exist_ok=True)
    llm_cfg = cfg.get("llm_rerank", {}) or {}
    llm_runtime_cfg = _resolve_llm_runtime_config(rerank_cfg=rerank_cfg, llm_cfg=llm_cfg)
    llm_prompt_cfg = llm_cfg.get("prompt", {}) or {}
    output_cfg = rerank_cfg.get("output", {}) or {}
    write_prompt_metadata = bool(output_cfg.get("write_prompt_metadata", True))

    dict_cfg = cfg.get("dict", {}) or {}
    sheet_names = dict_cfg.get("sheet_names", {}) or {}
    expredict_sheet = str(sheet_names.get("expredict", "expredict"))
    llm_examples_sheet = str(sheet_names.get("llm_examples", "llm_examples"))
    paths_cfg = cfg.get("paths", {}) or {}
    dict_xlsx_raw = paths_cfg.get("dict_xlsx") or paths_cfg.get("expredict_xlsx") or ""
    if dict_xlsx_raw:
        dict_xlsx_path = Path(dict_xlsx_raw)
    else:
        repo_root = Path(__file__).resolve().parents[2]
        dict_xlsx_path = repo_root / "data" / "dict" / "expredict.xlsx"
    expredict_meta_by_eid = _load_expredict_prompt_meta(
        dict_xlsx_path=dict_xlsx_path,
        expredict_sheet=expredict_sheet,
        logger=logger,
    )
    llm_examples_by_eid = _load_llm_examples_by_eid(
        dict_xlsx_path=dict_xlsx_path,
        llm_examples_sheet=llm_examples_sheet,
        logger=logger,
    )
    llm_examples_rows = sum(
        len(rows)
        for role_map in llm_examples_by_eid.values()
        for rows in role_map.values()
    )
    llm_transduction_cfg = llm_cfg.get("transduction", {}) or {}
    transduction_cfg = rerank_cfg.get("transduction", {}) or {}
    allow_multiple = bool(
        transduction_cfg.get(
            "allow_multiple",
            llm_transduction_cfg.get(
                "allow_multiple",
                rerank_cfg.get("allow_multiple", llm_cfg.get("allow_multiple", False)),
            ),
        )
    )
    stage_log_path = run_context.logs_dir / "stage.log"
    logger.info("infer_step2_rerank stage_log_path=%s outputs_dir=%s", stage_log_path, outputs_dir)
    logger.info("infer_step2_rerank rerank_run_id=%s", run_context.run_id)
    logger.info("infer_step2_rerank llm_examples loaded: eids=%s rows=%s", len(llm_examples_by_eid), llm_examples_rows)
    if bool(llm_prompt_cfg.get("include_examples", False)) and llm_examples_rows == 0:
        logger.warning(
            "infer_step2_rerank include_examples=true but llm_examples_rows=0 (dict_xlsx_path=%s)",
            dict_xlsx_path,
        )
        raise ConfigError(
            "llm_rerank.prompt.include_examples=true but llm_examples rows were not loaded. "
            f"resolved dict_xlsx_path={dict_xlsx_path}"
        )

    responses_loaded = 0
    responses_missing = 0
    call_llm_stats = {
        "responses_written": 0,
        "http_fail": 0,
        "http_401": 0,
        "parse_fail": 0,
        "guard_reject": 0,
        "empty_text": 0,
        "decision_line_non_empty": 0,
        "error_non_empty": 0,
        "decision_none_rows": 0,
        "n_prompts_in": 0,
        "n_batches": 0,
    }
    rerank_stats: dict[str, Any] = {
        "prompts_written": 0,
        "applied_selected": 0,
        "applied_selected_multi": 0,
        "n_selected_eids_total": 0,
        "fallback_used": 0,
        "n_records": 0,
        "n_input_candidates": 0,
        "n_to_llm_candidates": 0,
        "n_records_with_to_llm": 0,
        "missing_match_key_records": 0,
        "skipped_prompts": 0,
        "skipped_apply": 0,
    }
    report = {
        "created_at": iso_now(),
        "rerank_run_id": run_context.run_id,
        "input_pred_path": str(input_path),
        "rerank_output_path": None,
        "output_path": str(output_path),
        "llm_mode": llm_mode,
        "status": "running",
        "error": "",
        "dict_xlsx_path_resolved": str(dict_xlsx_path),
        "llm_examples_loaded_eids": len(llm_examples_by_eid),
        "llm_examples_loaded_rows": llm_examples_rows,
        "llm_model": str(llm_runtime_cfg.get("model") or ""),
        "llm_api_key_env": str(llm_runtime_cfg.get("api_key_env") or ""),
        "llm_base_url": str(llm_runtime_cfg.get("base_url") or ""),
    }

    try:
        if llm_mode == "apply_responses":
            (
                responses_map,
                responses_loaded,
                responses_empty_decision,
                responses_total_rows,
            ) = _load_decision_lines(llm_responses_path, allow_multiple=allow_multiple)
            logger.info(
                "infer_step2_rerank responses_loaded=%d responses_empty_decision=%d responses_total_rows=%d",
                responses_loaded,
                responses_empty_decision,
                responses_total_rows,
            )
            rerank_stats = _run_apply_responses(
                input_path=input_path,
                output_path=output_path,
                prompts_path=llm_prompts_path,
                responses_map=responses_map,
                allow_multiple=allow_multiple,
                write_prompt_metadata=write_prompt_metadata,
                expredict_meta_by_eid=expredict_meta_by_eid,
                llm_prompt_cfg=llm_prompt_cfg,
                llm_examples_by_eid=llm_examples_by_eid,
            )
            responses_missing = rerank_stats["responses_missing"]
            decision_none_rows = 0
            if llm_responses_path.exists():
                try:
                    for obj in _iter_jsonl(llm_responses_path):
                        decision_line = (
                            obj.get("decision_line")
                            or obj.get("decision")
                            or obj.get("output")
                            or obj.get("content")
                            or ""
                        )
                        if isinstance(decision_line, str) and decision_line.strip().startswith(
                            "DECISION: NONE"
                        ):
                            decision_none_rows += 1
                except Exception:
                    decision_none_rows = 0
            call_llm_stats["decision_line_non_empty"] = max(
                0, responses_total_rows - responses_empty_decision
            )
            call_llm_stats["decision_none_rows"] = decision_none_rows
        elif llm_mode == "call_llm":
            base_url = str(llm_runtime_cfg.get("base_url") or "")
            api_key_env = str(llm_runtime_cfg.get("api_key_env") or "OPENAI_API_KEY")
            model = str(llm_runtime_cfg.get("model") or "gpt-4.1-mini")
            temperature = float(llm_runtime_cfg.get("temperature", 0))
            max_tokens = int(llm_runtime_cfg.get("max_tokens", 64))
            batch_size = int(llm_runtime_cfg.get("batch_size", 20))
            retry = int(llm_runtime_cfg.get("retry", 2))
            backoff_seconds = float(llm_runtime_cfg.get("backoff_seconds", 1.0))
            timeout_seconds = float(llm_runtime_cfg.get("timeout_seconds", 45))
            rerank_stats = _run_export_only(
                input_path=input_path,
                output_path=output_path,
                prompts_path=llm_prompts_path,
                status_label="export_only",
                write_prompts=True,
                allow_multiple=allow_multiple,
                write_prompt_metadata=write_prompt_metadata,
                expredict_meta_by_eid=expredict_meta_by_eid,
                llm_prompt_cfg=llm_prompt_cfg,
                llm_examples_by_eid=llm_examples_by_eid,
            )
            prompts_rows = sum(1 for _ in _iter_jsonl(llm_prompts_path))
            if prompts_rows == 0:
                raise RuntimeError("infer_step2_rerank call_llm prompts_path empty")
            api_key = os.getenv(api_key_env)
            if not api_key:
                raise ConfigError("infer_step2_rerank.llm.api_key_env 환경변수에 API key가 없습니다.")
            call_llm_stats = _run_call_llm(
                prompts_path=llm_prompts_path,
                responses_path=llm_responses_path,
                base_url=base_url,
                api_key=api_key,
                model=model,
                temperature=temperature,
                max_tokens=max_tokens,
                batch_size=batch_size,
                retry=retry,
                backoff_seconds=backoff_seconds,
                timeout_seconds=timeout_seconds,
                allow_multiple=allow_multiple,
            )
            if (
                rerank_stats.get("prompts_written", 0) > 0
                and call_llm_stats.get("responses_written", 0) == 0
            ):
                raise RuntimeError("infer_step2_rerank full_run responses_rows empty")
        elif llm_mode == "full_run":
            if auto_selected:
                raise ConfigError("infer_step2_rerank.full_run에서는 input_pred_path가 필수입니다.")
            base_url = str(llm_runtime_cfg.get("base_url") or "")
            api_key_env = str(llm_runtime_cfg.get("api_key_env") or "OPENAI_API_KEY")
            model = str(llm_runtime_cfg.get("model") or "gpt-4.1-mini")
            temperature = float(llm_runtime_cfg.get("temperature", 0))
            max_tokens = int(llm_runtime_cfg.get("max_tokens", 64))
            batch_size = int(llm_runtime_cfg.get("batch_size", 20))
            retry = int(llm_runtime_cfg.get("retry", 2))
            backoff_seconds = float(llm_runtime_cfg.get("backoff_seconds", 1.0))
            timeout_seconds = float(llm_runtime_cfg.get("timeout_seconds", 45))
            rerank_stats = _run_export_only(
                input_path=input_path,
                output_path=output_path,
                prompts_path=llm_prompts_path,
                status_label="export_only",
                write_prompts=True,
                allow_multiple=allow_multiple,
                write_prompt_metadata=write_prompt_metadata,
                expredict_meta_by_eid=expredict_meta_by_eid,
                llm_prompt_cfg=llm_prompt_cfg,
                llm_examples_by_eid=llm_examples_by_eid,
            )
            prompts_rows = sum(1 for _ in _iter_jsonl(llm_prompts_path))
            if prompts_rows == 0:
                raise RuntimeError("infer_step2_rerank full_run prompts_path empty")
            api_key = os.getenv(api_key_env)
            if not api_key:
                raise ConfigError("infer_step2_rerank.llm.api_key_env 환경변수에 API key가 없습니다.")
            call_llm_stats = _run_call_llm(
                prompts_path=llm_prompts_path,
                responses_path=llm_responses_path,
                base_url=base_url,
                api_key=api_key,
                model=model,
                temperature=temperature,
                max_tokens=max_tokens,
                batch_size=batch_size,
                retry=retry,
                backoff_seconds=backoff_seconds,
                timeout_seconds=timeout_seconds,
                allow_multiple=allow_multiple,
            )
            (
                responses_map,
                responses_loaded,
                responses_empty_decision,
                responses_total_rows,
            ) = _load_decision_lines(llm_responses_path, allow_multiple=allow_multiple)
            rerank_stats = _run_apply_responses(
                input_path=input_path,
                output_path=output_path,
                prompts_path=llm_prompts_path,
                responses_map=responses_map,
                allow_multiple=allow_multiple,
                write_prompt_metadata=write_prompt_metadata,
                expredict_meta_by_eid=expredict_meta_by_eid,
                llm_prompt_cfg=llm_prompt_cfg,
                llm_examples_by_eid=llm_examples_by_eid,
            )
            responses_missing = rerank_stats["responses_missing"]
            call_llm_stats["decision_line_non_empty"] = max(
                0, responses_total_rows - responses_empty_decision
            )
        else:
            rerank_stats = _run_export_only(
                input_path=input_path,
                output_path=output_path,
                prompts_path=llm_prompts_path,
                status_label="export_only",
                write_prompts=True,
                allow_multiple=allow_multiple,
                write_prompt_metadata=write_prompt_metadata,
                expredict_meta_by_eid=expredict_meta_by_eid,
                llm_prompt_cfg=llm_prompt_cfg,
                llm_examples_by_eid=llm_examples_by_eid,
            )

        logger.info(
            "infer_step2_rerank inputs: n_input_candidates=%s n_to_llm_candidates=%s n_records_with_to_llm=%s",
            rerank_stats["n_input_candidates"],
            rerank_stats["n_to_llm_candidates"],
            rerank_stats["n_records_with_to_llm"],
        )
        logger.info(
            "infer_step2_rerank llm_prompts_path=%s llm_responses_path=%s",
            llm_prompts_path,
            llm_responses_path,
        )
        logger.info(
            "infer_step2_rerank llm_mode=%s prompts_written=%s",
            llm_mode,
            rerank_stats["prompts_written"],
        )
        if llm_mode in ("call_llm", "full_run"):
            responses_bytes = 0
            responses_rows = 0
            if llm_responses_path.exists():
                try:
                    responses_bytes = llm_responses_path.stat().st_size
                    responses_rows = sum(1 for _ in _iter_jsonl(llm_responses_path))
                except Exception:
                    responses_bytes = 0
                    responses_rows = 0
            logger.info(
                "infer_step2_rerank responses_path=%s responses_bytes=%s responses_rows=%s http_fail=%s http_401=%s parse_fail=%s guard_reject=%s empty_text=%s decision_line_non_empty=%s error_non_empty=%s decision_none_rows=%s n_prompts_in=%s n_batches=%s",
                llm_responses_path,
                responses_bytes,
                responses_rows,
                call_llm_stats["http_fail"],
                call_llm_stats["http_401"],
                call_llm_stats["parse_fail"],
                call_llm_stats["guard_reject"],
                call_llm_stats["empty_text"],
                call_llm_stats["decision_line_non_empty"],
                call_llm_stats["error_non_empty"],
                call_llm_stats["decision_none_rows"],
                call_llm_stats["n_prompts_in"],
                call_llm_stats["n_batches"],
            )
        else:
            logger.info(
                "infer_step2_rerank responses_loaded=%s missing=%s",
                responses_loaded,
                responses_missing,
            )
        logger.info(
            "infer_step2_rerank missing_match_key_records=%s skipped_prompts=%s skipped_apply=%s",
            rerank_stats["missing_match_key_records"],
            rerank_stats["skipped_prompts"],
            rerank_stats["skipped_apply"],
        )
        logger.info(
            "infer_step2_rerank applied_selected=%s fallback=%s no_response=%s missing=%s rerank_output_path=%s",
            rerank_stats["applied_selected"],
            rerank_stats["fallback_used"],
            responses_missing,
            responses_missing,
            output_path,
        )
        if llm_mode in ("call_llm", "apply_responses", "full_run") and llm_responses_path.exists():
            summary = _summarize_llm_responses(llm_responses_path)
            summary.update(
                {
                    "llm_mode": llm_mode,
                    "rerank_run_id": run_context.run_id,
                    "responses_path": str(llm_responses_path),
                    "prompts_path": str(llm_prompts_path),
                }
            )
            write_json(outputs_dir / "rerank_summary.json", summary, indent=2)

        logger.info(
            "infer_step2_rerank health prompts_written=%s responses_rows=%s http_fail=%s http_401=%s parse_fail=%s guard_reject=%s empty_text=%s decision_line_non_empty=%s error_non_empty=%s decision_none_rows=%s applied_selected=%s no_response=%s missing=%s",
            rerank_stats.get("prompts_written", 0),
            call_llm_stats.get("responses_written", 0),
            call_llm_stats.get("http_fail", 0),
            call_llm_stats.get("http_401", 0),
            call_llm_stats.get("parse_fail", 0),
            call_llm_stats.get("guard_reject", 0),
            call_llm_stats.get("empty_text", 0),
            call_llm_stats.get("decision_line_non_empty", 0),
            call_llm_stats.get("error_non_empty", 0),
            call_llm_stats.get("decision_none_rows", 0),
            rerank_stats.get("applied_selected", 0),
            responses_missing,
            responses_missing,
        )
        report["status"] = "ok"
    except Exception as exc:
        report["status"] = "failed"
        report["error"] = f"{type(exc).__name__}: {exc}"
        raise
    finally:
        responses_rows_report = 0
        if llm_responses_path.exists():
            try:
                responses_rows_report = sum(1 for _ in _iter_jsonl(llm_responses_path))
            except Exception:
                responses_rows_report = 0
        rerank_output_path = None
        if llm_mode in ("apply_responses", "full_run"):
            rerank_output_path = str(output_path)
        report.update(
            {
                "input_path": str(input_path),
                "input_path_source": input_path_source,
                "input_path_auto_selected": auto_selected,
                "output_path": str(output_path),
                "n_records": rerank_stats.get("n_records", 0),
                "n_input_candidates": rerank_stats.get("n_input_candidates", 0),
                "n_to_llm_candidates": rerank_stats.get("n_to_llm_candidates", 0),
                "n_records_with_to_llm": rerank_stats.get("n_records_with_to_llm", 0),
                "rerank_run_id": run_context.run_id,
                "input_pred_path": str(input_path),
                "rerank_output_path": rerank_output_path,
                "responses_path": str(llm_responses_path),
                "llm_mode": llm_mode,
                "prompts_written": rerank_stats.get("prompts_written", 0),
                "responses_rows": responses_rows_report,
                "decision_none_rows": call_llm_stats.get("decision_none_rows", 0),
                "applied_selected": rerank_stats.get("applied_selected", 0),
                "created_at": iso_now(),
            }
        )
        write_json(report_path, report, indent=2)
        logger.info(
            "[infer_step2_rerank][responses_path] path=%s exists=%s bytes=%s rows=%s",
            llm_responses_path,
            llm_responses_path.exists(),
            llm_responses_path.stat().st_size if llm_responses_path.exists() else 0,
            responses_rows_report,
        )
        logger.info(
            "[infer_step2_rerank][report_write] path=%s exists=%s bytes=%s",
            report_path,
            report_path.exists(),
            report_path.stat().st_size if report_path.exists() else 0,
        )
        logger.info(
            "[infer_step2_rerank][report_path] path=%s exists=%s bytes=%s",
            report_path,
            report_path.exists(),
            report_path.stat().st_size if report_path.exists() else 0,
        )
        logger.info(
            "[infer_step2_rerank][report_keys] keys=%s status=%s",
            list(report.keys()),
            report.get("status"),
        )
        _write_rerank_index_json(
            run_context=run_context,
            outputs_dir=outputs_dir,
            report_path=report_path,
            logger=logger,
        )
        _freeze_rerank_outputs(outputs_dir=outputs_dir, report_path=report_path, logger=logger)


def _resolve_match_key(record: dict[str, Any]) -> str | None:
    match_key = record.get("match_key")
    if isinstance(match_key, str) and match_key.strip():
        return match_key.strip()
    example_id = record.get("example_id")
    instance_id = record.get("instance_id")
    if isinstance(example_id, str) and example_id and isinstance(instance_id, (str, int)):
        return f"{example_id}#{instance_id}"
    return None


def _write_rerank_index_json(
    *,
    run_context: RunContext,
    outputs_dir: Path,
    report_path: Path,
    logger: logging.Logger,
) -> Path:
    report: dict[str, Any] = {}
    if report_path.exists() and report_path.stat().st_size > 0:
        try:
            report = json.loads(report_path.read_text(encoding="utf-8"))
        except Exception:
            report = {}

    def _artifact_entry(path_str: str, note: str) -> dict[str, Any]:
        path = Path(path_str) if path_str else Path("")
        exists = bool(path_str) and path.exists()
        size = path.stat().st_size if exists else 0
        return {
            "path": str(path_str),
            "exists": bool(exists),
            "bytes": int(size),
            "note": note,
        }

    input_pred_path = str(report.get("input_pred_path") or report.get("input_path") or "")
    rerank_output_path = str(report.get("rerank_output_path") or report.get("output_path") or "")
    responses_path = str(report.get("responses_path") or "")
    prompts_path = str(report.get("prompts_path") or "")

    artifacts: dict[str, dict[str, Any]] = {
        "report_json": _artifact_entry(
            str(report_path),
            "infer_step2_rerank 실행 요약(report)으로, eval pred_guard/llm_audit의 조인 기준입니다.",
        ),
        "input_pred": _artifact_entry(
            input_pred_path,
            "rerank 입력 pred(encoder 단계 산출물)로, 이 RUN이 무엇을 rerank했는지 증명합니다.",
        ),
        "rerank_output_pred": _artifact_entry(
            rerank_output_path,
            "rerank 적용 후 최종 pred 산출물로, eval에서 읽을 pred_path 후보입니다.",
        ),
        "responses_jsonl": _artifact_entry(
            responses_path,
            "LLM 응답 원본 JSONL로, llm_audit/decision 파싱의 근거입니다.",
        ),
    }
    if prompts_path:
        artifacts["prompts_jsonl"] = _artifact_entry(
            prompts_path,
            "LLM 호출 입력 프롬프트 JSONL로, responses와 1:1 대응 여부를 점검하는 근거입니다.",
        )

    run_summary = {
        "status": report.get("status"),
        "llm_mode": report.get("llm_mode"),
        "prompts_written": int(report.get("prompts_written") or 0),
        "responses_rows": int(report.get("responses_rows") or 0),
        "decision_none_rows": int(report.get("decision_none_rows") or 0),
        "applied_selected": int(report.get("applied_selected") or 0),
        "input_pred_path": input_pred_path or None,
        "rerank_output_path": str(report.get("rerank_output_path") or "") or None,
        "output_path": str(report.get("output_path") or "") or None,
        "responses_path": responses_path or None,
    }
    run_summary["run_summary_ko"] = (
        f"rerank RUN 요약: status={run_summary.get('status')}, "
        f"applied_selected={run_summary.get('applied_selected')}, "
        f"decision_none_rows={run_summary.get('decision_none_rows')}, "
        f"responses_rows={run_summary.get('responses_rows')} 입니다."
    )

    index_obj = {
        "schema_version": "rerank_index_v1",
        "generated_at": iso_now(),
        "run_dir": str(run_context.run_dir),
        "outputs_dir": str(outputs_dir),
        "run_summary": run_summary,
        "artifacts": artifacts,
    }
    out_path = outputs_dir / "INDEX.json"
    write_json(out_path, index_obj, indent=2)
    logger.info(
        "[infer_step2_rerank][index] wrote path=%s exists=%s bytes=%s",
        out_path,
        out_path.exists(),
        out_path.stat().st_size if out_path.exists() else 0,
    )
    return out_path


def _freeze_rerank_outputs(*, outputs_dir: Path, report_path: Path, logger: logging.Logger) -> None:
    logger.info(
        "[infer_step2_rerank][freeze] start schema=%s",
        RERANK_FREEZE_SPEC_V1.get("schema_version"),
    )
    if not report_path.exists() or report_path.stat().st_size <= 0:
        logger.info(
            "[infer_step2_rerank][freeze] FAIL reason=%s path=%s",
            "missing_report",
            report_path,
        )
        raise RuntimeError(f"infer_step2_rerank freeze failed: missing_report ({report_path})")
    try:
        report = json.loads(report_path.read_text(encoding="utf-8"))
    except Exception as exc:
        logger.info(
            "[infer_step2_rerank][freeze] FAIL reason=%s path=%s",
            "report_parse_fail",
            report_path,
        )
        raise RuntimeError(f"infer_step2_rerank freeze failed: report_parse_fail ({exc})")
    if not isinstance(report, dict):
        logger.info(
            "[infer_step2_rerank][freeze] FAIL reason=%s path=%s",
            "report_not_dict",
            report_path,
        )
        raise RuntimeError("infer_step2_rerank freeze failed: report_not_dict")
    for key in RERANK_FREEZE_SPEC_V1.get("required_report_keys", []):
        if key not in report:
            logger.info(
                "[infer_step2_rerank][freeze] FAIL reason=%s path=%s",
                "missing_report_key",
                report_path,
            )
            raise RuntimeError(f"infer_step2_rerank freeze failed: missing_report_key ({key})")

    input_pred_path = Path(str(report.get("input_pred_path") or ""))
    if not input_pred_path.exists() or input_pred_path.stat().st_size <= 0:
        logger.info(
            "[infer_step2_rerank][freeze] FAIL reason=%s path=%s",
            "input_pred_missing",
            input_pred_path,
        )
        raise RuntimeError(
            f"infer_step2_rerank freeze failed: input_pred_missing ({input_pred_path})"
        )

    out_pred_str = str(report.get("rerank_output_path") or report.get("output_path") or "")
    if not out_pred_str:
        logger.info(
            "[infer_step2_rerank][freeze] FAIL reason=%s path=%s",
            "missing_output_pred_key",
            report_path,
        )
        raise RuntimeError("infer_step2_rerank freeze failed: missing_output_pred_key")
    out_pred_path = Path(out_pred_str)
    if not out_pred_path.exists() or out_pred_path.stat().st_size <= 0:
        logger.info(
            "[infer_step2_rerank][freeze] FAIL reason=%s path=%s",
            "output_pred_missing",
            out_pred_path,
        )
        raise RuntimeError(
            f"infer_step2_rerank freeze failed: output_pred_missing ({out_pred_path})"
        )

    prompts_written = int(report.get("prompts_written") or 0)
    responses_path = Path(str(report.get("responses_path") or ""))
    if prompts_written > 0:
        if not responses_path.exists() or responses_path.stat().st_size <= 0:
            logger.info(
                "[infer_step2_rerank][freeze] FAIL reason=%s path=%s",
                "responses_missing",
                responses_path,
            )
            raise RuntimeError(
                f"infer_step2_rerank freeze failed: responses_missing ({responses_path})"
            )

    index_path = outputs_dir / str(RERANK_FREEZE_SPEC_V1.get("index_relpath") or "INDEX.json")
    if not index_path.exists() or index_path.stat().st_size <= 0:
        logger.info(
            "[infer_step2_rerank][freeze] FAIL reason=%s path=%s",
            "missing_index",
            index_path,
        )
        raise RuntimeError(f"infer_step2_rerank freeze failed: missing_index ({index_path})")
    try:
        index_obj = json.loads(index_path.read_text(encoding="utf-8"))
    except Exception as exc:
        logger.info(
            "[infer_step2_rerank][freeze] FAIL reason=%s path=%s",
            "index_parse_fail",
            index_path,
        )
        raise RuntimeError(f"infer_step2_rerank freeze failed: index_parse_fail ({exc})")
    if not isinstance(index_obj, dict):
        logger.info(
            "[infer_step2_rerank][freeze] FAIL reason=%s path=%s",
            "index_not_dict",
            index_path,
        )
        raise RuntimeError("infer_step2_rerank freeze failed: index_not_dict")
    for key in ["schema_version", "generated_at", "outputs_dir", "run_summary", "artifacts"]:
        if key not in index_obj:
            logger.info(
                "[infer_step2_rerank][freeze] FAIL reason=%s path=%s",
                "index_missing_key",
                index_path,
            )
            raise RuntimeError(f"infer_step2_rerank freeze failed: index_missing_key ({key})")
    artifacts = index_obj.get("artifacts")
    if not isinstance(artifacts, dict):
        logger.info(
            "[infer_step2_rerank][freeze] FAIL reason=%s path=%s",
            "index_missing_key",
            index_path,
        )
        raise RuntimeError("infer_step2_rerank freeze failed: index_missing_key (artifacts)")
    for key in ["report_json", "responses_jsonl", "rerank_output_pred"]:
        if key not in artifacts:
            logger.info(
                "[infer_step2_rerank][freeze] FAIL reason=%s path=%s",
                "index_missing_key",
                index_path,
            )
            raise RuntimeError(
                f"infer_step2_rerank freeze failed: index_missing_key (artifacts.{key})"
            )
    logger.info("[infer_step2_rerank][freeze] PASS")


def _iter_jsonl(path: Path) -> Iterable[dict[str, Any]]:
    if not path.exists():
        return []
    def _gen() -> Iterable[dict[str, Any]]:
        with path.open("r", encoding="utf-8") as fp:
            for line in fp:
                line = line.strip()
                if not line:
                    continue
                try:
                    obj = json.loads(line)
                except Exception:
                    continue
                if isinstance(obj, dict):
                    yield obj
    return _gen()


def _write_jsonl_atomic(path: Path, rows: Iterable[dict[str, Any]]) -> None:
    tmp_path = path.with_name(f"{path.name}.tmp")
    with tmp_path.open("w", encoding="utf-8") as fp:
        for row in rows:
            write_jsonl_line(fp, row)
    tmp_path.replace(path)


def _extract_decision_line(text: str, allow_multiple: bool) -> str:
    if not isinstance(text, str):
        return ""
    regex_none = r"^NONE\s*$"
    regex_single_num = r"^\d+\s*$"
    regex_multi_num = r"^\d+(?:\s*,\s*\d+)+\s*$"
    regex_single = r"^DECISION[:\s]+(e_id=(?P<eid>\w+)|NONE)\s*$"
    regex_multi = r"^DECISION[:\s]+e_id=(?P<eids>\w+(?:\s*,\s*\w+)*)\s*$"
    regex_multi_short = r"^DECISION:\s*(?P<eids>[a-z]{3}\d{3,}(?:\s*,\s*[a-z]{3}\d{3,})+)\s*$"
    regex_short = r"^DECISION:\s*(NONE|[a-z]{3}\d{3,})\s*$"
    regex_short_loose = r"^DECISION[:\s]+([a-z]{3}\d{3,}|NONE)\s*$"
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        if re.match(regex_none, line):
            return "NONE"
        if allow_multiple and re.match(regex_multi_num, line):
            nums = [n.strip() for n in line.split(',') if n.strip()]
            if nums:
                return ','.join(nums)
        if re.match(regex_single_num, line):
            return line.strip()
        match_short = re.match(regex_short, line)
        if match_short:
            val = match_short.group(1)
            if val == "NONE":
                return "DECISION: NONE"
            return f"DECISION: e_id={val}"
        match_loose = re.match(regex_short_loose, line)
        if match_loose:
            val = match_loose.group(1)
            if val == "NONE":
                return "DECISION: NONE"
            return f"DECISION: e_id={val}"
        if allow_multiple:
            match_multi_short = re.match(regex_multi_short, line)
            if match_multi_short:
                eids = [e.strip() for e in match_multi_short.group("eids").split(',') if e.strip()]
                if eids:
                    return f"DECISION: e_id={','.join(eids)}"
            match_multi = re.match(regex_multi, line)
            if match_multi:
                eids = [e.strip() for e in match_multi.group("eids").split(',') if e.strip()]
                if eids:
                    return f"DECISION: e_id={','.join(eids)}"
            match_single = re.match(regex_single, line)
            if match_single:
                if match_single.group("eid"):
                    return f"DECISION: e_id={match_single.group('eid')}"
                return "DECISION: NONE"
        else:
            match_single = re.match(regex_single, line)
            if match_single:
                if match_single.group("eid"):
                    return f"DECISION: e_id={match_single.group('eid')}"
                return "DECISION: NONE"
    return ""


def _validate_decision_eid(e_ids: list[str], candidate_eids: list[str]) -> bool:
    if not candidate_eids:
        return False
    cand_set = {str(eid) for eid in candidate_eids if eid}
    if not cand_set:
        return False
    for eid in e_ids:
        if str(eid) not in cand_set:
            return False
    return True


def _run_export_only(
    *,
    input_path: Path,
    output_path: Path,
    prompts_path: Path,
    status_label: str,
    write_prompts: bool,
    allow_multiple: bool,
    write_prompt_metadata: bool,
    expredict_meta_by_eid: dict[str, dict[str, str]] | None = None,
    llm_prompt_cfg: dict[str, Any] | None = None,
    llm_examples_by_eid: dict[str, dict[str, list[dict[str, str]]]] | None = None,
) -> dict[str, Any]:
    n_records = 0
    n_input_candidates = 0
    n_to_llm_candidates = 0
    n_records_with_to_llm = 0
    prompts_written = 0
    missing_match_key_records = 0
    skipped_prompts = 0
    skipped_apply = 0
    applied_selected = 0
    applied_selected_multi = 0
    applied_none = 0
    fallback_used = 0
    responses_missing = 0
    n_selected_eids_total = 0

    prompts_rows: list[dict[str, Any]] = []
    with input_path.open("r", encoding="utf-8") as fp, output_path.open(
        "w", encoding="utf-8"
    ) as out_fp:
        for line in fp:
            line = line.strip()
            if not line:
                continue
            try:
                record = json.loads(line)
            except Exception:
                continue
            if not isinstance(record, dict):
                continue
            candidates = record.get("candidates") or []
            if not isinstance(candidates, list):
                candidates = []

            match_key = _resolve_match_key(record)
            to_llm_with_rank = []
            for idx, cand in enumerate(candidates):
                if _is_to_llm_candidate(cand):
                    to_llm_with_rank.append((idx, cand))
            to_llm_with_rank.sort(
                key=lambda x: (
                    -(float(x[1].get("encoder_score") or 0.0)),
                    x[0],
                )
            )
            to_llm = [cand for _idx, cand in to_llm_with_rank]
            span_bundles = _group_to_llm_candidates_by_span(to_llm)
            span_bundle_summaries = _summarize_span_bundles(match_key, to_llm)
            n_input_candidates += len(candidates)
            n_to_llm_candidates += len(to_llm)
            if to_llm:
                n_records_with_to_llm += 1

            rerank_meta: dict[str, Any] = {
                "status": "no_candidates",
                "input_candidates": to_llm,
                "n_to_llm_candidates": len(to_llm),
                "span_bundle_keys": [item.get("span_bundle_key") for item in span_bundle_summaries],
                "span_bundle_key": _compose_span_bundle_key(match_key, _span_bundle_keys(to_llm)),
                "span_bundles": span_bundle_summaries,
                "n_span_bundles": len(span_bundles),
            }

            if to_llm:
                if not match_key:
                    missing_match_key_records += 1
                    skipped_prompts += 1
                    rerank_meta["status"] = "missing_match_key"
                else:
                    if write_prompts:
                        for span_key, bundle in span_bundles:
                            if not bundle:
                                continue
                            bundle_prompt = _build_llm_prompt(
                                record,
                                bundle,
                                _compose_span_bundle_key(match_key, [span_key]),
                                allow_multiple,
                                expredict_meta_by_eid=expredict_meta_by_eid,
                                llm_prompt_cfg=llm_prompt_cfg,
                                llm_examples_by_eid=llm_examples_by_eid,
                            )
                            if write_prompt_metadata:
                                prompts_rows.append(bundle_prompt)
                            else:
                                metadata = (
                                    bundle_prompt.get("metadata")
                                    if isinstance(bundle_prompt.get("metadata"), dict)
                                    else {}
                                )
                                prompts_rows.append(
                                    {
                                        "match_key": str(bundle_prompt.get("match_key") or ""),
                                        "system": str(bundle_prompt.get("system") or ""),
                                        "user": str(bundle_prompt.get("user") or ""),
                                        "metadata": {
                                            "span_bundle_key": str(metadata.get("span_bundle_key") or ""),
                                        },
                                    }
                                )
                            prompts_written += 1
                    rerank_meta["status"] = status_label

            output_record = {
                **record,
                "rerank": rerank_meta,
            }
            write_jsonl_line(out_fp, output_record)
            n_records += 1

    if write_prompts:
        _write_jsonl_atomic(prompts_path, prompts_rows)

    return {
        "n_records": n_records,
        "n_input_candidates": n_input_candidates,
        "n_to_llm_candidates": n_to_llm_candidates,
        "n_records_with_to_llm": n_records_with_to_llm,
        "prompts_written": prompts_written,
        "responses_missing": responses_missing,
        "applied_selected": applied_selected,
        "applied_selected_multi": applied_selected_multi,
        "n_selected_eids_total": n_selected_eids_total,
        "applied_none": applied_none,
        "fallback_used": fallback_used,
        "missing_match_key_records": missing_match_key_records,
        "skipped_prompts": skipped_prompts,
        "skipped_apply": skipped_apply,
    }


def _run_apply_responses(
    *,
    input_path: Path,
    output_path: Path,
    prompts_path: Path,
    responses_map: dict[str, str],
    allow_multiple: bool,
    write_prompt_metadata: bool,
    expredict_meta_by_eid: dict[str, dict[str, str]] | None = None,
    llm_prompt_cfg: dict[str, Any] | None = None,
    llm_examples_by_eid: dict[str, dict[str, list[dict[str, str]]]] | None = None,
) -> dict[str, Any]:
    n_records = 0
    n_input_candidates = 0
    n_to_llm_candidates = 0
    n_records_with_to_llm = 0
    prompts_written = 0
    missing_match_key_records = 0
    skipped_prompts = 0
    skipped_apply = 0
    applied_selected = 0
    applied_selected_multi = 0
    applied_none = 0
    fallback_used = 0
    responses_missing = 0
    n_selected_eids_total = 0

    prompts_rows: list[dict[str, Any]] = []
    with input_path.open("r", encoding="utf-8") as fp, output_path.open(
        "w", encoding="utf-8"
    ) as out_fp:
        for line in fp:
            line = line.strip()
            if not line:
                continue
            try:
                record = json.loads(line)
            except Exception:
                continue
            if not isinstance(record, dict):
                continue
            candidates = record.get("candidates") or []
            if not isinstance(candidates, list):
                candidates = []

            match_key = _resolve_match_key(record)
            to_llm_with_rank = []
            for idx, cand in enumerate(candidates):
                if _is_to_llm_candidate(cand):
                    to_llm_with_rank.append((idx, cand))
            to_llm_with_rank.sort(
                key=lambda x: (
                    -(float(x[1].get("encoder_score") or 0.0)),
                    x[0],
                )
            )
            to_llm = [cand for _idx, cand in to_llm_with_rank]
            span_bundles = _group_to_llm_candidates_by_span(to_llm)
            span_bundle_summaries = _summarize_span_bundles(match_key, to_llm)
            n_input_candidates += len(candidates)
            n_to_llm_candidates += len(to_llm)
            if to_llm:
                n_records_with_to_llm += 1

            rerank_meta: dict[str, Any] = {
                "status": "no_candidates",
                "input_candidates": to_llm,
                "n_to_llm_candidates": len(to_llm),
                "span_bundle_keys": [item.get("span_bundle_key") for item in span_bundle_summaries],
                "span_bundle_key": _compose_span_bundle_key(match_key, _span_bundle_keys(to_llm)),
                "span_bundles": span_bundle_summaries,
                "n_span_bundles": len(span_bundles),
                "span_reranks": [],
            }

            if to_llm:
                if not match_key:
                    missing_match_key_records += 1
                    skipped_prompts += 1
                    skipped_apply += 1
                    rerank_meta["status"] = "missing_match_key"
                else:
                    span_results: list[dict[str, Any]] = []
                    selected_union: list[str] = []
                    selected_seen: set[str] = set()
                    any_applied_or_fallback = False
                    all_protocol_ok = True
                    any_no_response = False
                    any_fallback = False
                    decision_lines_join: list[str] = []
                    for span_key, bundle in span_bundles:
                        if not bundle:
                            continue
                        bundle_match_key = _compose_span_bundle_key(match_key, [span_key])
                        prompt = _build_llm_prompt(
                            record,
                            bundle,
                            bundle_match_key,
                            allow_multiple,
                            expredict_meta_by_eid=expredict_meta_by_eid,
                            llm_prompt_cfg=llm_prompt_cfg,
                            llm_examples_by_eid=llm_examples_by_eid,
                        )
                        if write_prompt_metadata:
                            prompts_rows.append(prompt)
                        else:
                            metadata = (
                                prompt.get("metadata")
                                if isinstance(prompt.get("metadata"), dict)
                                else {}
                            )
                            prompts_rows.append(
                                {
                                    "match_key": str(prompt.get("match_key") or ""),
                                    "system": str(prompt.get("system") or ""),
                                    "user": str(prompt.get("user") or ""),
                                    "metadata": {
                                        "span_bundle_key": str(metadata.get("span_bundle_key") or ""),
                                    },
                                }
                            )
                        prompts_written += 1
                        decision_line = responses_map.get(bundle_match_key)
                        if decision_line is None or (isinstance(decision_line, str) and not decision_line.strip()):
                            # Backward compatibility: old response files may store record-level key
                            # without span suffix (sha#instance) instead of span_bundle_key.
                            legacy_keys: list[str] = []
                            record_key = str(match_key or "").strip()
                            if record_key:
                                legacy_keys.append(record_key)
                                if "||" in record_key:
                                    legacy_keys.append(record_key.split("||", 1)[0].strip())
                            bundle_base_key = str(bundle_match_key or "").split("||", 1)[0].strip()
                            if bundle_base_key:
                                legacy_keys.append(bundle_base_key)
                            seen_legacy: set[str] = set()
                            for lk in legacy_keys:
                                if not lk or lk in seen_legacy:
                                    continue
                                seen_legacy.add(lk)
                                cand = responses_map.get(lk)
                                if isinstance(cand, str) and cand.strip():
                                    decision_line = cand
                                    break
                        if decision_line is None or (isinstance(decision_line, str) and not decision_line.strip()):
                            bundle_result = {
                                "status": "no_response",
                                "decision_line": "",
                                "selected_eid": None,
                                "selected_eids": [],
                                "n_selected_eids": 0,
                                "protocol_ok": False,
                                "fallback_used": False,
                                "n_to_llm_candidates": len(bundle),
                            }
                            responses_missing += 1
                            any_no_response = True
                            all_protocol_ok = False
                        else:
                            bundle_result = _apply_decision_line_to_candidates(bundle, decision_line, allow_multiple)
                            if bundle_result.get("status") in {"applied", "fallback"}:
                                any_applied_or_fallback = True
                            if not bool(bundle_result.get("protocol_ok")):
                                all_protocol_ok = False
                            if bool(bundle_result.get("fallback_used")):
                                any_fallback = True
                            if bundle_result.get("decision_line"):
                                decision_lines_join.append(str(bundle_result.get("decision_line") or ""))
                            selected_eids_meta = bundle_result.get("selected_eids") or []
                            if selected_eids_meta:
                                for eid in selected_eids_meta:
                                    eid_str = str(eid or "").strip()
                                    if not eid_str or eid_str in selected_seen:
                                        continue
                                    selected_seen.add(eid_str)
                                    selected_union.append(eid_str)
                                applied_selected += 1
                                n_selected_eids_total += len(selected_eids_meta)
                                if len(selected_eids_meta) > 1:
                                    applied_selected_multi += 1
                            elif bundle_result.get("status") == "applied":
                                applied_none += 1
                            if bundle_result.get("status") == "fallback":
                                fallback_used += 1
                        span_results.append(
                            {
                                **bundle_result,
                                "span_key": span_key,
                                "span_bundle_key": bundle_match_key,
                            }
                        )
                    rerank_meta["span_reranks"] = span_results
                    rerank_meta["selected_eids"] = selected_union
                    rerank_meta["n_selected_eids"] = len(selected_union)
                    rerank_meta["selected_eid"] = ",".join(selected_union) or None
                    rerank_meta["decision_line"] = " || ".join([x for x in decision_lines_join if x])
                    rerank_meta["protocol_ok"] = all_protocol_ok if span_results else False
                    rerank_meta["fallback_used"] = any_fallback
                    if any_applied_or_fallback:
                        rerank_meta["status"] = "applied"
                    elif any_no_response:
                        rerank_meta["status"] = "no_response"
                    else:
                        rerank_meta["status"] = "no_candidates"

            output_record = {
                **record,
                "rerank": rerank_meta,
            }
            write_jsonl_line(out_fp, output_record)
            n_records += 1

    _write_jsonl_atomic(prompts_path, prompts_rows)

    return {
        "n_records": n_records,
        "n_input_candidates": n_input_candidates,
        "n_to_llm_candidates": n_to_llm_candidates,
        "n_records_with_to_llm": n_records_with_to_llm,
        "prompts_written": prompts_written,
        "responses_missing": responses_missing,
        "applied_selected": applied_selected,
        "applied_selected_multi": applied_selected_multi,
        "n_selected_eids_total": n_selected_eids_total,
        "applied_none": applied_none,
        "fallback_used": fallback_used,
        "missing_match_key_records": missing_match_key_records,
        "skipped_prompts": skipped_prompts,
        "skipped_apply": skipped_apply,
    }


def _run_call_llm(
    *,
    prompts_path: Path,
    responses_path: Path,
    base_url: str,
    api_key: str,
    model: str,
    temperature: float,
    max_tokens: int,
    batch_size: int,
    retry: int,
    backoff_seconds: float,
    timeout_seconds: float,
    allow_multiple: bool,
) -> dict[str, int]:
    prompts = list(_iter_jsonl(prompts_path))
    n_prompts_in = len(prompts)
    n_batches = 0
    responses: list[dict[str, Any]] = []
    responses_written = 0
    http_fail = 0
    http_401 = 0
    parse_fail = 0
    guard_reject = 0
    empty_text = 0
    raw_text_non_empty = 0
    decision_line_non_empty = 0
    error_non_empty = 0
    decision_none_rows = 0
    url = base_url.rstrip("/") + "/chat/completions"

    batch_size = max(batch_size, 1)
    for start in range(0, len(prompts), batch_size):
        batch = prompts[start : start + batch_size]
        n_batches += 1
        for prompt in batch:
            match_key = prompt.get("match_key")
            if not (isinstance(match_key, str) and match_key.strip()):
                match_key = ""
            else:
                match_key = match_key.strip()
            system = prompt.get("system") or ""
            user = prompt.get("user") or ""
            metadata = prompt.get("metadata") or {}
            candidate_eids = []
            candidate_e_ids_all = []
            candidate_span_keys_all = []
            candidate_encoder_ranks_all = []
            candidate_confidences = []
            candidate_ambiguous_flags = []
            span_bundle_key = ""
            span_bundle_keys = []
            n_span_bundles = 0
            n_candidates = 0
            if isinstance(metadata, dict):
                candidate_eids = metadata.get("candidate_e_ids") or []
                candidate_e_ids_all = metadata.get("candidate_e_ids_all") or []
                candidate_span_keys_all = metadata.get("candidate_span_keys_all") or []
                candidate_encoder_ranks_all = metadata.get("candidate_encoder_ranks_all") or []
                candidate_confidences = metadata.get("candidate_confidences") or []
                candidate_ambiguous_flags = metadata.get("candidate_ambiguous_flags") or []
                span_bundle_key = str(metadata.get("span_bundle_key") or "")
                span_bundle_keys = metadata.get("span_bundle_keys") or []
                n_span_bundles = metadata.get("n_span_bundles") or 0
                n_candidates = metadata.get("n_candidates") or 0
            if not isinstance(candidate_eids, list):
                candidate_eids = []
            if not isinstance(candidate_e_ids_all, list):
                candidate_e_ids_all = []
            if not isinstance(candidate_span_keys_all, list):
                candidate_span_keys_all = []
            if not isinstance(candidate_encoder_ranks_all, list):
                candidate_encoder_ranks_all = []
            if not isinstance(candidate_confidences, list):
                candidate_confidences = []
            if not isinstance(candidate_ambiguous_flags, list):
                candidate_ambiguous_flags = []
            if not isinstance(span_bundle_keys, list):
                span_bundle_keys = []
            if not isinstance(n_span_bundles, int):
                n_span_bundles = 0
            if not isinstance(n_candidates, int):
                n_candidates = 0

            attempts = 0
            error = ""
            raw_text = ""
            decision_line = ""
            start_time = time.time()
            for attempt in range(retry + 1):
                attempts = attempt + 1
                text, err, status_code = _call_openai_chat(
                    url=url,
                    api_key=api_key,
                    model=model,
                    system=system,
                    user=user,
                    temperature=temperature,
                    max_tokens=max_tokens,
                    timeout_seconds=timeout_seconds,
                )
                if err is None:
                    raw_text = text or ""
                    error = ""
                    break
                error = err
                if status_code == 401:
                    http_401 += 1
                    raise ConfigError("infer_step2_rerank call_llm auth failed (HTTP 401)")
                retryable = False
                if status_code is None:
                    retryable = True
                elif status_code == 429 or (500 <= status_code <= 599):
                    retryable = True
                if retryable and attempt < retry:
                    time.sleep(backoff_seconds * (2**attempt))
                    continue
                break
            elapsed_ms = int((time.time() - start_time) * 1000)
            if error:
                http_fail += 1
            else:
                if not raw_text.strip():
                    empty_text += 1
                else:
                    raw_text_non_empty += 1
                decision_line = _extract_decision_line(raw_text, allow_multiple)
                if decision_line.strip():
                    decision_line_non_empty += 1
                if decision_line.strip() in {"NONE", "DECISION: NONE"}:
                    decision_none_rows += 1
                if raw_text.strip() and (not decision_line) and (not error):
                    error = "parse_fail: no_decision_line"
                if not decision_line:
                    parse_fail += 1
                    if not error and "DECISION" in raw_text:
                        error = "parse_fail: no_decision_line"
                else:
                    parsed = _parse_decision_line(decision_line, allow_multiple, candidate_eids)
                    if not parsed.get("protocol_ok"):
                        parse_fail += 1
                        decision_line = ""
                        if not error:
                            error = "parse_fail: protocol_violation"
                    elif parsed.get("decision") == "EID":
                        eids = parsed.get("e_ids") or []
                        if not _validate_decision_eid(eids, candidate_eids):
                            guard_reject += 1
                            decision_line = ""
                            if not error:
                                error = "guard_reject: unknown_eid"
            raw_text_norm = raw_text.replace("\n", " ").replace("\r", " ")
            if error:
                error_non_empty += 1
            response_key = str(span_bundle_key or match_key or "").strip()
            response_row = {
                # Canonical response key must include span when available.
                "match_key": response_key,
                "record_match_key": str(match_key or "").strip(),
                "span_bundle_key": span_bundle_key,
                "span_bundle_keys": span_bundle_keys,
                "n_span_bundles": n_span_bundles,
                "decision_line": decision_line,
                "raw_text": raw_text_norm[:200],
                "error": error,
                "candidate_e_ids_all": candidate_e_ids_all,
                "candidate_span_keys_all": candidate_span_keys_all,
                "candidate_encoder_ranks_all": candidate_encoder_ranks_all,
                "candidate_confidences": candidate_confidences,
                "candidate_ambiguous_flags": candidate_ambiguous_flags,
                "n_candidates": n_candidates,
                "model": model,
                "attempts": attempts,
                "elapsed_ms": elapsed_ms,
            }
            responses.append(response_row)
            responses_written += 1

    responses_rows = len(responses)
    if http_401 > 0 or responses_rows == 0 or (responses_rows > 0 and raw_text_non_empty == 0):
        logger = logging.getLogger("kmwe")
        logger.info(
            "infer_step2_rerank fatal_call_llm_error=1 http_401=%s http_fail=%s responses_rows=%s raw_text_non_empty=%s",
            http_401,
            http_fail,
            responses_rows,
            raw_text_non_empty,
        )
        raise ConfigError("infer_step2_rerank call_llm fatal error (auth/empty responses)")

    _write_jsonl_atomic(responses_path, responses)
    return {
        "responses_written": responses_written,
        "http_fail": http_fail,
        "http_401": http_401,
        "parse_fail": parse_fail,
        "guard_reject": guard_reject,
        "empty_text": empty_text,
        "raw_text_non_empty": raw_text_non_empty,
        "decision_line_non_empty": decision_line_non_empty,
        "error_non_empty": error_non_empty,
        "decision_none_rows": decision_none_rows,
        "n_prompts_in": n_prompts_in,
        "n_batches": n_batches,
    }


def _build_llm_chat_messages(*, system: str, user: str) -> list[dict[str, str]]:
    return [
        {"role": "system", "content": str(system or "")},
        {"role": "user", "content": str(user or "")},
    ]


def _call_openai_chat(
    *,
    url: str,
    api_key: str,
    model: str,
    system: str,
    user: str,
    temperature: float,
    max_tokens: int,
    timeout_seconds: float,
) -> tuple[str | None, str | None, int | None]:
    payload = {
        "model": model,
        "messages": _build_llm_chat_messages(system=system, user=user),
        "temperature": temperature,
        "max_tokens": max_tokens,
    }
    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=data,
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=timeout_seconds) as resp:
            body = resp.read().decode("utf-8", errors="replace")
            status_code = resp.getcode()
    except urllib.error.HTTPError as exc:
        try:
            _ = exc.read()
        except Exception:
            pass
        return None, f"HTTP {exc.code}", exc.code
    except urllib.error.URLError as exc:
        return None, str(exc), None
    except Exception as exc:
        return None, str(exc), None
    try:
        obj = json.loads(body)
    except Exception:
        return None, "invalid_json", status_code
    try:
        text = obj["choices"][0]["message"]["content"]
    except Exception:
        return None, "missing_content", status_code
    return str(text), None, status_code


def _is_to_llm_candidate(candidate: dict[str, Any]) -> bool:
    if not isinstance(candidate, dict):
        return False
    if candidate.get("to_llm") is True:
        return True
    return False


def _span_bundle_keys(to_llm_candidates: list[dict[str, Any]]) -> list[str]:
    keys: list[str] = []
    seen: set[str] = set()
    for cand in to_llm_candidates:
        if not isinstance(cand, dict):
            continue
        span_key = str(cand.get("span_key") or "").strip()
        if not span_key or span_key in seen:
            continue
        seen.add(span_key)
        keys.append(span_key)
    return keys


def _compose_span_bundle_key(match_key: str, span_keys: list[str]) -> str:
    mk = str(match_key or "").strip()
    spans = [str(k).strip() for k in span_keys if str(k).strip()]
    if not mk:
        return ""
    if not spans:
        return mk
    return f"{mk}||{'__+__'.join(spans)}"


def _group_to_llm_candidates_by_span(
    to_llm_candidates: list[dict[str, Any]],
) -> list[tuple[str, list[dict[str, Any]]]]:
    grouped: list[tuple[str, list[dict[str, Any]]]] = []
    by_span: dict[str, list[dict[str, Any]]] = {}
    order: list[str] = []
    for idx, cand in enumerate(to_llm_candidates):
        if not isinstance(cand, dict):
            continue
        span_key = str(cand.get("span_key") or "").strip()
        if not span_key:
            span_key = f"__NO_SPAN__:{idx}"
        if span_key not in by_span:
            by_span[span_key] = []
            order.append(span_key)
        by_span[span_key].append(cand)
    for span_key in order:
        grouped.append((span_key, by_span[span_key]))
    return grouped


def _summarize_span_bundles(
    match_key: str,
    to_llm_candidates: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for span_key, bundle in _group_to_llm_candidates_by_span(to_llm_candidates):
        out.append(
            {
                "span_key": span_key,
                "span_bundle_key": _compose_span_bundle_key(match_key, [span_key]),
                "n_candidates": len(bundle),
                "candidate_e_ids": [str(c.get("e_id") or "") for c in bundle if isinstance(c, dict)],
            }
        )
    return out


def _normalize_example_role_order(role_order_raw: Any) -> list[str]:
    allowed = {"pos", "neg", "conf"}
    default_order = ["pos", "neg", "conf"]
    if not isinstance(role_order_raw, list):
        return default_order
    out: list[str] = []
    seen: set[str] = set()
    for item in role_order_raw:
        role = str(item or "").strip().lower()
        if not role or role in seen or role not in allowed:
            continue
        seen.add(role)
        out.append(role)
    if not out:
        return default_order
    for role in default_order:
        if role not in seen:
            out.append(role)
    return out


def _resolve_prompt_examples_config(llm_prompt_cfg: dict[str, Any] | None) -> dict[str, Any]:
    prompt_cfg = llm_prompt_cfg or {}
    include_examples = bool(prompt_cfg.get("include_examples", False))
    examples_cfg = prompt_cfg.get("examples", {}) or {}
    role_order = _normalize_example_role_order(examples_cfg.get("role_order"))
    max_per_role_cfg = examples_cfg.get("max_per_role", {}) or {}

    def _to_nonneg_int(val: Any, default: int) -> int:
        try:
            iv = int(val)
        except Exception:
            iv = int(default)
        return max(0, iv)

    max_per_role = {
        "pos": _to_nonneg_int(max_per_role_cfg.get("pos", 1), 1),
        "dispos": _to_nonneg_int(max_per_role_cfg.get("dispos", 2), 2),
        "neg": _to_nonneg_int(max_per_role_cfg.get("neg", 1), 1),
        "conf": _to_nonneg_int(max_per_role_cfg.get("conf", 2), 2),
    }
    max_examples_total = _to_nonneg_int(
        examples_cfg.get("max_examples_per_candidate", prompt_cfg.get("max_examples_per_candidate", 6)),
        6,
    )
    render_empty_slots = bool(examples_cfg.get("render_empty_slots", True))
    return {
        "include_examples": include_examples,
        "role_order": role_order,
        "max_per_role": max_per_role,
        "max_examples_total": max_examples_total,
        "render_empty_slots": render_empty_slots,
    }


def _parse_span_segments_value(span_segments: Any) -> list[tuple[int, int]]:
    if isinstance(span_segments, list):
        raw_items = span_segments
    else:
        text = str(span_segments or '').strip()
        if not text:
            return []
        try:
            raw_items = ast.literal_eval(text)
        except Exception:
            return []
    out: list[tuple[int, int]] = []
    for item in raw_items or []:
        if not isinstance(item, (list, tuple)) or len(item) != 2:
            continue
        try:
            start = int(item[0])
            end = int(item[1])
        except Exception:
            continue
        if end <= start:
            continue
        out.append((start, end))
    out.sort(key=lambda x: (x[0], x[1]))
    return out


def _build_marked_sentence(sentence: Any, span_segments: Any) -> str:
    text = str(sentence or '')
    spans = _parse_span_segments_value(span_segments)
    if not text or not spans:
        return text
    parts: list[str] = []
    cursor = 0
    for start, end in spans:
        if start < cursor or start < 0 or end > len(text):
            continue
        parts.append(text[cursor:start])
        parts.append('[SPAN]')
        parts.append(text[start:end])
        parts.append('[/SPAN]')
        cursor = end
    parts.append(text[cursor:])
    return ''.join(parts)


def _extract_target_span_text(sentence: Any, span_segments: Any, fallback: Any = '') -> str:
    text = str(sentence or '')
    spans = _parse_span_segments_value(span_segments)
    if text and spans:
        chunks = []
        for start, end in spans:
            if 0 <= start < end <= len(text):
                chunks.append(text[start:end])
        if chunks:
            return ' '.join(chunks)
    return str(fallback or '').strip()


def _build_candidate_example_lines(
    *,
    e_id: str,
    llm_examples_by_eid: dict[str, dict[str, list[dict[str, str]]]] | None,
    examples_cfg: dict[str, Any],
) -> tuple[list[str], dict[str, int]]:
    if not examples_cfg.get("include_examples"):
        return [], {}

    role_map = (llm_examples_by_eid or {}).get(str(e_id), {}) or {}
    role_order = examples_cfg.get("role_order") or ["pos", "neg", "conf"]
    max_per_role = examples_cfg.get("max_per_role") or {}
    max_examples_total = int(examples_cfg.get("max_examples_total") or 0)
    render_empty_slots = bool(examples_cfg.get("render_empty_slots", True))

    lines: list[str] = []
    used_total = 0
    counts: dict[str, int] = {}
    for role in role_order:
        pool = role_map.get(role, []) or []
        cap = int(max_per_role.get(role, 0) or 0)
        if cap <= 0:
            picked: list[dict[str, str]] = []
        else:
            remain = max_examples_total - used_total if max_examples_total > 0 else cap
            if max_examples_total > 0 and remain <= 0:
                picked = []
            else:
                take = cap if max_examples_total <= 0 else min(cap, remain)
                picked = [x for x in pool[:take] if isinstance(x, dict)]

        counts[role] = len(picked)
        if picked:
            lines.append(f"   {role}_examples:")
            for i, ex in enumerate(picked, start=1):
                ex_role = str(ex.get("example_role") or role).strip()
                raw_sentence = str(ex.get("raw_sentence") or "").strip()
                span_segments = str(ex.get("span_segments") or "").strip()
                span_text = _extract_target_span_text(
                    raw_sentence,
                    ex.get("span_segments"),
                    ex.get("span_text"),
                )
                marked_sentence = _build_marked_sentence(raw_sentence, ex.get("span_segments"))
                note = str(ex.get("note") or "").strip()
                line = (
                    f"    [{i}] example_role={ex_role} | target_sentence_marked={marked_sentence} | "
                    f"target_span_text={span_text} | span_segments={span_segments}"
                )
                if note:
                    line += f" | note={note}"
                lines.append(line)
            used_total += len(picked)
        elif render_empty_slots:
            lines.append(f"   {role}_examples: (none)")
    return lines, counts


def _build_llm_prompt(
    record: dict[str, Any],
    to_llm_candidates: list[dict[str, Any]],
    match_key: str,
    allow_multiple: bool,
    expredict_meta_by_eid: dict[str, dict[str, str]] | None = None,
    llm_prompt_cfg: dict[str, Any] | None = None,
    llm_examples_by_eid: dict[str, dict[str, list[dict[str, str]]]] | None = None,
) -> dict[str, Any]:
    target_sentence = record.get("target_sentence") or record.get("raw_sentence") or ""
    examples_cfg = _resolve_prompt_examples_config(llm_prompt_cfg)

    bundle_span_segments = None
    bundle_span_text = ""
    target_sentence_marked = str(target_sentence)
    if to_llm_candidates:
        bundle_span_segments = to_llm_candidates[0].get("span_segments")
        bundle_span_text = _extract_target_span_text(
            target_sentence,
            bundle_span_segments,
            to_llm_candidates[0].get("span_text"),
        )
        target_sentence_marked = _build_marked_sentence(target_sentence, bundle_span_segments)

    examples_counts_by_eid: dict[str, dict[str, int]] = {}
    numbered_lines: list[str] = []
    for idx, cand in enumerate(to_llm_candidates, start=1):
        e_id = str(cand.get("e_id") or "").strip()
        eid_meta = (expredict_meta_by_eid or {}).get(e_id, {}) if e_id else {}
        canonical_form = str(cand.get("canonical_form") or eid_meta.get("canonical_form") or "").strip()
        gloss = str(cand.get("gloss") or eid_meta.get("gloss") or "").strip()
        numbered_lines.append(
            f"{idx}) 대표형={canonical_form} | 뜻풀이={gloss}"
        )

        example_lines, example_counts = _build_candidate_example_lines(
            e_id=e_id,
            llm_examples_by_eid=llm_examples_by_eid,
            examples_cfg=examples_cfg,
        )
        if example_lines:
            numbered_lines.extend(example_lines)
        if example_counts:
            examples_counts_by_eid[e_id] = example_counts

    candidate_eids = [str(cand.get("e_id") or "") for cand in to_llm_candidates if cand.get("e_id")]
    span_bundle_keys = _span_bundle_keys(to_llm_candidates)
    span_bundle_key = (
        str(match_key or "")
        if "||" in str(match_key or "")
        else _compose_span_bundle_key(match_key, span_bundle_keys)
    )

    lines: list[str] = []
    lines.append(f"문장: {target_sentence_marked}")
    lines.append(f"표적 표현: {bundle_span_text}")
    lines.append("")
    lines.append("후보 의미:")
    lines.extend(numbered_lines)
    lines.append("")
    if allow_multiple:
        lines.append("위 문장의 표적 표현에 동시에 성립하는 후보가 있으면 함께 선택하라.")
        system = """당신은 한국어 표현문형 의미 선택기다.

표적 구간은 [SPAN] 와 [/SPAN] 사이에 표시된다.
주어진 문맥과 후보 의미를 비교하여, 표적 구간에 동시에 성립하는 후보가 여러 개이면 함께 선택하라.
어떤 후보도 문맥에 맞지 않으면 NONE을 선택하라.

출력:
- 한 줄만 출력한다.
- 문맥에 맞는 후보가 있으면 해당 번호들을 쉼표로만 연결해 쓴다.
- 어떤 후보도 맞지 않으면 NONE만 쓴다.
- 이유나 설명은 쓰지 않는다.
- 후보 목록에 없는 번호는 쓰지 않는다."""
    else:
        lines.append("위 문장의 표적 표현에 가장 알맞은 후보를 선택하라.")
        system = """당신은 한국어 표현문형 의미 선택기다.

표적 구간은 [SPAN] 와 [/SPAN] 사이에 표시된다.
주어진 문맥과 후보 의미를 비교하여, 표적 구간에 가장 알맞은 후보 하나만 선택하라.
어떤 후보도 문맥에 맞지 않으면 NONE을 선택하라.

출력:
- 문맥에 가장 알맞은 후보가 있으면 해당 번호 하나만 쓴다.
- 어떤 후보도 맞지 않으면 NONE만 쓴다.
- 출력에는 반드시 번호 하나 혹은 NONE만 기록되어야 한다.
- 이유나 설명은 쓰지 않는다.
- 후보 목록에 없는 번호는 쓰지 않는다."""
    user = "\n".join(lines)

    metadata = {
        "match_key": match_key,
        "span_bundle_key": span_bundle_key,
        "span_bundle_keys": span_bundle_keys,
        "n_span_bundles": len(span_bundle_keys),
        "n_candidates": len(to_llm_candidates),
        "candidate_e_ids": candidate_eids,
        "candidate_number_to_eid": {str(i + 1): eid for i, eid in enumerate(candidate_eids)},
        "candidate_e_ids_all": [str(cand.get("e_id") or "") for cand in to_llm_candidates],
        "candidate_span_keys_all": [str(cand.get("span_key") or "") for cand in to_llm_candidates],
        "candidate_encoder_ranks_all": [
            cand.get("encoder_rank") for cand in to_llm_candidates
        ],
        "candidate_confidences": [cand.get("confidence") for cand in to_llm_candidates],
        "candidate_ambiguous_flags": [
            cand.get("ambiguous") for cand in to_llm_candidates
        ],
        "include_examples": bool(examples_cfg.get("include_examples")),
        "examples_counts_by_eid": examples_counts_by_eid,
    }
    return {
        "match_key": match_key,
        "system": system,
        "user": user,
        "metadata": metadata,
    }


def _apply_decision_line_to_candidates(
    to_llm: list[dict[str, Any]], decision_line: str, allow_multiple: bool
) -> dict[str, Any]:
    parsed = _parse_decision_line(decision_line, allow_multiple, [str(c.get("e_id") or "") for c in to_llm])
    if not parsed["protocol_ok"]:
        selected_eid, fallback_used = _apply_fallback(to_llm)
        selected_eids = [selected_eid] if selected_eid else []
        return {
            "status": "fallback" if fallback_used else "no_response",
            "decision_line": decision_line,
            "selected_eid": selected_eid,
            "selected_eids": selected_eids,
            "n_selected_eids": len(selected_eids),
            "protocol_ok": False,
            "fallback_used": fallback_used,
            "n_to_llm_candidates": len(to_llm),
        }
    if parsed["decision"] == "NONE":
        return {
            "status": "applied",
            "decision_line": decision_line,
            "selected_eid": None,
            "selected_eids": [],
            "n_selected_eids": 0,
            "protocol_ok": True,
            "fallback_used": False,
            "n_to_llm_candidates": len(to_llm),
        }
    raw_selected_eids = parsed.get("e_ids") or []
    seen: set[str] = set()
    selected_eids: list[str] = []
    for eid in raw_selected_eids:
        eid_str = str(eid or "").strip()
        if not eid_str or eid_str in seen:
            continue
        seen.add(eid_str)
        selected_eids.append(eid_str)
    matched = False
    selected_eid_str = ",".join(selected_eids)
    for cand in to_llm:
        if str(cand.get("e_id") or "") in selected_eids:
            cand["triage"] = "confirm"
            matched = True
        else:
            cand["triage"] = "discard"
    if matched:
        return {
            "status": "applied",
            "decision_line": decision_line,
            "selected_eid": selected_eid_str or None,
            "selected_eids": selected_eids,
            "n_selected_eids": len(selected_eids),
            "protocol_ok": True,
            "fallback_used": False,
            "n_to_llm_candidates": len(to_llm),
        }
    selected_eid, fallback_used = _apply_fallback(to_llm)
    fallback_selected_eids = [selected_eid] if selected_eid else []
    return {
        "status": "fallback" if fallback_used else "no_response",
        "decision_line": decision_line,
        "selected_eid": selected_eid,
        "selected_eids": fallback_selected_eids,
        "n_selected_eids": len(fallback_selected_eids),
        "protocol_ok": True,
        "fallback_used": fallback_used,
        "n_to_llm_candidates": len(to_llm),
    }


def _apply_decision_line(
    record: dict[str, Any], decision_line: str, allow_multiple: bool
) -> dict[str, Any]:
    candidates = record.get("candidates") or []
    if not isinstance(candidates, list):
        candidates = []
    to_llm = [c for c in candidates if _is_to_llm_candidate(c)]
    return _apply_decision_line_to_candidates(to_llm, decision_line, allow_multiple)


def _parse_decision_line(decision_line: str, allow_multiple: bool, candidate_eids: list[str] | None = None) -> dict[str, Any]:
    if not isinstance(decision_line, str):
        return {"protocol_ok": False, "decision": None, "e_ids": []}
    lines = [line for line in decision_line.splitlines() if line.strip() != ""]
    if len(lines) != 1:
        return {"protocol_ok": False, "decision": None, "e_ids": []}
    line = lines[0].strip()
    candidate_eids = [str(x or "").strip() for x in (candidate_eids or []) if str(x or "").strip()]

    def _map_number_tokens(tokens: list[str]) -> dict[str, Any]:
        if not tokens:
            return {"protocol_ok": False, "decision": None, "e_ids": []}
        mapped: list[str] = []
        seen: set[str] = set()
        for tok in tokens:
            try:
                idx = int(tok)
            except Exception:
                return {"protocol_ok": False, "decision": None, "e_ids": []}
            if idx <= 0 or idx > len(candidate_eids):
                return {"protocol_ok": False, "decision": None, "e_ids": []}
            eid = candidate_eids[idx - 1]
            if eid in seen:
                continue
            seen.add(eid)
            mapped.append(eid)
        if not mapped:
            return {"protocol_ok": False, "decision": None, "e_ids": []}
        if (not allow_multiple) and len(mapped) > 1:
            return {"protocol_ok": False, "decision": None, "e_ids": []}
        return {"protocol_ok": True, "decision": "EID", "e_ids": mapped}

    if line == "NONE" or line == "DECISION: NONE":
        return {"protocol_ok": True, "decision": "NONE", "e_ids": []}
    if allow_multiple and re.match(r"^\d+(?:\s*,\s*\d+)+\s*$", line):
        return _map_number_tokens([x.strip() for x in line.split(',') if x.strip()])
    if re.match(r"^\d+\s*$", line):
        return _map_number_tokens([line.strip()])

    regex_short = r"^DECISION:\s*(NONE|[a-z]{3}\d{3,})\s*$"
    regex_short_loose = r"^DECISION[:\s]+([a-z]{3}\d{3,}|NONE)\s*$"
    match_short = re.match(regex_short, line)
    if match_short:
        val = match_short.group(1)
        if val == "NONE":
            return {"protocol_ok": True, "decision": "NONE", "e_ids": []}
        return {"protocol_ok": True, "decision": "EID", "e_ids": [val]}
    match_loose = re.match(regex_short_loose, line)
    if match_loose:
        val = match_loose.group(1)
        if val == "NONE":
            return {"protocol_ok": True, "decision": "NONE", "e_ids": []}
        return {"protocol_ok": True, "decision": "EID", "e_ids": [val]}
    if allow_multiple:
        regex_multi = r"^DECISION[:\s]+e_id=(?P<eids>\w+(?:\s*,\s*\w+)*)\s*$"
        regex_multi_short = r"^DECISION:\s*(?P<eids>[a-z]{3}\d{3,}(?:\s*,\s*[a-z]{3}\d{3,})+)\s*$"
        match_multi_short = re.match(regex_multi_short, line)
        if match_multi_short:
            eids = [e.strip() for e in match_multi_short.group("eids").split(',') if e.strip()]
            if not eids:
                return {"protocol_ok": False, "decision": None, "e_ids": []}
            return {"protocol_ok": True, "decision": "EID", "e_ids": eids}
        match_multi = re.match(regex_multi, line)
        if match_multi:
            eids = [e.strip() for e in match_multi.group("eids").split(',') if e.strip()]
            if not eids:
                return {"protocol_ok": False, "decision": None, "e_ids": []}
            return {"protocol_ok": True, "decision": "EID", "e_ids": eids}
    regex_single = r"^DECISION[:\s]+(e_id=(?P<eid>\w+)|NONE)\s*$"
    match_single = re.match(regex_single, line)
    if not match_single:
        return {"protocol_ok": False, "decision": None, "e_ids": []}
    if match_single.group("eid"):
        return {"protocol_ok": True, "decision": "EID", "e_ids": [match_single.group("eid")]}
    return {"protocol_ok": True, "decision": "NONE", "e_ids": []}


def _apply_fallback(
    to_llm: list[dict[str, Any]]
) -> tuple[str | None, bool]:
    for cand in to_llm:
        if str(cand.get("triage") or "") == "confirm":
            return str(cand.get("e_id") or ""), True
    best = None
    best_score = None
    for cand in to_llm:
        score = float(cand.get("encoder_score") or 0.0)
        if best is None or score > float(best_score or 0.0):
            best = cand
            best_score = score
    if best is not None:
        best["triage"] = "confirm"
        return str(best.get("e_id") or ""), True
    return None, False


def _load_decision_lines(
    path: Path,
    *,
    allow_multiple: bool,
) -> tuple[dict[str, str], int, int, int]:
    if not path.exists():
        return {}, 0, 0, 0
    logger = logging.getLogger("kmwe")
    mapping: dict[str, str] = {}
    loaded = 0
    responses_empty_decision = 0
    responses_total_rows = 0
    responses_recovered_from_raw = 0
    with path.open("r", encoding="utf-8") as fp:
        for line in fp:
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
            except Exception:
                continue
            if not isinstance(obj, dict):
                continue
            responses_total_rows += 1
            span_bundle_key = obj.get("span_bundle_key") if isinstance(obj, dict) else None
            match_key = obj.get("match_key") if isinstance(obj, dict) else None
            response_key = (
                str(span_bundle_key or "").strip()
                or str(match_key or "").strip()
            )
            # Backward compatibility for old response files that stored only
            # record-level match_key (sha#instance) without span suffix.
            if response_key and "||" not in response_key:
                span_bundle_keys_obj = obj.get("span_bundle_keys") if isinstance(obj, dict) else None
                recovered_span_bundle_key = ""
                if isinstance(span_bundle_keys_obj, list):
                    for item in span_bundle_keys_obj:
                        item_s = str(item or "").strip()
                        if item_s:
                            recovered_span_bundle_key = item_s
                            break
                if not recovered_span_bundle_key:
                    span_keys_obj = obj.get("candidate_span_keys_all") if isinstance(obj, dict) else None
                    if isinstance(span_keys_obj, list):
                        uniq: list[str] = []
                        seen: set[str] = set()
                        for item in span_keys_obj:
                            item_s = str(item or "").strip()
                            if not item_s or item_s in seen:
                                continue
                            seen.add(item_s)
                            uniq.append(item_s)
                        if len(uniq) == 1:
                            recovered_span_bundle_key = _compose_span_bundle_key(response_key, [uniq[0]])
                if recovered_span_bundle_key:
                    response_key = recovered_span_bundle_key
            if not response_key:
                continue
            decision_line = (
                obj.get("decision_line")
                or obj.get("decision")
                or obj.get("output")
                or obj.get("content")
                or ""
            )
            if not isinstance(decision_line, str):
                continue
            decision_line = decision_line.strip()
            if not decision_line:
                raw_text = obj.get("raw_text") or ""
                recovered = _extract_decision_line(str(raw_text), allow_multiple=allow_multiple)
                if recovered:
                    mapping[response_key] = recovered
                    loaded += 1
                    responses_recovered_from_raw += 1
                    continue
                responses_empty_decision += 1
                continue
            mapping[response_key] = decision_line
            loaded += 1
    logger.info(
        "responses_loaded=%s responses_empty_decision=%s responses_total_rows=%s responses_recovered_from_raw=%s",
        loaded,
        responses_empty_decision,
        responses_total_rows,
        responses_recovered_from_raw,
    )
    return mapping, loaded, responses_empty_decision, responses_total_rows


def _summarize_llm_responses(path: Path) -> dict[str, Any]:
    summary = {
        "responses_total_rows": 0,
        "decision_line_non_empty": 0,
        "error_non_empty": 0,
        "decision_none_rows": 0,
        "n_candidates_sum": 0,
        "n_candidates_count": 0,
    }
    if not path.exists():
        return summary
    for obj in _iter_jsonl(path):
        if not isinstance(obj, dict):
            continue
        summary["responses_total_rows"] += 1
        decision_line = str(obj.get("decision_line") or "")
        if decision_line.strip():
            summary["decision_line_non_empty"] += 1
            if decision_line.strip().startswith("DECISION: NONE"):
                summary["decision_none_rows"] += 1
        error = str(obj.get("error") or "")
        if error.strip():
            summary["error_non_empty"] += 1
        n_candidates = obj.get("n_candidates")
        if isinstance(n_candidates, int):
            summary["n_candidates_sum"] += n_candidates
            summary["n_candidates_count"] += 1
    if summary["n_candidates_count"] > 0:
        summary["n_candidates_avg"] = summary["n_candidates_sum"] / summary["n_candidates_count"]
    else:
        summary["n_candidates_avg"] = 0
    return summary


def _resolve_input_pred_path(
    *,
    cfg: dict[str, Any],
    run_context: RunContext,
    logger: logging.Logger,
    rerank_cfg: dict[str, Any],
) -> tuple[Path, str, bool]:
    explicit = rerank_cfg.get("input_pred_path") or rerank_cfg.get("pred_path") or rerank_cfg.get(
        "input_jsonl"
    )
    if explicit:
        input_path = Path(str(explicit))
        logger.info(
            "[paths] stage=infer_step2_rerank input_pred_path=%s explicit=true (NO RESELECT)",
            input_path,
        )
        return input_path, "infer_step2_rerank.input_pred_path", False

    allow_auto = bool(
        rerank_cfg.get("allow_auto_select")
        or rerank_cfg.get("auto_select")
        or rerank_cfg.get("allow_reselect")
    )
    if not allow_auto:
        raise ConfigError(
            "infer_step2_rerank.input_pred_path가 필요합니다 (reselect 금지 기본). "
            "자동 선택이 필요하면 infer_step2_rerank.allow_auto_select=true로 명시하세요."
        )

    artifacts_root = _artifacts_root_from_outputs_dir(run_context.outputs_dir, logger)
    candidate = _latest_infer_step1_output(artifacts_root, run_context.exp_id)
    if candidate is None:
        raise ConfigError("infer_step2_rerank input_pred_path 자동 선택 실패")
    logger.warning(
        "infer_step2_rerank 입력 경로 자동 선택됨 (reselect 주의): %s",
        candidate,
    )
    return candidate, "auto_latest_infer_step1", True


def _latest_infer_step1_output(artifacts_root: Path, exp_id: str) -> Path | None:
    artifacts_root = artifacts_root / exp_id / "infer_step1"
    if not artifacts_root.exists():
        return None
    run_dirs = sorted([p for p in artifacts_root.iterdir() if p.is_dir()])
    if not run_dirs:
        return None
    candidate = run_dirs[-1] / "outputs" / "infer_candidates.jsonl"
    if candidate.exists():
        return candidate
    return None


def _artifacts_root_from_outputs_dir(outputs_dir: Path, logger: logging.Logger) -> Path:
    outputs_dir = Path(outputs_dir)
    if len(outputs_dir.parents) < 4:
        raise ValueError(f"outputs_dir 경로 깊이가 부족합니다: {outputs_dir}")
    artifacts_root = outputs_dir.parents[3]
    logger.info("infer_step2_rerank artifacts_root(from outputs_dir): %s", artifacts_root)
    return artifacts_root
