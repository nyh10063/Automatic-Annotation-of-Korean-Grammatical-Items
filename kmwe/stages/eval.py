from __future__ import annotations

import ast
import csv
import json
import re
import logging
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import pandas as pd
import yaml

from kmwe.core.config_loader import ConfigError
from kmwe.core.run_context import RunContext
from kmwe.stages import build_silver as silver_loader
from kmwe.utils.jsonio import write_json, write_jsonl_line
from kmwe.core.utils import iso_now
from kmwe.core.fs_guard import assert_under_dir
from kmwe.utils.morph import analyze_with_kiwi

FOR_USERS_FREEZE_SPEC_V1 = {
    "schema_version": "for_users_freeze_v1",
    "required_files": [
        {"key": "eval_latest_rows_jsonl", "relpath": "for_users/eval_latest_rows.jsonl", "min_bytes": 1},
        {"key": "run_delta_summary_csv", "relpath": "for_users/run_delta_summary.csv", "min_bytes": 1},
        {"key": "eval_run_delta_csv", "relpath": "for_users/eval_run_delta.csv", "min_bytes": 1},
        {"key": "eval_run_delta_reasons_csv", "relpath": "for_users/eval_run_delta_reasons.csv", "min_bytes": 1},
        {"key": "eval_run_delta_samples_jsonl", "relpath": "for_users/eval_run_delta_samples.jsonl", "min_bytes": 1},
        {"key": "index_json", "relpath": "for_users/INDEX.json", "min_bytes": 1},
    ],
    "csv_headers": {
        "run_delta_summary_csv": ["metric", "curr", "prev", "delta", "note"],
        "eval_run_delta_reasons_csv": [
            "delta_kind",
            "delta_reason_bucket",
            "n_rows",
            "n_unique_match_key",
            "n_unique_gold_eid",
            "sample_match_keys",
        ],
        "eval_run_delta_csv_prefix": [
            "delta_kind",
            "delta_reason_bucket",
            "match_key",
            "gold_eid",
            "gold_span_key",
        ],
    },
    "json_required_top_keys": {
        "index_json": [
            "schema_version",
            "generated_at",
            "outputs_dir",
            "for_users_dir",
            "artifacts",
            "run_summary",
        ],
    },
}


def run_eval(*, cfg: dict[str, Any], run_context: RunContext) -> None:
    logger = logging.getLogger("kmwe")
    logger.info("[eval][run_eval_enter] called=true")
    logger.info("[eval] using module file: %s", __file__)
    outputs_dir = run_context.outputs_dir
    outputs_dir.mkdir(parents=True, exist_ok=True)

    artifacts_root = _artifacts_root_from_outputs_dir(outputs_dir, logger)

    gold_xlsx = Path(cfg.get("paths", {}).get("gold_xlsx") or "")
    if not gold_xlsx.exists():
        raise ConfigError(f"eval gold.xlsx 경로가 유효하지 않습니다: {gold_xlsx}")

    eval_cfg = cfg.get("eval", {}) or {}
    include_hold_cfg = bool(eval_cfg.get("include_hold", False))
    match_key_policy = str(eval_cfg.get("match_key_policy", "example_id_instance_or_sentence"))
    raw_allowed_splits = eval_cfg.get("allowed_splits", None)
    target_only = bool(eval_cfg.get("target_only", True))
    only_keys_in_gold = bool(eval_cfg.get("only_keys_in_gold", True))
    error_samples_max = int(eval_cfg.get("error_samples_max", 100))
    write_errors = bool(eval_cfg.get("write_errors", True))
    raw_report_views = eval_cfg.get("report_views", None)

    def _norm_views(x):
        if x is None:
            return []
        if isinstance(x, str):
            parts = [p.strip().lower() for p in x.replace(";", ",").split(",")]
            return [p for p in parts if p]
        if isinstance(x, (list, tuple)):
            out = []
            for v in x:
                if v is None:
                    continue
                out.append(str(v).strip().lower())
            return out
        try:
            return [str(v).strip().lower() for v in list(x)]
        except Exception:
            return []

    report_views = _norm_views(raw_report_views)
    report_views = [v for v in report_views if v in {"strict", "lenient"}]
    if not report_views:
        report_views = ["strict", "lenient"]
    if "strict" not in report_views:
        report_views = ["strict"] + report_views
    if "lenient" not in report_views:
        report_views = report_views + ["lenient"]
    seen = set()
    report_views = [v for v in report_views if not (v in seen or seen.add(v))]

    primary_view = str(eval_cfg.get("primary_view", "strict"))
    if primary_view not in report_views:
        primary_view = "strict"

    def _norm_splits(x):
        if x is None:
            return None
        if isinstance(x, str):
            parts = [p.strip().lower() for p in x.replace(";", ",").split(",")]
            out = [p for p in parts if p]
            return out or None
        if isinstance(x, (list, tuple, set)):
            out = []
            for v in x:
                if v is None:
                    continue
                s = str(v).strip().lower()
                if s:
                    out.append(s)
            return out or None
        s = str(x).strip().lower()
        return [s] if s else None

    allowed_splits = _norm_splits(raw_allowed_splits)

    gold_sheet_name = str(cfg.get("eval", {}).get("gold_sheet_name", "gold"))
    logger.info("[eval] allowed_splits=%s", allowed_splits or "ALL")
    gold_rows = _load_gold_rows(
        gold_xlsx,
        gold_sheet_name=gold_sheet_name,
        allowed_splits=allowed_splits,
    )
    neg_rows = _load_neg_rows(
        gold_xlsx,
        gold_sheet_name=gold_sheet_name,
        allowed_splits=allowed_splits,
    )
    neg_confusable_rows, _neg_confusable_occ = _load_neg_confusable_rows_openpyxl(
        gold_xlsx, gold_sheet_name=gold_sheet_name
    )
    if neg_confusable_rows:
        gold_rows.extend(neg_confusable_rows)
    nta_rows = _load_neg_target_absent_rows(
        gold_xlsx,
        gold_sheet_name=gold_sheet_name,
        allowed_splits=allowed_splits,
    )
    gold_schema = _detect_gold_schema_from_rows(gold_rows)
    gold_unique_keys = {
        (row.get("example_id"), row.get("instance_id")) for row in gold_rows
    }
    logger.info("[gold_occ] rows=%s", len(gold_rows))
    logger.info("[gold_occ] unique_keys=%s", len(gold_unique_keys))
    logger.info("[eval] gold_schema=%s", gold_schema)
    min_gold_rows = int(eval_cfg.get("min_gold_rows") or 0)
    if min_gold_rows > 0 and len(gold_rows) < min_gold_rows:
        raise ConfigError(
            f"gold rows too small: {len(gold_rows)} < min_gold_rows={min_gold_rows} "
            f"(check paths.gold_xlsx={gold_xlsx}; expected 운영 gold.xlsx(444행))"
        )
    if include_hold_cfg and raw_report_views is not None:
        logger.warning("eval.report_views가 설정되어 있어 eval.include_hold는 view별 설정으로 대체됩니다.")
    logger.info("[eval] report_views=%s primary_view=%s", report_views, primary_view)

    explicit_pred_path = bool(
        eval_cfg.get("pred_path")
        or eval_cfg.get("input_pred_path")
        or eval_cfg.get("pred_jsonl")
        or eval_cfg.get("input_jsonl")
    )
    logger.info("[eval][pred_guard_callsite] called=true")
    pred_path, pred_source = _resolve_pred_path(cfg, run_context, logger)
    if pred_source == "explicit":
        logger.info("[eval][pred_source] mode=explicit path=%s", pred_path)
    else:
        logger.info("[eval][pred_source] mode=autoselect path=%s", pred_path)
    logger.info("[eval][pred_guard_scan_enter] called=true mode=%s", pred_source)
    _latest_stage_output(artifacts_root, run_context.exp_id, "infer_step1")
    if not pred_path.exists():
        raise ConfigError(f"eval pred JSONL이 존재하지 않습니다: {pred_path}")
    pred_schema = _infer_pred_schema_from_path(pred_path, eval_cfg)
    pred_id_schema = detect_example_id_schema(pred_path)
    if (
        match_key_policy == "example_id_instance_or_sentence"
        and gold_schema in {"gold", "sha1"}
        and pred_id_schema in {"gold", "sha1"}
        and gold_schema != pred_id_schema
    ):
        logger.warning(
            "[eval][match_key_policy] auto_switch=true from=%s to=sentence_only gold_schema=%s pred_id_schema=%s",
            match_key_policy,
            gold_schema,
            pred_id_schema,
        )
        match_key_policy = "sentence_only"
    match_key_policy, gold_by_key, neg_by_key = _choose_best_match_key_policy(
        gold_rows=gold_rows,
        neg_rows=neg_rows,
        pred_path=pred_path,
        current_policy=match_key_policy,
        logger=logger,
    )
    cfg.setdefault("eval", {})["match_key_policy"] = match_key_policy
    if match_key_policy in ("example_id_instance_or_sentence", "example_id_instance_only"):
        collisions = {k: v for k, v in gold_by_key.items() if "#" not in str(k) and len(v) > 1}
        if collisions:
            raise ValueError(
                "Collision: duplicate example_id without instance_id. "
                "Under B plan, duplicates must be disambiguated by instance_id. "
                f"examples={list(collisions.keys())[:5]}"
            )
    if explicit_pred_path:
        logger.info("[eval][pred_guard_enter] called=true")
        logger.info("[eval] using explicit pred_path (no reselect): %s", pred_path)
        allow_stale_pred = bool(eval_cfg.get("allow_stale_pred", False))
        latest_pred_path = _latest_stage_output(artifacts_root, run_context.exp_id, "infer_step1")
        latest_res = None
        is_latest = False
        rerank_ok = False
        rerank_run_dir = None
        rerank_latest_run = None
        rerank_input_path = None
        report_output_path = None
        rerank_ok_source = "none"
        rerank_index_path = None
        rerank_index_loaded = False
        rerank_index_artifacts = None
        report_path = None
        report_loaded = False
        report_keys = None
        report_vals = None
        rerank_out_res_for_log = None
        logger.info(
            "[eval][pred_guard_pred] pred_path=%s pred_res=%s exists=%s bytes=%s",
            pred_path,
            pred_path.expanduser().resolve() if pred_path.exists() else "",
            pred_path.exists(),
            pred_path.stat().st_size if pred_path.exists() else 0,
        )
        if latest_pred_path is not None:
            try:
                is_latest = latest_pred_path.resolve() == pred_path.resolve()
            except Exception:
                is_latest = str(latest_pred_path) == str(pred_path)
            try:
                latest_res = latest_pred_path.expanduser().resolve()
            except Exception:
                latest_res = latest_pred_path
        logger.info(
            "[eval][pred_guard_latest] latest_infer_step1_pred_path=%s latest_res=%s",
            latest_pred_path,
            latest_res or "",
        )
        try:
            exp_root = artifacts_root / run_context.exp_id / "infer_step2_rerank"
            if exp_root.exists():
                run_dirs = [p for p in exp_root.iterdir() if p.is_dir()]
                run_id_pattern = re.compile(r"^\\d{8}_\\d{6}_[A-Za-z0-9]{6}$")
                run_dirs_filtered = [
                    p for p in run_dirs if run_id_pattern.match(p.name or "")
                ]
                if run_dirs_filtered:
                    rerank_latest_run = max(run_dirs_filtered, key=lambda p: p.stat().st_mtime)
                if pred_path.parent.name == "outputs":
                    rerank_run_dir = pred_path.parent.parent
                out_dir = pred_path.parent
                report_candidates = [
                    out_dir / "infer_step2_rerank_report.json",
                    out_dir.parent / "outputs" / "infer_step2_rerank_report.json",
                ]
                index_path = out_dir / "INDEX.json"
                rerank_index = _load_rerank_index(index_path, logger)
                rerank_index_path = index_path
                rerank_index_loaded = isinstance(rerank_index, dict)
                if rerank_index_loaded:
                    rerank_index_artifacts = rerank_index.get("artifacts")
                for cand in report_candidates:
                    if cand.exists() and cand.stat().st_size > 0:
                        report_path = cand
                        break
                logger.info(
                    "[eval][pred_guard_report_candidates] tried=%s",
                    report_candidates,
                )
                logger.info(
                    "[eval][pred_guard_report_path] path=%s exists=%s bytes=%s",
                    report_path,
                    report_path.exists() if report_path is not None else False,
                    report_path.stat().st_size if report_path is not None and report_path.exists() else 0,
                )
                report = {}
                if report_path is not None and report_path.exists():
                    try:
                        report = json.loads(report_path.read_text(encoding="utf-8"))
                        report_loaded = isinstance(report, dict)
                        report_keys = list(report.keys()) if isinstance(report, dict) else []
                        logger.info(
                            "[eval][pred_guard_report_keys] keys=%s",
                            list(report.keys()),
                        )
                    except Exception as exc:
                        logger.info("[eval][pred_guard_report_parse_fail] err=%s", exc)
                        report = {}
                input_path = report.get("input_pred_path") or report.get("input_path")
                report_output_path = report.get("rerank_output_path") or report.get("output_path")
                report_in_res = None
                report_out_res = None
                if input_path:
                    try:
                        report_in_res = Path(str(input_path)).expanduser().resolve()
                    except Exception:
                        report_in_res = Path(str(input_path))
                    rerank_input_path = Path(str(input_path))
                if report_output_path:
                    try:
                        report_out_res = Path(str(report_output_path)).expanduser().resolve()
                    except Exception:
                        report_out_res = Path(str(report_output_path))
                if report:
                    report_vals = {
                        "input_pred_path": input_path,
                        "rerank_output_path": report.get("rerank_output_path"),
                        "output_path": report.get("output_path"),
                        "responses_path": report.get("responses_path"),
                        "status": report.get("status"),
                    }
                    logger.info(
                        "[eval][pred_guard_report_vals] input_pred_path=%s input_res=%s rerank_output_path=%s output_res=%s",
                        input_path or "",
                        report_in_res or "",
                        report_output_path or "",
                        report_out_res or "",
                    )
                pred_res = None
                try:
                    pred_res = pred_path.expanduser().resolve()
                except Exception:
                    pred_res = pred_path
                index_used = False
                if isinstance(rerank_index, dict):
                    artifacts = rerank_index.get("artifacts")
                    if isinstance(artifacts, dict):
                        index_output_obj = artifacts.get("rerank_output_pred") or {}
                        index_resp_obj = artifacts.get("responses_jsonl") or {}
                        index_output_path_str = (
                            str(index_output_obj.get("path") or "")
                            if isinstance(index_output_obj, dict)
                            else ""
                        )
                        index_resp_path_str = (
                            str(index_resp_obj.get("path") or "")
                            if isinstance(index_resp_obj, dict)
                            else ""
                        )
                        logger.info(
                            "[eval][pred_guard_index_pred] path=%s",
                            pred_res or "",
                        )
                        logger.info(
                            "[eval][pred_guard_index_output] path=%s",
                            index_output_path_str,
                        )
                        try:
                            index_out_path = Path(index_output_path_str).expanduser().resolve()
                        except Exception:
                            index_out_path = Path(index_output_path_str) if index_output_path_str else None
                        index_out_exists = (
                            bool(index_output_path_str)
                            and index_out_path is not None
                            and Path(index_output_path_str).expanduser().exists()
                            and Path(index_output_path_str).expanduser().stat().st_size > 0
                        )
                        index_resp_exists = (
                            bool(index_resp_path_str)
                            and Path(index_resp_path_str).expanduser().exists()
                            and Path(index_resp_path_str).expanduser().stat().st_size > 0
                        )
                        if (
                            index_out_exists
                            and index_resp_exists
                            and index_out_path is not None
                            and pred_res is not None
                            and index_out_path == pred_res
                        ):
                            rerank_ok = True
                            index_used = True
                            rerank_ok_source = "index"
                            rerank_out_res_for_log = index_out_path
                            logger.info("[eval][rerank_ok] value=True source=index")
                        else:
                            logger.info(
                                "[eval][rerank_ok] value=False source=index reason=missing_artifact"
                            )
                if not index_used:
                    if (
                        report_in_res is not None
                        and report_out_res is not None
                        and latest_res is not None
                        and report_in_res == latest_res
                        and report_out_res == pred_res
                    ):
                        rerank_ok = True
                        rerank_ok_source = "report"
                        rerank_out_res_for_log = report_out_res
                        logger.info("[eval][rerank_ok] value=True source=report")
                    else:
                        rerank_ok_source = "report"
                        logger.info(
                            "[eval][rerank_ok] value=False source=report reason=legacy_check_failed"
                        )
        except Exception:
            rerank_ok = False
            rerank_ok_source = "exception"
        logger.info(
            "[eval][pred_guard_rerank_ok] rerank_ok=%s pred_res=%s report_out_res=%s latest_res=%s report_in_res=%s",
            rerank_ok,
            pred_path.expanduser().resolve() if pred_path.exists() else "",
            report_out_res if "report_out_res" in locals() else "",
            latest_res or "",
            report_in_res if "report_in_res" in locals() else "",
        )
        logger.info(
            "[eval][pred_guard] enabled=%s allow_stale_pred=%s latest_pred_path=%s given_pred_path=%s is_latest=%s rerank_ok=%s rerank_run_dir=%s rerank_latest_run=%s rerank_input_path=%s",
            True,
            allow_stale_pred,
            latest_pred_path,
            pred_path,
            is_latest,
            rerank_ok,
            rerank_run_dir,
            rerank_latest_run,
            rerank_input_path,
        )
        stale_decision = (not is_latest and not rerank_ok and not allow_stale_pred)
        stale_reason = "ok"
        try:
            pred_resolved_for_reason = pred_path.expanduser().resolve() if pred_path.exists() else None
        except Exception:
            pred_resolved_for_reason = pred_path
        if stale_decision:
            if latest_res is None:
                stale_reason = "missing_evidence"
            elif pred_resolved_for_reason != latest_res:
                stale_reason = "pred_not_latest"
            elif not rerank_ok:
                stale_reason = "rerank_not_ok"
            elif rerank_out_res_for_log is not None and pred_resolved_for_reason != rerank_out_res_for_log:
                stale_reason = "rerank_output_mismatch"
            else:
                stale_reason = "unknown"
        _log_pred_guard_evidence_v2(
            logger,
            pred_path=pred_path,
            latest_pred_path=latest_pred_path,
            rerank_ok=rerank_ok,
            rerank_ok_source=rerank_ok_source,
            rerank_index_path=rerank_index_path,
            rerank_index_loaded=rerank_index_loaded,
            rerank_index_artifacts=rerank_index_artifacts,
            report_path=report_path,
            report_loaded=report_loaded,
            report_keys=report_keys,
            report_vals=report_vals,
            stale=stale_decision,
            stale_reason=stale_reason,
            latest_resolved=latest_res,
            rerank_out_resolved=rerank_out_res_for_log,
        )
        if not is_latest and not rerank_ok and not allow_stale_pred:
            raise ConfigError(
                f"eval pred_path is stale (latest infer_step1={latest_pred_path}, given={pred_path})"
            )
    else:
        logger.info("[eval][pred_guard_skip] reason=no_pred_path")
    _log_pred_candidates_topk(artifacts_root, run_context.exp_id, logger, limit=5)
    if gold_schema != "unknown" and pred_id_schema != gold_schema:
        logger.warning(
            "[eval] gold_schema=%s but pred_id_schema=%s (join likely 0).",
            gold_schema,
            pred_id_schema,
        )
        if not explicit_pred_path:
            picked = pick_pred_path_for_gold(
                artifacts_root,
                run_context.exp_id,
                gold_schema,
                now_run_id=run_context.run_id,
            )
            if picked and picked != pred_path:
                pred_path = picked
                pred_schema = _infer_pred_schema_from_path(pred_path, eval_cfg)
                pred_id_schema = detect_example_id_schema(pred_path)
                logger.info("[eval] pred reselected: pred_schema=%s pred_path=%s", pred_schema, pred_path)
            else:
                logger.warning("[eval] pred reselect failed or unchanged; using %s", pred_path)

    logger.info(
        "[eval] final pred_schema=%s pred_path=%s pred_source=%s",
        pred_schema,
        pred_path,
        pred_source,
    )
    _log_join_key_samples(gold_by_key, pred_path, match_key_policy, logger)

    views: dict[str, dict[str, Any]] = {}
    all_errors: list[dict[str, Any]] = []
    view_cfg = eval_cfg.get("views", {}) or {}
    for view_name in report_views:
        include_hold_view = view_name == "lenient"
        view_override = view_cfg.get(view_name, {})
        if isinstance(view_override, dict) and "include_hold" in view_override:
            include_hold_view = bool(view_override.get("include_hold"))
        view_report, view_errors = _compute_eval_view(
            cfg=cfg,
            run_context=run_context,
            pred_path=pred_path,
            pred_schema=pred_schema,
            gold_rows=gold_rows,
            gold_by_key=gold_by_key,
            include_hold=include_hold_view,
            view_name=view_name,
        )
        views[view_name] = view_report
        all_errors.extend(view_errors)

    encoder_pred_path = None
    encoder_views: dict[str, dict[str, Any]] | None = None
    if pred_path.name == "infer_candidates.reranked.jsonl":
        report_path = pred_path.parent / "infer_step2_rerank_report.json"
        if report_path.exists():
            try:
                report_obj = json.loads(report_path.read_text(encoding="utf-8"))
                encoder_pred_path = report_obj.get("input_pred_path") or report_obj.get("input_path")
            except Exception:
                encoder_pred_path = None
        if encoder_pred_path:
            encoder_pred_path = Path(str(encoder_pred_path))
            if encoder_pred_path.exists():
                encoder_pred_schema = _infer_pred_schema_from_path(encoder_pred_path, eval_cfg)
                encoder_views = {}
                for view_name in report_views:
                    include_hold_view = view_name == "lenient"
                    view_override = view_cfg.get(view_name, {})
                    if isinstance(view_override, dict) and "include_hold" in view_override:
                        include_hold_view = bool(view_override.get("include_hold"))
                    view_report, _view_errors = _compute_eval_view(
                        cfg=cfg,
                        run_context=run_context,
                        pred_path=encoder_pred_path,
                        pred_schema=encoder_pred_schema,
                        gold_rows=gold_rows,
                        gold_by_key=gold_by_key,
                        include_hold=include_hold_view,
                        view_name=view_name,
                    )
                    encoder_views[view_name] = view_report

    missing = [v for v in ("strict", "lenient") if v not in views]
    if missing:
        raise RuntimeError(
            f"[eval] missing views={missing} "
            f"(raw_report_views={raw_report_views!r}, normalized={report_views}, file={__file__})"
        )

    primary_report = views.get(primary_view) or views.get("strict") or next(iter(views.values()))
    rows_csv = primary_report.pop("_rows_csv", [])
    gold_keys = set([str(k) for k in gold_by_key.keys()])
    pred_keys = _collect_pred_keys(pred_path, match_key_policy)
    join_intersection = len(gold_keys.intersection(pred_keys))
    joined, joined_sample = _count_joined_raw(pred_path, gold_by_key, match_key_policy)
    joined_fn = "compute_joined_v1"
    pred_records_count = _count_pred_records_simple(pred_path)
    report_counters = primary_report.get("counters", {})
    report_counters.update(
        {
            "pred_source": pred_source,
            "joined_fn": joined_fn,
            "n_gold_occurrences": len(gold_rows),
            "n_pred_records": pred_records_count,
            "gold_unique_keys": len(gold_keys),
            "pred_unique_keys": len(pred_keys),
            "join_intersection": join_intersection,
            "joined": joined,
            "tp_without_pred": 0,
        }
    )
    report = {
        "pred_path": str(pred_path),
        "pred_source": pred_source,
        "joined_fn": joined_fn,
        "pred_schema": pred_schema,
        "gold_xlsx": str(gold_xlsx),
        "overall": primary_report.get("overall", {}),
        "by_eid": primary_report.get("by_eid", []),
        "counters": report_counters,
        "policies": primary_report.get("policies", {}),
        "views": {
            name: {k: v for k, v in view.items() if k != "_rows_csv"} for name, view in views.items()
        },
    }
    _write_eval_run_meta(
        outputs_dir=outputs_dir,
        run_context=run_context,
        gold_xlsx=gold_xlsx,
        pred_path=pred_path,
        gold_rows=len(gold_rows),
        gold_occ_rows=len(gold_rows),
        pred_records=pred_records_count,
        gold_unique_keys=len(gold_keys),
        pred_unique_keys=len(pred_keys),
        join_intersection=join_intersection,
        joined=joined,
        explicit_pred_path=explicit_pred_path,
        pred_source=pred_source,
    )
    if joined_sample:
        logger.info(
            "[eval] joined_sample example_id=%s instance_id=%s e_id=%s span_key=%s pred_exists=%s",
            joined_sample.get("example_id"),
            joined_sample.get("instance_id"),
            joined_sample.get("e_id"),
            joined_sample.get("span_key"),
            joined_sample.get("pred_exists"),
        )
    logger.info("[eval] joined_fn=%s", joined_fn)
    if write_errors:
        errors_path = outputs_dir / "eval_errors.jsonl"
        with errors_path.open("w", encoding="utf-8") as fp_err:
            for item in all_errors:
                write_jsonl_line(fp_err, item)
    row_counters = {
        "n_gold_occurrence_rows": 0,
        "n_pred_ignored_rows": 0,
        "trap_fp_total": 0,
        "trap_fp_reason_counts": Counter(),
    }
    try:
        row_counters = _export_eval_for_users(
            cfg=cfg,
            run_context=run_context,
            outputs_dir=outputs_dir,
            pred_path=pred_path,
            gold_rows=gold_rows,
            nta_rows=nta_rows,
            gold_by_key=gold_by_key,
            neg_by_key=neg_by_key,
            views=views,
            report_views=report_views,
            export_variant="llm",
        )
    except Exception as exc:
        logger.error("eval export_for_users 실패: %s", exc, exc_info=True)
        raise
    if pred_path.name == "infer_candidates.reranked.jsonl":
        if encoder_pred_path and encoder_views:
            if Path(str(encoder_pred_path)).exists():
                try:
                    _export_eval_for_users(
                        cfg=cfg,
                        run_context=run_context,
                        outputs_dir=outputs_dir,
                        pred_path=Path(str(encoder_pred_path)),
                        gold_rows=gold_rows,
                        nta_rows=nta_rows,
                        gold_by_key=gold_by_key,
                        neg_by_key=neg_by_key,
                        views=encoder_views,
                        report_views=report_views,
                        export_variant="encoder",
                    )
                except Exception as exc:
                    logger.error("eval export_for_users(encoder) 실패: %s", exc, exc_info=True)
                    raise
            else:
                logger.info(
                    "[for_users][min_export] variant=encoder skipped=true reason=encoder_pred_missing path=%s",
                    encoder_pred_path,
                )
        else:
            report_path = pred_path.parent / "infer_step2_rerank_report.json"
            logger.info(
                "[for_users][min_export] variant=encoder skipped=true reason=missing_rerank_report path=%s",
                report_path,
            )
    elif pred_path.name == "infer_candidates.jsonl":
        try:
            _export_eval_for_users(
                cfg=cfg,
                run_context=run_context,
                outputs_dir=outputs_dir,
                pred_path=pred_path,
                gold_rows=gold_rows,
                nta_rows=nta_rows,
                gold_by_key=gold_by_key,
                neg_by_key=neg_by_key,
                views=views,
                report_views=report_views,
                export_variant="encoder",
            )
        except Exception as exc:
            logger.error("eval export_for_users(encoder direct) 실패: %s", exc, exc_info=True)
            raise
    try:
        _write_detect_coverage_summary(
            cfg=cfg,
            run_context=run_context,
            outputs_dir=outputs_dir,
            logger=logger,
        )
        _write_llm_none_breakdown(
            cfg=cfg,
            run_context=run_context,
            outputs_dir=outputs_dir,
            logger=logger,
        )
        _write_polyset_ambiguity_stats(
            cfg=cfg,
            run_context=run_context,
            outputs_dir=outputs_dir,
            logger=logger,
        )
        _write_rerank_health_dashboard(
            report_path=pred_path.parent / "infer_step2_rerank_report.json"
            if pred_path is not None
            else None,
            out_dir=outputs_dir,
            eval_run_id=run_context.run_id,
            pred_path=pred_path,
            cfg=cfg,
            logger=logger,
        )
        _append_rerank_health_trend(
            dashboard_path=outputs_dir / "for_users" / "rerank_health_dashboard.json",
            trend_path=outputs_dir / "for_users" / "rerank_health_trend.jsonl",
            logger=logger,
        )
        _write_llm_audit_summary(
            run_context=run_context,
            outputs_dir=outputs_dir,
            pred_path=pred_path,
            logger=logger,
        )
        _write_rerank_health_summary(
            run_context=run_context,
            outputs_dir=outputs_dir,
            pred_path=pred_path,
            logger=logger,
        )
        _write_fn_analysis_summary(
            cfg=cfg,
            run_context=run_context,
            outputs_dir=outputs_dir,
            logger=logger,
        )
        _write_run_delta_summary_csv(
            run_context=run_context,
            outputs_dir=outputs_dir,
            logger=logger,
        )
        _write_eval_run_delta_csv(
            run_context=run_context,
            outputs_dir=outputs_dir,
            logger=logger,
        )
        _write_eval_run_delta_reasons_csv(
            outputs_dir=outputs_dir,
            logger=logger,
        )
        _write_eval_run_delta_samples_jsonl(
            outputs_dir=outputs_dir,
            logger=logger,
            n_per_kind=20,
        )
        _write_for_users_index_json(
            run_context=run_context,
            outputs_dir=outputs_dir,
            logger=logger,
        )
        _freeze_for_users_artifacts(
            outputs_dir=outputs_dir,
            logger=logger,
        )
    except Exception as exc:
        logger.error("eval fn/coverage summary 실패: %s", exc, exc_info=True)
        raise
    report_counters.update(row_counters)
    report_counters.setdefault("trap_fp_total", 0)
    report_counters.setdefault("trap_fp_reason_counts", Counter())
    if "pred_ignored_reason_counts" in report_counters:
        report_counters["pred_ignored_reason_counts"] = dict(
            report_counters["pred_ignored_reason_counts"]
        )
    if "trap_fp_reason_counts" in report_counters:
        report_counters["trap_fp_reason_counts"] = dict(
            report_counters["trap_fp_reason_counts"]
        )
    pred_ignored_rows = report_counters.get("pred_ignored_rows", 0)
    if pred_ignored_rows > 0:
        logger.warning(
            "[eval] pred_ignored rows=%s (check only_keys_in_gold/triage policies)",
            pred_ignored_rows,
        )
        logger.warning(
            "[eval] pred_ignored_reason_counts=%s",
            report_counters.get("pred_ignored_reason_counts", {}),
        )
    write_json(outputs_dir / "eval_report.json", report, indent=2)
    logger.info("eval 완료: %s", outputs_dir)


def _load_gold_rows(
    gold_xlsx: Path,
    *,
    gold_sheet_name: str = "gold",
    allowed_splits: list[str] | None = None,
) -> list[dict[str, Any]]:
    logger = logging.getLogger("kmwe")
    gold_df = pd.read_excel(gold_xlsx, sheet_name=gold_sheet_name, engine="openpyxl")
    logger.info("[gold] rows=%s cols=%s", len(gold_df), list(gold_df.columns))
    if "split" in gold_df.columns:
        logger.info(
            "[gold] split_counts=%s",
            gold_df["split"].fillna("NA").value_counts().to_dict(),
        )
    if "gold_example_role" in gold_df.columns:
        logger.info(
            "[gold] role_counts=%s",
            gold_df["gold_example_role"].fillna("NA").value_counts().to_dict(),
        )
    if "span_segments" in gold_df.columns:
        s = gold_df["span_segments"]
        has_span = s.notna() & (s.astype(str).str.strip().str.len() > 2)
        logger.info("[gold] has_span_rows=%s", int(has_span.sum()))
    rows: list[dict[str, Any]] = []
    for _, row in gold_df.iterrows():
        if allowed_splits is not None:
            split_val = row.get("split")
            split_norm = "" if pd.isna(split_val) else str(split_val).strip().lower()
            if split_norm not in allowed_splits:
                continue
        span_raw = row.get("span_segments")
        target_sentence = row.get("target_sentence")
        e_id = row.get("e_id")
        if pd.isna(span_raw) or pd.isna(target_sentence) or pd.isna(e_id):
            continue
        role = str(row.get("gold_example_role", "") or "")
        if not role.startswith("pos"):
            continue
        if not isinstance(span_raw, str):
            span_raw = str(span_raw)
        if not span_raw or span_raw.strip().lower() == "nan":
            continue
        try:
            span_segments = ast.literal_eval(span_raw)
        except Exception:
            continue
        if not span_segments:
            continue
        span_key = silver_loader._span_key_from_segments(span_segments)
        rows.append(
            {
                "e_id": str(e_id),
                "example_id": row.get("example_id"),
                "gold_example_role": role,
                "target_sentence": str(target_sentence),
                "span_segments": span_segments,
                "span_key": span_key,
                "instance_id": row.get("instance_id"),
                "doc_id": row.get("doc_id"),
                "sent_index": row.get("sent_index"),
            }
        )
    return rows


def _load_neg_rows(
    gold_xlsx: Path,
    *,
    gold_sheet_name: str = "gold",
    allowed_splits: list[str] | None = None,
) -> list[dict[str, Any]]:
    gold_df = pd.read_excel(gold_xlsx, sheet_name=gold_sheet_name, engine="openpyxl")
    rows: list[dict[str, Any]] = []
    skipped_non_single = 0
    skipped_missing_eids = 0
    fallback_legacy_eid = 0
    expanded_conf_eids = 0
    role_counts: Counter[str] = Counter()
    for _, row in gold_df.iterrows():
        if allowed_splits is not None:
            split_val = row.get("split")
            split_norm = "" if pd.isna(split_val) else str(split_val).strip().lower()
            if split_norm not in allowed_splits:
                continue
        span_raw = row.get("span_segments")
        target_sentence = row.get("target_sentence")
        if pd.isna(span_raw) or pd.isna(target_sentence):
            continue
        role = str(row.get("gold_example_role", "") or "")
        if not role.startswith("neg"):
            continue
        if not isinstance(span_raw, str):
            span_raw = str(span_raw)
        if not span_raw or span_raw.strip().lower() == "nan":
            continue
        try:
            span_segments = ast.literal_eval(span_raw)
        except Exception:
            continue
        if not span_segments:
            continue
        if (
            not isinstance(span_segments, (list, tuple))
            or len(span_segments) != 1
        ):
            skipped_non_single += 1
            continue
        if role not in {"neg_confusable", "neg_boundary", "neg_target_absent"}:
            continue
        role_counts[role] += 1
        conf_e_id = row.get("conf_e_id")
        conf_e_id_raw = ""
        if conf_e_id is not None and not pd.isna(conf_e_id):
            conf_e_id_raw = str(conf_e_id)
        target_eids: list[str] = []
        if conf_e_id_raw:
            for part in conf_e_id_raw.split(";"):
                part = part.strip()
                if part:
                    target_eids.append(part)
        if not target_eids:
            e_id = row.get("e_id")
            if e_id is not None and not pd.isna(e_id):
                target_eids = [str(e_id).strip()]
                if target_eids and target_eids[0]:
                    fallback_legacy_eid += 1
        if not target_eids:
            skipped_missing_eids += 1
            continue
        span_key = silver_loader._span_key_from_segments(span_segments)
        for eid in target_eids:
            if not eid:
                continue
            rows.append(
                {
                    "e_id": str(eid),
                    "example_id": row.get("example_id"),
                    "target_sentence": str(target_sentence),
                    "span_segments": span_segments,
                    "span_key": span_key,
                    "instance_id": row.get("instance_id"),
                    "doc_id": row.get("doc_id"),
                    "sent_index": row.get("sent_index"),
                    "gold_example_role": role,
                    "conf_e_id_raw": conf_e_id_raw,
                }
            )
            expanded_conf_eids += 1
    if skipped_non_single > 0:
        logging.getLogger("kmwe").info(
            "[eval] neg span_segments skipped (non_single)=%s", skipped_non_single
        )
    logging.getLogger("kmwe").info(
        "[eval] neg role_counts=%s expanded_conf_eids=%s fallback_legacy_eid=%s skipped_missing_eids=%s skipped_non_single=%s",
        dict(role_counts),
        expanded_conf_eids,
        fallback_legacy_eid,
        skipped_missing_eids,
        skipped_non_single,
    )
    return rows


def _iter_jsonl_lines(path: Path) -> Iterable[dict[str, Any]]:
    if not path.exists():
        return []
    def _gen() -> Iterable[dict[str, Any]]:
        with path.open("r", encoding="utf-8") as fp:
            for idx, line in enumerate(fp, start=1):
                line = line.strip()
                if not line:
                    continue
                try:
                    obj = json.loads(line)
                except Exception as exc:
                    raise ConfigError(f"jsonl parse failed at line {idx}: {path}") from exc
                if isinstance(obj, dict):
                    yield obj
    return _gen()


def _write_fn_analysis_summary(
    *,
    cfg: dict[str, Any],
    run_context: RunContext,
    outputs_dir: Path,
    logger: logging.Logger,
) -> None:
    for_users_dir = outputs_dir / "for_users"
    rows_jsonl_path = for_users_dir / "eval_latest_rows.jsonl"
    if not rows_jsonl_path.exists():
        raise ConfigError(f"eval_latest_rows.jsonl missing: {rows_jsonl_path}")
    uncertainty_cfg = cfg.get("infer_step1", {}).get("uncertainty", {}) or {}
    try:
        low_conf_threshold = float(uncertainty_cfg.get("low_conf_threshold", 0.55))
    except Exception:
        low_conf_threshold = 0.55
    try:
        margin_threshold = float(uncertainty_cfg.get("margin_threshold", 0.10))
    except Exception:
        margin_threshold = 0.10
    logger.info("[eval][fn_summary] start rows_path=%s", rows_jsonl_path)
    fn_reason_counts = Counter()
    none_reason_counts = Counter()
    missing_field_counts = Counter()
    n_rows_total = 0
    n_fn_total = 0
    known_reason = 0
    llm_none_total = 0
    n_fn_unknown = 0
    llm_none_examples: list[dict[str, Any]] = []
    other_fn_examples: list[dict[str, Any]] = []
    for item in _iter_jsonl_lines(rows_jsonl_path):
        if not isinstance(item, dict):
            continue
        n_rows_total += 1
        if item.get("row_kind") != "gold_occurrence" or item.get("view") != "strict":
            continue
        fn_reason = str(item.get("fn_reason_code") or "")
        if not fn_reason:
            continue
        n_fn_total += 1
        fn_reason_counts[fn_reason] += 1
        known_reason += 1
        none_reason = str(item.get("none_reason") or "")
        if fn_reason == "LLM_NONE":
            llm_none_total += 1
            if none_reason:
                none_reason_counts[none_reason] += 1
            else:
                none_reason_counts["UNKNOWN"] += 1
        if fn_reason == "UNCLASSIFIED_FN":
            n_fn_unknown += 1
            missing_fields = item.get("evidence_missing_fields") or []
            if isinstance(missing_fields, list):
                for field in missing_fields:
                    if field:
                        missing_field_counts[str(field)] += 1
        full = item.get("full") or {}
        record = full.get("record") or {}
        gold_item = full.get("gold_item") or {}
        rerank = record.get("rerank") or {}
        decision_line = item.get("decision_line") or rerank.get("decision_line") or ""
        protocol_ok = item.get("protocol_ok")
        if protocol_ok is None:
            protocol_ok = rerank.get("protocol_ok")
        example = {
            "match_key": item.get("match_key") or "",
            "example_id": record.get("example_id") or "",
            "instance_id": record.get("instance_id") or "",
            "gold_eid": gold_item.get("e_id") or "",
            "pred_eid": (
                item.get("pred_eid")
                or rerank.get("selected_eid")
                or record.get("pred_eid")
                or ""
            ),
            "fn_reason_code": fn_reason,
            "none_reason": none_reason or "",
            "decision_line": str(decision_line),
            "protocol_ok": protocol_ok if protocol_ok is not None else "",
        }
        n_candidates = item.get("n_candidates")
        n_to_llm_candidates = item.get("n_to_llm_candidates")
        ambiguous_flag = item.get("ambiguous_flag")
        if n_candidates is not None:
            example["n_candidates"] = n_candidates
        if n_to_llm_candidates is not None:
            example["n_to_llm_candidates"] = n_to_llm_candidates
        if ambiguous_flag is not None:
            example["ambiguous_flag"] = ambiguous_flag
        if fn_reason == "LLM_NONE":
            llm_none_examples.append(example)
        else:
            other_fn_examples.append(example)
    top_examples: list[dict[str, Any]] = []
    top_examples.extend(llm_none_examples[:20])
    if len(top_examples) < 20:
        top_examples.extend(other_fn_examples[: (20 - len(top_examples))])
    coverage = 0.0
    if n_fn_total > 0:
        coverage = float(known_reason) / float(n_fn_total)
    summary = {
        "created_at": iso_now(),
        "eval_run_id": run_context.run_id,
        "n_rows_total": n_rows_total,
        "n_fn_total": n_fn_total,
        "fn_reason_counts": dict(fn_reason_counts),
        "llm_none_total": llm_none_total,
        "llm_none_reason_counts": dict(none_reason_counts),
        "fn_reason_coverage": coverage,
        "thresholds_snapshot": {
            "low_conf_threshold": low_conf_threshold,
            "margin_threshold": margin_threshold,
        },
        "top_examples": top_examples,
    }
    out_path = for_users_dir / "fn_analysis_summary.json"
    write_json(out_path, summary, indent=2)
    logger.info(
        "[eval][fn_reason_coverage] n_fn_total=%s known=%s coverage=%.4f",
        n_fn_total,
        known_reason,
        coverage,
    )
    logger.info(
        "[eval][llm_none_reason_counts] total=%s counts=%s",
        llm_none_total,
        dict(none_reason_counts),
    )
    logger.info(
        "[eval][unknown_bucket] n_fn_unknown=%s missing_top=%s",
        n_fn_unknown,
        missing_field_counts.most_common(3),
    )
    try:
        bytes_written = out_path.stat().st_size
    except Exception:
        bytes_written = 0
    logger.info(
        "[eval][fn_summary] wrote path=%s bytes=%s top_examples=%s",
        out_path,
        bytes_written,
        len(top_examples),
    )
    try:
        if llm_none_examples:
            csv_path = for_users_dir / "fn_examples_llm_none.csv"
            with csv_path.open("w", encoding="utf-8", newline="") as fp:
                writer = csv.DictWriter(
                    fp,
                    fieldnames=[
                        "match_key",
                        "example_id",
                        "instance_id",
                        "gold_eid",
                        "pred_eid",
                        "none_reason",
                        "decision_line",
                    ],
                )
                writer.writeheader()
                for ex in llm_none_examples[:100]:
                    writer.writerow(
                        {
                            "match_key": ex.get("match_key", ""),
                            "example_id": ex.get("example_id", ""),
                            "instance_id": ex.get("instance_id", ""),
                            "gold_eid": ex.get("gold_eid", ""),
                            "pred_eid": ex.get("pred_eid", ""),
                            "none_reason": ex.get("none_reason", ""),
                            "decision_line": ex.get("decision_line", ""),
                        }
                    )
    except Exception:
        logger.warning("[eval] failed to write fn_examples_llm_none.csv", exc_info=True)


def _write_detect_coverage_summary(
    *,
    cfg: dict[str, Any],
    run_context: RunContext,
    outputs_dir: Path,
    logger: logging.Logger,
) -> None:
    _ = cfg
    for_users_dir = outputs_dir / "for_users"
    rows_jsonl_path = for_users_dir / "eval_latest_rows.jsonl"
    if not rows_jsonl_path.exists():
        raise ConfigError(f"eval_latest_rows.jsonl missing: {rows_jsonl_path}")
    by_gold_eid: dict[str, dict[str, Any]] = {}
    n_rows_scanned = 0
    n_rows_eligible = 0
    total_detect_hit = 0
    for item in _iter_jsonl_lines(rows_jsonl_path):
        if not isinstance(item, dict):
            continue
        n_rows_scanned += 1
        if item.get("row_kind") != "gold_occurrence" or item.get("view") != "strict":
            continue
        if item.get("eval_tag") == "non_evaluable":
            continue
        full = item.get("full") or {}
        record = full.get("record") or {}
        gold_item = full.get("gold_item") or {}
        gold_eid = str(gold_item.get("e_id") or "")
        if not gold_eid:
            continue
        gold_span_key = _canonical_span_key(
            gold_item.get("span_key"), gold_item.get("span_segments")
        )
        candidates = record.get("candidates") or []
        if not isinstance(candidates, list):
            candidates = []
        cand_eids = {
            str(c.get("e_id") or "")
            for c in candidates
            if isinstance(c, dict) and c.get("e_id") is not None
        }
        n_rows_eligible += 1
        entry = by_gold_eid.setdefault(
            gold_eid,
            {
                "n_total": 0,
                "n_detect_hit": 0,
                "n_detect_miss": 0,
                "n_boundary_hit": 0,
                "n_boundary_miss": 0,
            },
        )
        entry["n_total"] += 1
        if gold_eid in cand_eids:
            entry["n_detect_hit"] += 1
            total_detect_hit += 1
            if gold_span_key:
                cand_span_keys_for_gold = [
                    _canonical_span_key(c.get("span_key"), c.get("span_segments"))
                    for c in candidates
                    if isinstance(c, dict) and str(c.get("e_id") or "") == gold_eid
                ]
                cand_span_keys_for_gold = [
                    k for k in cand_span_keys_for_gold if k
                ]
                if cand_span_keys_for_gold:
                    if gold_span_key in cand_span_keys_for_gold:
                        entry["n_boundary_hit"] += 1
                    else:
                        entry["n_boundary_miss"] += 1
        else:
            entry["n_detect_miss"] += 1
    overall_hit_rate = (
        float(total_detect_hit) / float(n_rows_eligible) if n_rows_eligible > 0 else 0.0
    )
    for gold_eid, entry in by_gold_eid.items():
        n_total = entry["n_total"]
        n_detect_hit = entry["n_detect_hit"]
        entry["detect_hit_rate"] = (
            float(n_detect_hit) / float(n_total) if n_total > 0 else 0.0
        )
        n_boundary_hit = entry.get("n_boundary_hit", 0)
        entry["boundary_hit_rate"] = (
            float(n_boundary_hit) / float(n_detect_hit) if n_detect_hit > 0 else 0.0
        )
    top_miss = sorted(
        by_gold_eid.items(),
        key=lambda kv: int(kv[1].get("n_detect_miss", 0)),
        reverse=True,
    )[:20]
    top_miss_eids = [{**{"gold_eid": k}, **v} for k, v in top_miss]
    summary = {
        "created_at": iso_now(),
        "eval_run_id": run_context.run_id,
        "n_rows_scanned": n_rows_scanned,
        "n_rows_eligible": n_rows_eligible,
        "overall_detect_hit_rate": overall_hit_rate,
        "by_gold_eid": by_gold_eid,
        "top_miss_eids": top_miss_eids,
    }
    out_path = for_users_dir / "detect_coverage.json"
    write_json(out_path, summary, indent=2)
    try:
        bytes_written = out_path.stat().st_size
    except Exception:
        bytes_written = 0
    logger.info(
        "[eval][detect_coverage] wrote path=%s bytes=%s n_eids=%s overall_hit_rate=%.4f top_miss=%s",
        out_path,
        bytes_written,
        len(by_gold_eid),
        overall_hit_rate,
        [e.get("gold_eid") for e in top_miss_eids[:5]],
    )


def _write_llm_none_breakdown(
    *,
    cfg: dict[str, Any],
    run_context: RunContext,
    outputs_dir: Path,
    logger: logging.Logger,
) -> None:
    _ = cfg
    for_users_dir = outputs_dir / "for_users"
    rows_jsonl_path = for_users_dir / "eval_latest_rows.jsonl"
    if not rows_jsonl_path.exists():
        raise ConfigError(f"eval_latest_rows.jsonl missing: {rows_jsonl_path}")
    by_reason_counts: Counter[str] = Counter()
    stats_by_reason: dict[str, dict[str, Any]] = {}
    n_eval_total = 0
    n_eval_llm_none = 0
    present_counts = {
        "n_candidates": 0,
        "conf": 0,
        "margin": 0,
        "polyset_id": 0,
        "ambiguous": 0,
    }
    for item in _iter_jsonl_lines(rows_jsonl_path):
        if not isinstance(item, dict):
            continue
        if item.get("row_kind") != "gold_occurrence" or item.get("view") != "strict":
            continue
        if item.get("eval_tag") == "non_evaluable":
            continue
        n_eval_total += 1
        if str(item.get("fn_reason_code") or "") != "LLM_NONE":
            continue
        n_eval_llm_none += 1
        none_reason = str(item.get("none_reason") or "UNKNOWN")
        by_reason_counts[none_reason] += 1
        bucket = stats_by_reason.setdefault(
            none_reason,
            {
                "count": 0,
                "sum_n_candidates": 0,
                "n_candidates_count": 0,
                "sum_top1_conf": 0.0,
                "top1_conf_count": 0,
                "sum_margin": 0.0,
                "margin_count": 0,
                "ambiguous_true": 0,
                "ambiguous_count": 0,
                "polyset_counts": Counter(),
            },
        )
        full = item.get("full") or {}
        record = full.get("record") or {}
        candidates = record.get("candidates") or []
        if not isinstance(candidates, list):
            candidates = []
        n_candidates = record.get("n_to_llm_candidates")
        if not isinstance(n_candidates, int):
            n_candidates = len(candidates)
        conf_top1 = record.get("candidate_encoder_conf_top1")
        if conf_top1 is None:
            conf_top1 = record.get("encoder_conf_top1")
        if conf_top1 is None and candidates:
            cand0 = candidates[0] if isinstance(candidates[0], dict) else None
            if cand0:
                conf_top1 = cand0.get("encoder_conf")
        margin = record.get("candidate_margin_top1_top2")
        if margin is None:
            margin = record.get("margin_top1_top2")
        ambiguous = record.get("is_ambiguous")
        if ambiguous is None:
            ambiguous = record.get("ambiguous")
        polyset_id = record.get("polyset_id")
        if polyset_id is None:
            polyset_id = record.get("gold_polyset_id")
        bucket["count"] += 1
        if isinstance(n_candidates, int):
            bucket["sum_n_candidates"] += n_candidates
            bucket["n_candidates_count"] += 1
            present_counts["n_candidates"] += 1
        if conf_top1 is not None:
            try:
                bucket["sum_top1_conf"] += float(conf_top1)
                bucket["top1_conf_count"] += 1
                present_counts["conf"] += 1
            except Exception:
                pass
        if margin is not None:
            try:
                bucket["sum_margin"] += float(margin)
                bucket["margin_count"] += 1
                present_counts["margin"] += 1
            except Exception:
                pass
        if isinstance(ambiguous, bool):
            bucket["ambiguous_count"] += 1
            if ambiguous:
                bucket["ambiguous_true"] += 1
            present_counts["ambiguous"] += 1
        if polyset_id is not None:
            bucket["polyset_counts"].update([str(polyset_id)])
            present_counts["polyset_id"] += 1

    stats_by_none_reason = {}
    for reason, bucket in stats_by_reason.items():
        count = bucket["count"]
        avg_n_candidates = (
            float(bucket["sum_n_candidates"]) / float(bucket["n_candidates_count"])
            if bucket["n_candidates_count"] > 0
            else 0.0
        )
        avg_top1_conf = (
            float(bucket["sum_top1_conf"]) / float(bucket["top1_conf_count"])
            if bucket["top1_conf_count"] > 0
            else None
        )
        avg_margin = (
            float(bucket["sum_margin"]) / float(bucket["margin_count"])
            if bucket["margin_count"] > 0
            else None
        )
        ambiguous_rate = (
            float(bucket["ambiguous_true"]) / float(bucket["ambiguous_count"])
            if bucket["ambiguous_count"] > 0
            else 0.0
        )
        stats_by_none_reason[reason] = {
            "n": count,
            "avg_n_candidates": avg_n_candidates,
            "avg_encoder_conf_top1": avg_top1_conf,
            "avg_margin_top1_top2": avg_margin,
            "ambiguous_rate": ambiguous_rate,
            "polyset_id_topk": bucket["polyset_counts"].most_common(10),
        }
    summary = {
        "created_at": iso_now(),
        "eval_run_id": run_context.run_id,
        "n_eval_total": n_eval_total,
        "n_eval_llm_none": n_eval_llm_none,
        "by_none_reason": dict(by_reason_counts),
        "stats_by_none_reason": stats_by_none_reason,
    }
    out_path = for_users_dir / "llm_none_breakdown.json"
    logger.info(
        "[eval][llm_none_breakdown_counts] n_eval_llm_none=%s by_none_reason=%s",
        n_eval_llm_none,
        dict(by_reason_counts),
    )
    logger.info(
        "[eval][llm_none_breakdown_stats] keys_present=%s",
        present_counts,
    )
    write_json(out_path, summary, indent=2)
    try:
        bytes_written = out_path.stat().st_size
    except Exception:
        bytes_written = 0
    top_reasons = sorted(
        stats_by_none_reason.items(), key=lambda kv: kv[1].get("n", 0), reverse=True
    )[:3]
    logger.info(
        "[eval][llm_none_breakdown] wrote path=%s bytes=%s n_total=%s top_reasons=%s",
        out_path,
        bytes_written,
        n_eval_llm_none,
        [k for k, _ in top_reasons],
    )


def _write_polyset_ambiguity_stats(
    *, cfg: dict[str, Any], run_context: RunContext, outputs_dir: Path, logger: logging.Logger
) -> None:
    for_users_dir = outputs_dir / "for_users"
    rows_jsonl_path = for_users_dir / "eval_latest_rows.jsonl"
    if not rows_jsonl_path.exists():
        raise ConfigError(f"eval_latest_rows.jsonl missing: {rows_jsonl_path}")
    margin_threshold = (
        ((cfg.get("infer") or {}).get("uncertainty") or {}).get("margin_threshold", 0.10)
    )
    try:
        margin_threshold = float(margin_threshold)
    except Exception:
        margin_threshold = 0.10

    n_eval_total = 0
    n_llm_none = 0
    n_llm_none_ambiguous_polyset = 0
    n_records_seen = 0
    n_records_with_candidates = 0
    n_groups_with_polyset_id = 0
    n_groups_with_top2 = 0
    skipped_no_candidates = 0
    skipped_no_top2 = 0

    margins_all: list[float] = []
    global_pair_counts: Counter[str] = Counter()
    polyset_stats: dict[str, dict[str, Any]] = {}

    for item in _iter_jsonl_lines(rows_jsonl_path):
        if not isinstance(item, dict):
            continue
        if item.get("row_kind") != "gold_occurrence" or item.get("view") != "strict":
            continue
        if item.get("eval_tag") == "non_evaluable":
            continue
        n_eval_total += 1
        if str(item.get("fn_reason_code") or "") != "LLM_NONE":
            continue
        n_llm_none += 1
        if str(item.get("none_reason") or "") != "AMBIGUOUS_POLYSET":
            continue
        n_llm_none_ambiguous_polyset += 1
        n_records_seen += 1
        full = item.get("full") or {}
        record = full.get("record") or {}
        candidates = record.get("candidates")
        if candidates is None:
            candidates_json = record.get("candidates_json")
            if isinstance(candidates_json, str) and candidates_json.strip():
                try:
                    candidates = json.loads(candidates_json)
                except Exception:
                    candidates = None
        if not isinstance(candidates, list):
            skipped_no_candidates += 1
            continue
        n_records_with_candidates += 1
        match_key = item.get("match_key") or record.get("match_key") or ""
        span_key = record.get("span_key") or ""
        groups: dict[tuple[str, str, str], list[tuple[str, float]]] = {}
        for cand in candidates:
            if not isinstance(cand, dict):
                continue
            eid = cand.get("e_id")
            if eid is None:
                eid = cand.get("eid")
            if eid is None:
                continue
            polyset_id = cand.get("polyset_id")
            if polyset_id is None:
                continue
            conf = None
            for key in ("encoder_conf", "encoder_prob", "score"):
                if key in cand:
                    conf = cand.get(key)
                    break
            if conf is None:
                continue
            try:
                conf_val = float(conf)
            except Exception:
                continue
            gkey = (str(match_key), str(span_key), str(polyset_id))
            groups.setdefault(gkey, []).append((str(eid), conf_val))
        if not groups:
            skipped_no_candidates += 1
            continue
        for (_, _, polyset_id), items in groups.items():
            n_groups_with_polyset_id += 1
            items_sorted = sorted(items, key=lambda x: x[1], reverse=True)
            if len(items_sorted) < 2:
                skipped_no_top2 += 1
                continue
            n_groups_with_top2 += 1
            eid1, conf1 = items_sorted[0]
            eid2, conf2 = items_sorted[1]
            margin = conf1 - conf2
            margins_all.append(margin)
            ambiguous = margin < margin_threshold
            pair_key = f"{eid1}|{eid2}"
            global_pair_counts[f"{polyset_id}:{pair_key}"] += 1
            bucket = polyset_stats.setdefault(
                str(polyset_id),
                {
                    "n_groups": 0,
                    "n_ambiguous": 0,
                    "margins": [],
                    "pair_counts": Counter(),
                },
            )
            bucket["n_groups"] += 1
            if ambiguous:
                bucket["n_ambiguous"] += 1
            bucket["margins"].append(margin)
            bucket["pair_counts"][pair_key] += 1

    def _percentile(vals: list[float], pct: float) -> float | None:
        if not vals:
            return None
        vals_sorted = sorted(vals)
        idx = int(round((len(vals_sorted) - 1) * pct))
        idx = max(0, min(idx, len(vals_sorted) - 1))
        return float(vals_sorted[idx])

    by_polyset_id = []
    for polyset_id, bucket in polyset_stats.items():
        n_groups = bucket["n_groups"]
        n_ambiguous = bucket["n_ambiguous"]
        margins = bucket["margins"]
        margin_mean = float(sum(margins) / float(len(margins))) if margins else None
        by_polyset_id.append(
            {
                "polyset_id": polyset_id,
                "n_groups": n_groups,
                "n_ambiguous": n_ambiguous,
                "ambiguous_rate": float(n_ambiguous) / float(n_groups) if n_groups > 0 else 0.0,
                "margin_summary": {
                    "mean": margin_mean,
                    "p10": _percentile(margins, 0.10),
                    "p50": _percentile(margins, 0.50),
                    "p90": _percentile(margins, 0.90),
                },
                "top_pairs": bucket["pair_counts"].most_common(10),
            }
        )
    by_polyset_id = sorted(by_polyset_id, key=lambda x: x["n_groups"], reverse=True)[:50]

    bins = [round(i * 0.02, 2) for i in range(0, 16)]
    hist_counts = [0 for _ in range(len(bins) + 1)]
    for val in margins_all:
        placed = False
        for i in range(len(bins) - 1):
            if bins[i] <= val < bins[i + 1]:
                hist_counts[i] += 1
                placed = True
                break
        if not placed:
            hist_counts[-1] += 1
    margin_histogram = []
    for i in range(len(bins) - 1):
        margin_histogram.append(
            {"bin_left": bins[i], "bin_right": bins[i + 1], "count": hist_counts[i]}
        )
    margin_histogram.append({"bin_left": bins[-1], "bin_right": None, "count": hist_counts[-1]})

    summary = {
        "created_at": iso_now(),
        "eval_run_id": run_context.run_id,
        "margin_threshold": margin_threshold,
        "n_eval_total": n_eval_total,
        "n_llm_none": n_llm_none,
        "n_llm_none_ambiguous_polyset": n_llm_none_ambiguous_polyset,
        "coverage": {
            "n_records_seen": n_records_seen,
            "n_records_with_candidates": n_records_with_candidates,
            "n_groups_with_polyset_id": n_groups_with_polyset_id,
            "n_groups_with_top2": n_groups_with_top2,
        },
        "by_polyset_id": by_polyset_id,
        "top_pairs_global": global_pair_counts.most_common(50),
        "margin_histogram": margin_histogram,
    }
    if n_groups_with_top2 == 0:
        summary["status"] = "insufficient_data"

    out_path = for_users_dir / "polyset_ambiguity_stats.json"
    write_json(out_path, summary, indent=2)
    try:
        bytes_written = out_path.stat().st_size
    except Exception:
        bytes_written = 0
    logger.info(
        "[eval][polyset_ambiguity] wrote path=%s bytes=%s n_llm_none_ambiguous_polyset=%s n_groups_with_top2=%s",
        out_path,
        bytes_written,
        n_llm_none_ambiguous_polyset,
        n_groups_with_top2,
    )
    logger.info(
        "[eval][polyset_ambiguity_coverage] candidates=%s groups_with_polyset_id=%s groups_with_top2=%s skipped_no_candidates=%s skipped_no_top2=%s",
        n_records_with_candidates,
        n_groups_with_polyset_id,
        n_groups_with_top2,
        skipped_no_candidates,
        skipped_no_top2,
    )


def _write_rerank_health_dashboard(
    *,
    report_path: Path | None,
    out_dir: Path,
    eval_run_id: str | None,
    pred_path: Path | None,
    cfg: dict[str, Any] | None,
    logger: logging.Logger,
) -> None:
    _ = cfg
    for_users_dir = out_dir / "for_users"
    for_users_dir.mkdir(parents=True, exist_ok=True)
    if report_path is None and pred_path is not None:
        candidate = pred_path.parent / "infer_step2_rerank_report.json"
        if candidate.exists() and candidate.stat().st_size > 0:
            report_path = candidate
    report = {}
    report_status = "insufficient_data"
    if report_path is not None and report_path.exists() and report_path.stat().st_size > 0:
        try:
            report = json.loads(report_path.read_text(encoding="utf-8"))
            report_status = str(report.get("status") or "ok")
        except Exception:
            report = {}
            report_status = "insufficient_data"
    def _get_int(key: str) -> int:
        try:
            return int(report.get(key) or 0)
        except Exception:
            return 0
    counts = {
        "n_records": _get_int("n_records"),
        "n_input_candidates": _get_int("n_input_candidates"),
        "n_to_llm_candidates": _get_int("n_to_llm_candidates"),
        "n_records_with_to_llm": _get_int("n_records_with_to_llm"),
        "prompts_written": _get_int("prompts_written"),
        "responses_rows": _get_int("responses_rows"),
        "decision_none_rows": _get_int("decision_none_rows"),
        "applied_selected": _get_int("applied_selected"),
        "no_response": _get_int("no_response"),
        "missing": _get_int("missing"),
        "http_fail": _get_int("http_fail"),
        "http_401": _get_int("http_401"),
        "parse_fail": _get_int("parse_fail"),
        "guard_reject": _get_int("guard_reject"),
        "empty_text": _get_int("empty_text"),
        "decision_line_non_empty": _get_int("decision_line_non_empty"),
        "error_non_empty": _get_int("error_non_empty"),
    }
    prompts_written = counts["prompts_written"]
    responses_rows = counts["responses_rows"]
    decision_none_rows = counts["decision_none_rows"]
    applied_selected = counts["applied_selected"]
    no_response = counts["no_response"]
    missing = counts["missing"]
    rates = {
        "decision_none_ratio": (
            float(decision_none_rows) / float(responses_rows) if responses_rows > 0 else None
        ),
        "apply_rate_per_prompt": (
            float(applied_selected) / float(prompts_written) if prompts_written > 0 else None
        ),
        "response_ok_rate": (
            float(responses_rows) / float(prompts_written) if prompts_written > 0 else None
        ),
        "no_response_rate": (
            float(no_response) / float(prompts_written) if prompts_written > 0 else None
        ),
        "missing_rate": (
            float(missing) / float(prompts_written) if prompts_written > 0 else None
        ),
    }
    payload = {
        "created_at": iso_now(),
        "eval_run_id": eval_run_id or "",
        "pred_path": str(pred_path) if pred_path is not None else "",
        "rerank_report": {
            "report_path": str(report_path) if report_path is not None else None,
            "rerank_run_id": report.get("rerank_run_id"),
            "report_created_at": report.get("created_at"),
            "llm_mode": report.get("llm_mode"),
            "status": report_status,
            "error": report.get("error") or "",
            "input_pred_path": report.get("input_pred_path") or report.get("input_path"),
            "rerank_output_path": report.get("rerank_output_path") or report.get("output_path"),
            "responses_path": report.get("responses_path"),
        },
        "counts": counts,
        "rates": rates,
    }
    out_path = for_users_dir / "rerank_health_dashboard.json"
    try:
        write_json(out_path, payload, indent=2)
        bytes_written = out_path.stat().st_size if out_path.exists() else 0
        logger.info(
            "[eval][rerank_health_dashboard] wrote path=%s bytes=%s status=%s prompts_written=%s responses_rows=%s applied_selected=%s decision_none_rows=%s",
            out_path,
            bytes_written,
            report_status,
            prompts_written,
            responses_rows,
            applied_selected,
            decision_none_rows,
        )
        logger.info(
            "[eval][rerank_health_dashboard_rates] decision_none_ratio=%s apply_rate_per_prompt=%s response_ok_rate=%s no_response_rate=%s missing_rate=%s",
            rates.get("decision_none_ratio"),
            rates.get("apply_rate_per_prompt"),
            rates.get("response_ok_rate"),
            rates.get("no_response_rate"),
            rates.get("missing_rate"),
        )
    except Exception as exc:
        logger.info(
            "[eval][rerank_health_dashboard] skipped=true reason=%s",
            f"{type(exc).__name__}: {exc}",
        )


def _append_rerank_health_trend(
    *, dashboard_path: Path, trend_path: Path, logger: logging.Logger
) -> None:
    if not dashboard_path.exists() or dashboard_path.stat().st_size == 0:
        logger.info(
            "[eval][rerank_health_trend] skipped=true reason=missing_dashboard path=%s",
            dashboard_path,
        )
        return
    try:
        dashboard = json.loads(dashboard_path.read_text(encoding="utf-8"))
    except Exception as exc:
        logger.info(
            "[eval][rerank_health_trend] skipped=true reason=%s path=%s",
            f"{type(exc).__name__}: {exc}",
            dashboard_path,
        )
        return
    rerank_report = dashboard.get("rerank_report") or {}
    counts = dashboard.get("counts") or {}
    rates = dashboard.get("rates") or {}
    created_at = dashboard.get("created_at") or rerank_report.get("report_created_at")
    eval_run_id = str(dashboard.get("eval_run_id") or "")
    existing_eval_ids: set[str] = set()
    if trend_path.exists():
        try:
            with trend_path.open("r", encoding="utf-8") as fp:
                tail_lines = fp.readlines()[-200:]
            for line in tail_lines:
                line = line.strip()
                if not line:
                    continue
                try:
                    obj = json.loads(line)
                except Exception:
                    continue
                eid = obj.get("eval_run_id")
                if isinstance(eid, str) and eid:
                    existing_eval_ids.add(eid)
        except Exception:
            existing_eval_ids = set()
    if eval_run_id and eval_run_id in existing_eval_ids:
        logger.info(
            "[eval][rerank_health_trend] skipped=true reason=duplicate_eval_run_id path=%s",
            trend_path,
        )
        return
    record = {
        "created_at": created_at,
        "eval_run_id": eval_run_id,
        "pred_path": dashboard.get("pred_path") or "",
        "rerank_run_id": rerank_report.get("rerank_run_id"),
        "llm_mode": rerank_report.get("llm_mode"),
        "status": rerank_report.get("status"),
        "error": rerank_report.get("error"),
        "prompts_written": counts.get("prompts_written"),
        "responses_rows": counts.get("responses_rows"),
        "decision_none_rows": counts.get("decision_none_rows"),
        "applied_selected": counts.get("applied_selected"),
        "no_response": counts.get("no_response"),
        "missing": counts.get("missing"),
        "http_fail": counts.get("http_fail") or 0,
        "parse_fail": counts.get("parse_fail") or 0,
        "decision_none_ratio": rates.get("decision_none_ratio"),
        "apply_rate_per_prompt": rates.get("apply_rate_per_prompt"),
        "response_ok_rate": rates.get("response_ok_rate"),
        "no_response_rate": rates.get("no_response_rate"),
        "missing_rate": rates.get("missing_rate"),
    }
    try:
        trend_path.parent.mkdir(parents=True, exist_ok=True)
        with trend_path.open("a", encoding="utf-8") as fp:
            fp.write(json.dumps(record, ensure_ascii=False) + "\n")
        bytes_written = trend_path.stat().st_size if trend_path.exists() else 0
        try:
            with trend_path.open("r", encoding="utf-8") as fp:
                line_count = sum(1 for _ in fp)
        except Exception:
            line_count = 0
        logger.info(
            "[eval][rerank_health_trend] appended path=%s bytes=%s line_count=%s eval_run_id=%s",
            trend_path,
            bytes_written,
            line_count,
            eval_run_id,
        )
        logger.info(
            "[eval][rerank_health_trend_tail] tail_eval_run_id=%s decision_none_ratio=%s response_ok_rate=%s apply_rate_per_prompt=%s",
            eval_run_id,
            record.get("decision_none_ratio"),
            record.get("response_ok_rate"),
            record.get("apply_rate_per_prompt"),
        )
    except Exception as exc:
        logger.info(
            "[eval][rerank_health_trend] skipped=true reason=%s path=%s",
            f"{type(exc).__name__}: {exc}",
            trend_path,
        )


def _compute_eval_core_counts(rows_path: Path) -> dict[str, Any]:
    counts: dict[str, Any] = {
        "n_eval": 0,
        "tp": 0,
        "fp": 0,
        "fn": 0,
        "fn_reason_counts": Counter(),
        "llm_none": 0,
        "decision_none_rows": 0,
        "applied_selected": 0,
    }
    for item in _iter_jsonl_lines(rows_path):
        if not isinstance(item, dict):
            continue
        if item.get("row_kind") != "gold_occurrence" or item.get("view") != "strict":
            continue
        if item.get("eval_tag") == "non_evaluable":
            continue
        counts["n_eval"] += 1
        status = str(item.get("status") or "")
        if status == "TP":
            counts["tp"] += 1
        elif status == "FP":
            counts["fp"] += 1
        elif status == "FN":
            counts["fn"] += 1
            code = str(item.get("fn_reason_code") or "UNCLASSIFIED_FN")
            counts["fn_reason_counts"][code] += 1
            if code == "LLM_NONE":
                counts["llm_none"] += 1
        decision_line = str(item.get("rerank_decision_line") or "")
        if decision_line.startswith("DECISION: NONE"):
            counts["decision_none_rows"] += 1
        selected_eid = item.get("rerank_selected_eid")
        if selected_eid is not None and str(selected_eid).strip() != "":
            counts["applied_selected"] += 1
    return counts


def _write_run_delta_summary_csv(
    *, run_context: RunContext, outputs_dir: Path, logger: logging.Logger
) -> None:
    try:
        curr_run_dir = run_context.run_dir
        eval_root = curr_run_dir.parent
        run_dirs = [p for p in eval_root.iterdir() if p.is_dir()]
        run_dirs_sorted = sorted(run_dirs, key=lambda p: p.stat().st_mtime, reverse=True)
        if len(run_dirs_sorted) < 2:
            logger.info(
                "[eval][run_delta] skipped=true reason=no_prev_eval_run curr_rows=%s prev_rows=%s",
                curr_run_dir / "outputs" / "for_users" / "eval_latest_rows.jsonl",
                "",
            )
            return
        curr_run = run_dirs_sorted[0]
        prev_run = run_dirs_sorted[1]
        curr_rows = curr_run / "outputs" / "for_users" / "eval_latest_rows.jsonl"
        prev_rows = prev_run / "outputs" / "for_users" / "eval_latest_rows.jsonl"
        if not curr_rows.exists() or not prev_rows.exists():
            logger.info(
                "[eval][run_delta] skipped=true reason=missing_rows_jsonl curr_rows=%s prev_rows=%s",
                curr_rows,
                prev_rows,
            )
            return
        curr_counts = _compute_eval_core_counts(curr_rows)
        prev_counts = _compute_eval_core_counts(prev_rows)
        out_path = outputs_dir / "for_users" / "run_delta_summary.csv"
        out_path.parent.mkdir(parents=True, exist_ok=True)
        fieldnames = ["metric", "curr", "prev", "delta", "note"]
        rows = []
        def _delta(a: Any, b: Any) -> Any:
            if isinstance(a, (int, float)) and isinstance(b, (int, float)):
                return a - b
            return ""
        n_eval_curr = curr_counts.get("n_eval") or 0
        n_eval_prev = prev_counts.get("n_eval") or 0
        tp_curr = curr_counts.get("tp") or 0
        tp_prev = prev_counts.get("tp") or 0
        fn_curr = curr_counts.get("fn") or 0
        fn_prev = prev_counts.get("fn") or 0
        rows.append({"metric": "n_eval", "curr": n_eval_curr, "prev": n_eval_prev, "delta": _delta(n_eval_curr, n_eval_prev), "note": ""})
        rows.append({"metric": "tp", "curr": tp_curr, "prev": tp_prev, "delta": _delta(tp_curr, tp_prev), "note": ""})
        rows.append({"metric": "fp", "curr": curr_counts.get("fp") or 0, "prev": prev_counts.get("fp") or 0, "delta": _delta(curr_counts.get("fp") or 0, prev_counts.get("fp") or 0), "note": ""})
        rows.append({"metric": "fn", "curr": fn_curr, "prev": fn_prev, "delta": _delta(fn_curr, fn_prev), "note": ""})
        tp_rate_curr = (float(tp_curr) / float(n_eval_curr)) if n_eval_curr > 0 else None
        tp_rate_prev = (float(tp_prev) / float(n_eval_prev)) if n_eval_prev > 0 else None
        fn_rate_curr = (float(fn_curr) / float(n_eval_curr)) if n_eval_curr > 0 else None
        fn_rate_prev = (float(fn_prev) / float(n_eval_prev)) if n_eval_prev > 0 else None
        rows.append({"metric": "tp_rate", "curr": tp_rate_curr, "prev": tp_rate_prev, "delta": _delta(tp_rate_curr, tp_rate_prev), "note": ""})
        rows.append({"metric": "fn_rate", "curr": fn_rate_curr, "prev": fn_rate_prev, "delta": _delta(fn_rate_curr, fn_rate_prev), "note": ""})
        curr_fn_reasons: Counter[str] = curr_counts.get("fn_reason_counts") or Counter()
        prev_fn_reasons: Counter[str] = prev_counts.get("fn_reason_counts") or Counter()
        top_reasons = [k for k, _ in curr_fn_reasons.most_common(10)]
        for reason in top_reasons:
            curr_val = curr_fn_reasons.get(reason, 0)
            prev_val = prev_fn_reasons.get(reason, 0)
            rows.append(
                {
                    "metric": f"fn_reason.{reason}",
                    "curr": curr_val,
                    "prev": prev_val,
                    "delta": _delta(curr_val, prev_val),
                    "note": "",
                }
            )
        if "decision_none_rows" in curr_counts:
            rows.append(
                {
                    "metric": "rerank.decision_none_rows",
                    "curr": curr_counts.get("decision_none_rows"),
                    "prev": prev_counts.get("decision_none_rows"),
                    "delta": _delta(curr_counts.get("decision_none_rows"), prev_counts.get("decision_none_rows")),
                    "note": "",
                }
            )
        if "applied_selected" in curr_counts:
            rows.append(
                {
                    "metric": "rerank.applied_selected",
                    "curr": curr_counts.get("applied_selected"),
                    "prev": prev_counts.get("applied_selected"),
                    "delta": _delta(curr_counts.get("applied_selected"), prev_counts.get("applied_selected")),
                    "note": "",
                }
            )
        with out_path.open("w", encoding="utf-8", newline="") as fp:
            writer = csv.DictWriter(fp, fieldnames=fieldnames)
            writer.writeheader()
            for row in rows:
                writer.writerow(row)
        bytes_written = out_path.stat().st_size if out_path.exists() else 0
        logger.info(
            "[eval][run_delta] wrote path=%s bytes=%s curr_run=%s prev_run=%s",
            out_path,
            bytes_written,
            curr_run.name,
            prev_run.name,
        )
        all_reasons = set(curr_fn_reasons.keys()) | set(prev_fn_reasons.keys())
        top_delta_reason = ""
        top_delta_val = 0
        for reason in all_reasons:
            delta_val = curr_fn_reasons.get(reason, 0) - prev_fn_reasons.get(reason, 0)
            if top_delta_reason == "" or abs(delta_val) > abs(top_delta_val):
                top_delta_reason = reason
                top_delta_val = delta_val
        logger.info(
            "[eval][run_delta_tpfn] tp=%s->%s delta=%s fn=%s->%s delta=%s top_fn_reason_delta=%s:%s",
            tp_prev,
            tp_curr,
            tp_curr - tp_prev,
            fn_prev,
            fn_curr,
            fn_curr - fn_prev,
            top_delta_reason,
            top_delta_val,
        )
    except Exception as exc:
        logger.info(
            "[eval][run_delta] skipped=true reason=%s curr_rows=%s prev_rows=%s",
            f"{type(exc).__name__}: {exc}",
            str((run_context.run_dir / "outputs" / "for_users" / "eval_latest_rows.jsonl")),
            "",
        )


def _write_eval_run_delta_csv(
    *, run_context: RunContext, outputs_dir: Path, logger: logging.Logger
) -> None:
    try:
        curr_run_dir = run_context.run_dir
        eval_root = curr_run_dir.parent
        run_dirs = [p for p in eval_root.iterdir() if p.is_dir()]
        run_dirs_sorted = sorted(run_dirs, key=lambda p: p.stat().st_mtime, reverse=True)
        if len(run_dirs_sorted) < 2:
            logger.info(
                "[eval][eval_run_delta] skipped=true reason=no_prev_eval_run curr_rows=%s prev_rows=%s",
                curr_run_dir / "outputs" / "for_users" / "eval_latest_rows.jsonl",
                "",
            )
            return
        curr_run = run_dirs_sorted[0]
        prev_run = run_dirs_sorted[1]
        curr_rows = curr_run / "outputs" / "for_users" / "eval_latest_rows.jsonl"
        prev_rows = prev_run / "outputs" / "for_users" / "eval_latest_rows.jsonl"
        if not curr_rows.exists() or not prev_rows.exists():
            logger.info(
                "[eval][eval_run_delta] skipped=true reason=missing_rows_jsonl curr_rows=%s prev_rows=%s",
                curr_rows,
                prev_rows,
            )
            return

        def _iter_strict_rows(path: Path) -> dict[tuple[str, str, str], dict[str, Any]]:
            rows_by_key: dict[tuple[str, str, str], dict[str, Any]] = {}
            for item in _iter_jsonl_lines(path):
                if not isinstance(item, dict):
                    continue
                if item.get("row_kind") != "gold_occurrence" or item.get("view") != "strict":
                    continue
                if item.get("eval_tag") == "non_evaluable":
                    continue
                match_key = str(item.get("match_key") or "")
                gold_eid = str(item.get("gold_eid") or "")
                gold_span_key = str(item.get("gold_span_key") or "")
                join_key = (match_key, gold_eid, gold_span_key)
                rows_by_key[join_key] = item
            return rows_by_key

        curr_map = _iter_strict_rows(curr_rows)
        prev_map = _iter_strict_rows(prev_rows)

        def _decision_line_present(row: dict[str, Any]) -> bool:
            text = str(
                row.get("rerank_decision_line")
                or row.get("decision_line")
                or ""
            )
            return bool(text.strip())

        def _decision_line_prefix(row: dict[str, Any]) -> str:
            text = str(
                row.get("rerank_decision_line")
                or row.get("decision_line")
                or ""
            ).strip()
            return text[:40]

        def _has_to_llm(row: dict[str, Any]) -> str:
            val = row.get("has_to_llm")
            if isinstance(val, bool):
                return "true" if val else "false"
            return ""

        def _evidence_list(row: dict[str, Any]) -> str:
            val = row.get("evidence_missing_fields")
            if isinstance(val, list):
                return "|".join([str(x) for x in val])
            return ""

        fieldnames = [
            "join_key",
            "match_key",
            "gold_eid",
            "gold_span_key",
            "prev_status",
            "cur_status",
            "prev_pred_eid",
            "cur_pred_eid",
            "prev_fn_reason_code",
            "cur_fn_reason_code",
            "prev_none_reason",
            "cur_none_reason",
            "prev_rerank_status",
            "cur_rerank_status",
            "prev_decision_line_present",
            "cur_decision_line_present",
            "prev_protocol_ok",
            "cur_protocol_ok",
            "prev_has_to_llm",
            "cur_has_to_llm",
            "delta_kind",
            "delta_reason_bucket",
            "prev_evidence_missing_fields",
            "cur_evidence_missing_fields",
            "prev_rerank_decision_line_prefix",
            "cur_rerank_decision_line_prefix",
            "prev_pred_span_key",
            "cur_pred_span_key",
            "prev_pred_score",
            "cur_pred_score",
        ]
        if not fieldnames:
            raise RuntimeError("eval_run_delta fieldnames missing")

        rows_out = []
        tp_drop = 0
        tp_gain = 0
        fn_new = 0
        fn_fixed = 0
        total_changed = 0
        tp_drop_keys: list[str] = []

        all_keys = set(curr_map.keys()) | set(prev_map.keys())
        for join_key in sorted(all_keys):
            prev_row = prev_map.get(join_key)
            cur_row = curr_map.get(join_key)
            if prev_row is None:
                prev_status = ""
                cur_status = str(cur_row.get("status") or "")
                delta_kind = "new_row"
            elif cur_row is None:
                prev_status = str(prev_row.get("status") or "")
                cur_status = ""
                delta_kind = "removed_row"
            else:
                prev_status = str(prev_row.get("status") or "")
                cur_status = str(cur_row.get("status") or "")
                if prev_status == "TP" and cur_status == "FN":
                    delta_kind = "tp_drop"
                elif prev_status == "FN" and cur_status == "TP":
                    delta_kind = "tp_gain"
                elif prev_status == "FN" and cur_status == "FN":
                    prev_code = str(prev_row.get("fn_reason_code") or "")
                    cur_code = str(cur_row.get("fn_reason_code") or "")
                    if prev_code != cur_code:
                        delta_kind = "fn_reason_changed"
                    else:
                        delta_kind = "unchanged"
                elif prev_status == "TP" and cur_status == "TP":
                    prev_pred_eid = str(prev_row.get("pred_eid") or "")
                    cur_pred_eid = str(cur_row.get("pred_eid") or "")
                    if prev_pred_eid != cur_pred_eid:
                        delta_kind = "tp_pred_changed"
                    else:
                        delta_kind = "unchanged"
                else:
                    delta_kind = "unchanged"

            def _bucket(prev_row: dict[str, Any] | None, cur_row: dict[str, Any] | None) -> str:
                if prev_row is None or cur_row is None:
                    return "unknown_change"
                prev_code = str(prev_row.get("fn_reason_code") or "")
                cur_code = str(cur_row.get("fn_reason_code") or "")
                if prev_code != cur_code:
                    return f"reason_code_changed:{prev_code}->{cur_code}"
                prev_rerank = str(prev_row.get("rerank_status") or "")
                cur_rerank = str(cur_row.get("rerank_status") or "")
                if prev_rerank != cur_rerank:
                    return f"rerank_status_changed:{prev_rerank}->{cur_rerank}"
                prev_none = str(prev_row.get("none_reason") or "")
                cur_none = str(cur_row.get("none_reason") or "")
                if prev_none != cur_none:
                    return f"none_reason_changed:{prev_none}->{cur_none}"
                prev_dec = _decision_line_present(prev_row)
                cur_dec = _decision_line_present(cur_row)
                if prev_dec != cur_dec:
                    return f"decision_line_presence_changed:{prev_dec}->{cur_dec}"
                prev_proto = str(prev_row.get("protocol_ok") or "")
                cur_proto = str(cur_row.get("protocol_ok") or "")
                if prev_proto != cur_proto:
                    return f"protocol_ok_changed:{prev_proto}->{cur_proto}"
                prev_pred = str(prev_row.get("pred_eid") or "")
                cur_pred = str(cur_row.get("pred_eid") or "")
                if prev_pred != cur_pred:
                    return f"pred_eid_changed:{prev_pred}->{cur_pred}"
                prev_to_llm = _has_to_llm(prev_row)
                cur_to_llm = _has_to_llm(cur_row)
                if prev_to_llm != cur_to_llm:
                    return f"to_llm_presence_changed:{prev_to_llm}->{cur_to_llm}"
                return "unknown_change"

            delta_reason_bucket = _bucket(prev_row, cur_row)
            if delta_kind != "unchanged":
                total_changed += 1
            if delta_kind == "tp_drop":
                tp_drop += 1
                tp_drop_keys.append(join_key[0])
            if delta_kind == "tp_gain":
                tp_gain += 1
            if delta_kind == "fn_new":
                fn_new += 1
            if delta_kind == "fn_fixed":
                fn_fixed += 1
            if prev_status != "FN" and cur_status == "FN":
                fn_new += 1
            if prev_status == "FN" and cur_status != "FN" and cur_status != "":
                fn_fixed += 1

            row = {
                "join_key": "|".join(join_key),
                "match_key": join_key[0],
                "gold_eid": join_key[1],
                "gold_span_key": join_key[2],
                "prev_status": prev_status,
                "cur_status": cur_status,
                "prev_pred_eid": str(prev_row.get("pred_eid") or "") if prev_row else "",
                "cur_pred_eid": str(cur_row.get("pred_eid") or "") if cur_row else "",
                "prev_fn_reason_code": str(prev_row.get("fn_reason_code") or "") if prev_row else "",
                "cur_fn_reason_code": str(cur_row.get("fn_reason_code") or "") if cur_row else "",
                "prev_none_reason": str(prev_row.get("none_reason") or "") if prev_row else "",
                "cur_none_reason": str(cur_row.get("none_reason") or "") if cur_row else "",
                "prev_rerank_status": str(prev_row.get("rerank_status") or "") if prev_row else "",
                "cur_rerank_status": str(cur_row.get("rerank_status") or "") if cur_row else "",
                "prev_decision_line_present": str(_decision_line_present(prev_row)) if prev_row else "",
                "cur_decision_line_present": str(_decision_line_present(cur_row)) if cur_row else "",
                "prev_protocol_ok": str(prev_row.get("protocol_ok") or "") if prev_row else "",
                "cur_protocol_ok": str(cur_row.get("protocol_ok") or "") if cur_row else "",
                "prev_has_to_llm": _has_to_llm(prev_row) if prev_row else "",
                "cur_has_to_llm": _has_to_llm(cur_row) if cur_row else "",
                "delta_kind": delta_kind,
                "delta_reason_bucket": delta_reason_bucket,
                "prev_evidence_missing_fields": _evidence_list(prev_row) if prev_row else "",
                "cur_evidence_missing_fields": _evidence_list(cur_row) if cur_row else "",
                "prev_rerank_decision_line_prefix": _decision_line_prefix(prev_row) if prev_row else "",
                "cur_rerank_decision_line_prefix": _decision_line_prefix(cur_row) if cur_row else "",
                "prev_pred_span_key": str(prev_row.get("pred_span_key") or "") if prev_row else "",
                "cur_pred_span_key": str(cur_row.get("pred_span_key") or "") if cur_row else "",
                "prev_pred_score": str(prev_row.get("pred_score") or "") if prev_row else "",
                "cur_pred_score": str(cur_row.get("pred_score") or "") if cur_row else "",
            }
            rows_out.append(row)

        out_path = outputs_dir / "for_users" / "eval_run_delta.csv"
        out_path.parent.mkdir(parents=True, exist_ok=True)
        with out_path.open("w", encoding="utf-8", newline="") as fp:
            writer = csv.DictWriter(fp, fieldnames=fieldnames)
            writer.writeheader()
            for row in rows_out:
                writer.writerow(row)
        bytes_written = out_path.stat().st_size if out_path.exists() else 0
        logger.info(
            "[eval][eval_run_delta] wrote path=%s bytes=%s curr_run=%s prev_run=%s",
            out_path,
            bytes_written,
            curr_run.name,
            prev_run.name,
        )
        logger.info(
            "[eval][eval_run_delta_counts] tp_drop=%s tp_gain=%s fn_new=%s fn_fixed=%s total_changed=%s",
            tp_drop,
            tp_gain,
            fn_new,
            fn_fixed,
            total_changed,
        )
        tp_drop_sample = ";".join(sorted(tp_drop_keys)[:20])
        logger.info(
            "[eval][eval_run_delta_top_tp_drop] n=%s sample=%s",
            min(20, len(tp_drop_keys)),
            tp_drop_sample,
        )
    except Exception as exc:
        logger.info(
            "[eval][eval_run_delta] skipped=true reason=%s curr_rows=%s prev_rows=%s",
            f"{type(exc).__name__}: {exc}",
            str((run_context.run_dir / "outputs" / "for_users" / "eval_latest_rows.jsonl")),
            "",
        )


def _write_eval_run_delta_reasons_csv(*, outputs_dir: Path, logger: logging.Logger) -> None:
    delta_path = outputs_dir / "for_users" / "eval_run_delta.csv"
    out_path = outputs_dir / "for_users" / "eval_run_delta_reasons.csv"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    if not delta_path.exists():
        logger.info(
            "[eval][eval_run_delta_reasons] skipped=true reason=missing_delta_csv path=%s",
            delta_path,
        )
        return
    groups: dict[tuple[str, str], dict[str, Any]] = {}
    tp_drop_rows = 0
    fn_new_rows = 0
    try:
        with delta_path.open("r", encoding="utf-8", newline="") as fp:
            reader = csv.DictReader(fp)
            for row in reader:
                if not isinstance(row, dict):
                    continue
                delta_kind = str(row.get("delta_kind") or "")
                if delta_kind not in {"tp_drop", "fn_new"}:
                    continue
                delta_reason_bucket = str(row.get("delta_reason_bucket") or "")
                match_key = str(row.get("match_key") or "")
                gold_eid = str(row.get("gold_eid") or "")
                gkey = (delta_kind, delta_reason_bucket)
                bucket = groups.setdefault(
                    gkey,
                    {
                        "n_rows": 0,
                        "match_keys": set(),
                        "gold_eids": set(),
                    },
                )
                bucket["n_rows"] += 1
                if match_key:
                    bucket["match_keys"].add(match_key)
                if gold_eid:
                    bucket["gold_eids"].add(gold_eid)
                if delta_kind == "tp_drop":
                    tp_drop_rows += 1
                elif delta_kind == "fn_new":
                    fn_new_rows += 1
    except Exception as exc:
        logger.info(
            "[eval][eval_run_delta_reasons] skipped=true reason=%s path=%s",
            f"{type(exc).__name__}: {exc}",
            delta_path,
        )
        return

    rows_out = []
    for (delta_kind, delta_reason_bucket), bucket in groups.items():
        sample_match_keys = ";".join(sorted(bucket["match_keys"])[:5])
        rows_out.append(
            {
                "delta_kind": delta_kind,
                "delta_reason_bucket": delta_reason_bucket,
                "n_rows": bucket["n_rows"],
                "n_unique_match_key": len(bucket["match_keys"]),
                "n_unique_gold_eid": len(bucket["gold_eids"]),
                "sample_match_keys": sample_match_keys,
            }
        )
    rows_out = sorted(
        rows_out,
        key=lambda r: (str(r.get("delta_kind") or ""), -int(r.get("n_rows") or 0), str(r.get("delta_reason_bucket") or "")),
    )
    fieldnames = [
        "delta_kind",
        "delta_reason_bucket",
        "n_rows",
        "n_unique_match_key",
        "n_unique_gold_eid",
        "sample_match_keys",
    ]
    with out_path.open("w", encoding="utf-8", newline="") as fp:
        writer = csv.DictWriter(fp, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows_out:
            writer.writerow(row)
    bytes_written = out_path.stat().st_size if out_path.exists() else 0
    logger.info(
        "[eval][eval_run_delta_reasons] wrote path=%s bytes=%s",
        out_path,
        bytes_written,
    )
    logger.info(
        "[eval][eval_run_delta_reasons_counts] tp_drop_rows=%s fn_new_rows=%s n_groups=%s",
        tp_drop_rows,
        fn_new_rows,
        len(rows_out),
    )
    tp_drop_top3 = [
        f"{r['delta_reason_bucket']}({r['n_rows']})"
        for r in rows_out
        if r.get("delta_kind") == "tp_drop"
    ][:3]
    fn_new_top3 = [
        f"{r['delta_reason_bucket']}({r['n_rows']})"
        for r in rows_out
        if r.get("delta_kind") == "fn_new"
    ][:3]
    logger.info(
        "[eval][eval_run_delta_reasons_top] tp_drop_top3=%s fn_new_top3=%s",
        ";".join(tp_drop_top3),
        ";".join(fn_new_top3),
    )


def _write_eval_run_delta_samples_jsonl(
    *, outputs_dir: Path, logger: logging.Logger, n_per_kind: int = 20
) -> None:
    delta_path = outputs_dir / "for_users" / "eval_run_delta.csv"
    out_path = outputs_dir / "for_users" / "eval_run_delta_samples.jsonl"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    if not delta_path.exists():
        logger.info(
            "[eval][eval_run_delta_samples] skipped=true reason=missing_delta_csv path=%s",
            delta_path,
        )
        return
    rows_by_kind: dict[str, list[dict[str, str]]] = {"tp_drop": [], "fn_new": []}
    try:
        with delta_path.open("r", encoding="utf-8", newline="") as fp:
            reader = csv.DictReader(fp)
            for row in reader:
                if not isinstance(row, dict):
                    continue
                kind = str(row.get("delta_kind") or "")
                if kind not in ("tp_drop", "fn_new"):
                    continue
                row2 = {str(k): str(v or "") for k, v in row.items()}
                rows_by_kind[kind].append(row2)
    except Exception as exc:
        logger.info(
            "[eval][eval_run_delta_samples] skipped=true reason=%s path=%s",
            f"{type(exc).__name__}: {exc}",
            delta_path,
        )
        return

    def _sort_key(r: dict[str, str]) -> tuple[str, str, str, str]:
        return (
            str(r.get("delta_reason_bucket") or ""),
            str(r.get("match_key") or ""),
            str(r.get("gold_eid") or ""),
            str(r.get("gold_span_key") or ""),
        )

    tp_drop_sorted = sorted(rows_by_kind["tp_drop"], key=_sort_key)[: max(0, int(n_per_kind))]
    fn_new_sorted = sorted(rows_by_kind["fn_new"], key=_sort_key)[: max(0, int(n_per_kind))]
    out_rows: list[dict[str, Any]] = []

    def _to_sample(kind: str, r: dict[str, str]) -> dict[str, Any]:
        match_key = str(r.get("match_key") or "")
        gold_eid = str(r.get("gold_eid") or "")
        gold_span_key = str(r.get("gold_span_key") or "")
        return {
            "sample_kind": kind,
            "delta_reason_bucket": str(r.get("delta_reason_bucket") or ""),
            "match_key": match_key,
            "gold_eid": gold_eid,
            "gold_span_key": gold_span_key,
            "prev": {
                "status": str(r.get("prev_status") or ""),
                "pred_eid": str(r.get("prev_pred_eid") or ""),
                "fn_reason_code": str(r.get("prev_fn_reason_code") or ""),
                "none_reason": str(r.get("prev_none_reason") or ""),
                "rerank_status": str(r.get("prev_rerank_status") or ""),
                "decision_line_present": str(r.get("prev_decision_line_present") or ""),
                "protocol_ok": str(r.get("prev_protocol_ok") or ""),
                "has_to_llm": str(r.get("prev_has_to_llm") or ""),
                "decision_line_prefix": str(r.get("prev_rerank_decision_line_prefix") or ""),
            },
            "cur": {
                "status": str(r.get("cur_status") or ""),
                "pred_eid": str(r.get("cur_pred_eid") or ""),
                "fn_reason_code": str(r.get("cur_fn_reason_code") or ""),
                "none_reason": str(r.get("cur_none_reason") or ""),
                "rerank_status": str(r.get("cur_rerank_status") or ""),
                "decision_line_present": str(r.get("cur_decision_line_present") or ""),
                "protocol_ok": str(r.get("cur_protocol_ok") or ""),
                "has_to_llm": str(r.get("cur_has_to_llm") or ""),
                "decision_line_prefix": str(r.get("cur_rerank_decision_line_prefix") or ""),
            },
            "join_key": f"{match_key}|{gold_eid}|{gold_span_key}",
        }

    for r in tp_drop_sorted:
        out_rows.append(_to_sample("tp_drop", r))
    for r in fn_new_sorted:
        out_rows.append(_to_sample("fn_new", r))

    with out_path.open("w", encoding="utf-8") as fp:
        if not out_rows:
            write_jsonl_line(
                fp,
                {
                    "note": "no_tp_drop_or_fn_new",
                    "tp_drop": 0,
                    "fn_new": 0,
                },
            )
        else:
            for row in out_rows:
                write_jsonl_line(fp, row)

    bytes_written = out_path.stat().st_size if out_path.exists() else 0
    logger.info(
        "[eval][eval_run_delta_samples] wrote path=%s bytes=%s",
        out_path,
        bytes_written,
    )
    logger.info(
        "[eval][eval_run_delta_samples_counts] tp_drop=%s fn_new=%s n_per_kind=%s",
        len(tp_drop_sorted),
        len(fn_new_sorted),
        n_per_kind,
    )
    tp_first = out_rows[0]["join_key"] if out_rows and out_rows[0]["sample_kind"] == "tp_drop" else ""
    fn_first = ""
    for row in out_rows:
        if row.get("sample_kind") == "fn_new":
            fn_first = str(row.get("join_key") or "")
            break
    logger.info(
        "[eval][eval_run_delta_samples_top] tp_drop_first=%s fn_new_first=%s",
        tp_first,
        fn_first,
    )


def _write_for_users_index_json(
    *, run_context: RunContext, outputs_dir: Path, logger: logging.Logger
) -> None:
    for_users_dir = outputs_dir / "for_users"
    if not for_users_dir.exists():
        logger.info(
            "[eval][for_users_index] skipped=true reason=missing_for_users_dir path=%s",
            for_users_dir,
        )
        return
    artifact_specs = {
        "eval_latest_rows_jsonl": "eval_latest_rows.jsonl",
        "run_delta_summary_csv": "run_delta_summary.csv",
        "eval_run_delta_csv": "eval_run_delta.csv",
        "eval_run_delta_reasons_csv": "eval_run_delta_reasons.csv",
        "eval_run_delta_samples_jsonl": "eval_run_delta_samples.jsonl",
        "llm_none_breakdown_json": "llm_none_breakdown.json",
        "rerank_health_dashboard_json": "rerank_health_dashboard.json",
        "rerank_health_trend_jsonl": "rerank_health_trend.jsonl",
        "polyset_ambiguity_stats_json": "polyset_ambiguity_stats.json",
        "llm_audit_json": "llm_audit.json",
    }
    artifact_note_ko: dict[str, str] = {
        "eval_latest_rows_jsonl": "이번 eval RUN의 SSOT 결과(rows_jsonl)로, 모든 판단/분석의 출발점입니다.",
        "run_delta_summary_csv": "직전 2개 eval RUN의 변화 요약(Δ)만 간단히 집계한 CSV(projection)입니다.",
        "eval_run_delta_csv": "직전 2개 eval RUN을 join해 TP/FN 변화와 증거 필드의 차이를 행 단위로 기록한 CSV입니다.",
        "eval_run_delta_reasons_csv": "tp_drop/fn_new 변화의 원인 버킷(delta_reason_bucket)을 집계한 요약 CSV입니다.",
        "eval_run_delta_samples_jsonl": "tp_drop/fn_new 변화 케이스를 결정적으로 상위 N개 샘플로 고정한 JSONL(재현용)입니다.",
        "llm_none_breakdown_json": "LLM_NONE 케이스를 none_reason별로 분해·통계한 관측 JSON입니다.",
        "rerank_health_dashboard_json": "rerank 실행의 health 지표(응답/적용/프로토콜 등)를 1회 RUN 기준으로 요약한 JSON입니다.",
        "rerank_health_trend_jsonl": "rerank health 지표를 RUN마다 1줄씩 누적 기록한 trend JSONL입니다.",
        "polyset_ambiguity_stats_json": "polyset ambiguous 관련 margin/분포를 읽기 전용으로 집계한 통계 JSON입니다.",
        "llm_audit_json": "LLM responses/decision/protocol을 조인해 감사(audit) 관점으로 점검하기 위한 JSON입니다.",
    }
    default_note = "이번 RUN의 for_users 산출물 인덱스 항목입니다."
    artifacts: dict[str, dict[str, Any]] = {}
    n_exists = 0
    missing_keys: list[str] = []
    generated_at = iso_now()
    for key, rel_name in artifact_specs.items():
        path = for_users_dir / rel_name
        exists = path.exists()
        size = path.stat().st_size if exists else 0
        if exists:
            n_exists += 1
        else:
            missing_keys.append(key)
        artifacts[key] = {
            "path": str(path),
            "exists": bool(exists),
            "bytes": int(size),
            "note": artifact_note_ko.get(key, default_note),
        }
    rows_path = for_users_dir / "eval_latest_rows.jsonl"
    n_strict = 0
    n_non_evaluable = 0
    n_eval = 0
    tp = 0
    fp = 0
    fn = 0
    if rows_path.exists() and rows_path.stat().st_size > 0:
        for item in _iter_jsonl_lines(rows_path):
            if not isinstance(item, dict):
                continue
            if item.get("row_kind") != "gold_occurrence" or item.get("view") != "strict":
                continue
            n_strict += 1
            if item.get("eval_tag") == "non_evaluable":
                n_non_evaluable += 1
                continue
            n_eval += 1
            status = str(item.get("status") or "")
            if status == "TP":
                tp += 1
            elif status == "FP":
                fp += 1
            elif status == "FN":
                fn += 1
    rerank_applied_selected = 0
    rerank_decision_none_rows = 0
    rerank_responses_rows = 0
    dashboard_path = for_users_dir / "rerank_health_dashboard.json"
    if dashboard_path.exists() and dashboard_path.stat().st_size > 0:
        try:
            dashboard = json.loads(dashboard_path.read_text(encoding="utf-8"))
            counts = dashboard.get("counts") or {}
            rerank_applied_selected = int(counts.get("applied_selected") or 0)
            rerank_decision_none_rows = int(counts.get("decision_none_rows") or 0)
            rerank_responses_rows = int(counts.get("responses_rows") or 0)
        except Exception:
            rerank_applied_selected = 0
            rerank_decision_none_rows = 0
            rerank_responses_rows = 0
    run_summary = {
        "eval_run_dir": str(run_context.run_dir),
        "generated_at": generated_at,
        "n_eval": int(n_eval),
        "tp": int(tp),
        "fp": int(fp),
        "fn": int(fn),
        "n_strict": int(n_strict),
        "n_non_evaluable": int(n_non_evaluable),
        "rerank_applied_selected": int(rerank_applied_selected),
        "rerank_decision_none_rows": int(rerank_decision_none_rows),
        "rerank_responses_rows": int(rerank_responses_rows),
        "delta_summary_is_delta": True,
        "run_summary_ko": f"이번 RUN 요약: n_eval={n_eval}, TP={tp}, FP={fp}, FN={fn} (strict 기준).",
    }
    payload = {
        "schema_version": "for_users_index_v1",
        "generated_at": generated_at,
        "eval_run_dir": str(run_context.run_dir),
        "outputs_dir": str(outputs_dir),
        "for_users_dir": str(for_users_dir),
        "run_summary": run_summary,
        "artifacts": artifacts,
    }
    out_path = for_users_dir / "INDEX.json"
    write_json(out_path, payload, indent=2)
    bytes_written = out_path.stat().st_size if out_path.exists() else 0
    logger.info(
        "[eval][for_users_index] wrote path=%s bytes=%s",
        out_path,
        bytes_written,
    )
    logger.info(
        "[eval][for_users_index_counts] n_total=%s n_exists=%s",
        len(artifact_specs),
        n_exists,
    )
    logger.info("[eval][for_users_index_notes] applied n=%s", len(artifacts))
    logger.info(
        "[eval][for_users_index_summary] n_eval=%s tp=%s fp=%s fn=%s",
        n_eval,
        tp,
        fp,
        fn,
    )
    logger.info(
        "[eval][for_users_index_missing] missing=%s",
        ";".join(missing_keys[:5]),
    )


def _freeze_for_users_artifacts(*, outputs_dir: Path, logger: logging.Logger) -> None:
    spec = FOR_USERS_FREEZE_SPEC_V1
    csv_headers = spec.get("csv_headers") or {}
    logger.info("[eval][for_users_freeze] start schema=%s", spec.get("schema_version"))
    for_users_dir = outputs_dir / "for_users"
    if not for_users_dir.exists():
        logger.info(
            "[eval][for_users_freeze] FAIL reason=%s path=%s",
            "missing_for_users_dir",
            for_users_dir,
        )
        raise RuntimeError(f"for_users freeze failed: missing_for_users_dir ({for_users_dir})")

    required_files = spec.get("required_files") or []
    for item in required_files:
        key = str(item.get("key") or "")
        relpath = str(item.get("relpath") or "")
        min_bytes = int(item.get("min_bytes") or 0)
        path = outputs_dir / relpath
        if not path.exists() and key in {
            "run_delta_summary_csv",
            "eval_run_delta_csv",
            "eval_run_delta_reasons_csv",
            "eval_run_delta_samples_jsonl",
        }:
            path.parent.mkdir(parents=True, exist_ok=True)
            if key == "run_delta_summary_csv":
                header = list((csv_headers.get("run_delta_summary_csv") or []))
                with path.open("w", encoding="utf-8", newline="") as fp:
                    writer = csv.writer(fp)
                    writer.writerow(header)
            elif key == "eval_run_delta_reasons_csv":
                header = list((csv_headers.get("eval_run_delta_reasons_csv") or []))
                with path.open("w", encoding="utf-8", newline="") as fp:
                    writer = csv.writer(fp)
                    writer.writerow(header)
            elif key == "eval_run_delta_csv":
                header = list((csv_headers.get("eval_run_delta_csv_prefix") or []))
                with path.open("w", encoding="utf-8", newline="") as fp:
                    writer = csv.writer(fp)
                    writer.writerow(header)
            elif key == "eval_run_delta_samples_jsonl":
                path.write_text('{"delta_kind":"","delta_reason_bucket":""}\n', encoding="utf-8")
            logger.info(
                "[eval][for_users_freeze] created_placeholder=true key=%s path=%s",
                key,
                path,
            )
        if not path.exists():
            logger.info("[eval][for_users_freeze] FAIL reason=%s path=%s", "missing_file", path)
            raise RuntimeError(f"for_users freeze failed: missing_file ({key}:{path})")
        size = path.stat().st_size
        if size < min_bytes:
            logger.info("[eval][for_users_freeze] FAIL reason=%s path=%s", "bytes_too_small", path)
            raise RuntimeError(
                f"for_users freeze failed: bytes_too_small ({key}:{path}, {size}<{min_bytes})"
            )

    for item in required_files:
        key = str(item.get("key") or "")
        relpath = str(item.get("relpath") or "")
        path = outputs_dir / relpath
        if key == "run_delta_summary_csv" and key in csv_headers:
            expected = list(csv_headers.get(key) or [])
            with path.open("r", encoding="utf-8", newline="") as fp:
                reader = csv.reader(fp)
                actual = next(reader, [])
            if actual != expected:
                logger.info("[eval][for_users_freeze] FAIL reason=%s path=%s", "csv_header_mismatch", path)
                logger.info("[eval][for_users_freeze] expected_header=%s", expected)
                logger.info("[eval][for_users_freeze] actual_header=%s", actual)
                raise RuntimeError(f"for_users freeze failed: csv_header_mismatch ({key})")
        if key == "eval_run_delta_reasons_csv" and key in csv_headers:
            expected = list(csv_headers.get(key) or [])
            with path.open("r", encoding="utf-8", newline="") as fp:
                reader = csv.reader(fp)
                actual = next(reader, [])
            if actual != expected:
                logger.info("[eval][for_users_freeze] FAIL reason=%s path=%s", "csv_header_mismatch", path)
                logger.info("[eval][for_users_freeze] expected_header=%s", expected)
                logger.info("[eval][for_users_freeze] actual_header=%s", actual)
                raise RuntimeError(f"for_users freeze failed: csv_header_mismatch ({key})")
        if key == "eval_run_delta_csv":
            expected_prefix = list(csv_headers.get("eval_run_delta_csv_prefix") or [])
            with path.open("r", encoding="utf-8", newline="") as fp:
                reader = csv.reader(fp)
                actual = next(reader, [])
            if not all(col in actual for col in expected_prefix):
                logger.info("[eval][for_users_freeze] FAIL reason=%s path=%s", "csv_header_mismatch", path)
                logger.info("[eval][for_users_freeze] expected_header=%s", expected_prefix)
                logger.info("[eval][for_users_freeze] actual_header=%s", actual)
                raise RuntimeError("for_users freeze failed: csv_header_mismatch (eval_run_delta_csv)")

    index_path = outputs_dir / "for_users" / "INDEX.json"
    try:
        index_obj = json.loads(index_path.read_text(encoding="utf-8"))
    except Exception as exc:
        logger.info("[eval][for_users_freeze] FAIL reason=%s path=%s", "index_parse_fail", index_path)
        raise RuntimeError(f"for_users freeze failed: index_parse_fail ({exc})")
    required_index_keys = (spec.get("json_required_top_keys") or {}).get("index_json") or []
    for key in required_index_keys:
        if key not in index_obj:
            logger.info(
                "[eval][for_users_freeze] FAIL reason=%s path=%s",
                "index_missing_key",
                index_path,
            )
            raise RuntimeError(f"for_users freeze failed: index_missing_key ({key})")
    artifacts = index_obj.get("artifacts")
    if not isinstance(artifacts, dict):
        logger.info(
            "[eval][for_users_freeze] FAIL reason=%s path=%s",
            "index_missing_key",
            index_path,
        )
        raise RuntimeError("for_users freeze failed: index_missing_key (artifacts)")
    required_artifact_keys = [
        "eval_latest_rows_jsonl",
        "run_delta_summary_csv",
        "eval_run_delta_csv",
        "eval_run_delta_reasons_csv",
        "eval_run_delta_samples_jsonl",
    ]
    for key in required_artifact_keys:
        if key not in artifacts:
            logger.info(
                "[eval][for_users_freeze] FAIL reason=%s path=%s",
                "index_missing_key",
                index_path,
            )
            raise RuntimeError(f"for_users freeze failed: index_missing_key (artifacts.{key})")
    logger.info("[eval][for_users_freeze] PASS n_required=%s", len(required_files))


def _write_rerank_health_summary(
    *,
    run_context: RunContext,
    outputs_dir: Path,
    pred_path: Path,
    logger: logging.Logger,
) -> None:
    report_path = pred_path.parent / "infer_step2_rerank_report.json"
    if not report_path.exists():
        logger.info(
            "[eval][rerank_health_summary] skipped=true reason=missing_report path=%s",
            report_path,
        )
        return
    try:
        report = json.loads(report_path.read_text(encoding="utf-8"))
    except Exception:
        report = {}
    subset_keys = [
        "created_at",
        "rerank_run_id",
        "llm_mode",
        "input_pred_path",
        "rerank_output_path",
        "prompts_written",
        "responses_rows",
        "decision_none_rows",
        "applied_selected",
        "n_records",
        "n_input_candidates",
        "n_to_llm_candidates",
        "n_records_with_to_llm",
    ]
    rerank_report = {k: report.get(k) for k in subset_keys}
    summary = {
        "created_at": iso_now(),
        "eval_run_id": run_context.run_id,
        "report_path": str(report_path),
        "rerank_report": rerank_report,
    }
    out_path = outputs_dir / "for_users" / "rerank_health_summary.json"
    write_json(out_path, summary, indent=2)
    try:
        bytes_written = out_path.stat().st_size
    except Exception:
        bytes_written = 0
    logger.info(
        "[eval][rerank_health_summary] wrote path=%s bytes=%s",
        out_path,
        bytes_written,
    )


def _load_rerank_index(index_path: Path, logger: logging.Logger) -> dict[str, Any] | None:
    exists = index_path.exists()
    size = index_path.stat().st_size if exists else 0
    logger.info(
        "[eval][rerank_index] loaded path=%s exists=%s bytes=%s",
        index_path,
        exists,
        size,
    )
    if not exists or size <= 0:
        return None
    try:
        obj = json.loads(index_path.read_text(encoding="utf-8"))
    except Exception as exc:
        logger.warning("[eval][rerank_index] invalid_json path=%s err=%s", index_path, exc)
        return None
    if not isinstance(obj, dict):
        logger.warning("[eval][rerank_index] invalid_type path=%s", index_path)
        return None
    if "schema_version" not in obj or "artifacts" not in obj:
        logger.warning("[eval][rerank_index] missing_required_keys path=%s", index_path)
        return None
    return obj


def _log_pred_guard_evidence_v2(
    logger: logging.Logger,
    *,
    pred_path: Path,
    latest_pred_path: Path | None,
    rerank_ok: bool,
    rerank_ok_source: str,
    rerank_index_path: Path | None,
    rerank_index_loaded: bool,
    rerank_index_artifacts: dict[str, Any] | None,
    report_path: Path | None,
    report_loaded: bool,
    report_keys: list[str] | None,
    report_vals: dict[str, Any] | None,
    stale: bool,
    stale_reason: str,
    latest_resolved: Path | None,
    rerank_out_resolved: Path | None,
) -> None:
    try:
        pred_res = pred_path.expanduser().resolve() if pred_path.exists() else None
    except Exception:
        pred_res = pred_path
    try:
        latest_res = (
            latest_pred_path.expanduser().resolve()
            if latest_pred_path is not None and latest_pred_path.exists()
            else None
        )
    except Exception:
        latest_res = latest_pred_path
    index_exists = False
    index_bytes = 0
    if rerank_index_path is not None:
        try:
            index_exists = rerank_index_path.exists()
            index_bytes = rerank_index_path.stat().st_size if index_exists else 0
        except Exception:
            index_exists = False
            index_bytes = 0
    report_exists = False
    report_bytes = 0
    if report_path is not None:
        try:
            report_exists = report_path.exists()
            report_bytes = report_path.stat().st_size if report_exists else 0
        except Exception:
            report_exists = False
            report_bytes = 0
    idx_output = {}
    idx_resp = {}
    idx_report = {}
    if isinstance(rerank_index_artifacts, dict):
        idx_output = rerank_index_artifacts.get("rerank_output_pred") or {}
        idx_resp = rerank_index_artifacts.get("responses_jsonl") or {}
        idx_report = rerank_index_artifacts.get("report_json") or {}
    logger.info(
        "[eval][pred_guard_v2_pred] path=%s exists=%s bytes=%s",
        pred_res if pred_res is not None else "None",
        pred_path.exists(),
        pred_path.stat().st_size if pred_path.exists() else 0,
    )
    logger.info(
        "[eval][pred_guard_v2_latest] path=%s exists=%s bytes=%s",
        latest_res if latest_res is not None else "None",
        latest_pred_path.exists() if latest_pred_path is not None else False,
        latest_pred_path.stat().st_size
        if latest_pred_path is not None and latest_pred_path.exists()
        else 0,
    )
    logger.info(
        "[eval][pred_guard_v2_rerank_ok] value=%s source=%s",
        rerank_ok,
        rerank_ok_source or "none",
    )
    logger.info(
        "[eval][pred_guard_v2_index] path=%s loaded=%s exists=%s bytes=%s",
        str(rerank_index_path) if rerank_index_path is not None else "None",
        str(bool(rerank_index_loaded)).lower(),
        index_exists,
        index_bytes,
    )
    logger.info(
        "[eval][pred_guard_v2_index_artifacts] rerank_output_pred.exists=%s bytes=%s responses_jsonl.exists=%s bytes=%s report_json.exists=%s bytes=%s",
        bool(idx_output.get("exists")) if isinstance(idx_output, dict) else False,
        int(idx_output.get("bytes") or 0) if isinstance(idx_output, dict) else 0,
        bool(idx_resp.get("exists")) if isinstance(idx_resp, dict) else False,
        int(idx_resp.get("bytes") or 0) if isinstance(idx_resp, dict) else 0,
        (
            bool(idx_report.get("exists"))
            if isinstance(idx_report, dict) and idx_report
            else "n/a"
        ),
        (
            int(idx_report.get("bytes") or 0)
            if isinstance(idx_report, dict) and idx_report
            else "n/a"
        ),
    )
    logger.info(
        "[eval][pred_guard_v2_report] path=%s loaded=%s exists=%s bytes=%s",
        str(report_path) if report_path is not None else "None",
        str(bool(report_loaded)).lower(),
        report_exists,
        report_bytes,
    )
    vals = report_vals or {}
    logger.info(
        "[eval][pred_guard_v2_report_vals] input_pred_path=%s rerank_output_path=%s output_path=%s responses_path=%s status=%s",
        vals.get("input_pred_path", None),
        vals.get("rerank_output_path", None),
        vals.get("output_path", None),
        vals.get("responses_path", None),
        vals.get("status", None),
    )
    logger.info(
        "[eval][pred_guard_v2_decision] stale=%s reason=%s pred_resolved=%s latest_resolved=%s rerank_out_resolved=%s",
        stale,
        stale_reason,
        pred_res if pred_res is not None else "None",
        latest_resolved if latest_resolved is not None else "None",
        rerank_out_resolved if rerank_out_resolved is not None else "None",
    )


def _can_use_rerank_index_fast_path(rerank_index: dict[str, Any] | None) -> tuple[bool, str]:
    if not isinstance(rerank_index, dict):
        return False, "missing_index"
    if not rerank_index.get("schema_version"):
        return False, "missing_schema_version"
    artifacts = rerank_index.get("artifacts")
    if not isinstance(artifacts, dict):
        return False, "missing_artifacts"
    responses = artifacts.get("responses_jsonl")
    if not isinstance(responses, dict):
        return False, "missing_responses"
    if not bool(responses.get("exists")) or int(responses.get("bytes") or 0) <= 0:
        return False, "missing_responses"
    rerank_output = artifacts.get("rerank_output_pred")
    if not isinstance(rerank_output, dict):
        return False, "missing_output_pred"
    if not bool(rerank_output.get("exists")) or int(rerank_output.get("bytes") or 0) <= 0:
        return False, "missing_output_pred"
    report_obj = artifacts.get("report_json")
    if isinstance(report_obj, dict):
        if (not bool(report_obj.get("exists"))) or int(report_obj.get("bytes") or 0) <= 0:
            return False, "missing_report"
    return True, "ok"


def _write_llm_audit_summary(
    *,
    run_context: RunContext,
    outputs_dir: Path,
    pred_path: Path,
    logger: logging.Logger,
) -> None:
    for_users_dir = outputs_dir / "for_users"
    rows_jsonl_path = for_users_dir / "eval_latest_rows.jsonl"
    if not rows_jsonl_path.exists():
        raise ConfigError(f"eval_latest_rows.jsonl missing: {rows_jsonl_path}")
    out_path = for_users_dir / "llm_audit.json"
    index_path = pred_path.parent / "INDEX.json"
    rerank_index = _load_rerank_index(index_path, logger)
    fast_ok, fast_reason = _can_use_rerank_index_fast_path(rerank_index)
    logger.info(
        "[eval][llm_audit_fast_path] enabled=%s reason=%s",
        "true" if fast_ok else "false",
        fast_reason,
    )
    if fast_ok:
        artifacts = (rerank_index or {}).get("artifacts") or {}
        responses_obj = artifacts.get("responses_jsonl") or {}
        responses_path_fast = str(responses_obj.get("path") or "")
        responses_bytes_fast = int(responses_obj.get("bytes") or 0)
        payload_fast = {
            "created_at": iso_now(),
            "eval_run_id": run_context.run_id,
            "mode": "fast_path",
            "skipped_heavy_join": True,
            "rerank_index_path": str(index_path),
            "responses_path": responses_path_fast,
            "responses_bytes": responses_bytes_fast,
            "note": "INDEX 기반 fast_path로 llm_audit의 대량 응답 파싱/조인을 생략했습니다.",
            "inputs": {
                "eval_rows_path": str(rows_jsonl_path),
                "responses_path": responses_path_fast,
                "report_path": "",
            },
            "counts": {
                "n_eval_llm_none": 0,
                "n_joined": 0,
                "missing_in_responses": 0,
                "by_audit_reason": {
                    "DECISION_NONE": 0,
                    "NO_RESPONSE": 0,
                    "HTTP_FAIL": 0,
                    "PROTOCOL_OR_PARSE_FAIL": 0,
                    "EMPTY_TEXT": 0,
                    "OTHER": 0,
                },
            },
            "samples": {
                "DECISION_NONE": [],
                "NO_RESPONSE": [],
                "HTTP_FAIL": [],
                "PROTOCOL_OR_PARSE_FAIL": [],
                "EMPTY_TEXT": [],
                "OTHER": [],
            },
        }
        write_json(out_path, payload_fast, indent=2)
        try:
            bytes_written = out_path.stat().st_size
        except Exception:
            bytes_written = 0
        logger.info("[eval][llm_audit_fast_path] skipped_heavy_join=true")
        logger.info(
            "[eval][llm_audit] wrote path=%s bytes=%s n_eval_llm_none=%s n_joined=%s",
            out_path,
            bytes_written,
            0,
            0,
        )
        return
    logger.info("[eval][llm_audit_fast_path] skipped_heavy_join=false")
    report_path = pred_path.parent / "infer_step2_rerank_report.json"
    responses_path = None
    if isinstance(rerank_index, dict):
        artifacts = rerank_index.get("artifacts") or {}
        if isinstance(artifacts, dict):
            resp_obj = artifacts.get("responses_jsonl") or {}
            if isinstance(resp_obj, dict):
                resp_path_raw = resp_obj.get("path")
                if resp_path_raw:
                    resp_path = Path(str(resp_path_raw))
                    if resp_path.exists() and resp_path.stat().st_size > 0:
                        responses_path = resp_path
    if report_path.exists():
        try:
            report_obj = json.loads(report_path.read_text(encoding="utf-8"))
            if responses_path is None:
                responses_path = report_obj.get("responses_path")
        except Exception:
            if responses_path is None:
                responses_path = None
    if responses_path:
        responses_path = Path(str(responses_path))
    if responses_path is None or not responses_path.exists():
        search_root = run_context.run_dir / "outputs" / "infer_step2_rerank"
        candidates = []
        if search_root.exists():
            candidates = list(search_root.glob("*responses*.jsonl"))
        if candidates:
            responses_path = max(candidates, key=lambda p: p.stat().st_mtime)
        else:
            logger.info(
                "[eval][llm_audit] skipped=true reason=missing_responses path=%s",
                responses_path,
            )
            logger.info(
                "[eval][llm_audit_join] n_eval_llm_none=%s responses_found=%s n_joined=%s reason=%s",
                0,
                False,
                0,
                "responses_missing",
            )
            return
    responses_by_match_key: dict[str, dict[str, Any]] = {}
    for resp in _iter_jsonl_lines(responses_path):
        if not isinstance(resp, dict):
            continue
        match_key = resp.get("match_key")
        if not isinstance(match_key, str) or not match_key.strip():
            continue
        responses_by_match_key[str(match_key)] = {
            "decision_line": resp.get("decision_line") or "",
            "raw_text": resp.get("raw_text") or "",
            "error": resp.get("error") or "",
        }
    by_reason = {
        "DECISION_NONE": 0,
        "NO_RESPONSE": 0,
        "HTTP_FAIL": 0,
        "PROTOCOL_OR_PARSE_FAIL": 0,
        "EMPTY_TEXT": 0,
        "OTHER": 0,
    }
    samples: dict[str, list[dict[str, Any]]] = {k: [] for k in by_reason.keys()}
    n_eval_llm_none = 0
    n_joined = 0
    missing_in_responses = 0
    for item in _iter_jsonl_lines(rows_jsonl_path):
        if not isinstance(item, dict):
            continue
        if item.get("row_kind") != "gold_occurrence" or item.get("view") != "strict":
            continue
        if item.get("eval_tag") == "non_evaluable":
            continue
        if str(item.get("fn_reason_code") or "") != "LLM_NONE":
            continue
        match_key = item.get("match_key")
        if not isinstance(match_key, str) or not match_key.strip():
            continue
        n_eval_llm_none += 1
        full = item.get("full") or {}
        record = full.get("record") or {}
        gold_item = full.get("gold_item") or {}
        protocol_ok = None
        rerank = record.get("rerank") or {}
        if isinstance(rerank, dict):
            protocol_ok = rerank.get("protocol_ok")
        resp = responses_by_match_key.get(match_key)
        audit_reason = "OTHER"
        decision_line = ""
        raw_text = ""
        error = ""
        if resp is None:
            audit_reason = "NO_RESPONSE"
            missing_in_responses += 1
        else:
            n_joined += 1
            decision_line = str(resp.get("decision_line") or "")
            raw_text = str(resp.get("raw_text") or "")
            error = str(resp.get("error") or "")
            if error != "":
                if error.startswith("HTTP"):
                    audit_reason = "HTTP_FAIL"
                else:
                    audit_reason = "PROTOCOL_OR_PARSE_FAIL"
            elif raw_text.strip() == "":
                audit_reason = "EMPTY_TEXT"
            elif decision_line.strip().startswith("DECISION: NONE"):
                audit_reason = "DECISION_NONE"
            else:
                audit_reason = "OTHER"
        by_reason[audit_reason] += 1
        if len(samples[audit_reason]) < 5:
            samples[audit_reason].append(
                {
                    "match_key": match_key,
                    "gold_eid": gold_item.get("e_id") or "",
                    "none_reason": item.get("none_reason") or "",
                    "decision_line": decision_line,
                    "error": error,
                    "protocol_ok": protocol_ok,
                    "raw_text_snippet": raw_text[:200],
                }
            )
    if n_joined == 0:
        if n_eval_llm_none == 0:
            reason = "no_llm_none_in_eval"
        else:
            reason = "match_key_mismatch"
        logger.info(
            "[eval][llm_audit_join] n_eval_llm_none=%s responses_found=%s n_joined=%s reason=%s",
            n_eval_llm_none,
            bool(responses_path and responses_path.exists()),
            n_joined,
            reason,
        )
    payload = {
        "created_at": iso_now(),
        "eval_run_id": run_context.run_id,
        "mode": "full",
        "skipped_heavy_join": False,
        "inputs": {
            "eval_rows_path": str(rows_jsonl_path),
            "responses_path": str(responses_path),
            "report_path": str(report_path),
        },
        "counts": {
            "n_eval_llm_none": n_eval_llm_none,
            "n_joined": n_joined,
            "missing_in_responses": missing_in_responses,
            "by_audit_reason": by_reason,
        },
        "samples": samples,
    }
    write_json(out_path, payload, indent=2)
    try:
        bytes_written = out_path.stat().st_size
    except Exception:
        bytes_written = 0
    logger.info(
        "[eval][llm_audit] wrote path=%s bytes=%s n_eval_llm_none=%s n_joined=%s",
        out_path,
        bytes_written,
        n_eval_llm_none,
        n_joined,
    )
    logger.info(
        "[eval][llm_audit_counts] DECISION_NONE=%s NO_RESPONSE=%s HTTP_FAIL=%s PROTOCOL_OR_PARSE_FAIL=%s EMPTY_TEXT=%s OTHER=%s",
        by_reason["DECISION_NONE"],
        by_reason["NO_RESPONSE"],
        by_reason["HTTP_FAIL"],
        by_reason["PROTOCOL_OR_PARSE_FAIL"],
        by_reason["EMPTY_TEXT"],
        by_reason["OTHER"],
    )
    sample_keys = {
        k: ",".join([s.get("match_key", "") for s in v]) for k, v in samples.items()
    }
    logger.info(
        "[eval][llm_audit_samples] DECISION_NONE=%s NO_RESPONSE=%s HTTP_FAIL=%s PROTOCOL_OR_PARSE_FAIL=%s EMPTY_TEXT=%s OTHER=%s",
        sample_keys.get("DECISION_NONE", ""),
        sample_keys.get("NO_RESPONSE", ""),
        sample_keys.get("HTTP_FAIL", ""),
        sample_keys.get("PROTOCOL_OR_PARSE_FAIL", ""),
        sample_keys.get("EMPTY_TEXT", ""),
        sample_keys.get("OTHER", ""),
    )


def _load_neg_confusable_rows_openpyxl(
    gold_xlsx: Path, *, gold_sheet_name: str = "gold"
) -> tuple[list[dict[str, Any]], int]:
    rows: list[dict[str, Any]] = []
    total_occ = 0
    try:
        from openpyxl import load_workbook

        wb = load_workbook(gold_xlsx, read_only=True, data_only=True)
        ws = wb[gold_sheet_name]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        header = [str(h).strip() for h in (header_row or [])]
        idx_map = {name: i for i, name in enumerate(header) if name}
        required = {"example_id", "gold_example_role", "span_segments", "target_sentence"}
        if not required.issubset(idx_map.keys()):
            raise ConfigError(
                "gold.xlsx missing required columns for neg_confusable: "
                f"{sorted(required - set(idx_map.keys()))}"
            )
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            role = str(row[idx_map["gold_example_role"]] or "").strip()
            if role != "neg_confusable":
                continue
            example_id = (
                row[idx_map["example_id"]] if idx_map["example_id"] < len(row) else None
            )
            if example_id is None or str(example_id).strip() == "":
                continue
            total_occ += 1
            span_raw = (
                row[idx_map["span_segments"]] if idx_map["span_segments"] < len(row) else None
            )
            target_sentence = (
                row[idx_map["target_sentence"]]
                if idx_map["target_sentence"] < len(row)
                else None
            )
            if span_raw is None or target_sentence is None:
                continue
            if not isinstance(span_raw, str):
                span_raw = str(span_raw)
            if not span_raw or span_raw.strip().lower() == "nan":
                continue
            try:
                span_segments = ast.literal_eval(span_raw)
            except Exception:
                continue
            if not span_segments:
                continue
            span_key = silver_loader._span_key_from_segments(span_segments)
            e_id = None
            conf_e_id_raw = ""
            if "conf_e_id" in idx_map and idx_map["conf_e_id"] < len(row):
                conf_e_id = row[idx_map["conf_e_id"]]
                if conf_e_id is not None:
                    conf_e_id_raw = str(conf_e_id).strip()
            if conf_e_id_raw:
                parts = [p.strip() for p in conf_e_id_raw.split(";") if p.strip()]
                if parts:
                    e_id = parts[0]
            if not e_id and "e_id" in idx_map and idx_map["e_id"] < len(row):
                eid_val = row[idx_map["e_id"]]
                if eid_val is not None:
                    e_id = str(eid_val).strip()
            if not e_id:
                continue
            rows.append(
                {
                    "e_id": str(e_id),
                    "example_id": example_id,
                    "target_sentence": str(target_sentence),
                    "span_segments": span_segments,
                    "span_key": span_key,
                    "instance_id": row[idx_map["instance_id"]]
                    if "instance_id" in idx_map and idx_map["instance_id"] < len(row)
                    else None,
                    "doc_id": row[idx_map["doc_id"]]
                    if "doc_id" in idx_map and idx_map["doc_id"] < len(row)
                    else None,
                    "sent_index": row[idx_map["sent_index"]]
                    if "sent_index" in idx_map and idx_map["sent_index"] < len(row)
                    else None,
                    "gold_example_role": "neg_confusable",
                    "conf_e_id_raw": conf_e_id_raw,
                }
            )
        wb.close()
    except Exception as exc:
        logging.getLogger("kmwe").warning(
            "[eval][neg_confusable_trace] failed to read gold.xlsx: %s", exc
        )
    return rows, total_occ


def _load_neg_target_absent_rows(
    gold_xlsx: Path,
    *,
    gold_sheet_name: str = "gold",
    allowed_splits: list[str] | None = None,
) -> list[dict[str, Any]]:
    gold_df = pd.read_excel(gold_xlsx, sheet_name=gold_sheet_name, engine="openpyxl")
    rows: list[dict[str, Any]] = []
    for _, row in gold_df.iterrows():
        if allowed_splits is not None:
            split_val = row.get("split")
            split_norm = "" if pd.isna(split_val) else str(split_val).strip().lower()
            if split_norm not in allowed_splits:
                continue
        role = str(row.get("gold_example_role", "") or "")
        if role != "neg_target_absent":
            continue
        span_segments = row.get("span_segments")
        if isinstance(span_segments, str):
            try:
                span_segments = ast.literal_eval(span_segments)
            except Exception:
                pass
        span_key = _canonical_span_key(row.get("span_key"), span_segments)
        rows.append(
            {
                "example_id": row.get("example_id"),
                "instance_id": row.get("instance_id"),
                "target_sentence": row.get("target_sentence"),
                "gold_example_role": role,
                "e_id": row.get("e_id"),
                "split": row.get("split"),
                "doc_id": row.get("doc_id"),
                "sent_index": row.get("sent_index"),
                "span_key": span_key,
            }
        )
    return rows


def _detect_gold_schema_from_rows(rows: list[dict[str, Any]], n: int = 50) -> str:
    gold = 0
    sha1 = 0
    for row in rows[:n]:
        ex = row.get("example_id")
        if ex is None:
            continue
        ex = str(ex).strip()
        if re.match(r"^g\d{4}$", ex):
            gold += 1
        elif re.match(r"^[0-9a-f]{40}$", ex):
            sha1 += 1
    if gold > sha1 and gold > 0:
        return "gold"
    if sha1 > gold and sha1 > 0:
        return "sha1"
    return "unknown"


def detect_example_id_schema(path: Path, n: int = 50) -> str:
    gold = 0
    sha1 = 0
    try:
        with Path(path).open("r", encoding="utf-8") as fp:
            for line_no, line in enumerate(fp, start=1):
                if line_no > n:
                    break
                line = line.strip()
                if not line:
                    continue
                try:
                    obj = json.loads(line)
                except Exception:
                    continue
                ex = obj.get("example_id")
                if ex is None:
                    continue
                ex = str(ex).strip()
                if re.match(r"^g\d{4}$", ex):
                    gold += 1
                elif re.match(r"^[0-9a-f]{40}$", ex):
                    sha1 += 1
    except Exception:
        return "unknown"
    if gold > sha1 and gold > 0:
        return "gold"
    if sha1 > gold and sha1 > 0:
        return "sha1"
    return "unknown"


def _infer_pred_schema_from_path(pred_path: Path, eval_cfg: dict[str, Any]) -> str:
    if eval_cfg.get("pred_schema"):
        return str(eval_cfg.get("pred_schema"))
    parts = list(pred_path.parts)
    if "infer_step2_rerank" in parts or pred_path.name == "infer_candidates.reranked.jsonl":
        return "infer_step2_rerank"
    if "infer_step1" in parts or pred_path.name == "infer_candidates.jsonl":
        return "infer_step1"
    if "build_silver" in parts or pred_path.name == "silver.jsonl":
        return "build_silver"
    return "unknown"


def _iter_pred_records_simple(pred_path: Path):
    with pred_path.open("r", encoding="utf-8") as fp:
        for line in fp:
            line = line.strip()
            if not line:
                continue
            yield json.loads(line)


def _extract_raw_candidates(record: dict[str, Any]) -> list[dict[str, Any]]:
    if record.get("silver_labels"):
        candidates = list(record.get("silver_labels") or [])
    else:
        candidates = list(record.get("candidates") or [])
    return [_normalize_pred_candidate(c) for c in candidates]


def _canonical_span_key(span_key: Any, span_segments: Any) -> str:
    if span_segments:
        try:
            return silver_loader._span_key_from_segments(span_segments)
        except Exception:
            pass
    if span_key is None:
        return ""
    return str(span_key)


def _summarize_neg_items(neg_items: list[dict[str, Any]]) -> dict[str, Any]:
    roles = Counter()
    roles_set: set[str] = set()
    span_keys: set[str] = set()
    conf_e_ids: set[str] = set()
    legacy_e_ids: set[str] = set()
    neg_pairs_map: dict[tuple[str, str], str] = {}
    for item in neg_items:
        role = str(item.get("gold_example_role") or "")
        if role:
            roles[role] += 1
            roles_set.add(role)
        span_key = _canonical_span_key(item.get("span_key"), item.get("span_segments"))
        if span_key:
            span_keys.add(span_key)
        conf_raw = str(item.get("conf_e_id_raw") or "").strip()
        if conf_raw:
            for part in conf_raw.split(";"):
                part = part.strip()
                if part:
                    conf_e_ids.add(part)
        else:
            legacy = str(item.get("e_id") or "").strip()
            if legacy:
                legacy_e_ids.add(legacy)
        eid = str(item.get("e_id") or "").strip()
        if eid and span_key:
            neg_pairs_map[(eid, span_key)] = role
    roles_summary = ";".join(f"{k}:{v}" for k, v in roles.items()) if roles else ""
    span_keys_summary = ";".join(sorted(span_keys)) if span_keys else ""
    conf_summary = ";".join(sorted(conf_e_ids)) if conf_e_ids else ""
    legacy_summary = ";".join(sorted(legacy_e_ids)) if legacy_e_ids else ""
    return {
        "neg_gold_roles": roles_summary,
        "neg_gold_span_keys": span_keys_summary,
        "neg_gold_conf_e_ids": conf_summary,
        "neg_gold_legacy_e_ids": legacy_summary,
        "neg_gold_pairs_n": len(neg_pairs_map),
        "neg_pairs_map": neg_pairs_map,
        "neg_roles_set": roles_set,
    }


def _candidate_in_neg_gold(
    role: str,
    gold_eid: str,
    gold_span_key: str,
    pred_eid: str,
    pred_span_key: str,
    default_value: Any = "",
) -> Any:
    if not role:
        return default_value
    gold_eid_set = {p.strip() for p in str(gold_eid or "").split(";") if p.strip()}
    pair_match = (
        str(pred_span_key or "") == str(gold_span_key or "")
        and str(pred_eid or "") in gold_eid_set
    )
    if role == "neg_confusable":
        return pair_match
    if role == "neg_boundary":
        return not pair_match
    return default_value


def _tf(x) -> str:
    if x is None:
        return ""
    if isinstance(x, bool):
        return "TRUE" if x else "FALSE"
    s = str(x).strip()
    if s == "":
        return ""
    sl = s.lower()
    if sl in ("true", "t", "1", "yes", "y"):
        return "TRUE"
    if sl in ("false", "f", "0", "no", "n"):
        return "FALSE"
    return ""

def _compute_joined(
    gold_items: list[dict[str, Any]],
    pred_candidates: list[dict[str, Any]],
) -> dict[str, Any]:
    gold_pairs: list[tuple[str, str]] = []
    for g in gold_items:
        ge = str(g.get("e_id") or "")
        gk = _canonical_span_key(g.get("span_key"), g.get("span_segments"))
        if ge and gk:
            gold_pairs.append((ge, gk))
    pred_pairs: set[tuple[str, str]] = set()
    for cand in pred_candidates:
        ge = str(cand.get("e_id") or "")
        gk = _canonical_span_key(cand.get("span_key"), cand.get("span_segments"))
        if ge and gk:
            pred_pairs.add((ge, gk))
    matched = [pair for pair in gold_pairs if pair in pred_pairs]
    return {
        "joined_count": len(matched),
        "matched_pairs": matched,
    }


def _build_rows_by_key(
    rows: list[dict[str, Any]],
    match_key_policy: str,
) -> dict[str, list[dict[str, Any]]]:
    rows_by_key: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        match_key = _match_key(row, match_key_policy)
        if match_key:
            rows_by_key[str(match_key)].append(row)
    return rows_by_key


def _choose_best_match_key_policy(
    *,
    gold_rows: list[dict[str, Any]],
    neg_rows: list[dict[str, Any]],
    pred_path: Path,
    current_policy: str,
    logger: logging.Logger,
) -> tuple[str, dict[str, list[dict[str, Any]]], dict[str, list[dict[str, Any]]]]:
    candidate_policies: list[str] = []
    for policy in (current_policy, "sentence_only"):
        if policy and policy not in candidate_policies:
            candidate_policies.append(policy)

    best_policy = current_policy
    best_gold_by_key = _build_rows_by_key(gold_rows, current_policy)
    best_neg_by_key = _build_rows_by_key(neg_rows, current_policy)
    best_intersection = len(set(best_gold_by_key.keys()) & _collect_pred_keys(pred_path, current_policy))
    diagnostics: list[tuple[str, int, int, int]] = [
        (current_policy, best_intersection, len(best_gold_by_key), len(_collect_pred_keys(pred_path, current_policy)))
    ]

    for policy in candidate_policies[1:]:
        gold_by_key = _build_rows_by_key(gold_rows, policy)
        neg_by_key = _build_rows_by_key(neg_rows, policy)
        pred_keys = _collect_pred_keys(pred_path, policy)
        intersection = len(set(gold_by_key.keys()) & pred_keys)
        diagnostics.append((policy, intersection, len(gold_by_key), len(pred_keys)))
        if intersection > best_intersection:
            best_policy = policy
            best_gold_by_key = gold_by_key
            best_neg_by_key = neg_by_key
            best_intersection = intersection

    logger.info(
        "[eval][match_key_probe] candidates=%s",
        [
            {
                "policy": policy,
                "intersection": intersection,
                "gold_keys": gold_keys,
                "pred_keys": pred_keys,
            }
            for policy, intersection, gold_keys, pred_keys in diagnostics
        ],
    )
    if best_policy != current_policy:
        logger.warning(
            "[eval][match_key_policy] auto_reselect=true from=%s to=%s reason=better_join current_intersection=%s best_intersection=%s",
            current_policy,
            best_policy,
            diagnostics[0][1] if diagnostics else 0,
            best_intersection,
        )
    return best_policy, best_gold_by_key, best_neg_by_key


def _find_raw_candidate_for_gold(gold_e_id: str, gold_span_key: str, record: dict[str, Any]) -> dict[str, Any] | None:
    if not gold_e_id or not gold_span_key:
        return None
    for cand in _extract_raw_candidates(record):
        if str(cand.get("e_id") or "") == gold_e_id and str(cand.get("span_key") or "") == gold_span_key:
            return cand
    return None


def _collect_pred_keys(pred_path: Path, match_key_policy: str) -> set[str]:
    keys: set[str] = set()
    for record in _iter_pred_records_simple(pred_path):
        sentence = record.get("target_sentence") or record.get("raw_sentence")
        match_key = _match_key(record, match_key_policy, fallback_sentence=sentence)
        if match_key:
            keys.add(str(match_key))
    return keys


def _count_joined_raw(
    pred_path: Path,
    gold_by_key: dict[str, list[dict[str, Any]]],
    match_key_policy: str,
) -> tuple[int, dict[str, Any] | None]:
    joined = 0
    sample: dict[str, Any] | None = None
    for record in _iter_pred_records_simple(pred_path):
        sentence = record.get("target_sentence") or record.get("raw_sentence")
        match_key = _match_key(record, match_key_policy, fallback_sentence=sentence)
        gold_items = list(gold_by_key.get(match_key, []))
        if not gold_items:
            continue
        raw_preds = _extract_raw_candidates(record)
        join_info = _compute_joined(gold_items, raw_preds)
        joined += int(join_info.get("joined_count") or 0)
        if sample is None and gold_items:
            g0 = gold_items[0]
            ge = str(g0.get("e_id") or "")
            gk = _canonical_span_key(g0.get("span_key"), g0.get("span_segments"))
            pred_exists = (ge, gk) in set(join_info.get("matched_pairs") or [])
            sample = {
                "example_id": record.get("example_id"),
                "instance_id": record.get("instance_id"),
                "e_id": ge,
                "span_key": gk,
                "pred_exists": pred_exists,
            }
    return joined, sample


def _count_pred_records_simple(pred_path: Path) -> int:
    count = 0
    for _ in _iter_pred_records_simple(pred_path):
        count += 1
    return count


def _write_eval_run_meta(
    *,
    outputs_dir: Path,
    run_context: RunContext,
    gold_xlsx: Path,
    pred_path: Path,
    gold_rows: int,
    gold_occ_rows: int,
    pred_records: int,
    gold_unique_keys: int,
    pred_unique_keys: int,
    join_intersection: int,
    joined: int,
    explicit_pred_path: bool,
    pred_source: str,
) -> None:
    git_commit = None
    try:
        manifest_path = run_context.manifest_path()
        if manifest_path.exists():
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            git_commit = manifest.get("git_commit")
    except Exception:
        git_commit = None
    run_meta = {
        "stage": run_context.stage,
        "exp_id": run_context.exp_id,
        "run_id": run_context.run_id,
        "git_commit": git_commit,
        "timestamp": iso_now(),
        "paths": {
            "gold_xlsx": str(gold_xlsx),
            "pred_path": str(pred_path),
        },
        "counters": {
            "gold_rows": gold_rows,
            "gold_occ_rows": gold_occ_rows,
            "pred_records": pred_records,
            "gold_unique_keys": gold_unique_keys,
            "pred_unique_keys": pred_unique_keys,
            "join_intersection": join_intersection,
            "joined": joined,
        },
        "explicit_pred_path": explicit_pred_path,
        "pred_source": pred_source,
    }
    write_json(outputs_dir / "run_meta.json", run_meta, indent=2)


def pick_pred_path_for_gold(
    artifacts_root: Path,
    exp_id: str,
    gold_schema: str,
    now_run_id: str | None = None,
) -> Path | None:
    candidates = _list_stage_candidates(artifacts_root, exp_id, "build_silver")
    for cand in candidates:
        if detect_example_id_schema(cand) == gold_schema:
            return cand
    return _latest_stage_output(artifacts_root, exp_id, "build_silver")


def _list_stage_candidates(artifacts_root: Path, exp_id: str, stage: str) -> list[Path]:
    root = artifacts_root / exp_id / stage
    if not root.exists():
        return []
    if stage == "build_silver":
        output_name = "silver.jsonl"
    elif stage == "infer_step2_rerank":
        output_name = "infer_candidates.reranked.jsonl"
    else:
        output_name = "infer_candidates.jsonl"
    out: list[tuple[float, Path]] = []
    for run_dir in root.iterdir():
        if not run_dir.is_dir():
            continue
        cand = run_dir / "outputs" / output_name
        if cand.exists():
            try:
                out.append((cand.stat().st_mtime, cand))
            except Exception:
                continue
    out.sort(key=lambda x: x[0], reverse=True)
    return [p for _, p in out]


def _log_pred_candidates_topk(
    artifacts_root: Path, exp_id: str, logger: logging.Logger, limit: int = 5
) -> None:
    cands = _list_stage_candidates(artifacts_root, exp_id, "build_silver")[:limit]
    if not cands:
        logger.warning("[eval] pred candidates: none under build_silver")
        return
    msg_parts = []
    for p in cands:
        msg_parts.append(f"{p} (schema={detect_example_id_schema(p)})")
    logger.info("[eval] pred candidates (top-%s, build_silver): %s", limit, msg_parts)


def _sample_pred_keys(path: Path, n: int = 3) -> list[str]:
    keys: list[str] = []
    try:
        with path.open("r", encoding="utf-8") as fp:
            for line in fp:
                line = line.strip()
                if not line:
                    continue
                rec = json.loads(line)
                ex = rec.get("example_id")
                inst = _normalize_instance_id(rec.get("instance_id"))
                if ex is None or inst is None:
                    continue
                keys.append(f"{str(ex)}#{inst}")
                if len(keys) >= n:
                    break
    except Exception:
        return []
    return keys


def _log_join_key_samples(
    gold_by_key: dict[str, list[dict[str, Any]]],
    pred_path: Path,
    match_key_policy: str,
    logger: logging.Logger,
) -> None:
    if match_key_policy not in ("example_id_instance_or_sentence", "example_id_instance_only"):
        return
    gold_keys = [str(k) for k in list(gold_by_key.keys())[:3]]
    pred_keys = _sample_pred_keys(pred_path, n=3)
    logger.info("[eval] gold_key samples=%s", gold_keys)
    logger.info("[eval] pred_key samples=%s", pred_keys)


def _compute_eval_view(
    *,
    cfg: dict[str, Any],
    run_context: RunContext,
    pred_path: Path,
    gold_rows: list[dict[str, Any]],
    gold_by_key: dict[str, list[dict[str, Any]]],
    include_hold: bool,
    view_name: str,
    pred_schema: str | None = None,
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    _ = pred_schema
    error_samples_max = int(cfg.get("eval", {}).get("error_samples_max", 100))
    target_only = bool(cfg.get("eval", {}).get("target_only", True))
    only_keys_in_gold = bool(cfg.get("eval", {}).get("only_keys_in_gold", True))
    match_key_policy = str(cfg.get("eval", {}).get("match_key_policy", "example_id_instance_or_sentence"))
    span_scope_policy = str(cfg.get("eval", {}).get("span_scope_policy", "gold_only"))

    tp = fp = fn = 0
    by_eid = defaultdict(lambda: {"tp": 0, "fp": 0, "fn": 0, "support": 0})
    error_samples: list[dict[str, Any]] = []
    n_pred_ignored_non_target = 0
    n_pred_ignored_extra_target_spans = 0
    rows_csv: list[dict[str, Any]] = []
    seen_match_keys: set[str] = set()

    for record in _iter_pred_records_simple(pred_path):
        sentence = record.get("target_sentence") or record.get("raw_sentence")
        if not sentence:
            continue

        match_key = _match_key(record, match_key_policy, fallback_sentence=sentence)
        if match_key:
            seen_match_keys.add(str(match_key))
        gold_items = list(gold_by_key.get(match_key, []))
        if only_keys_in_gold and not gold_items:
            continue
        target_eid = None
        if target_only:
            if len(gold_items) != 1:
                raise ConfigError(
                    f"partial-label eval: match_key={match_key} 에 gold 행이 {len(gold_items)}개입니다. "
                    f"(example_id#instance_id 기준으로 1개여야 합니다)"
                )
            target_eid = str(gold_items[0].get("e_id") or "")

        gold_keys = {(g["e_id"], g["span_key"]) for g in gold_items}
        gold_target_span_keys = {g["span_key"] for g in gold_items if g.get("span_key")}
        is_positive = bool(gold_target_span_keys)
        for g in gold_items:
            by_eid[g["e_id"]]["support"] += 1

        preds_all = _extract_pred_candidates(record, include_hold=include_hold)
        if target_only and target_eid:
            preds = [p for p in preds_all if str(p.get("e_id") or "") == target_eid]
            ignored = len(preds_all) - len(preds)
        else:
            preds = preds_all
            ignored = 0
        n_pred_ignored_non_target += ignored

        used_keys: set[tuple[str, str]] = set()
        tp_i = fp_i = fn_i = 0
        for pred in preds:
            cand_key = (str(pred.get("e_id", "")), pred.get("span_key", ""))
            if cand_key in gold_keys and cand_key not in used_keys:
                tp += 1
                tp_i += 1
                used_keys.add(cand_key)
                by_eid[cand_key[0]]["tp"] += 1
            elif (
                span_scope_policy == "gold_only"
                and target_eid
                and cand_key[0] == target_eid
                and cand_key[1] not in gold_target_span_keys
                and is_positive
            ):
                n_pred_ignored_extra_target_spans += 1
            else:
                fp += 1
                fp_i += 1
                by_eid[cand_key[0]]["fp"] += 1
                if len(error_samples) < error_samples_max:
                    error_samples.append(
                        {
                            "type": "FP",
                            "e_id": cand_key[0],
                            "target_sentence": sentence,
                            "match_key": match_key,
                            "cand_key": cand_key,
                            "target_eid": target_eid,
                            "n_ignored_non_target_preds": ignored,
                            "span_key": cand_key[1],
                            "span_segments": pred.get("span_segments"),
                            "view": view_name,
                        }
                    )

        for g in gold_items:
            gkey = (g["e_id"], g["span_key"])
            if gkey not in used_keys:
                fn += 1
                fn_i += 1
                by_eid[g["e_id"]]["fn"] += 1
                if len(error_samples) < error_samples_max:
                    error_samples.append(
                        {
                            "type": "FN",
                            "e_id": g["e_id"],
                            "target_sentence": sentence,
                            "match_key": match_key,
                            "cand_key": gkey,
                            "target_eid": target_eid,
                            "n_ignored_non_target_preds": ignored,
                            "span_key": g["span_key"],
                            "span_segments": g["span_segments"],
                            "view": view_name,
                        }
                    )

        gold_item = gold_items[0] if gold_items else {}
        gold_span_segments = gold_item.get("span_segments")
        gold_span_text = _span_text_from_segments(sentence, gold_span_segments)
        pred_target_span_keys = ";".join([p.get("span_key", "") for p in preds if p.get("span_key")])
        pred_target_span_texts = ";".join(
            [
                _span_text_from_segments(sentence, p.get("span_segments"))
                for p in preds
                if p.get("span_segments")
            ]
        )
        rows_csv.append(
            {
                "match_key": match_key,
                "doc_id": record.get("doc_id"),
                "sent_index": record.get("sent_index"),
                "example_id": record.get("example_id"),
                "instance_id": record.get("instance_id"),
                "target_sentence": sentence,
                "gold_eid": gold_item.get("e_id"),
                "gold_span_key": gold_item.get("span_key"),
                "gold_span_segments": json.dumps(gold_span_segments, ensure_ascii=False)
                if gold_span_segments is not None
                else "",
                "gold_span_text": gold_span_text,
                "pred_target_candidates_json": json.dumps(preds, ensure_ascii=False),
                "pred_target_span_keys": pred_target_span_keys,
                "pred_target_span_texts": pred_target_span_texts,
                "n_pred_all": len(preds_all),
                "n_pred_target": len(preds),
                "n_ignored_non_target": ignored,
                "n_tp": tp_i,
                "n_fp": fp_i,
                "n_fn": fn_i,
                "status": "TP" if tp_i > 0 else ("FP" if fp_i > 0 else "FN"),
                "status_detail": f"tp={tp_i};fp={fp_i};fn={fn_i}",
                "debug_detect_json": json.dumps((record.get("debug") or {}).get("detect"), ensure_ascii=False),
                "debug_verify_json": json.dumps((record.get("debug") or {}).get("verify"), ensure_ascii=False),
                "debug_context_json": json.dumps((record.get("debug") or {}).get("context"), ensure_ascii=False),
            }
        )

    neg_confusable_rows = [
        row
        for row in gold_rows
        if str(row.get("gold_example_role") or "").strip() == "neg_confusable"
    ]
    if neg_confusable_rows:
        for g in neg_confusable_rows:
            match_key = _match_key(g, match_key_policy)
            if not match_key or str(match_key) in seen_match_keys:
                continue
            eid = str(g.get("e_id") or "")
            if not eid:
                continue
            by_eid[eid]["support"] += 1
            by_eid[eid]["fn"] += 1
            fn += 1
            if len(error_samples) < error_samples_max:
                error_samples.append(
                    {
                        "type": "FN",
                        "e_id": eid,
                        "target_sentence": g.get("target_sentence"),
                        "match_key": match_key,
                        "cand_key": (eid, g.get("span_key")),
                        "target_eid": eid,
                        "n_ignored_non_target_preds": 0,
                        "span_key": g.get("span_key"),
                        "span_segments": g.get("span_segments"),
                        "view": view_name,
                    }
                )
            gold_span_segments = g.get("span_segments")
            gold_span_text = _span_text_from_segments(g.get("target_sentence"), gold_span_segments)
            rows_csv.append(
                {
                    "match_key": match_key,
                    "doc_id": g.get("doc_id"),
                    "sent_index": g.get("sent_index"),
                    "example_id": g.get("example_id"),
                    "instance_id": g.get("instance_id"),
                    "target_sentence": g.get("target_sentence"),
                    "gold_eid": eid,
                    "gold_span_key": g.get("span_key"),
                    "gold_span_segments": json.dumps(gold_span_segments, ensure_ascii=False)
                    if gold_span_segments is not None
                    else "",
                    "gold_span_text": gold_span_text,
                    "pred_target_candidates_json": "[]",
                    "pred_target_span_keys": "",
                    "pred_target_span_texts": "",
                    "n_pred_all": 0,
                    "n_pred_target": 0,
                    "n_ignored_non_target": 0,
                    "n_tp": 0,
                    "n_fp": 0,
                    "n_fn": 1,
                    "status": "FN",
                    "status_detail": "tp=0;fp=0;fn=1;reason=missing_pred_record",
                    "debug_detect_json": "{}",
                    "debug_verify_json": "{}",
                    "debug_context_json": "{}",
                }
            )

    precision = tp / (tp + fp) if tp + fp > 0 else 0.0
    recall = tp / (tp + fn) if tp + fn > 0 else 0.0
    f1 = (2 * precision * recall) / (precision + recall) if precision + recall > 0 else 0.0

    topn = int(cfg.get("eval", {}).get("report_topn_eid", 20))
    by_eid_sorted = sorted(
        [
            {
                "e_id": eid,
                "support": stats["support"],
                "tp": stats["tp"],
                "fp": stats["fp"],
                "fn": stats["fn"],
                "precision": stats["tp"] / (stats["tp"] + stats["fp"])
                if stats["tp"] + stats["fp"] > 0
                else 0.0,
                "recall": stats["tp"] / (stats["tp"] + stats["fn"]) if stats["tp"] + stats["fn"] > 0 else 0.0,
            }
            for eid, stats in by_eid.items()
        ],
        key=lambda item: (item["support"], item["e_id"]),
        reverse=True,
    )[:topn]

    report = {
        "overall": {"tp": tp, "fp": fp, "fn": fn, "precision": precision, "recall": recall, "f1": f1},
        "by_eid": by_eid_sorted,
        "counters": {
            "n_pred_ignored_non_target": n_pred_ignored_non_target,
            "n_pred_ignored_extra_target_spans": n_pred_ignored_extra_target_spans,
        },
        "policies": {
            "target_only": target_only,
            "only_keys_in_gold": only_keys_in_gold,
            "match_key_policy": match_key_policy,
            "include_hold": include_hold,
            "span_scope_policy": span_scope_policy,
        },
        "_rows_csv": rows_csv,
    }
    return report, error_samples


def _match_key(record: dict[str, Any], policy: str, *, fallback_sentence: str | None = None) -> str:
    example_id = record.get("example_id")
    instance_id = _normalize_instance_id(record.get("instance_id"))
    doc_id = record.get("doc_id")
    sent_index = _normalize_instance_id(record.get("sent_index"))
    sentence = record.get("target_sentence") or fallback_sentence

    if policy in ("example_id_instance_or_sentence", "example_id_instance_only"):
        if example_id is not None and instance_id is not None:
            return f"{str(example_id)}#{instance_id}"
        if example_id is not None and instance_id is None:
            raise ValueError(
                f"missing instance_id for example_id under policy={policy}: {record}"
            )
        if policy == "example_id_instance_only":
            raise ValueError(
                f"missing example_id/instance_id under policy={policy}: {record}"
            )
        if doc_id is not None and sent_index is not None:
            return f"doc:{doc_id}#sent:{sent_index}"
        if example_id is not None:
            return str(example_id)
        return str(sentence or "")
    if policy == "example_id_or_sentence":
        if example_id is not None:
            return str(example_id)
        if doc_id is not None and sent_index is not None:
            return f"doc:{doc_id}#sent:{sent_index}"
        return str(sentence or "")
    if policy == "sentence_only":
        return str(sentence or "")
    return str(sentence or "")


def _normalize_instance_id(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, float) and pd.isna(value):
        return None
    if isinstance(value, int):
        return value
    text = str(value).strip()
    if text == "":
        return None
    try:
        return int(text)
    except Exception as exc:
        raise ValueError(f"instance_id/sent_index must be int-compatible: {value}") from exc


def _span_text_from_segments(sentence: str, span_segments: Any) -> str:
    if not sentence or not span_segments:
        return ""
    try:
        parts: list[str] = []
        for start, end in span_segments:
            if not isinstance(start, int) or not isinstance(end, int):
                continue
            if start < 0 or end > len(sentence) or start >= end:
                continue
            parts.append(sentence[start:end])
        if not parts:
            return ""
        return "⟪" + "⟫⟪".join(parts) + "⟫"
    except Exception:
        return ""


def _write_eval_latest_csv(path: Path, rows: list[dict[str, Any]], overall: dict[str, Any]) -> None:
    fieldnames = [
        "match_key",
        "doc_id",
        "sent_index",
        "example_id",
        "instance_id",
        "target_sentence",
        "gold_eid",
        "gold_span_key",
        "gold_span_segments",
        "gold_span_text",
        "pred_target_candidates_json",
        "pred_target_span_keys",
        "pred_target_span_texts",
        "n_pred_all",
        "n_pred_target",
        "n_ignored_non_target",
        "n_tp",
        "n_fp",
        "n_fn",
        "status",
        "status_detail",
        "eval_tag",
        "eval_tag_reason",
        "debug_detect_json",
        "debug_verify_json",
        "debug_context_json",
        "overall_precision",
        "overall_recall",
        "overall_f1",
        "overall_tp",
        "overall_fp",
        "overall_fn",
    ]
    with path.open("w", encoding="utf-8", newline="") as fp:
        writer = csv.DictWriter(fp, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
        writer.writerow(
            {
                "match_key": "__OVERALL__",
                "overall_precision": overall.get("precision"),
                "overall_recall": overall.get("recall"),
                "overall_f1": overall.get("f1"),
                "overall_tp": overall.get("tp"),
                "overall_fp": overall.get("fp"),
                "overall_fn": overall.get("fn"),
            }
        )


def _summarize_candidate(cand: dict[str, Any], sentence: str) -> dict[str, Any]:
    return {
        "e_id": cand.get("e_id"),
        "span_key": cand.get("span_key", ""),
        "span_segments": cand.get("span_segments"),
        "span_text": _span_text_from_segments(sentence, cand.get("span_segments")),
        "score": cand.get("score"),
        "triage": cand.get("triage"),
        "hard_fail_triggered": cand.get("hard_fail_triggered"),
        "hard_fail_reasons": cand.get("hard_fail_reasons"),
        "stage_hits": cand.get("stage_hits"),
    }


def _export_eval_for_users(
    *,
    cfg: dict[str, Any],
    run_context: RunContext,
    outputs_dir: Path,
    pred_path: Path,
    gold_rows: list[dict[str, Any]],
    nta_rows: list[dict[str, Any]] | None = None,
    gold_by_key: dict[str, list[dict[str, Any]]],
    neg_by_key: dict[str, list[dict[str, Any]]],
    views: dict[str, dict[str, Any]],
    report_views: list[str],
    export_variant: str | None = None,
) -> dict[str, int]:
    import logging
    logger = logging.getLogger("kmwe")
    eval_cfg = cfg.get("eval", {}) or {}
    if not bool(eval_cfg.get("export_for_users", True)):
        return {"n_gold_occurrence_rows": 0, "n_pred_ignored_rows": 0}
    variant = export_variant or "llm"
    write_full_csv = variant != "encoder"
    export_path = str(eval_cfg.get("export_for_users_path", "for_users/eval_latest.csv"))
    csv_path = Path(export_path)
    if not csv_path.is_absolute():
        csv_path = outputs_dir / csv_path
    assert_under_dir(csv_path, outputs_dir, what="eval for_users csv")
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    rows_jsonl_path = csv_path.with_name("eval_latest_rows.jsonl")
    readme_path = csv_path.with_name("eval_latest_README.md")
    assert_under_dir(rows_jsonl_path, outputs_dir, what="eval for_users rows jsonl")
    assert_under_dir(readme_path, outputs_dir, what="eval for_users readme")
    include_ignored = bool(eval_cfg.get("export_for_users_include_ignored", True))
    # SSOT: for_users에는 pred_ignored도 반드시 포함
    include_ignored = True
    write_rows_jsonl = bool(eval_cfg.get("export_for_users_write_rows_jsonl", True))
    write_readme = bool(eval_cfg.get("export_for_users_write_readme", True))
    if not write_full_csv:
        write_rows_jsonl = False
        write_readme = False
    max_json_chars = int(eval_cfg.get("export_for_users_max_json_cell_chars", 5000))
    match_key_policy = str(eval_cfg.get("match_key_policy", "example_id_instance_or_sentence"))
    target_only = bool(eval_cfg.get("target_only", True))
    only_keys_in_gold = bool(eval_cfg.get("only_keys_in_gold", True))
    span_scope_policy = str(eval_cfg.get("span_scope_policy", "gold_only"))

    neg_confusable_example_ids: list[str] = []
    neg_confusable_total = 0
    neg_confusable_missing_example_id = 0
    gold_total_rows = 0
    try:
        from openpyxl import load_workbook

        gold_xlsx_path = Path(cfg.get("paths", {}).get("gold_xlsx") or "")
        wb = load_workbook(gold_xlsx_path, read_only=True, data_only=True)
        ws = wb["gold"]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        header = [str(h).strip() for h in (header_row or [])]
        idx_map = {name: i for i, name in enumerate(header) if name}
        if "example_id" not in idx_map or "gold_example_role" not in idx_map:
            raise ConfigError("gold.xlsx missing required columns: example_id, gold_example_role")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            gold_total_rows += 1
            role = str(row[idx_map["gold_example_role"]] or "").strip()
            if role != "neg_confusable":
                continue
            ex = row[idx_map["example_id"]] if idx_map["example_id"] < len(row) else None
            if ex is None or str(ex).strip() == "":
                neg_confusable_missing_example_id += 1
                continue
            neg_confusable_total += 1
            neg_confusable_example_ids.append(str(ex))
        wb.close()
    except Exception as exc:
        logger.warning("[eval][neg_confusable_trace] failed to read gold.xlsx: %s", exc)
    logger.info(
        "[eval][neg_confusable_trace] gold_total=%s gold_occ=%s",
        gold_total_rows,
        neg_confusable_total,
    )

    rows, rows_jsonl, row_counters = _build_eval_export_rows(
        cfg=cfg,
        pred_path=pred_path,
        gold_by_key=gold_by_key,
        neg_by_key=neg_by_key,
        views=views,
        report_views=report_views,
        match_key_policy=match_key_policy,
        target_only=target_only,
        only_keys_in_gold=only_keys_in_gold,
        span_scope_policy=span_scope_policy,
        include_ignored=include_ignored,
        max_json_chars=max_json_chars,
    )
    # === SSOT ordering for for_users export ===
    # evaluable -> pred_ignored -> non_evaluable(neg_target_absent)
    evaluable_rows: list[dict[str, Any]] = []
    pred_ignored_rows: list[dict[str, Any]] = []
    non_eval_rows: list[dict[str, Any]] = []
    for r in rows:
        role = r.get("gold_example_role") or (r.get("full") or {}).get("record", {}).get(
            "gold_example_role"
        )
        rk = r.get("row_kind")
        if role == "neg_target_absent":
            non_eval_rows.append(r)
        elif rk == "pred_ignored":
            pred_ignored_rows.append(r)
        else:
            evaluable_rows.append(r)
    rows = evaluable_rows + pred_ignored_rows + non_eval_rows
    logger.info(
        "[eval][for_users_ordering] evaluable=%d pred_ignored=%d non_eval=%d",
        len(evaluable_rows),
        len(pred_ignored_rows),
        len(non_eval_rows),
    )
    pred_ignored_counts = Counter()
    pred_ignored_rows = 0
    for row in rows_jsonl:
        if row.get("row_kind") != "pred_ignored":
            continue
        pred_ignored_rows += 1
        full = row.get("full") or {}
        reason = full.get("ignored_reason")
        if reason:
            pred_ignored_counts[str(reason)] += 1
    row_counters["pred_ignored_rows"] = pred_ignored_rows
    row_counters["pred_ignored_reason_counts"] = pred_ignored_counts

    fn_reason_counts = Counter()
    none_reason_counts = Counter()
    none_reason_examples = []
    n_fn_total = 0
    rerank_status_counts = Counter()
    n_decision_none = 0
    n_selected_eid_null = 0
    for row in rows:
        if row.get("row_kind") != "gold_occurrence" or row.get("view") != "strict":
            continue
        rerank_status = row.get("rerank_status") or ""
        if rerank_status:
            rerank_status_counts[str(rerank_status)] += 1
        decision_line = str(row.get("rerank_decision_line") or "")
        if decision_line.startswith("DECISION: NONE"):
            n_decision_none += 1
        selected_eid = row.get("rerank_selected_eid")
        if selected_eid is None or str(selected_eid).strip() == "":
            n_selected_eid_null += 1
        role = str(row.get("gold_example_role") or "")
        if row.get("status") == "FN" and role != "neg_target_absent":
            n_fn_total += 1
            code = row.get("fn_reason_code") or "UNCLASSIFIED_FN"
            fn_reason_counts[str(code)] += 1
            if code == "LLM_NONE":
                nr = row.get("none_reason") or "UNKNOWN"
                none_reason_counts[str(nr)] += 1
                if len(none_reason_examples) < 5:
                    none_reason_examples.append(
                        {
                            "reason": str(nr),
                            "match_key": str(row.get("match_key") or ""),
                            "rerank_status": str(row.get("rerank_status") or ""),
                            "decision_line": str(row.get("rerank_decision_line") or ""),
                        }
                    )
    logger.info(
        "[eval][fn_reason_bucket] n_fn_total=%s counts=%s",
        n_fn_total,
        dict(fn_reason_counts),
    )
    logger.info("[eval][none_reason_counts] %s", dict(none_reason_counts))
    if none_reason_examples:
        logger.info(
            "[eval][none_reason_examples] %s",
            "; ".join(
                (
                    f"reason={ex['reason']} match_key={ex['match_key']} rerank_status={ex['rerank_status']} decision_line={ex['decision_line']}"
                )
                for ex in none_reason_examples
            ),
        )
    else:
        logger.info("[eval][none_reason_examples] (none)")
    logger.info(
        "[eval][none_reason_bucket] n_fn_llm_none=%s counts=%s",
        fn_reason_counts.get("LLM_NONE", 0),
        dict(none_reason_counts),
    )
    logger.info(
        "[eval][rerank_health] status_counts=%s n_decision_none=%s n_selected_eid_null=%s",
        dict(rerank_status_counts),
        n_decision_none,
        n_selected_eid_null,
    )

    neg_confusable_set = set(neg_confusable_example_ids)
    pred_example_ids: set[str] = set()
    filtered_only_keys_example_ids: set[str] = set()
    filtered_triage_example_ids: set[str] = set()
    try:
        with pred_path.open("r", encoding="utf-8") as fp_pred:
            for line in fp_pred:
                line = line.strip()
                if not line:
                    continue
                try:
                    record = json.loads(line)
                except Exception:
                    continue
                example_id = record.get("example_id")
                if example_id is None and isinstance(record.get("meta"), dict):
                    example_id = record["meta"].get("example_id")
                if example_id is None:
                    continue
                example_id = str(example_id)
                if example_id not in neg_confusable_set:
                    continue
                pred_example_ids.add(example_id)
                sentence = record.get("target_sentence") or record.get("raw_sentence") or ""
                match_key = _match_key(record, match_key_policy, fallback_sentence=sentence)
                if only_keys_in_gold and not gold_by_key.get(match_key, []):
                    filtered_only_keys_example_ids.add(example_id)
                    continue
                rerank = record.get("rerank") or {}
                rerank_status = rerank.get("status")
                if rerank_status in ("applied", "fallback") and record.get("candidates"):
                    raw_candidates = list(record.get("candidates") or [])
                elif record.get("silver_labels"):
                    raw_candidates = list(record.get("silver_labels") or [])
                else:
                    raw_candidates = list(record.get("candidates") or [])
                if raw_candidates:
                    triage_candidates = _extract_pred_candidates(record, include_hold=False)
                    if not triage_candidates:
                        filtered_triage_example_ids.add(example_id)
    except Exception as exc:
        logger.warning("[eval][neg_confusable_trace] failed to scan pred_path: %s", exc)

    dropped_gold_occ = neg_confusable_missing_example_id
    joined_count = sum(1 for ex in neg_confusable_example_ids if ex in pred_example_ids)
    dropped_join = len(neg_confusable_example_ids) - joined_count
    dropped_only_keys = sum(1 for ex in neg_confusable_example_ids if ex in filtered_only_keys_example_ids)
    dropped_triage = sum(1 for ex in neg_confusable_example_ids if ex in filtered_triage_example_ids)
    logger.info(
        "[eval][neg_confusable_trace] dropped_at=gold_occ count=%s",
        dropped_gold_occ,
    )
    logger.info(
        "[eval][neg_confusable_trace] dropped_at=join count=%s",
        dropped_join,
    )
    logger.info(
        "[eval][neg_confusable_trace] dropped_at=filter_only_keys count=%s",
        dropped_only_keys,
    )
    logger.info(
        "[eval][neg_confusable_trace] dropped_at=filter_triage count=%s",
        dropped_triage,
    )

    existing_neg_confusable_example_ids: set[str] = set()
    for row in rows_jsonl:
        if row.get("row_kind") != "gold_occurrence" or row.get("view") != "strict":
            continue
        full = row.get("full") or {}
        gold_item = full.get("gold_item") or {}
        if str(gold_item.get("gold_example_role") or "").strip() != "neg_confusable":
            continue
        ex = gold_item.get("example_id")
        if ex is not None:
            existing_neg_confusable_example_ids.add(str(ex))

    included_in_rows = sum(
        1 for ex in neg_confusable_example_ids if ex in existing_neg_confusable_example_ids
    )
    included_example_ids_min: set[str] = set()
    for row in rows:
        if row.get("row_kind") != "gold_occurrence" or row.get("view") != "strict":
            continue
        if str(row.get("gold_example_role") or "").strip() == "neg_confusable":
            ex = row.get("example_id")
            if ex is not None:
                included_example_ids_min.add(str(ex))
    included_in_min = sum(1 for ex in neg_confusable_example_ids if ex in included_example_ids_min)
    after_join = len(neg_confusable_example_ids) - dropped_join
    after_only_keys = max(0, after_join - dropped_only_keys)
    after_triage = max(0, after_only_keys - dropped_triage)
    dropped_projection = max(0, after_triage - included_in_rows)
    logger.info(
        "[eval][neg_confusable_trace] dropped_at=projection count=%s",
        dropped_projection,
    )
    logger.info(
        "[eval][neg_confusable_trace] included_in_rows=%s included_in_min=%s",
        included_in_rows,
        included_in_min,
    )

    neg_confusable_eval_rows = [
        row
        for row in rows
        if row.get("row_kind") == "gold_occurrence"
        and row.get("view") == "strict"
        and str(row.get("gold_example_role") or "").strip() == "neg_confusable"
    ]
    fn_count_in_metrics = sum(1 for row in neg_confusable_eval_rows if row.get("status") == "FN")
    tp_count_in_metrics = sum(1 for row in neg_confusable_eval_rows if row.get("status") == "TP")
    fp_count_in_metrics = sum(1 for row in neg_confusable_eval_rows if row.get("status") == "FP")
    appended_gold_only_eval_rows = int(row_counters.get("neg_confusable_appended_eval_rows") or 0)
    logger.info(
        "[eval][neg_confusable_eval] gold_occ=%s included_in_eval_rows=%s appended_gold_only_eval_rows=%s",
        len(neg_confusable_example_ids),
        len(neg_confusable_eval_rows),
        appended_gold_only_eval_rows,
    )
    logger.info(
        "[eval][neg_confusable_eval] fn_count_in_metrics=%s tp_count_in_metrics=%s fp_count_in_metrics=%s",
        fn_count_in_metrics,
        tp_count_in_metrics,
        fp_count_in_metrics,
    )

    fieldnames = [
        "row_kind",
        "view",
        "match_key",
        "match_key_raw",
        "example_id",
        "example_id_full",
        "instance_id",
        "doc_id",
        "sent_index",
        "target_sentence",
        "gold_example_role",
        "gold_eid",
        "gold_span_key",
        "gold_span_segments",
        "gold_span_text",
        "pred_eid",
        "no_pred_reason",
        "pred_span_key",
        "pred_span_segments",
        "pred_span_text",
        "pred_score",
        "pred_triage",
        "pred_ignored_reason",
        "pred_ignored_reason_detail",
        "neg_gold_roles",
        "neg_gold_span_keys",
        "neg_gold_conf_e_ids",
        "neg_gold_legacy_e_ids",
        "neg_gold_pairs_n",
        "candidate_e_id",
        "candidate_span_key",
        "candidate_e_ids_all",
        "candidate_span_keys_all",
        "candidate_encoder_ranks_all",
        "candidate_confidences_all",
        "candidates_json",
        "candidate_pair",
        "candidate_in_neg_gold",
        "pred_hard_fail_triggered",
        "pred_hard_fail_reasons",
        "pred_stage_hits_json",
        "pred_target_candidates_json",
        "pred_target_span_keys",
        "pred_target_span_texts",
        "n_pred_all",
        "n_pred_target",
        "n_ignored_non_target",
        "n_ignored_out_of_scope",
        "n_ignored_by_triage",
        "n_tp",
        "n_fp",
        "n_fn",
        "status",
        "status_detail",
        "eval_tag",
        "eval_tag_reason",
        "debug_detect_json",
        "debug_verify_json",
        "debug_context_json",
        "overall_precision",
        "overall_recall",
        "overall_f1",
        "overall_tp",
        "overall_fp",
        "overall_fn",
        "morph_info",
        "bridge_info",
        "thing_bridge_info",
        "detect_components_info",
        "policies_json",
    ]
    required_for_users_cols = [
        "rerank_status",
        "rerank_decision_line",
        "rerank_selected_eid",
        "fn_reason_code",
        "none_reason",
        "decision_line",
        "protocol_ok",
    ]
    for col in required_for_users_cols:
        if col not in fieldnames:
            fieldnames.append(col)

    nta_gold_rows = nta_rows or []
    gold_neg_target_absent_occ = len(nta_gold_rows)
    gold_only_rows: list[dict[str, Any]] = []
    gold_only_rows_jsonl: list[dict[str, Any]] = []
    for idx, gold_item in enumerate(nta_gold_rows, start=1):
        example_id = gold_item.get("example_id")
        instance_id = gold_item.get("instance_id")
        match_key = ""
        if example_id is not None and instance_id is not None:
            match_key = f"{example_id}#{instance_id}"
        row = {
            "row_kind": "gold_only",
            "view": "strict",
            "match_key": match_key,
            "example_id": example_id if example_id is not None else "",
            "instance_id": instance_id if instance_id is not None else "",
            "doc_id": gold_item.get("doc_id") or "",
            "sent_index": gold_item.get("sent_index") or "",
            "target_sentence": gold_item.get("target_sentence") or "",
            "gold_example_role": "neg_target_absent",
            "gold_eid": gold_item.get("e_id") or "",
            "gold_span_key": gold_item.get("span_key") or "",
            "gold_span_segments": gold_item.get("span_segments") or "",
            "gold_span_text": "",
            "pred_eid": "",
            "pred_span_key": "",
            "pred_span_segments": "",
            "pred_span_text": "",
            "pred_score": "",
            "pred_triage": "",
            "pred_ignored_reason": "",
            "pred_ignored_reason_detail": "",
            "neg_gold_roles": "",
            "neg_gold_span_keys": "",
            "neg_gold_conf_e_ids": "",
            "neg_gold_legacy_e_ids": "",
            "neg_gold_pairs_n": "",
            "candidate_e_id": "",
            "candidate_span_key": "",
            "candidate_pair": "",
            "candidate_in_neg_gold": "",
            "pred_hard_fail_triggered": "",
            "pred_hard_fail_reasons": "",
            "pred_stage_hits_json": "",
            "pred_target_candidates_json": "",
            "pred_target_span_keys": "",
            "pred_target_span_texts": "",
            "n_pred_all": "",
            "n_pred_target": "",
            "n_ignored_non_target": "",
            "n_ignored_out_of_scope": "",
            "n_ignored_by_triage": "",
            "n_tp": "",
            "n_fp": "",
            "n_fn": "",
            "status": "NEG_GOLD_ONLY",
            "status_detail": "",
            "eval_tag": "non_evaluable",
            "eval_tag_reason": "gold_role_neg_target_absent",
            "debug_detect_json": "",
            "debug_verify_json": "",
            "debug_context_json": "",
            "overall_precision": "",
            "overall_recall": "",
            "overall_f1": "",
            "overall_tp": "",
            "overall_fp": "",
            "overall_fn": "",
            "morph_info": "",
            "bridge_info": "",
            "thing_bridge_info": "",
            "detect_components_info": "",
            "policies_json": "",
        }
        gold_only_rows.append(row)
        gold_only_rows_jsonl.append(
            {
                "row_id": f"row_gold_only_{idx:06d}",
                "row_kind": "gold_only",
                "view": "strict",
                "eval_tag": "non_evaluable",
                "eval_tag_reason": "gold_role_neg_target_absent",
                "gold_example_role": "neg_target_absent",
                "full": {"gold_item": gold_item},
            }
        )
    if gold_only_rows:
        rows.extend(gold_only_rows)
        rows_jsonl.extend(gold_only_rows_jsonl)
    included_in_rows = sum(
        1
        for row in rows
        if str(row.get("gold_example_role") or "").strip() == "neg_target_absent"
    )
    rows_written_with_tag = sum(1 for row in rows if row.get("eval_tag") == "non_evaluable")
    excluded_from_metrics = 0
    excluded_from_fn = 0
    logger.info(
        "[eval][non_evaluable] gold_neg_target_absent_occ=%s included_in_rows=%s",
        gold_neg_target_absent_occ,
        included_in_rows,
    )
    logger.info(
        "[eval][non_evaluable] excluded_from_metrics=%s excluded_from_fn=%s",
        excluded_from_metrics,
        excluded_from_fn,
    )
    logger.info("[eval][non_evaluable] tag_field=eval_tag value=non_evaluable")
    logger.info(
        "[eval][non_evaluable] rows_written_with_tag=%s",
        rows_written_with_tag,
    )
    if gold_neg_target_absent_occ > 0 and included_in_rows == 0:
        logger.warning(
            "[eval][non_evaluable] neg_target_absent found but not added to rows"
        )

    if write_full_csv:
        with csv_path.open("w", encoding="utf-8-sig", newline="") as fp:
            writer = csv.DictWriter(fp, fieldnames=fieldnames)
            writer.writeheader()
            for row in rows:
                writer.writerow(row)

    if write_rows_jsonl:
        with rows_jsonl_path.open("w", encoding="utf-8") as fp_rows:
            for row in rows_jsonl:
                if "eval_tag" not in row:
                    if row.get("row_kind") == "gold_occurrence":
                        full = row.get("full") or {}
                        gold_item = full.get("gold_item") or {}
                        role = str(gold_item.get("gold_example_role") or "").strip()
                        if role == "neg_target_absent":
                            row["eval_tag"] = "non_evaluable"
                            row["eval_tag_reason"] = "gold_role_neg_target_absent"
                        else:
                            row["eval_tag"] = ""
                            row["eval_tag_reason"] = ""
                    else:
                        row["eval_tag"] = ""
                        row["eval_tag_reason"] = ""
                write_jsonl_line(fp_rows, row)

    def _parse_span_key(span_key: Any) -> tuple[int, int] | None:
        if not span_key or not isinstance(span_key, str):
            return None
        if ":" not in span_key:
            return None
        left, right = span_key.split(":", 1)
        if not (left.isdigit() and right.isdigit()):
            return None
        start = int(left)
        end = int(right)
        if start < 0 or end < 0 or start > end:
            return None
        return (start, end)

    def _coerce_span_segments(span_segments: Any) -> list[list[int]] | None:
        if not span_segments:
            return None
        value = span_segments
        if isinstance(value, str):
            text = value.strip()
            if not text or text.lower() == "nan":
                return None
            try:
                value = ast.literal_eval(text)
            except Exception:
                return None
        if not isinstance(value, (list, tuple)):
            return None
        out: list[list[int]] = []
        for seg in value:
            if (
                isinstance(seg, (list, tuple))
                and len(seg) == 2
                and str(seg[0]).isdigit()
                and str(seg[1]).isdigit()
            ):
                start = int(seg[0])
                end = int(seg[1])
                if 0 <= start < end:
                    out.append([start, end])
        return out or None

    def _gold_span_text_from_key(sentence: Any, span_key: Any, span_segments: Any = None) -> str:
        if not isinstance(sentence, str):
            return ""
        segments = _coerce_span_segments(span_segments)
        if segments:
            parts = []
            for start, end in segments:
                if start >= len(sentence) or end > len(sentence):
                    continue
                parts.append(sentence[start:end])
            if parts:
                return " ... ".join(parts)
        parsed = _parse_span_key(span_key)
        if not parsed:
            return ""
        start, end = parsed
        if end > len(sentence) or start >= len(sentence):
            return ""
        if start >= end:
            return ""
        return sentence[start:end]

    def _gold_span_window(sentence: Any, span_key: Any, span_segments: Any = None, window: int = 20) -> str:
        if not isinstance(sentence, str):
            return ""
        segments = _coerce_span_segments(span_segments)
        if segments:
            segments = sorted(segments, key=lambda x: x[0])
            left_bound = max(0, segments[0][0] - window)
            right_bound = min(len(sentence), segments[-1][1] + window)
            snippet = sentence[left_bound:right_bound]
            for start, end in sorted(segments, key=lambda x: x[0], reverse=True):
                rel_start = start - left_bound
                rel_end = end - left_bound
                if rel_start < 0 or rel_end > len(snippet) or rel_start >= rel_end:
                    continue
                snippet = snippet[:rel_end] + "⟧" + snippet[rel_end:]
                snippet = snippet[:rel_start] + "⟦" + snippet[rel_start:]
            return snippet
        parsed = _parse_span_key(span_key)
        if not parsed:
            return ""
        start, end = parsed
        if end > len(sentence) or start < 0 or start >= len(sentence):
            return ""
        if start >= end:
            return ""
        left = sentence[max(0, start - window) : start]
        span = sentence[start:end]
        right = sentence[end : end + window]
        return f"{left}⟦{span}⟧{right}"

    def _fn_bucket(
        *, gold_span_key: Any, gold_span_len: Any, neg_gold_conf_e_ids: Any
    ) -> str:
        if not gold_span_key:
            return "fn_no_span"
        if _parse_span_key(gold_span_key) is None:
            return ""
        try:
            span_len = int(gold_span_len) if gold_span_len != "" else None
        except Exception:
            return ""
        if span_len is not None and span_len <= 2:
            return "fn_boundary_suspect"
        if neg_gold_conf_e_ids:
            return "fn_confusable_suspect"
        return "fn_other"

    if variant == "encoder":
        min_csv_path = csv_path.with_name("eval_min_encoder.csv")
    else:
        min_csv_path = csv_path.with_name("eval_min_encoder_llm.csv")
    min_fieldnames = [
        "row_kind",
        "view",
        "match_key",
        "example_id",
        "instance_id",
        "example_key_full",
        "group",
        "gold_example_role",
        "target_sentence",
        "gold_eid",
        "gold_span_key",
        "pred_eid",
        "no_pred_reason",
        "pred_span_key",
        "rerank_status",
        "rerank_decision_line",
        "rerank_selected_eid",
        "decision_line",
        "protocol_ok",
        "candidate_e_ids_all",
        "candidate_span_keys_all",
        "candidate_confidences_all",
        "group_a_target_dropped",
        "group_a_target_dropped_encoder_score",
        "group_a_target_dropped_confidence",
        "group_a_target_dropped_reason",
        "candidate_encoder_ranks_all",
        "candidates_json",
        "status",
        "pred_triage",
        "pred_score",
        "pred_ignored_reason",
        "neg_gold_span_keys",
        "neg_gold_roles",
        "fn_reason_code",
        "none_reason",
    ]
    diag_cols = [
        "rerank_status",
        "rerank_decision_line",
        "rerank_selected_eid",
        "decision_line",
        "protocol_ok",
        "fn_reason_code",
        "none_reason",
    ]
    summary_strict_cols = [
        "metric_strict_precision",
        "metric_strict_recall",
        "metric_strict_f1",
        "count_strict_tp",
        "count_strict_fp",
        "count_strict_fn",
        "count_strict_gold_occ",
        "count_strict_gold_occ_rows",
        "count_strict_tp_rows",
        "count_strict_fn_rows",
        "count_strict_pred_ignored_rows",
    ]
    summary_lenient_cols = [
        "metric_lenient_precision_a",
        "metric_lenient_recall_a",
        "metric_lenient_f1_a",
        "count_lenient_tp",
        "count_lenient_tn",
        "count_lenient_fp",
        "count_lenient_fn",
        "count_lenient_gold_occ",
        "count_lenient_gold_occ_rows",
        "count_lenient_tp_rows",
        "count_lenient_fn_rows",
        "count_lenient_pred_ignored_rows",
    ]
    summary_trap_cols = [
        "count_trap_fp",
        "count_trap_fp_reason_counts_json",
    ]
    tail_cols = [
        "gold_span_text",
        "gold_span_window",
        "gold_span_len",
        "fn_bucket",
        "fn_subbucket",
        "label_fn_main",
        "label_fn_detail",
        "label_action_hint",
        "label_owner",
        "label_done",
    ]
    summary_cols = summary_strict_cols + summary_lenient_cols + summary_trap_cols
    min_export_drop_cols = {
        "eval_tag",
        "eval_tag_reason",
        "neg_gold_conf_e_ids",
        "candidate_in_neg_gold",
        "gold_span_len",
        "fn_subbucket",
        "label_fn_main",
        "label_fn_detail",
        "label_action_hint",
        "label_owner",
        "label_done",
        "metric_strict_precision",
        "metric_strict_recall",
        "metric_strict_f1",
        "count_strict_tp",
        "count_strict_fp",
        "count_strict_fn",
        "count_strict_gold_occ",
        "count_strict_gold_occ_rows",
        "count_strict_tp_rows",
        "count_strict_fn_rows",
        "count_strict_pred_ignored_rows",
        "count_lenient_gold_occ",
        "count_lenient_gold_occ_rows",
        "count_lenient_tp_rows",
        "count_lenient_fn_rows",
        "count_lenient_pred_ignored_rows",
    }
    raw_min_export_views = eval_cfg.get("min_export_views", "lenient")
    if isinstance(raw_min_export_views, str):
        min_export_views = [
            v.strip() for v in raw_min_export_views.split(",") if v.strip() in {"strict", "lenient"}
        ]
    elif isinstance(raw_min_export_views, (list, tuple)):
        min_export_views = [str(v).strip() for v in raw_min_export_views if str(v).strip() in {"strict", "lenient"}]
    else:
        min_export_views = ["lenient"]
    if not min_export_views:
        min_export_views = ["lenient"]
    base_cols: list[str] = []
    for col in (min_fieldnames + tail_cols):
        if col in diag_cols or col in summary_cols:
            continue
        if col not in base_cols:
            base_cols.append(col)
    min_fieldnames_ssot = [
        c for c in (base_cols + diag_cols + summary_cols) if c not in min_export_drop_cols
    ]
    min_write_fieldnames: list[str] = []
    for col in min_fieldnames_ssot:
        if col == "status":
            continue
        min_write_fieldnames.append(col)
        if col == "target_sentence":
            min_write_fieldnames.append("status_encoder")
        if col == "rerank_selected_eid":
            min_write_fieldnames.append("status_llm")
    logger.info("[DEBUG][mincsv-open] path=%s", min_csv_path)
    triage_blocked_keys: set[tuple[str, str, str]] = set()
    strict_fn_keys: set[tuple[str, str, str]] = set()
    lenient_fn_keys: set[tuple[str, str, str]] = set()
    non_target_confusion: dict[str, dict[tuple[str, str], set[tuple[str, str]]]] = {
        "strict": {},
        "lenient": {},
    }
    explain_counts = {
        "strict": {
            "gold_occ_rows": 0,
            "tp_rows": 0,
            "fn_rows": 0,
            "pred_ignored_rows": 0,
        },
        "lenient": {
            "gold_occ_rows": 0,
            "tp_rows": 0,
            "fn_rows": 0,
            "pred_ignored_rows": 0,
        },
    }
    for row in rows:
        try:
            view = row.get("view")
            if view not in ("strict", "lenient"):
                continue
            row_kind = row.get("row_kind")
            if row_kind == "gold_occurrence":
                explain_counts[view]["gold_occ_rows"] += 1
                if row.get("status") == "TP":
                    explain_counts[view]["tp_rows"] += 1
                elif row.get("status") == "FN":
                    explain_counts[view]["fn_rows"] += 1
                    match_key = row.get("match_key")
                    gold_eid = row.get("gold_eid")
                    gold_span_key = row.get("gold_span_key")
                    if match_key and gold_eid and gold_span_key:
                        key = (str(match_key), str(gold_eid), str(gold_span_key))
                        if view == "strict":
                            strict_fn_keys.add(key)
                        else:
                            lenient_fn_keys.add(key)
            elif row_kind == "pred_ignored":
                explain_counts[view]["pred_ignored_rows"] += 1
                match_key = row.get("match_key")
                cand_eid = row.get("candidate_e_id")
                cand_span_key = row.get("candidate_span_key")
                if (
                    view == "strict"
                    and row.get("pred_ignored_reason") == "triage_filtered"
                    and match_key
                    and cand_eid
                    and cand_span_key
                ):
                    triage_blocked_keys.add(
                        (str(match_key), str(cand_eid), str(cand_span_key))
                    )
                if (
                    row.get("pred_ignored_reason") == "non_target"
                    and match_key
                    and cand_span_key
                ):
                    pred_eid = row.get("pred_eid") or ""
                    bucket = non_target_confusion[view].setdefault(
                        (str(match_key), str(cand_span_key)), set()
                    )
                    bucket.add((str(cand_eid or ""), str(pred_eid)))
        except Exception:
            continue
    fn_labels_path = csv_path.with_name("eval_latest_fn_labels.csv")
    fn_labels_fieldnames = [
        "row_kind",
        "view",
        "match_key",
        "example_id",
        "instance_id",
        "target_sentence",
        "gold_eid",
        "gold_span_key",
        "status",
        "gold_span_text",
        "pred_eid",
        "no_pred_reason",
        "pred_span_key",
        "gold_span_window",
        "gold_span_len",
        "fn_bucket",
        "fn_subbucket",
        "strict_has_triage_filtered_same_key",
        "label_fn_main",
        "label_fn_detail",
        "label_action_hint",
        "label_owner",
        "label_done",
    ]
    header_required = {
        "candidate_e_ids_all",
        "candidate_span_keys_all",
        "candidate_encoder_ranks_all",
        "candidates_json",
        "rerank_status",
        "rerank_decision_line",
        "rerank_selected_eid",
        "decision_line",
        "protocol_ok",
        "fn_reason_code",
        "none_reason",
        *summary_strict_cols,
        *summary_lenient_cols,
        *summary_trap_cols,
    }
    if variant == "encoder":
        header_required = header_required.difference(
            {
                "rerank_status",
                "rerank_decision_line",
                "rerank_selected_eid",
                "decision_line",
                "protocol_ok",
            }
        )
    header_required = header_required.difference(min_export_drop_cols)
    header_has_cols = header_required.issubset(set(min_fieldnames_ssot))
    logger.info(
        "[eval][for_users_min_candidates_all] header_has_cols=%s n_fieldnames=%s",
        header_has_cols,
        len(min_fieldnames_ssot),
    )
    if not header_has_cols:
        raise ConfigError("min_fieldnames missing required columns")
    ssot: dict[str, dict[str, Any]] = {}
    ssot_payload_by_mk: dict[str, dict[str, Any]] = {}
    ssot_gold_key_by_mk: dict[str, str] = {}
    ssot_pred_key_by_gold_key: dict[str, str] = {}
    ssot_match_keys = 0
    ssot_match_keys_with_candidates = 0
    for item in rows_jsonl:
        full = item.get("full") or {}
        record = full.get("record") or {}
        sentence = record.get("target_sentence") or record.get("raw_sentence") or ""
        primary_match_key = _match_key(record, match_key_policy, fallback_sentence=sentence)
        example_id = record.get("example_id")
        instance_id = record.get("instance_id", 1)
        fallback_match_key = ""
        if example_id is not None:
            fallback_match_key = f"{example_id}#{instance_id}"
        key_candidates = []
        for key in (primary_match_key, fallback_match_key, str(sentence or "")):
            key = str(key or "")
            if key and key not in key_candidates:
                key_candidates.append(key)
        if not key_candidates:
            continue
        match_key = key_candidates[0]
        ssot_match_keys += 1
        rerank = record.get("rerank") or {}
        selected_eid = rerank.get("selected_eid")
        candidates = record.get("candidates") or []
        if not isinstance(candidates, list):
            candidates = []
        sortable_cands: list[dict[str, Any]] = [c for c in candidates if isinstance(c, dict)]
        if any(c.get("encoder_rank") is not None for c in sortable_cands):
            indexed = list(enumerate(sortable_cands))
            indexed.sort(
                key=lambda x: (
                    0 if x[1].get("encoder_rank") is not None else 1,
                    int(x[1].get("encoder_rank") or 0),
                    x[0],
                )
            )
            sorted_cands = [c for _, c in indexed]
        else:
            sorted_cands = sortable_cands
        cand_e_ids_all = "|".join([str(c.get("e_id") or "") for c in sorted_cands])
        cand_span_keys_all = "|".join([str(c.get("span_key") or "") for c in sorted_cands])
        cand_encoder_ranks_all = "|".join(
            [str(c.get("encoder_rank") or "") for c in sorted_cands]
        )
        cand_confidences_all = "|".join(
            [str(c.get("confidence") or "") for c in sorted_cands]
        )
        cand_json = json.dumps(sorted_cands, ensure_ascii=False)
        selected_span_key = None
        if selected_eid and isinstance(rerank.get("input_candidates"), list):
            for cand in rerank.get("input_candidates") or []:
                if (
                    isinstance(cand, dict)
                    and str(cand.get("e_id") or "") == str(selected_eid)
                    and cand.get("span_key")
                ):
                    selected_span_key = str(cand.get("span_key"))
                    break
        if selected_span_key is None and selected_eid:
            for cand in candidates:
                if (
                    isinstance(cand, dict)
                    and str(cand.get("e_id") or "") == str(selected_eid)
                    and cand.get("span_key")
                ):
                    selected_span_key = str(cand.get("span_key"))
                    break
        if selected_span_key is None:
            for cand in candidates:
                if isinstance(cand, dict) and cand.get("encoder_rank") == 1 and cand.get("span_key"):
                    selected_span_key = str(cand.get("span_key"))
                    break
        if selected_span_key is None and sorted_cands:
            first = sorted_cands[0]
            if isinstance(first, dict) and first.get("span_key"):
                selected_span_key = str(first.get("span_key"))
        cand_eid = None
        cand_span_key = None
        candidate_primary = None
        for cand in sorted_cands:
            if isinstance(cand, dict) and cand.get("encoder_rank") == 1:
                candidate_primary = cand
                break
        if candidate_primary is None and sorted_cands:
            candidate_primary = sorted_cands[0] if isinstance(sorted_cands[0], dict) else None
        if candidate_primary is not None:
            cand_eid = candidate_primary.get("e_id")
            cand_span_key = candidate_primary.get("span_key")
        has_candidates = bool(sorted_cands)
        dropped_candidates = record.get("dropped_candidates") or []
        if not isinstance(dropped_candidates, list):
            dropped_candidates = []
        dropped_sortable: list[dict[str, Any]] = [
            d for d in dropped_candidates if isinstance(d, dict)
        ]
        dropped_by_eid: dict[str, dict[str, Any]] = {}
        for dropped in dropped_sortable:
            dropped_eid = str(dropped.get("e_id") or "")
            if dropped_eid and dropped_eid not in dropped_by_eid:
                dropped_by_eid[dropped_eid] = dropped
        has_dropped_candidates = bool(dropped_sortable)
        gold_item = full.get("gold_item") or {}
        gold_example_id = gold_item.get("example_id")
        gold_instance_id = gold_item.get("instance_id")
        gold_example_key = ""
        if gold_example_id is not None:
            gold_example_key = f"{gold_example_id}#{gold_instance_id if gold_instance_id is not None else 1}"
        pred_example_key = ""
        if example_id is not None:
            pred_example_key = f"{example_id}#{instance_id if instance_id is not None else 1}"
        if pred_example_key and pred_example_key not in key_candidates:
            key_candidates.append(pred_example_key)
        if gold_example_key:
            for key in key_candidates:
                ssot_gold_key_by_mk[key] = gold_example_key
            if pred_example_key:
                ssot_pred_key_by_gold_key.setdefault(gold_example_key, pred_example_key)

        if has_candidates:
            ssot_match_keys_with_candidates += 1
        if has_candidates or has_dropped_candidates:
            rerank_payload = record.get("rerank") or {}
            payload = {
                "candidate_e_ids_all": cand_e_ids_all.replace("|", ";"),
                "candidate_span_keys_all": cand_span_keys_all.replace("|", ";"),
                "candidate_encoder_ranks_all": cand_encoder_ranks_all.replace("|", ";"),
                "candidate_confidences_all": cand_confidences_all.replace("|", ";"),
                "candidates_json": cand_json,
                "_dropped_by_eid": dropped_by_eid,
                "_span_reranks": list(rerank_payload.get("span_reranks") or []),
            }
            for key in key_candidates:
                ssot_payload_by_mk[key] = payload
            if match_key == "g0001#1":
                logger.info(
                    "[eval][debug_candidates_payload_ssot] mk=g0001#1 n_cands=%s e_ids_all=%s span_keys_all=%s ranks_all=%s json_len=%s dropped=%s",
                    len(sorted_cands),
                    payload.get("candidate_e_ids_all"),
                    payload.get("candidate_span_keys_all"),
                    payload.get("candidate_encoder_ranks_all"),
                    len(payload.get("candidates_json") or ""),
                    len(dropped_sortable),
                )
        meta = {
            "selected_eid": selected_eid,
            "selected_span_key": selected_span_key,
            "cand_eid": cand_eid,
            "cand_span_key": cand_span_key,
            "has_candidates": has_candidates,
        }
        for key in key_candidates:
            ssot[key] = meta
    logger.info(
        "[eval][for_users_min_candidates_all] ssot_match_keys=%s ssot_match_keys_with_candidates=%s",
        ssot_match_keys,
        ssot_match_keys_with_candidates,
    )

    def _normalize_scalar(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and pd.isna(value):
            return ""
        text = str(value)
        if text.strip().lower() in {"none", "nan"}:
            return ""
        return text

    def _normalize_display_example_id(value: Any) -> str:
        text = _normalize_scalar(value)
        if re.fullmatch(r"[0-9a-f]{40}", text):
            return f"sha1:{text[:8]}"
        return text

    def _normalize_min_out_row(
        row: dict[str, Any],
        *,
        ssot_row: dict[str, Any] | None,
        match_key_hint: str,
        group_hint: str = "",
    ) -> dict[str, Any]:
        out = dict(row)
        for key in (
            "row_kind",
            "view",
            "match_key",
            "match_key_raw",
            "example_id",
            "example_id_full",
            "instance_id",
            "example_key_full",
            "target_sentence",
            "gold_example_role",
            "gold_eid",
            "gold_span_key",
            "pred_eid",
            "pred_span_key",
            "candidate_e_id",
            "candidate_span_key",
            "status",
            "eval_tag",
            "eval_tag_reason",
            "pred_triage",
            "pred_score",
            "pred_ignored_reason",
            "neg_gold_conf_e_ids",
            "neg_gold_span_keys",
            "neg_gold_roles",
            "group",
            "fn_reason_code",
            "none_reason",
        ):
            out[key] = _normalize_scalar(out.get(key, ""))

        if not out.get("match_key_raw"):
            out["match_key_raw"] = _normalize_scalar(out.get("match_key", ""))
        if not out.get("example_id_full"):
            out["example_id_full"] = _normalize_scalar(out.get("example_id", ""))
        if not out.get("example_key_full"):
            raw_example = _normalize_scalar(out.get("example_id", ""))
            raw_instance = _normalize_scalar(out.get("instance_id", ""))
            if raw_example:
                suffix = raw_instance if raw_instance else "1"
                out["example_key_full"] = f"{raw_example}#{suffix}"
            else:
                out["example_key_full"] = _normalize_scalar(out.get("match_key", ""))

        out["example_id"] = _normalize_display_example_id(out.get("example_id", ""))

        if not out.get("match_key") and match_key_hint:
            out["match_key"] = match_key_hint
        if not out.get("example_id") and "#" in out.get("match_key", ""):
            out["example_id"] = _normalize_display_example_id(
                out["match_key"].split("#", 1)[0]
            )
        if not out.get("instance_id") and "#" in out.get("match_key", ""):
            out["instance_id"] = out["match_key"].split("#", 1)[1]
        if (
            out.get("target_sentence")
            and out.get("match_key") == out.get("target_sentence")
            and out.get("example_id")
        ):
            suffix = f"#{out.get('instance_id')}" if out.get("instance_id") else ""
            out["match_key"] = f"{out['example_id']}{suffix}"

        if not out.get("group") and group_hint:
            out["group"] = _normalize_scalar(group_hint)

        if ssot_row:
            if not out.get("pred_eid"):
                out["pred_eid"] = _normalize_scalar(ssot_row.get("selected_eid"))
            if not out.get("pred_span_key"):
                out["pred_span_key"] = _normalize_scalar(ssot_row.get("selected_span_key"))
            if not out.get("candidate_e_id"):
                out["candidate_e_id"] = _normalize_scalar(ssot_row.get("cand_eid"))
            if not out.get("candidate_span_key"):
                out["candidate_span_key"] = _normalize_scalar(ssot_row.get("cand_span_key"))

        out["gold_span_key"] = _canonical_span_key(out.get("gold_span_key"), None)
        out["pred_span_key"] = _canonical_span_key(out.get("pred_span_key"), None)
        out["candidate_span_key"] = _canonical_span_key(out.get("candidate_span_key"), None)
        return out

    fn_label_rows: list[dict[str, Any]] = []
    fn_labels_missing_key_counts = Counter()
    example_id_counts: Counter[str] = Counter()
    example_key_counts: Counter[str] = Counter()
    for row in rows:
        example_id = row.get("example_id")
        instance_id = row.get("instance_id")
        ex_key = str(example_id) if example_id is not None else ""
        example_id_counts[ex_key] += 1
        if instance_id is not None and str(instance_id) != "":
            example_key_counts[f"{ex_key}#{instance_id}"] += 1
    total_rows_in_export = len(rows)
    unique_example_id = len(example_id_counts)
    max_rows_per_example_id = max(example_id_counts.values()) if example_id_counts else 0
    if example_key_counts:
        max_rows_per_example_id = max(max_rows_per_example_id, max(example_key_counts.values()))
    n_example_id_with_ge2_rows = sum(1 for v in example_id_counts.values() if v >= 2)
    logger.info(
        "[eval][for_users_multirow] total_rows_in_export=%d",
        total_rows_in_export,
    )
    logger.info(
        "[eval][for_users_multirow] unique_example_id=%d max_rows_per_example_id=%d",
        unique_example_id,
        max_rows_per_example_id,
    )
    logger.info(
        "[eval][for_users_multirow] n_example_id_with_ge2_rows=%d",
        n_example_id_with_ge2_rows,
    )

    expredict_by_eid: dict[str, dict[str, Any]] = {}
    try:
        from openpyxl import load_workbook

        expredict_xlsx_path = Path(cfg["paths"]["dict_xlsx"])
        if expredict_xlsx_path.exists():
            dict_cfg = cfg.get("dict", {}) or {}
            sheet_names = dict_cfg.get("sheet_names", {}) or {}
            expredict_sheet = sheet_names.get("expredict", "expredict")
            wb = load_workbook(expredict_xlsx_path, read_only=True, data_only=True)
            if expredict_sheet in wb.sheetnames:
                ws = wb[expredict_sheet]
            else:
                ws = wb[wb.sheetnames[0]]
            expredict_header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            if "e_id" in expredict_header:
                idx_eid = expredict_header.index("e_id")
                idx_group = expredict_header.index("group") if "group" in expredict_header else None
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row:
                        continue
                    eid = row[idx_eid] if idx_eid < len(row) else None
                    if eid is None:
                        continue
                    eid_key = str(eid)
                    if eid_key in expredict_by_eid:
                        continue
                    group_value = row[idx_group] if idx_group is not None and idx_group < len(row) else None
                    expredict_by_eid[eid_key] = {"group": group_value}
            wb.close()
    except Exception as exc:
        logger.warning("[eval][for_users_min_group] expredict preload failed: %s", exc)

    def _split_selected_eids(raw_value: Any) -> list[str]:
        text_val = str(raw_value or "").strip()
        if not text_val:
            return []
        out: list[str] = []
        seen: set[str] = set()
        for part in text_val.replace(";", ",").split(","):
            eid = part.strip()
            if not eid or eid in seen:
                continue
            seen.add(eid)
            out.append(eid)
        return out

    def _group_hint_for_min_row(row: dict[str, Any]) -> str:
        rerank_selected_eids = _split_selected_eids(row.get("rerank_selected_eid"))
        group_lookup_eid = (
            row.get("gold_eid")
            or row.get("pred_eid")
            or row.get("candidate_e_id")
            or (rerank_selected_eids[0] if rerank_selected_eids else "")
            or ""
        )
        if not group_lookup_eid:
            return ""
        return str((expredict_by_eid.get(str(group_lookup_eid)) or {}).get("group") or "")

    gold_key_by_sentence: dict[str, str] = {}
    gold_key_by_sentence_instance: dict[tuple[str, str], str] = {}
    for gold_item in list(gold_rows or []) + list(nta_gold_rows or []):
        if not isinstance(gold_item, dict):
            continue
        sentence_key = str(gold_item.get("target_sentence") or "")
        example_id_val = gold_item.get("example_id")
        instance_id_val = gold_item.get("instance_id")
        instance_norm = str(instance_id_val if instance_id_val is not None else 1)
        if not sentence_key or example_id_val is None:
            continue
        gold_key = f"{example_id_val}#{instance_norm}"
        gold_key_by_sentence.setdefault(sentence_key, gold_key)
        gold_key_by_sentence_instance.setdefault((sentence_key, instance_norm), gold_key)

    try:
        with pred_path.open("r", encoding="utf-8") as fp_pred_src:
            for line in fp_pred_src:
                line = line.strip()
                if not line:
                    continue
                try:
                    pred_record = json.loads(line)
                except Exception:
                    continue
                if not isinstance(pred_record, dict):
                    continue
                pred_sentence = str(
                    pred_record.get("target_sentence")
                    or pred_record.get("raw_sentence")
                    or ""
                )
                pred_example_id = pred_record.get("example_id")
                pred_instance_id = pred_record.get("instance_id", 1)
                pred_instance_norm = str(pred_instance_id if pred_instance_id is not None else 1)
                pred_key = (
                    f"{pred_example_id}#{pred_instance_norm}" if pred_example_id is not None else ""
                )
                gold_key = ""
                if pred_sentence:
                    gold_key = (
                        gold_key_by_sentence_instance.get((pred_sentence, pred_instance_norm))
                        or gold_key_by_sentence.get(pred_sentence)
                        or ""
                    )
                if gold_key and pred_key:
                    ssot_pred_key_by_gold_key.setdefault(gold_key, pred_key)

                pred_candidates = pred_record.get("candidates") or []
                if not isinstance(pred_candidates, list):
                    pred_candidates = []
                pred_sortable: list[dict[str, Any]] = [
                    c for c in pred_candidates if isinstance(c, dict)
                ]
                if any(c.get("encoder_rank") is not None for c in pred_sortable):
                    indexed = list(enumerate(pred_sortable))
                    indexed.sort(
                        key=lambda x: (
                            0 if x[1].get("encoder_rank") is not None else 1,
                            int(x[1].get("encoder_rank") or 0),
                            x[0],
                        )
                    )
                    pred_sorted_cands = [c for _, c in indexed]
                else:
                    pred_sorted_cands = pred_sortable
                pred_dropped = pred_record.get("dropped_candidates") or []
                if not isinstance(pred_dropped, list):
                    pred_dropped = []
                pred_dropped_by_eid: dict[str, dict[str, Any]] = {}
                for dropped in pred_dropped:
                    if not isinstance(dropped, dict):
                        continue
                    dropped_eid = str(dropped.get("e_id") or "")
                    if dropped_eid and dropped_eid not in pred_dropped_by_eid:
                        pred_dropped_by_eid[dropped_eid] = dropped
                pred_rerank = pred_record.get("rerank") or {}
                pred_span_reranks = list(pred_rerank.get("span_reranks") or [])
                if pred_key and pred_key not in ssot_payload_by_mk and (pred_sorted_cands or pred_dropped_by_eid):
                    ssot_payload_by_mk[pred_key] = {
                        "candidate_e_ids_all": ";".join([str(c.get("e_id") or "") for c in pred_sorted_cands]),
                        "candidate_span_keys_all": ";".join([str(c.get("span_key") or "") for c in pred_sorted_cands]),
                        "candidate_encoder_ranks_all": ";".join([str(c.get("encoder_rank") or "") for c in pred_sorted_cands]),
                        "candidate_confidences_all": ";".join([str(c.get("confidence") or "") for c in pred_sorted_cands]),
                        "candidates_json": json.dumps(pred_sorted_cands, ensure_ascii=False),
                        "_dropped_by_eid": pred_dropped_by_eid,
                        "_span_reranks": pred_span_reranks,
                    }
                elif pred_key and pred_span_reranks:
                    existing_payload = ssot_payload_by_mk.get(pred_key)
                    if isinstance(existing_payload, dict) and not existing_payload.get("_span_reranks"):
                        existing_payload["_span_reranks"] = pred_span_reranks
    except Exception as exc:
        logger.warning("[eval][for_users_min_pred_preload] failed: %s", exc)

    summary_rows_for_min = [r for r in rows if r.get("row_kind") == "overall_summary"]
    detail_rows_for_min = [r for r in rows if r.get("row_kind") != "overall_summary"]
    detail_rows_for_min.sort(
        key=lambda r: (
            0 if _group_hint_for_min_row(r) == "a" else 1 if _group_hint_for_min_row(r) == "b" else 2,
            str(r.get("gold_eid") or "~"),
            str(r.get("example_id") or ""),
            str(r.get("instance_id") or ""),
            str(r.get("target_sentence") or ""),
        )
    )
    rows_for_min = summary_rows_for_min + detail_rows_for_min

    def _payload_for_min_row_lookup(row: dict[str, Any]) -> dict[str, Any] | None:
        mk = row.get("match_key")
        if not mk:
            example_id = row.get("example_id")
            instance_id = row.get("instance_id")
            if example_id is not None:
                mk = f"{example_id}#{instance_id if instance_id is not None else 1}"
        payload = ssot_payload_by_mk.get(str(mk)) if mk is not None else None
        if payload is None:
            gold_key_lookup = str(mk or "")
            if not gold_key_lookup:
                ex_for_gold_key = row.get("example_id_full") or row.get("example_id")
                inst_for_gold_key = row.get("instance_id")
                if ex_for_gold_key:
                    gold_key_lookup = f"{ex_for_gold_key}#{inst_for_gold_key if inst_for_gold_key is not None and str(inst_for_gold_key) != '' else 1}"
            pred_key_lookup = ssot_pred_key_by_gold_key.get(gold_key_lookup) or ""
            if pred_key_lookup:
                payload = ssot_payload_by_mk.get(str(pred_key_lookup))
        return payload if isinstance(payload, dict) else None

    def _status_llm_for_summary(row: dict[str, Any], group_hint: str) -> str:
        if group_hint != "b":
            return ""
        gold_role = str(row.get("gold_example_role") or "").strip()
        rerank_selected = set(_split_selected_eids(row.get("rerank_selected_eid")))
        gold_eid_val = str(row.get("gold_eid") or "")
        pos_roles = {"pos_conti", "pos_disconti"}
        neg_roles = {"neg_target_absent", "neg_confusable", "neg_boundary"}
        if gold_role in pos_roles:
            if gold_eid_val and gold_eid_val in rerank_selected:
                return "TP"
            return "FN" if not rerank_selected else "FP"
        if gold_role in neg_roles:
            return "FP" if rerank_selected else "TN"
        return ""

    def _effective_min_status_for_summary(row: dict[str, Any]) -> str:
        include_in_min = row.get("view") in min_export_views
        if (
            not include_in_min
            and row.get("row_kind") == "gold_only"
            and str(row.get("gold_example_role") or "").strip() == "neg_target_absent"
        ):
            include_in_min = True
        if not include_in_min or row.get("row_kind") == "overall_summary":
            return ""
        group_hint = _group_hint_for_min_row(row)
        if group_hint == "a":
            gold_role = str(row.get("gold_example_role") or "").strip()
            if gold_role == "neg_target_absent":
                payload = _payload_for_min_row_lookup(row)
                live_candidate_eids: set[str] = set()
                if payload is not None:
                    live_candidate_eids = {
                        part.strip()
                        for part in str(payload.get("candidate_e_ids_all") or "").split(";")
                        if part and part.strip()
                    }
                gold_eid_val = str(row.get("gold_eid") or "")
                return "FP" if gold_eid_val and gold_eid_val in live_candidate_eids else "TN"
            return str(row.get("status") or "")
        if group_hint == "b":
            return _status_llm_for_summary(row, group_hint)
        return ""

    user_lenient_tp = 0
    user_lenient_tn = 0
    user_lenient_fp = 0
    user_lenient_fn = 0
    # min CSV lenient summary is A-group-only.
    # B-group is evaluated in the span min export.
    for _summary_row in rows_for_min:
        # Group must be decided by gold_eid to avoid cross-contamination from
        # candidate/pred hints and to avoid pre-normalization blank group fields.
        _gold_eid_for_group = str(_summary_row.get("gold_eid") or "")
        _gold_group = str((expredict_by_eid.get(_gold_eid_for_group) or {}).get("group") or "")
        if _gold_group != "a":
            continue
        # Align summary counts with min CSV exported rows.
        # pred_ignored is internally generated but excluded from min CSV output.
        if str(_summary_row.get("row_kind") or "") not in {"gold_occurrence", "gold_only"}:
            continue
        _status_for_summary = _effective_min_status_for_summary(_summary_row)
        if _status_for_summary == "TP":
            user_lenient_tp += 1
        elif _status_for_summary == "TN":
            user_lenient_tn += 1
        elif _status_for_summary == "FP":
            user_lenient_fp += 1
        elif _status_for_summary == "FN":
            user_lenient_fn += 1

    def _safe_ratio(num: int, den: int) -> str:
        if den <= 0:
            return ""
        return str(float(num) / float(den))

    span_min_csv_path = None
    span_min_fieldnames = [
        "row_kind",
        "view",
        "match_key",
        "example_id",
        "instance_id",
        "example_key_full",
        "group",
        "gold_example_role",
        "target_sentence",
        "gold_span_window",
        "gold_eid",
        "gold_span_key",
        "status_llm",
        "span_rerank_decision_line",
        "rerank_selected_eid",
        "candidate_e_ids_all",
        "candidate_span_keys_all",
        "span_rerank_span_key",
        "target_matched_span",
        "include_in_b_metrics",
        "rerank_status",
        "rerank_decision_line",
        "span_rerank_status",
        "eid_no_pred_reason",
        "span_rerank_selected_eid",
        "span_rerank_selected_eids",
        "span_rerank_protocol_ok",
        "span_rerank_none_reason",
        "num_tp",
        "num_tn",
        "num_fp",
        "num_fn",
        "metric_precision",
        "metric_recall",
        "metric_f1",
        "count_b_gold_unique_examples",
        "count_b_span_unique_examples",
        "note",
    ]
    span_min_rows: list[dict[str, Any]] = []

    with min_csv_path.open("w", encoding="utf-8", newline="") as fp_min:
        writer = csv.DictWriter(fp_min, fieldnames=min_write_fieldnames)
        def _writerow_min(writer, row: dict):
            row2 = {k: row.get(k, "") for k in writer.fieldnames or []}
            row2["status_encoder"] = row.get("status", "")
            row2["status_llm"] = row.get("status_llm", "")
            if "candidate_in_neg_gold" in row2:
                row2["candidate_in_neg_gold"] = _tf(row2.get("candidate_in_neg_gold"))
                assert row2["candidate_in_neg_gold"] in ("", "TRUE", "FALSE"), row2[
                    "candidate_in_neg_gold"
                ]
                assert isinstance(row2["candidate_in_neg_gold"], str), type(
                    row2["candidate_in_neg_gold"]
                )
            writer.writerow(row2)
        writer.writeheader()
        min_count = 0
        min_type_counts = Counter()
        min_norm_counts = Counter()
        fn_subbucket_counts = Counter()
        metric_strict_sample: dict[str, Any] | None = None
        missing_counter_keys: set[str] = set()
        true_miss_keys = strict_fn_keys & lenient_fn_keys
        backfilled_pred_eid = 0
        backfilled_pred_span_key = 0
        backfilled_candidate_e_id = 0
        backfilled_candidate_span_key = 0
        min_has_pred_or_candidate: dict[str, bool] = {}
        filled_all_cols = 0
        rows_with_candidates = 0
        rows_with_candidates_but_empty_all = 0
        b_gold_unique_example_keys_for_span: set[str] = set()
        for row in rows_for_min:
            match_key = row.get("match_key")
            if not match_key:
                example_id = row.get("example_id")
                instance_id = row.get("instance_id")
                if example_id is not None:
                    match_key = f"{example_id}#{instance_id if instance_id is not None else 1}"
            ssot_row = ssot.get(str(match_key)) if match_key is not None else None
            if ssot_row:
                if not row.get("pred_eid") and ssot_row.get("selected_eid"):
                    row["pred_eid"] = ssot_row.get("selected_eid")
                    backfilled_pred_eid += 1
                if not row.get("pred_span_key") and ssot_row.get("selected_span_key"):
                    row["pred_span_key"] = ssot_row.get("selected_span_key")
                    backfilled_pred_span_key += 1
                if not row.get("candidate_e_id") and ssot_row.get("cand_eid"):
                    row["candidate_e_id"] = ssot_row.get("cand_eid")
                    backfilled_candidate_e_id += 1
                if not row.get("candidate_span_key") and ssot_row.get("cand_span_key"):
                    row["candidate_span_key"] = ssot_row.get("cand_span_key")
                    backfilled_candidate_span_key += 1
                if match_key:
                    has_value = bool(row.get("pred_eid") or row.get("candidate_e_id"))
                    min_has_pred_or_candidate[str(match_key)] = (
                        min_has_pred_or_candidate.get(str(match_key), False) or has_value
                    )
            out_row = {k: row.get(k, "") for k in min_fieldnames_ssot}
            mk = out_row.get("match_key")
            if not mk:
                example_id = out_row.get("example_id")
                instance_id = out_row.get("instance_id")
                if example_id is not None:
                    mk = f"{example_id}#{instance_id if instance_id is not None else 1}"
            payload = ssot_payload_by_mk.get(str(mk)) if mk is not None else None
            gold_key_lookup = str(mk or "")
            if not gold_key_lookup:
                ex_for_gold_key = out_row.get("example_id_full") or out_row.get("example_id")
                inst_for_gold_key = out_row.get("instance_id")
                if ex_for_gold_key:
                    gold_key_lookup = f"{ex_for_gold_key}#{inst_for_gold_key if inst_for_gold_key is not None and str(inst_for_gold_key) != '' else 1}"
            pred_key_lookup = ssot_pred_key_by_gold_key.get(gold_key_lookup) or ""
            if pred_key_lookup:
                pred_payload = ssot_payload_by_mk.get(str(pred_key_lookup))
                if payload is None:
                    payload = pred_payload
                elif variant == "llm":
                    payload_span_reranks = payload.get("_span_reranks") if isinstance(payload, dict) else None
                    pred_span_reranks = pred_payload.get("_span_reranks") if isinstance(pred_payload, dict) else None
                    if (not payload_span_reranks) and pred_span_reranks:
                        payload = pred_payload
            target_drop: dict[str, Any] | None = None
            if payload:
                out_row["candidate_e_ids_all"] = payload.get("candidate_e_ids_all") or ""
                out_row["candidate_span_keys_all"] = payload.get("candidate_span_keys_all") or ""
                out_row["candidate_encoder_ranks_all"] = payload.get("candidate_encoder_ranks_all") or ""
                out_row["candidate_confidences_all"] = payload.get("candidate_confidences_all") or ""
                out_row["candidates_json"] = payload.get("candidates_json") or ""
                dropped_lookup = payload.get("_dropped_by_eid") or {}
                if isinstance(dropped_lookup, dict):
                    maybe_target_drop = dropped_lookup.get(str(out_row.get("gold_eid") or ""))
                    if isinstance(maybe_target_drop, dict):
                        target_drop = maybe_target_drop
                if str(mk) == "g0001#1":
                    logger.info(
                        "[eval][debug_candidates_payload_min] mk=%s has_payload=%s e_ids_all=%s json_len=%s",
                        mk,
                        True,
                        out_row.get("candidate_e_ids_all"),
                        len(out_row.get("candidates_json") or ""),
                    )
                rows_with_candidates += 1
                if str(out_row.get("candidate_e_ids_all") or "").strip():
                    filled_all_cols += 1
                else:
                    rows_with_candidates_but_empty_all += 1
            else:
                out_row["candidate_e_ids_all"] = out_row.get("candidate_e_ids_all") or ""
                out_row["candidate_span_keys_all"] = out_row.get("candidate_span_keys_all") or ""
                out_row["candidate_encoder_ranks_all"] = out_row.get("candidate_encoder_ranks_all") or ""
                out_row["candidate_confidences_all"] = out_row.get("candidate_confidences_all") or ""
                out_row["candidates_json"] = out_row.get("candidates_json") or ""
                if str(mk) == "g0001#1":
                    logger.info(
                        "[eval][debug_candidates_payload_min] mk=%s has_payload=%s e_ids_all=%s json_len=%s",
                        mk,
                        False,
                        out_row.get("candidate_e_ids_all"),
                        len(out_row.get("candidates_json") or ""),
                    )
            group_lookup_eid = (
                row.get("gold_eid")
                or row.get("pred_eid")
                or row.get("candidate_e_id")
                or row.get("rerank_selected_eid")
                or ""
            )
            group_hint = ""
            if group_lookup_eid:
                group_hint = str(
                    (expredict_by_eid.get(str(group_lookup_eid)) or {}).get("group") or ""
                )

            out_row = _normalize_min_out_row(
                out_row,
                ssot_row=ssot_row,
                match_key_hint=str(mk or ""),
                group_hint=group_hint,
            )
            gold_key_hint = ssot_gold_key_by_mk.get(str(mk)) if mk is not None else ""
            if not gold_key_hint:
                sentence_key = str(out_row.get("target_sentence") or "")
                if sentence_key:
                    gold_key_hint = (
                        ssot_gold_key_by_mk.get(sentence_key)
                        or gold_key_by_sentence.get(sentence_key)
                        or ""
                    )
            if gold_key_hint:
                out_row["example_key_full"] = gold_key_hint
            elif str(out_row.get("gold_example_role") or "").strip().startswith("neg"):
                out_row["example_key_full"] = str(
                    out_row.get("match_key_raw") or out_row.get("example_key_full") or ""
                )

            out_row["group_a_target_dropped"] = ""
            out_row["group_a_target_dropped_encoder_score"] = ""
            out_row["group_a_target_dropped_confidence"] = ""
            out_row["group_a_target_dropped_reason"] = ""
            out_row["no_pred_reason"] = ""
            if target_drop:
                out_row["group_a_target_dropped"] = "TRUE"
                out_row["group_a_target_dropped_encoder_score"] = _normalize_scalar(
                    target_drop.get("encoder_score", "")
                )
                out_row["group_a_target_dropped_confidence"] = _normalize_scalar(
                    target_drop.get("confidence", "")
                )
                out_row["group_a_target_dropped_reason"] = _normalize_scalar(
                    target_drop.get("routing_reason", "")
                )
                if not str(out_row.get("pred_eid") or "").strip():
                    out_row["no_pred_reason"] = "gold-matched eid rejected by encoder"
            elif (
                str(out_row.get("group") or "") == "a"
                and str(out_row.get("gold_example_role") or "").strip() == "neg_target_absent"
            ):
                live_candidate_eids = {
                    part.strip()
                    for part in str(out_row.get("candidate_e_ids_all") or "").split(";")
                    if part and part.strip()
                }
                gold_eid_for_a_neg = str(out_row.get("gold_eid") or "")
                out_row["group_a_target_dropped"] = (
                    "other reason"
                    if gold_eid_for_a_neg and gold_eid_for_a_neg in live_candidate_eids
                    else "filtered by rule"
                )

            # Fallback only from infer payload (no synthetic span generation).
            if not str(out_row.get("gold_span_key") or "").strip() and isinstance(payload, dict):
                inferred_span_key = ""
                gold_eid_for_span = str(out_row.get("gold_eid") or "")

                if gold_eid_for_span:
                    try:
                        raw_candidates = payload.get("candidates_json") or ""
                        parsed_candidates = json.loads(raw_candidates) if isinstance(raw_candidates, str) and raw_candidates.strip() else []
                        if isinstance(parsed_candidates, list):
                            for cand in parsed_candidates:
                                if not isinstance(cand, dict):
                                    continue
                                if str(cand.get("e_id") or "") != gold_eid_for_span:
                                    continue
                                cand_span = str(cand.get("span_key") or "").strip()
                                if cand_span:
                                    inferred_span_key = cand_span
                                    break
                    except Exception:
                        inferred_span_key = ""

                if not inferred_span_key:
                    raw_span_reranks = payload.get("_span_reranks") or []
                    if isinstance(raw_span_reranks, list):
                        for sr in raw_span_reranks:
                            if not isinstance(sr, dict):
                                continue
                            sr_span = str(sr.get("span_key") or "").strip()
                            if sr_span:
                                inferred_span_key = sr_span
                                break

                if not inferred_span_key:
                    span_keys_all = str(out_row.get("candidate_span_keys_all") or "").strip()
                    if span_keys_all:
                        inferred_span_key = span_keys_all.split(";")[0].strip()

                if inferred_span_key:
                    out_row["gold_span_key"] = inferred_span_key

            gold_span_text = _gold_span_text_from_key(
                out_row.get("target_sentence"),
                out_row.get("gold_span_key"),
                row.get("gold_span_segments"),
            )
            gold_span_len = len(gold_span_text) if gold_span_text else ""
            out_row["gold_span_text"] = gold_span_text
            out_row["gold_span_len"] = gold_span_len
            out_row["gold_span_window"] = _gold_span_window(
                out_row.get("target_sentence"),
                out_row.get("gold_span_key"),
                row.get("gold_span_segments"),
            )
            out_row["fn_bucket"] = _fn_bucket(
                gold_span_key=out_row.get("gold_span_key"),
                gold_span_len=gold_span_len,
                neg_gold_conf_e_ids=out_row.get("neg_gold_conf_e_ids"),
            )
            fn_subbucket = ""
            try:
                if (
                    out_row.get("row_kind") == "gold_occurrence"
                    and out_row.get("view") == "strict"
                    and out_row.get("status") == "FN"
                    and out_row.get("gold_eid")
                    and out_row.get("gold_span_key")
                ):
                    match_key = out_row.get("match_key")
                    key = (
                        str(match_key),
                        str(out_row.get("gold_eid")),
                        str(out_row.get("gold_span_key")),
                    )
                    if match_key and key in triage_blocked_keys:
                        fn_subbucket = "fn_triage_blocked"
                    elif match_key and key in true_miss_keys:
                        fn_subbucket = "fn_true_miss"
                    elif match_key:
                        nt_key = (str(match_key), str(out_row.get("gold_span_key")))
                        nt_items = non_target_confusion["strict"].get(nt_key, set())
                        gold_eid = str(out_row.get("gold_eid"))
                        for cand_eid, pred_eid in nt_items:
                            if cand_eid and cand_eid != gold_eid and pred_eid != gold_eid:
                                fn_subbucket = "fn_non_target_confusion"
                                break
            except Exception:
                fn_subbucket = ""
            out_row["fn_subbucket"] = fn_subbucket
            if fn_subbucket:
                fn_subbucket_counts[fn_subbucket] += 1
            metric_strict_precision = ""
            metric_strict_recall = ""
            metric_strict_f1 = ""
            metric_lenient_precision = ""
            metric_lenient_recall = ""
            metric_lenient_f1 = ""
            count_strict_tp = ""
            count_strict_fp = ""
            count_strict_fn = ""
            count_lenient_tp = ""
            count_lenient_tn = ""
            count_lenient_fp = ""
            count_lenient_fn = ""
            count_strict_gold_occ = ""
            count_lenient_gold_occ = ""
            count_strict_gold_occ_rows = ""
            count_lenient_gold_occ_rows = ""
            count_strict_tp_rows = ""
            count_lenient_tp_rows = ""
            count_strict_fn_rows = ""
            count_lenient_fn_rows = ""
            count_strict_pred_ignored_rows = ""
            count_lenient_pred_ignored_rows = ""
            count_trap_fp = ""
            count_trap_fp_reason_counts_json = ""
            label_fn_main = ""
            label_fn_detail = ""
            label_action_hint = ""
            label_owner = ""
            label_done = ""
            if out_row.get("row_kind") == "overall_summary":
                if out_row.get("view") == "strict":
                    metric_strict_precision = row.get("overall_precision", "")
                    metric_strict_recall = row.get("overall_recall", "")
                    metric_strict_f1 = row.get("overall_f1", "")
                    count_strict_tp = row.get("overall_tp", "")
                    count_strict_fp = row.get("overall_fp", "")
                    count_strict_fn = row.get("overall_fn", "")
                    if count_strict_tp == "":
                        missing_counter_keys.add("count_strict_tp")
                    if count_strict_fp == "":
                        missing_counter_keys.add("count_strict_fp")
                    if count_strict_fn == "":
                        missing_counter_keys.add("count_strict_fn")
                    count_strict_gold_occ = row.get("overall_gold_occ", "")
                    if count_strict_gold_occ == "":
                        missing_counter_keys.add("count_strict_gold_occ")
                    count_strict_gold_occ_rows = str(
                        explain_counts["strict"]["gold_occ_rows"]
                    )
                    count_strict_tp_rows = str(explain_counts["strict"]["tp_rows"])
                    count_strict_fn_rows = str(explain_counts["strict"]["fn_rows"])
                    count_strict_pred_ignored_rows = str(
                        explain_counts["strict"]["pred_ignored_rows"]
                    )
                    count_trap_fp = str(row_counters.get("trap_fp_total", 0))
                    count_trap_fp_reason_counts_json = json.dumps(
                        dict(row_counters.get("trap_fp_reason_counts") or {}),
                        ensure_ascii=False,
                    )
                    metric_strict_sample = {
                        "precision": metric_strict_precision,
                        "recall": metric_strict_recall,
                        "f1": metric_strict_f1,
                    }
                elif out_row.get("view") == "lenient":
                    count_lenient_tp = str(user_lenient_tp)
                    count_lenient_tn = str(user_lenient_tn)
                    count_lenient_fp = str(user_lenient_fp)
                    count_lenient_fn = str(user_lenient_fn)
                    metric_lenient_precision = _safe_ratio(
                        user_lenient_tp,
                        user_lenient_tp + user_lenient_fp,
                    )
                    metric_lenient_recall = _safe_ratio(
                        user_lenient_tp,
                        user_lenient_tp + user_lenient_fn,
                    )
                    if metric_lenient_precision and metric_lenient_recall:
                        p_val = float(metric_lenient_precision)
                        r_val = float(metric_lenient_recall)
                        metric_lenient_f1 = (
                            str((2.0 * p_val * r_val) / (p_val + r_val))
                            if (p_val + r_val) > 0
                            else ""
                        )
                    else:
                        metric_lenient_f1 = ""
                    count_lenient_gold_occ = row.get("overall_gold_occ", "")
                    if count_lenient_gold_occ == "":
                        missing_counter_keys.add("count_lenient_gold_occ")
                    count_lenient_gold_occ_rows = str(
                        explain_counts["lenient"]["gold_occ_rows"]
                    )
                    count_lenient_tp_rows = str(explain_counts["lenient"]["tp_rows"])
                    count_lenient_fn_rows = str(explain_counts["lenient"]["fn_rows"])
                    count_lenient_pred_ignored_rows = str(
                        explain_counts["lenient"]["pred_ignored_rows"]
                    )
            out_row["metric_strict_precision"] = metric_strict_precision
            out_row["metric_strict_recall"] = metric_strict_recall
            out_row["metric_strict_f1"] = metric_strict_f1
            out_row["metric_lenient_precision_a"] = metric_lenient_precision
            out_row["metric_lenient_recall_a"] = metric_lenient_recall
            out_row["metric_lenient_f1_a"] = metric_lenient_f1
            out_row["count_strict_tp"] = count_strict_tp
            out_row["count_strict_fp"] = count_strict_fp
            out_row["count_strict_fn"] = count_strict_fn
            out_row["count_lenient_tp"] = count_lenient_tp
            out_row["count_lenient_tn"] = count_lenient_tn
            out_row["count_lenient_fp"] = count_lenient_fp
            out_row["count_lenient_fn"] = count_lenient_fn
            out_row["count_strict_gold_occ"] = count_strict_gold_occ
            out_row["count_lenient_gold_occ"] = count_lenient_gold_occ
            out_row["count_strict_gold_occ_rows"] = count_strict_gold_occ_rows
            out_row["count_lenient_gold_occ_rows"] = count_lenient_gold_occ_rows
            out_row["count_strict_tp_rows"] = count_strict_tp_rows
            out_row["count_lenient_tp_rows"] = count_lenient_tp_rows
            out_row["count_strict_fn_rows"] = count_strict_fn_rows
            out_row["count_lenient_fn_rows"] = count_lenient_fn_rows
            out_row["count_strict_pred_ignored_rows"] = count_strict_pred_ignored_rows
            out_row["count_lenient_pred_ignored_rows"] = count_lenient_pred_ignored_rows
            out_row["count_trap_fp"] = count_trap_fp
            out_row["count_trap_fp_reason_counts_json"] = count_trap_fp_reason_counts_json
            if (
                str(out_row.get("gold_example_role") or "").strip() == "neg_target_absent"
                and str(out_row.get("group") or "") == "a"
            ):
                gold_eid_for_neg = str(out_row.get("gold_eid") or "")
                live_candidate_eids = {
                    part.strip()
                    for part in str(out_row.get("candidate_e_ids_all") or "").split(";")
                    if part and part.strip()
                }
                out_row["pred_eid"] = (
                    gold_eid_for_neg
                    if gold_eid_for_neg and gold_eid_for_neg in live_candidate_eids
                    else ""
                )
            if str(out_row.get("gold_example_role") or "").strip() == "neg_target_absent":
                gold_eid_for_neg = str(out_row.get("gold_eid") or "")
                pred_eid_for_neg = str(out_row.get("pred_eid") or "")
                if pred_eid_for_neg and pred_eid_for_neg == gold_eid_for_neg:
                    out_row["status"] = "FP"
                else:
                    out_row["status"] = "TN"

            status_llm = ""
            if str(out_row.get("group") or "") == "b":
                gold_role = str(out_row.get("gold_example_role") or "").strip()
                rerank_selected = set(_split_selected_eids(out_row.get("rerank_selected_eid")))
                gold_eid_val = str(out_row.get("gold_eid") or "")
                pos_roles = {"pos_conti", "pos_disconti"}
                neg_roles = {"neg_target_absent", "neg_confusable", "neg_boundary"}
                if gold_role in pos_roles:
                    if gold_eid_val and gold_eid_val in rerank_selected:
                        status_llm = "TP"
                    else:
                        status_llm = "FN" if not rerank_selected else "FP"
                elif gold_role in neg_roles:
                    status_llm = "FP" if rerank_selected else "TN"
            out_row["status_llm"] = status_llm

            # no_pred_reason taxonomy for empty pred_eid (min export user-facing diagnosis)
            if (
                out_row.get("row_kind") != "overall_summary"
                and not str(out_row.get("pred_eid") or "").strip()
            ):
                group_val = str(out_row.get("group") or "").strip()
                row_kind_val = str(out_row.get("row_kind") or "").strip()
                gold_role_val = str(out_row.get("gold_example_role") or "").strip()
                drop_flag_val = str(out_row.get("group_a_target_dropped") or "").strip()
                drop_reason_val = str(out_row.get("group_a_target_dropped_reason") or "").strip()
                current_reason = str(out_row.get("no_pred_reason") or "").strip()

                if group_val == "a":
                    if current_reason == "gold-matched eid rejected by encoder":
                        pass
                    elif drop_flag_val == "TRUE" or drop_reason_val:
                        out_row["no_pred_reason"] = "gold-matched eid rejected by encoder"
                    elif gold_role_val == "neg_target_absent" and drop_flag_val == "filtered by rule":
                        out_row["no_pred_reason"] = "a_no_target_candidate(filtered by rule)"
                    else:
                        out_row["no_pred_reason"] = "rule_could_not_detect"
                elif group_val == "b":
                    if row_kind_val == "gold_only":
                        out_row["no_pred_reason"] = "b_gold_only_no_information"

            if (
                out_row.get("row_kind") == "gold_occurrence"
                and out_row.get("view") == "strict"
                and out_row.get("status") == "FN"
            ):
                if out_row.get("fn_subbucket") == "fn_triage_blocked":
                    label_fn_main = "triage_blocked"
                elif out_row.get("fn_subbucket") == "fn_non_target_confusion":
                    label_fn_main = "non_target_confusion"
                elif out_row.get("fn_subbucket") == "fn_true_miss":
                    label_fn_main = "true_miss"
                elif out_row.get("fn_bucket") == "fn_boundary_suspect":
                    label_fn_main = "boundary_suspect"
                else:
                    label_fn_main = "other"
            out_row["label_fn_main"] = label_fn_main
            out_row["label_fn_detail"] = label_fn_detail
            out_row["label_action_hint"] = label_action_hint
            out_row["label_owner"] = label_owner
            out_row["label_done"] = label_done
            if (
                out_row.get("row_kind") == "gold_occurrence"
                and out_row.get("view") == "strict"
                and out_row.get("status") == "FN"
            ):
                for key in ("status", "pred_eid", "pred_span_key"):
                    if key not in out_row:
                        fn_labels_missing_key_counts[key] += 1
                strict_has_triage_filtered_same_key = ""
                if (
                    out_row.get("match_key")
                    and out_row.get("gold_eid")
                    and out_row.get("gold_span_key")
                ):
                    strict_has_triage_filtered_same_key = (
                        "TRUE"
                        if (
                            str(out_row.get("match_key")),
                            str(out_row.get("gold_eid")),
                            str(out_row.get("gold_span_key")),
                        )
                        in triage_blocked_keys
                        else "FALSE"
                    )
                fn_label_rows.append(
                    {
                        "row_kind": out_row.get("row_kind", ""),
                        "view": out_row.get("view", ""),
                        "match_key": out_row.get("match_key", ""),
                        "example_id": out_row.get("example_id", ""),
                        "instance_id": out_row.get("instance_id", ""),
                        "target_sentence": out_row.get("target_sentence", ""),
                        "gold_eid": out_row.get("gold_eid", ""),
                        "gold_span_key": out_row.get("gold_span_key", ""),
                        "status": out_row.get("status", ""),
                        "gold_span_text": out_row.get("gold_span_text", ""),
                        "pred_eid": out_row.get("pred_eid", ""),
                        "pred_span_key": out_row.get("pred_span_key", ""),
                        "gold_span_window": out_row.get("gold_span_window", ""),
                        "gold_span_len": out_row.get("gold_span_len", ""),
                        "fn_bucket": out_row.get("fn_bucket", ""),
                        "fn_subbucket": out_row.get("fn_subbucket", ""),
                        "strict_has_triage_filtered_same_key": strict_has_triage_filtered_same_key,
                        "label_fn_main": out_row.get("label_fn_main", ""),
                        "label_fn_detail": out_row.get("label_fn_detail", ""),
                        "label_action_hint": out_row.get("label_action_hint", ""),
                        "label_owner": out_row.get("label_owner", ""),
                        "label_done": out_row.get("label_done", ""),
                    }
                )
            include_in_min = out_row.get("view") in min_export_views
            if (
                not include_in_min
                and out_row.get("row_kind") == "gold_only"
                and str(out_row.get("gold_example_role") or "").strip() == "neg_target_absent"
            ):
                include_in_min = True
            if not include_in_min:
                continue
            # Keep pred_ignored for internal diagnostics, but hide it in min user CSV.
            if out_row.get("row_kind") == "pred_ignored":
                continue
            min_count += 1
            if str(out_row.get("group") or "") == "b":
                _b_key = str(out_row.get("example_key_full") or "").strip()
                if _b_key:
                    b_gold_unique_example_keys_for_span.add(_b_key)
            min_type_counts[type(out_row.get("candidate_in_neg_gold")).__name__] += 1
            normalized = _tf(out_row.get("candidate_in_neg_gold"))
            min_norm_counts[normalized] += 1
            _writerow_min(writer, out_row)
            if variant == "llm":
                span_reranks = []
                if isinstance(payload, dict):
                    raw_span_reranks = payload.get("_span_reranks") or []
                    if isinstance(raw_span_reranks, list):
                        span_reranks = [sr for sr in raw_span_reranks if isinstance(sr, dict)]
                if str(out_row.get("group") or "") == "b":
                    gold_role = str(out_row.get("gold_example_role") or "").strip()
                    gold_eid_val = str(out_row.get("gold_eid") or "")
                    pos_roles = {"pos_conti", "pos_disconti"}
                    neg_roles = {"neg_target_absent", "neg_confusable", "neg_boundary"}
                    if span_reranks:
                        for sr in span_reranks:
                            selected_raw = sr.get("selected_eids") or []
                            if isinstance(selected_raw, str):
                                selected_list = [part.strip() for part in selected_raw.split(",") if part.strip()]
                            elif isinstance(selected_raw, list):
                                selected_list = [str(part).strip() for part in selected_raw if str(part).strip()]
                            else:
                                selected_list = []
                            selected_set = set(selected_list)
                            span_status_llm = ""
                            if gold_role in pos_roles:
                                if gold_eid_val and gold_eid_val in selected_set:
                                    span_status_llm = "TP"
                                else:
                                    span_status_llm = "FN" if not selected_set else "FP"
                            elif gold_role in neg_roles:
                                span_status_llm = "FP" if selected_set else "TN"
                            span_min_rows.append({
                                "row_kind": out_row.get("row_kind", ""),
                                "view": out_row.get("view", ""),
                                "match_key": out_row.get("match_key", ""),
                                "example_id": out_row.get("example_id", ""),
                                "instance_id": out_row.get("instance_id", ""),
                                "example_key_full": out_row.get("example_key_full", ""),
                                "group": out_row.get("group", ""),
                                "gold_example_role": out_row.get("gold_example_role", ""),
                                "target_sentence": out_row.get("target_sentence", ""),
                                "gold_eid": out_row.get("gold_eid", ""),
                                "gold_span_key": out_row.get("gold_span_key", "") or sr.get("span_key", ""),
                                "gold_span_window": out_row.get("gold_span_window", ""),
                                "rerank_status": out_row.get("rerank_status", ""),
                                "rerank_decision_line": out_row.get("rerank_decision_line", ""),
                                "rerank_selected_eid": out_row.get("rerank_selected_eid", ""),
                                "status_llm": span_status_llm,
                                "eid_no_pred_reason": "",
                                "candidate_e_ids_all": out_row.get("candidate_e_ids_all", ""),
                                "candidate_span_keys_all": out_row.get("candidate_span_keys_all", ""),
                                "span_rerank_span_key": sr.get("span_key", ""),
                                "span_rerank_match_key": sr.get("span_bundle_key", ""),
                                "span_rerank_status": sr.get("status", ""),
                                "span_rerank_decision_line": sr.get("decision_line", ""),
                                "span_rerank_selected_eid": sr.get("selected_eid", ""),
                                "span_rerank_selected_eids": ";".join(selected_list),
                                "span_rerank_protocol_ok": _tf(sr.get("protocol_ok")),
                                "span_rerank_none_reason": sr.get("none_reason", ""),
                            })
                    elif str(out_row.get("row_kind") or "") == "gold_only":
                        if gold_role == "neg_target_absent":
                            # Placeholder span row for B-group neg_target_absent gold-only entries.
                            span_min_rows.append({
                                "row_kind": out_row.get("row_kind", ""),
                                "view": out_row.get("view", ""),
                                "match_key": out_row.get("match_key", ""),
                                "example_id": out_row.get("example_id", ""),
                                "instance_id": out_row.get("instance_id", ""),
                                "example_key_full": out_row.get("example_key_full", ""),
                                "group": out_row.get("group", ""),
                                "gold_example_role": out_row.get("gold_example_role", ""),
                                "target_sentence": out_row.get("target_sentence", ""),
                                "gold_eid": out_row.get("gold_eid", ""),
                                "gold_span_key": out_row.get("gold_span_key", "") or (
                                    (str(out_row.get("candidate_span_keys_all") or "").split(";")[0])
                                    if str(out_row.get("candidate_span_keys_all") or "")
                                    else ""
                                ),
                                "gold_span_window": out_row.get("gold_span_window", ""),
                                "rerank_status": out_row.get("rerank_status", ""),
                                "rerank_decision_line": out_row.get("rerank_decision_line", ""),
                                "rerank_selected_eid": out_row.get("rerank_selected_eid", ""),
                                "status_llm": out_row.get("status_llm", ""),
                                "eid_no_pred_reason": "",
                                "candidate_e_ids_all": out_row.get("candidate_e_ids_all", ""),
                                "candidate_span_keys_all": out_row.get("candidate_span_keys_all", ""),
                                "span_rerank_span_key": "",
                                "span_rerank_match_key": str(out_row.get("match_key") or ""),
                                "span_rerank_status": "gold_only_placeholder",
                                "span_rerank_decision_line": "",
                                "span_rerank_selected_eid": "",
                                "span_rerank_selected_eids": "",
                                "span_rerank_protocol_ok": "",
                                "span_rerank_none_reason": "",
                            })
                        elif gold_role in pos_roles:
                            # Explicit no-detection marker for B-group positives where
                            # no span_reranks were produced by rule detect.
                            span_min_rows.append({
                                "row_kind": out_row.get("row_kind", ""),
                                "view": out_row.get("view", ""),
                                "match_key": out_row.get("match_key", ""),
                                "example_id": out_row.get("example_id", ""),
                                "instance_id": out_row.get("instance_id", ""),
                                "example_key_full": out_row.get("example_key_full", ""),
                                "group": out_row.get("group", ""),
                                "gold_example_role": out_row.get("gold_example_role", ""),
                                "target_sentence": out_row.get("target_sentence", ""),
                                "gold_eid": out_row.get("gold_eid", ""),
                                "gold_span_key": out_row.get("gold_span_key", "") or (
                                    (str(out_row.get("candidate_span_keys_all") or "").split(";")[0])
                                    if str(out_row.get("candidate_span_keys_all") or "")
                                    else ""
                                ),
                                "gold_span_window": out_row.get("gold_span_window", ""),
                                "rerank_status": out_row.get("rerank_status", ""),
                                "rerank_decision_line": out_row.get("rerank_decision_line", ""),
                                "rerank_selected_eid": "",
                                "status_llm": "FN",
                                "eid_no_pred_reason": "",
                                "candidate_e_ids_all": out_row.get("candidate_e_ids_all", ""),
                                "candidate_span_keys_all": out_row.get("candidate_span_keys_all", ""),
                                "span_rerank_span_key": "",
                                "span_rerank_match_key": str(out_row.get("match_key") or ""),
                                "span_rerank_status": "rule_could_not_detect",
                                "span_rerank_decision_line": "",
                                "span_rerank_selected_eid": "",
                                "span_rerank_selected_eids": "",
                                "span_rerank_protocol_ok": "",
                                "span_rerank_none_reason": "",
                            })
        still_empty_but_has_candidates = 0
        for key, entry in ssot.items():
            if entry.get("has_candidates") and not min_has_pred_or_candidate.get(key, False):
                still_empty_but_has_candidates += 1
        logger.info(
            "[eval][for_users_min_backfill] backfilled_pred_eid=%s backfilled_pred_span_key=%s",
            backfilled_pred_eid,
            backfilled_pred_span_key,
        )
        logger.info(
            "[eval][for_users_min_backfill] backfilled_candidate_e_id=%s backfilled_candidate_span_key=%s",
            backfilled_candidate_e_id,
            backfilled_candidate_span_key,
        )
        logger.info(
            "[eval][for_users_min_backfill] still_empty_but_has_candidates=%s",
            still_empty_but_has_candidates,
        )
        logger.info(
            "[eval][for_users_min_candidates_all] filled_all_cols=%s rows_with_candidates=%s",
            filled_all_cols,
            rows_with_candidates,
        )
        logger.info(
            "[eval][for_users_min_candidates_all] rows_with_candidates_but_empty_all=%s",
            rows_with_candidates_but_empty_all,
        )
        logger.info("[eval][for_users_min_projection] rows_written=%d", min_count)
        logger.info("[eval][for_users_min_projection] views=%s", min_export_views)
        logger.info(
            "[for_users][min_export] variant=%s path=%s rows=%s",
            variant,
            min_csv_path,
            len(rows),
        )
        strict_overall = (views.get("strict") or {}).get("overall", {})
        logger.info(
            "[for_users][min_export] variant=%s tp=%s fp=%s fn=%s pred_path=%s",
            variant,
            strict_overall.get("tp", ""),
            strict_overall.get("fp", ""),
            strict_overall.get("fn", ""),
            pred_path,
        )
        min_bytes = min_csv_path.stat().st_size
        logger.info(
            "[eval][for_users_min_candidates_all] wrote_min_csv=%s bytes=%s",
            min_csv_path,
            min_bytes,
        )
        if variant == "llm":
            # Mark span rows that are valid for B-metric aggregation.
            for _span_row in span_min_rows:
                _span_key = str(_span_row.get("span_rerank_span_key") or "")
                _gold_span_key = str(_span_row.get("gold_span_key") or "")
                _gold_role = str(_span_row.get("gold_example_role") or "")
                if _gold_role == "neg_target_absent":
                    if _span_key == "":
                        _span_row["target_matched_span"] = "correctly_selected_no_span"
                    elif _gold_span_key and _span_key == _gold_span_key:
                        _span_row["target_matched_span"] = "wrongly_matched"
                    else:
                        _span_row["target_matched_span"] = "mismatched_span"
                else:
                    _span_row["target_matched_span"] = "yes" if (_span_key == "" or _span_key == _gold_span_key) else "no"

            for _span_row in span_min_rows:
                _tm = str(_span_row.get("target_matched_span") or "")
                _span_row["include_in_b_metrics"] = "true" if _tm in {"yes", "correctly_selected_no_span", "wrongly_matched"} else "false"

            for _span_row in span_min_rows:
                _selected = str(_span_row.get("span_rerank_selected_eid") or "").strip()
                _status = str(_span_row.get("span_rerank_status") or "").strip()
                _decision = str(_span_row.get("span_rerank_decision_line") or "").strip().upper()
                _reason = ""
                if not _selected:
                    if _status in {"gold_only_placeholder", "rule_could_not_detect"}:
                        _reason = "rule_could_not_detect"
                    elif "DECISION: NONE" in _decision:
                        _reason = "rule_detected_llm_decision_none"
                    elif _status == "no_response":
                        _reason = "llm_no_response"
                    elif _status == "applied":
                        _reason = "applied_but_no_selected_eid"
                _span_row["eid_no_pred_reason"] = _reason

            b_num_tp = b_num_tn = b_num_fp = b_num_fn = 0
            for _span_row in span_min_rows:
                if str(_span_row.get("group") or "") != "b":
                    continue
                if str(_span_row.get("include_in_b_metrics") or "").lower() != "true":
                    continue
                _st = str(_span_row.get("status_llm") or "")
                if _st == "TP":
                    b_num_tp += 1
                elif _st == "TN":
                    b_num_tn += 1
                elif _st == "FP":
                    b_num_fp += 1
                elif _st == "FN":
                    b_num_fn += 1

            b_precision = _safe_ratio(b_num_tp, b_num_tp + b_num_fp)
            b_recall = _safe_ratio(b_num_tp, b_num_tp + b_num_fn)
            if b_precision and b_recall:
                _bp = float(b_precision)
                _br = float(b_recall)
                b_f1 = str((2.0 * _bp * _br) / (_bp + _br)) if (_bp + _br) > 0 else ""
            else:
                b_f1 = ""

            b_span_unique_example_keys = {
                str(_span_row.get("example_key_full") or "").strip()
                for _span_row in span_min_rows
                if str(_span_row.get("group") or "") == "b"
                and str(_span_row.get("example_key_full") or "").strip()
            }

            # Write summary values in first row only (spreadsheet-friendly top-right view).
            if span_min_rows:
                span_min_rows[0]["num_tp"] = str(b_num_tp)
                span_min_rows[0]["num_tn"] = str(b_num_tn)
                span_min_rows[0]["num_fp"] = str(b_num_fp)
                span_min_rows[0]["num_fn"] = str(b_num_fn)
                span_min_rows[0]["metric_precision"] = b_precision
                span_min_rows[0]["metric_recall"] = b_recall
                span_min_rows[0]["metric_f1"] = b_f1
                span_min_rows[0]["count_b_gold_unique_examples"] = str(len(b_gold_unique_example_keys_for_span))
                span_min_rows[0]["count_b_span_unique_examples"] = str(len(b_span_unique_example_keys))
                span_min_rows[0]["note"] = "(1) 성능(recall, precision, f1, num_tp, num_tn, num_fp, num_fn)을 점검하려면 include_in_b_metrics=true 로 필터링한 후 집계하면 됨"
                if len(span_min_rows) > 1:
                    span_min_rows[1]["note"] = "(2) 골드예문 수와 그룹 b의 문장 수를 비교하고 싶을 때에는 count_b_gold_unique_examples와 count_b_span_unique_examples를 비교하면 됨."

            span_min_csv_path = min_csv_path.with_name("eval_min_encoder_llm_spans.csv")
            with span_min_csv_path.open("w", encoding="utf-8", newline="") as fp_span:
                span_writer = csv.DictWriter(fp_span, fieldnames=span_min_fieldnames)
                span_writer.writeheader()
                for span_row in span_min_rows:
                    span_writer.writerow({k: span_row.get(k, "") for k in span_min_fieldnames})
            logger.info(
                "[for_users][min_export_spans] path=%s rows=%s b_num_tp=%s b_num_tn=%s b_num_fp=%s b_num_fn=%s",
                span_min_csv_path,
                len(span_min_rows),
                b_num_tp,
                b_num_tn,
                b_num_fp,
                b_num_fn,
            )
        logger.info(
            "[DEBUG][mincsv-write] rows=%s type_counts=%s norm_counts=%s",
            min_count,
            dict(min_type_counts),
            dict(min_norm_counts),
        )
        if fn_subbucket_counts:
            logger.info(
                "[for_users][SSOT] fn_subbucket_counts=%s",
                dict(fn_subbucket_counts),
            )
        if metric_strict_sample:
            logger.info(
                "[for_users][SSOT] overall_summary(strict) metrics=%s",
                metric_strict_sample,
            )
        if missing_counter_keys:
            logger.info(
                "[for_users][SSOT] missing counter key=%s",
                sorted(missing_counter_keys),
            )
        for view in ("strict", "lenient"):
            tp_rows = explain_counts[view]["tp_rows"]
            fn_rows = explain_counts[view]["fn_rows"]
            gold_occ_rows = explain_counts[view]["gold_occ_rows"]
            if tp_rows + fn_rows != gold_occ_rows:
                logger.warning(
                    "[for_users][SSOT] gold_occ sanity mismatch view=%s tp+fn=%s gold=%s",
                    view,
                    tp_rows + fn_rows,
                    gold_occ_rows,
                )
    try:
        with fn_labels_path.open("w", encoding="utf-8-sig", newline="") as fp_labels:
            writer = csv.DictWriter(fp_labels, fieldnames=fn_labels_fieldnames)
            writer.writeheader()
            for row in fn_label_rows:
                writer.writerow(row)
        size_bytes = fn_labels_path.stat().st_size
        logger.info(
            "[for_users][fn_labels] path=%s size_bytes=%s",
            fn_labels_path,
            size_bytes,
        )
        logger.info(
            "[for_users][fn_labels] header head/tail=%r/%r",
            fn_labels_fieldnames[:6],
            fn_labels_fieldnames[-6:],
        )
        logger.info(
            "[for_users][fn_labels] missing_key_counts=%s",
            dict(fn_labels_missing_key_counts),
        )
        logger.info(
            "[for_users][fn_labels] rows_written=%s",
            len(fn_label_rows),
        )
    except Exception as exc:
        logger.warning("[for_users][fn_labels] write failed: %s", exc)
    try:
        raw_unique: set[str] = set()
        with min_csv_path.open("r", encoding="utf-8", errors="replace", newline="") as fp_chk:
            reader = csv.DictReader(fp_chk)
            logger.info(
                "[for_users][SSOT] mincsv columns(sample)=%r",
                (reader.fieldnames or [])[:15],
            )
            if reader.fieldnames:
                logger.info(
                    "[for_users][SSOT] mincsv header tail=%r",
                    reader.fieldnames[-10:],
                )
            expected_new = [
                "fn_subbucket",
                "metric_strict_precision",
                "metric_strict_recall",
                "metric_strict_f1",
                "metric_lenient_precision_a",
                "metric_lenient_recall_a",
                "metric_lenient_f1_a",
                "count_strict_tp",
                "count_strict_fp",
                "count_strict_fn",
                "count_lenient_tp",
                "count_lenient_tn",
                "count_lenient_fp",
                "count_lenient_fn",
                "count_strict_gold_occ",
                "count_lenient_gold_occ",
                "count_strict_gold_occ_rows",
                "count_lenient_gold_occ_rows",
                "count_strict_tp_rows",
                "count_lenient_tp_rows",
                "count_strict_fn_rows",
                "count_lenient_fn_rows",
                "count_strict_pred_ignored_rows",
                "count_lenient_pred_ignored_rows",
                "label_fn_main",
                "label_fn_detail",
                "label_action_hint",
                "label_owner",
                "label_done",
            ]
            missing = [name for name in expected_new if name not in (reader.fieldnames or [])]
            if missing:
                logger.warning("[for_users][SSOT] mincsv missing columns=%r", missing)
            counter_cols = [
                "count_strict_tp",
                "count_strict_fp",
                "count_strict_fn",
                "count_lenient_tp",
                "count_lenient_tn",
                "count_lenient_fp",
                "count_lenient_fn",
                "count_strict_gold_occ",
                "count_lenient_gold_occ",
                "count_strict_gold_occ_rows",
                "count_lenient_gold_occ_rows",
                "count_strict_tp_rows",
                "count_lenient_tp_rows",
                "count_strict_fn_rows",
                "count_lenient_fn_rows",
                "count_strict_pred_ignored_rows",
                "count_lenient_pred_ignored_rows",
            ]
            label_cols = [
                "label_fn_main",
                "label_fn_detail",
                "label_action_hint",
                "label_owner",
                "label_done",
            ]
            bad_counter_rows = 0
            label_main_filled = 0
            for row in reader:
                val = (row.get("candidate_in_neg_gold") or "").strip()
                if val:
                    raw_unique.add(val)
                filled_counter = any((row.get(c) or "").strip() for c in counter_cols)
                if row.get("row_kind") == "overall_summary":
                    if not filled_counter and any(c in (reader.fieldnames or []) for c in counter_cols):
                        bad_counter_rows += 1
                else:
                    if filled_counter:
                        bad_counter_rows += 1
                if (row.get("label_fn_main") or "").strip():
                    label_main_filled += 1
            if bad_counter_rows:
                logger.warning(
                    "[for_users][SSOT] mincsv counter fill rule violated rows=%s",
                    bad_counter_rows,
                )
            if reader.fieldnames:
                logger.info(
                    "[for_users][SSOT] mincsv label header tail=%r",
                    reader.fieldnames[-10:],
                )
            logger.info(
                "[for_users][SSOT] label_fn_main filled=%s",
                label_main_filled,
            )
        raw_list = sorted(raw_unique)[:10]
        logger.info(
            "[for_users][SSOT] candidate_in_neg_gold unique=%r",
            raw_list,
        )
        bad = [v for v in raw_unique if v not in ("TRUE", "FALSE")]
        if bad:
            raise RuntimeError(f"mincsv candidate_in_neg_gold bad={bad[:10]}")
    except Exception as exc:
        logger.warning(
            "[for_users][SSOT] mincsv candidate_in_neg_gold unique(sample) failed: %s",
            exc,
        )
    try:
        with fn_labels_path.open("r", encoding="utf-8-sig", errors="replace", newline="") as fp_labels_chk:
            reader = csv.DictReader(fp_labels_chk)
            fields = reader.fieldnames or []
            label_tail = [
                "label_fn_main",
                "label_fn_detail",
                "label_action_hint",
                "label_owner",
                "label_done",
            ]
            missing_labels = [name for name in label_tail if name not in fields]
            if missing_labels:
                logger.warning("[for_users][fn_labels] missing labels=%r", missing_labels)
            required_cols = ["status", "pred_eid", "pred_span_key"]
            missing_required = [name for name in required_cols if name not in fields]
            if missing_required:
                logger.warning("[for_users][fn_labels] missing required=%r", missing_required)
            if fields and fields[-5:] != label_tail:
                logger.warning(
                    "[for_users][fn_labels] label tail mismatch=%r",
                    fields[-5:],
                )
            sample_rows: list[tuple[str, str, str]] = []
            label_main_filled = 0
            status_unique: set[str] = set()
            for row in reader:
                if (row.get("label_fn_main") or "").strip():
                    label_main_filled += 1
                val = (row.get("status") or "").strip()
                if val:
                    status_unique.add(val)
                if len(sample_rows) < 3:
                    sample_rows.append(
                        (
                            str(row.get("match_key") or ""),
                            str(row.get("status") or ""),
                            str(row.get("gold_eid") or ""),
                            str(row.get("gold_span_key") or ""),
                            str(row.get("pred_eid") or ""),
                            str(row.get("pred_span_key") or ""),
                        )
                    )
            logger.info(
                "[for_users][fn_labels] sample_rows=%s",
                sample_rows,
            )
            logger.info(
                "[for_users][fn_labels] label_fn_main filled=%s",
                label_main_filled,
            )
            if status_unique:
                logger.info(
                    "[for_users][fn_labels] status_unique=%r",
                    sorted(status_unique),
                )
    except Exception as exc:
        logger.warning("[for_users][fn_labels] read check failed: %s", exc)

    fn_breakdown_path = csv_path.with_name("eval_latest_fn_breakdown.csv")
    fn_categories = [
        "no_candidates",
        "candidate_but_low_conf",
        "to_llm_but_not_confirmed",
        "rerank_confirm_but_still_fn",
    ]
    fn_counts = {k: 0 for k in fn_categories}
    fn_samples: dict[str, list[str]] = {k: [] for k in fn_categories}
    rows_jsonl_by_key: dict[tuple[str, str], dict[str, Any]] = {}
    n_key_miss = 0
    n_rows_key_missing = 0
    fn_keys: list[tuple[str, str]] = []
    rows_source = str(rows_jsonl_path)

    if rows_jsonl:
        for item in rows_jsonl:
            if item.get("row_kind") != "gold_occurrence" or item.get("view") != "strict":
                continue
            full = item.get("full") or {}
            record = full.get("record") or {}
            ex_id = record.get("example_id")
            inst_id = record.get("instance_id")
            if ex_id is None or inst_id is None:
                n_rows_key_missing += 1
                continue
            rows_jsonl_by_key[(str(ex_id), str(inst_id))] = record

    try:
        with fn_labels_path.open("r", encoding="utf-8-sig", errors="replace", newline="") as fp_labels:
            reader = csv.DictReader(fp_labels)
            for row in reader:
                if row.get("view") != "strict":
                    continue
                if row.get("row_kind") != "gold_occurrence":
                    continue
                if row.get("status") != "FN":
                    continue
                ex_id = row.get("example_id")
                inst_id = row.get("instance_id")
                if ex_id is None or inst_id is None:
                    continue
                fn_keys.append((str(ex_id), str(inst_id)))
    except Exception:
        fn_keys = []

    n_fn_total = len(fn_keys)
    for ex_id, inst_id in fn_keys:
        category = "candidate_but_low_conf"
        record = rows_jsonl_by_key.get((ex_id, inst_id))
        if record is None:
            n_key_miss += 1
            category = "candidate_but_low_conf"
        else:
            candidates = record.get("candidates") or record.get("silver_labels") or []
            if not isinstance(candidates, list):
                candidates = []
            if not candidates:
                category = "no_candidates"
            else:
                rerank = record.get("rerank") or {}
                rerank_status = rerank.get("status")
                selected_eid = rerank.get("selected_eid")
                decision_line = rerank.get("decision_line")
                has_to_llm = any(
                    isinstance(cand, dict) and cand.get("to_llm") is True for cand in candidates
                )
                if rerank_status == "applied" and selected_eid:
                    category = "rerank_confirm_but_still_fn"
                elif has_to_llm and (
                    rerank_status in {"no_response", "guard_reject", "parse_fail", "applied_none"}
                    or not decision_line
                ):
                    category = "to_llm_but_not_confirmed"
                else:
                    for cand in candidates:
                        reason = cand.get("routing_reason")
                        if isinstance(reason, str) and "low_conf" in reason.lower():
                            category = "candidate_but_low_conf"
                            break
        fn_counts[category] += 1
        match_key = f"{ex_id}#{inst_id}"
        if match_key and len(fn_samples[category]) < 5:
            fn_samples[category].append(match_key)

    try:
        with fn_breakdown_path.open("w", encoding="utf-8-sig", newline="") as fp_breakdown:
            writer = csv.DictWriter(
                fp_breakdown,
                fieldnames=["view", "fn_category", "count", "sample_match_keys"],
            )
            writer.writeheader()
            for key in fn_categories:
                writer.writerow(
                    {
                        "view": "strict",
                        "fn_category": key,
                        "count": fn_counts.get(key, 0),
                        "sample_match_keys": ";".join(fn_samples.get(key, [])),
                    }
                )
    except Exception:
        pass

    breakdown_sum = sum(fn_counts.values())
    logger.info(
        "[eval][fn_breakdown] strict_fn_total=%s breakdown_sum=%s n_key_miss=%s n_rows_key_missing=%s rows_source=%s fn_labels_path=%s",
        n_fn_total,
        breakdown_sum,
        n_key_miss,
        n_rows_key_missing,
        rows_source,
        fn_labels_path,
    )

    fn_breakdown_json_path = outputs_dir / "fn_breakdown.json"
    fn_samples_jsonl_path = outputs_dir / "fn_samples.jsonl"
    fn_buckets = [
        "no_candidate",
        "candidate_filtered",
        "boundary_error",
        "rerank_not_applied",
        "unknown",
    ]
    bucket_counts = {k: 0 for k in fn_buckets}
    bucket_samples: dict[str, list[list[str | None]]] = {k: [] for k in fn_buckets}
    pred_source_label = "unknown"
    if pred_path.name == "infer_candidates.reranked.jsonl":
        pred_source_label = "infer_step2_rerank"
    elif pred_path.name == "infer_candidates.jsonl":
        pred_source_label = "infer_step1"
    elif pred_path.name == "silver.jsonl":
        pred_source_label = "build_silver"

    record_by_key: dict[tuple[str, str], dict[str, Any]] = {}
    for item in rows_jsonl:
        if item.get("row_kind") != "gold_occurrence" or item.get("view") != "strict":
            continue
        full = item.get("full") or {}
        record = full.get("record") or {}
        ex_id = record.get("example_id")
        inst_id = record.get("instance_id")
        if ex_id is not None and inst_id is not None:
            record_by_key[(str(ex_id), str(inst_id))] = record

    def _parse_segments(val: Any) -> Any:
        if not val:
            return None
        if isinstance(val, str):
            try:
                return json.loads(val)
            except Exception:
                return None
        return val

    fn_samples: list[dict[str, Any]] = []
    strict_fn_rows = [
        row
        for row in rows
        if row.get("row_kind") == "gold_occurrence"
        and row.get("view") == "strict"
        and row.get("status") == "FN"
    ]
    for row in strict_fn_rows:
        example_id = row.get("example_id")
        instance_id = row.get("instance_id")
        gold_eid = row.get("gold_eid")
        pred_eid = row.get("pred_eid")
        gold_span_key = row.get("gold_span_key")
        pred_span_key = row.get("pred_span_key")
        bucket = "unknown"
        if not pred_eid:
            bucket = "no_candidate"
        elif gold_eid and pred_eid != gold_eid:
            bucket = "candidate_filtered"
        elif gold_eid and pred_eid == gold_eid and gold_span_key and pred_span_key and pred_span_key != gold_span_key:
            bucket = "boundary_error"
        else:
            record = None
            if example_id is not None and instance_id is not None:
                record = record_by_key.get((str(example_id), str(instance_id)))
            if pred_source_label == "infer_step2_rerank" and record:
                rerank = record.get("rerank") or {}
                rerank_status = rerank.get("status")
                if rerank_status in {"no_response", "guard_reject", "parse_fail", "applied_none"}:
                    bucket = "rerank_not_applied"
        bucket_counts[bucket] += 1
        if len(bucket_samples[bucket]) < 20:
            bucket_samples[bucket].append([str(example_id) if example_id is not None else None, str(gold_eid) if gold_eid is not None else None])
        record = None
        split = None
        if example_id is not None and instance_id is not None:
            record = record_by_key.get((str(example_id), str(instance_id)))
            if record:
                split = record.get("split")
        fn_samples.append(
            {
                "example_id": example_id if example_id is not None else None,
                "e_id": gold_eid if gold_eid is not None else None,
                "split": split if split is not None else None,
                "bucket": bucket,
                "pred_source": pred_source_label,
                "pred_e_id": pred_eid if pred_eid is not None and pred_eid != "" else None,
                "span_gold": _parse_segments(row.get("gold_span_segments")),
                "span_pred": _parse_segments(row.get("pred_span_segments")),
            }
        )

    if strict_fn_rows and not fn_samples:
        raise ConfigError("fn_samples.jsonl empty while strict FN exists")

    try:
        write_json(
            fn_breakdown_json_path,
            {
                "total_fn_strict": len(strict_fn_rows),
                "buckets": bucket_counts,
                "top_samples": bucket_samples,
            },
            indent=2,
        )
        with fn_samples_jsonl_path.open("w", encoding="utf-8") as fp_fn:
            for row in fn_samples:
                write_jsonl_line(fp_fn, row)
    except Exception as exc:
        raise ConfigError(f"fn_breakdown write failed: {exc}")

    candidate_filtered_breakdown_path = outputs_dir / "candidate_filtered_breakdown.json"
    candidate_filtered_samples_path = outputs_dir / "candidate_filtered_samples.jsonl"
    candidate_filtered_reason_counts: Counter[str] = Counter()
    candidate_filtered_samples: list[dict[str, Any]] = []
    candidate_filtered_sample_keys: dict[str, list[str]] = defaultdict(list)
    candidate_filtered_total = 0
    hard_fail_triggered_count = 0
    json_parse_fail_counts: Counter[str] = Counter()

    def _is_truthy(value: Any) -> bool:
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return value != 0
        if isinstance(value, str):
            return value.strip().lower() in {"true", "1", "yes", "y"}
        return False

    def _lower_text(value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip().lower()

    def _safe_parse_json(value: Any, *, key: str) -> None:
        if value is None:
            return
        if isinstance(value, str) and value.strip() == "":
            return
        try:
            json.loads(value)
        except Exception:
            json_parse_fail_counts[key] += 1

    for row in strict_fn_rows:
        pred_eid = row.get("pred_eid")
        gold_eid = row.get("gold_eid")
        if not pred_eid:
            continue
        if not gold_eid or pred_eid == gold_eid:
            continue

        candidate_filtered_total += 1
        match_key = row.get("match_key") or ""
        example_id = row.get("example_id")
        instance_id = row.get("instance_id")
        gold_span_key = row.get("gold_span_key")
        pred_span_key = row.get("pred_span_key")
        pred_hard_fail_triggered = _is_truthy(row.get("pred_hard_fail_triggered"))
        if pred_hard_fail_triggered:
            hard_fail_triggered_count += 1

        pred_hard_fail_reasons = row.get("pred_hard_fail_reasons") or ""
        pred_stage_hits_json = row.get("pred_stage_hits_json") or ""
        debug_verify_json = row.get("debug_verify_json") or ""
        debug_context_json = row.get("debug_context_json") or ""
        pred_ignored_reason = row.get("pred_ignored_reason") or ""
        pred_triage = row.get("pred_triage") or ""

        _safe_parse_json(pred_hard_fail_reasons, key="hard_fail_reasons")
        _safe_parse_json(pred_stage_hits_json, key="stage_hits")
        _safe_parse_json(debug_verify_json, key="debug_verify")
        _safe_parse_json(debug_context_json, key="debug_context")

        text = (
            f"{pred_hard_fail_reasons} {pred_stage_hits_json} {debug_verify_json} "
            f"{debug_context_json} {pred_ignored_reason} {pred_triage}"
        ).lower()
        reason_code = "unknown"
        reason_detail = "unknown"

        if pred_hard_fail_triggered or (
            "hard_fail" in text
            or "hard fail" in text
            or "n_candidates_discarded_by_hard_fail" in text
        ):
            reason_code = "verify_hard_fail"
            reason_detail = "hard_fail_triggered" if pred_hard_fail_triggered else "hard_fail in text"
        elif (
            "context_neg" in text
            or "context_negative" in text
            or ("context" in text and "neg" in text)
        ):
            reason_code = "context_neg_hit"
            if "context_neg" in _lower_text(debug_context_json) or "context_negative" in _lower_text(
                debug_context_json
            ):
                reason_detail = "context_neg_hit in debug_context_json"
            else:
                reason_detail = "context_neg_hit in text"
        elif (
            "allowed_sense" in text
            or "disallowed_sense" in text
            or ("sense" in text and "constraint" in text)
        ):
            reason_code = "sense_constraint"
            reason_detail = "sense_constraint in text"
        elif "pos" in text and (
            "mismatch" in text or "not match" in text or "seq" in text
        ):
            reason_code = "pos_mismatch"
            reason_detail = "pos_mismatch in text"
        elif "lemma" in text and ("mismatch" in text or "not match" in text):
            reason_code = "lemma_mismatch"
            reason_detail = "lemma_mismatch in text"
        elif "triage_filtered" in text or (
            "triage" in text and ("filtered" in text or "blocked" in text)
        ):
            reason_code = "triage_filtered"
            reason_detail = "triage_filtered in text"
        elif "non_target" in text or "out_of_scope" in text:
            reason_code = "non_target"
            reason_detail = "non_target in text"

        candidate_filtered_reason_counts[reason_code] += 1
        if match_key and len(candidate_filtered_sample_keys[reason_code]) < 3:
            if match_key not in candidate_filtered_sample_keys[reason_code]:
                candidate_filtered_sample_keys[reason_code].append(str(match_key))

        hint_source = ""
        for field_val in (
            pred_hard_fail_reasons,
            pred_stage_hits_json,
            debug_verify_json,
            debug_context_json,
            pred_ignored_reason,
            pred_triage,
        ):
            if isinstance(field_val, str) and field_val.strip():
                hint_source = field_val.strip()
                break
            if field_val:
                hint_source = str(field_val).strip()
                break
        hint = hint_source.replace("\n", " ").replace("\t", " ").strip()
        if len(hint) > 400:
            hint = hint[:400]

        if len(candidate_filtered_samples) < 200:
            candidate_filtered_samples.append(
                {
                    "match_key": match_key,
                    "example_id": example_id,
                    "instance_id": instance_id,
                    "gold_eid": gold_eid,
                    "pred_eid": pred_eid,
                    "gold_span_key": gold_span_key,
                    "pred_span_key": pred_span_key,
                    "fn_bucket": "candidate_filtered",
                    "fn_subbucket": row.get("fn_subbucket") or "",
                    "reason_code": reason_code,
                    "reason_detail": reason_detail,
                    "hint": hint,
                }
            )

    unknown_after_norm = candidate_filtered_reason_counts.get("unknown", 0)
    unknown_ratio = (
        unknown_after_norm / candidate_filtered_total if candidate_filtered_total > 0 else 0.0
    )
    breakdown_payload = {
        "total_candidate_filtered": candidate_filtered_total,
        "reason_code_counts": dict(candidate_filtered_reason_counts),
        "unknown_after_norm": unknown_after_norm,
        "unknown_ratio": unknown_ratio,
        "hard_fail_triggered_count": hard_fail_triggered_count,
        "json_parse_fail_counts": dict(json_parse_fail_counts),
        "sample_keys_by_reason": dict(candidate_filtered_sample_keys),
    }
    try:
        write_json(candidate_filtered_breakdown_path, breakdown_payload, indent=2)
    except Exception as exc:
        raise ConfigError(f"candidate_filtered_breakdown write failed: {exc}") from exc
    try:
        with candidate_filtered_samples_path.open("w", encoding="utf-8") as fp_cf:
            if candidate_filtered_samples:
                for row in candidate_filtered_samples:
                    write_jsonl_line(fp_cf, row)
    except Exception as exc:
        raise ConfigError(f"candidate_filtered_samples write failed: {exc}") from exc

    breakdown_bytes = candidate_filtered_breakdown_path.stat().st_size
    samples_bytes = candidate_filtered_samples_path.stat().st_size
    logger.info(
        "[eval][candidate_filtered_diag] total=%s unknown_after_norm=%s unknown_ratio=%.4f hard_fail_triggered=%s",
        candidate_filtered_total,
        unknown_after_norm,
        unknown_ratio,
        hard_fail_triggered_count,
    )
    logger.info(
        "[eval][candidate_filtered_diag] wrote breakdown=%s bytes=%s unique_reason_codes=%s",
        candidate_filtered_breakdown_path,
        breakdown_bytes,
        len(candidate_filtered_reason_counts),
    )
    logger.info(
        "[eval][candidate_filtered_diag] wrote samples=%s lines=%s bytes=%s",
        candidate_filtered_samples_path,
        len(candidate_filtered_samples),
        samples_bytes,
    )

    fn_no_candidate_map_breakdown_path = outputs_dir / "fn_no_candidate_map_breakdown.json"
    fn_no_candidate_map_samples_path = outputs_dir / "fn_no_candidate_map_samples.jsonl"
    no_candidate_fn_samples = [row for row in fn_samples if row.get("bucket") == "no_candidate"]
    total_no_candidate_fn = len(no_candidate_fn_samples)
    map_buckets = [
        "mapped_to_infer_step1_no_candidate",
        "has_candidates_but_no_pred",
        "missing_infer_step1_artifact",
    ]
    map_counts = {k: 0 for k in map_buckets}
    map_top_samples: dict[str, list[list[str | None]]] = {k: [] for k in map_buckets}
    infer_step1_candidates_path: Path | None = None
    infer_step1_run_dir_used: Path | None = None

    def _latest_infer_step1_output(artifacts_root: Path, exp_id: str) -> Path | None:
        root = artifacts_root / exp_id / "infer_step1"
        if not root.exists():
            return None
        run_dirs = sorted([p for p in root.iterdir() if p.is_dir()])
        if not run_dirs:
            return None
        candidate = run_dirs[-1] / "outputs" / "infer_candidates.jsonl"
        return candidate if candidate.exists() else None

    infer_step1_pick_reason = None
    if pred_path.name == "infer_candidates.jsonl":
        infer_step1_candidates_path = pred_path
        infer_step1_pick_reason = "direct_infer_step1"
    elif pred_path.name == "infer_candidates.reranked.jsonl":
        rerank_run_dir = pred_path.parent.parent
        config_path = rerank_run_dir / "config_resolved.yaml"
        input_pred_path = None
        if config_path.exists():
            try:
                rerank_cfg = yaml.safe_load(config_path.read_text(encoding="utf-8")) or {}
                rerank_block = rerank_cfg.get("infer_step2_rerank", {}) or {}
                input_pred_path = (
                    rerank_block.get("input_pred_path")
                    or rerank_block.get("pred_path")
                    or rerank_block.get("input_jsonl")
                )
            except Exception as exc:
                logger.warning("[eval][fn_map] failed to read rerank config: %s", exc)
        if input_pred_path:
            infer_step1_candidates_path = Path(str(input_pred_path))
            infer_step1_pick_reason = "rerank_trace"
        else:
            artifacts_root = _artifacts_root_from_outputs_dir(outputs_dir, logger)
            fallback = None
            infer_root = artifacts_root / run_context.exp_id / "infer_step1"
            if infer_root.exists():
                run_dirs = sorted([p for p in infer_root.iterdir() if p.is_dir()], reverse=True)
                for run_dir in run_dirs:
                    breakdown_path = run_dir / "outputs" / "no_candidate_breakdown.json"
                    if breakdown_path.exists():
                        fallback = run_dir / "outputs" / "infer_candidates.jsonl"
                        infer_step1_pick_reason = "fallback_has_no_candidate_breakdown"
                        break
                if fallback is None and run_dirs:
                    fallback = run_dirs[0] / "outputs" / "infer_candidates.jsonl"
                    infer_step1_pick_reason = "fallback_latest"
            if fallback is not None:
                infer_step1_candidates_path = fallback
                logger.info(
                    "[eval][fn_map] fallback_infer_step1_latest=true run_dir=%s",
                    fallback.parent.parent,
                )
    if infer_step1_candidates_path is None:
        artifacts_root = _artifacts_root_from_outputs_dir(outputs_dir, logger)
        fallback = None
        infer_root = artifacts_root / run_context.exp_id / "infer_step1"
        if infer_root.exists():
            run_dirs = sorted([p for p in infer_root.iterdir() if p.is_dir()], reverse=True)
            for run_dir in run_dirs:
                breakdown_path = run_dir / "outputs" / "no_candidate_breakdown.json"
                if breakdown_path.exists():
                    fallback = run_dir / "outputs" / "infer_candidates.jsonl"
                    infer_step1_pick_reason = "fallback_has_no_candidate_breakdown"
                    break
            if fallback is None and run_dirs:
                fallback = run_dirs[0] / "outputs" / "infer_candidates.jsonl"
                infer_step1_pick_reason = "fallback_latest"
        if fallback is not None:
            infer_step1_candidates_path = fallback
            logger.info(
                "[eval][fn_map] fallback_infer_step1_latest=true run_dir=%s",
                fallback.parent.parent,
            )

    if infer_step1_candidates_path and infer_step1_candidates_path.exists():
        infer_step1_run_dir_used = infer_step1_candidates_path.parent.parent
    else:
        infer_step1_candidates_path = None
    if infer_step1_run_dir_used is not None and infer_step1_pick_reason:
        logger.info(
            "[eval][fn_map] infer_step1_pick_reason=%s run_dir=%s",
            infer_step1_pick_reason,
            infer_step1_run_dir_used,
        )

    candidate_counts_by_example: dict[str, Counter[str]] = defaultdict(Counter)
    needed_example_ids = {
        str(row.get("example_id"))
        for row in no_candidate_fn_samples
        if row.get("example_id") is not None
    }
    if infer_step1_candidates_path:
        try:
            with infer_step1_candidates_path.open("r", encoding="utf-8") as fp:
                for line in fp:
                    line = line.strip()
                    if not line:
                        continue
                    record = json.loads(line)
                    example_id = record.get("example_id")
                    if example_id is None:
                        continue
                    example_id_key = str(example_id)
                    if example_id_key not in needed_example_ids:
                        continue
                    candidates = record.get("candidates") or []
                    if not isinstance(candidates, list):
                        continue
                    for cand in candidates:
                        if not isinstance(cand, dict):
                            continue
                        e_id = cand.get("e_id")
                        if e_id is None:
                            continue
                        candidate_counts_by_example[example_id_key][str(e_id)] += 1
        except Exception as exc:
            logger.warning("[eval][fn_map] failed to read infer_step1 candidates: %s", exc)
            infer_step1_candidates_path = None
            infer_step1_run_dir_used = None

    fn_map_samples: list[dict[str, Any]] = []
    for row in no_candidate_fn_samples:
        example_id = row.get("example_id")
        e_id = row.get("e_id")
        infer_has_any = None
        infer_count = None
        mapped_bucket = "missing_infer_step1_artifact"
        note = None
        if infer_step1_candidates_path is None:
            mapped_bucket = "missing_infer_step1_artifact"
            note = "infer_step1 artifact missing"
        elif example_id is not None and e_id is not None:
            example_key = str(example_id)
            e_id_key = str(e_id)
            infer_count = int(candidate_counts_by_example.get(example_key, {}).get(e_id_key, 0))
            infer_has_any = infer_count > 0
            if infer_has_any:
                mapped_bucket = "has_candidates_but_no_pred"
                note = "infer_step1 has candidate for eid"
            else:
                mapped_bucket = "mapped_to_infer_step1_no_candidate"
                note = "infer_step1 no candidate for eid"
        map_counts[mapped_bucket] += 1
        if len(map_top_samples[mapped_bucket]) < 20:
            map_top_samples[mapped_bucket].append(
                [
                    str(example_id) if example_id is not None else None,
                    str(e_id) if e_id is not None else None,
                ]
            )
        fn_map_samples.append(
            {
                "example_id": example_id if example_id is not None else None,
                "e_id": e_id if e_id is not None else None,
                "bucket_eval": "no_candidate",
                "infer_step1_run_dir_used": str(infer_step1_run_dir_used)
                if infer_step1_run_dir_used is not None
                else None,
                "infer_step1_has_any_candidate_for_eid": infer_has_any,
                "infer_step1_candidate_count_for_eid": infer_count,
                "mapped_bucket": mapped_bucket,
                "note": note,
            }
        )

    if sum(map_counts.values()) != total_no_candidate_fn:
        raise ConfigError(
            "fn_no_candidate_map: breakdown sum mismatch "
            f"sum={sum(map_counts.values())} total={total_no_candidate_fn}"
        )
    try:
        write_json(
            fn_no_candidate_map_breakdown_path,
            {
                "total_no_candidate_fn": total_no_candidate_fn,
                "mapped_to_infer_step1_no_candidate": map_counts["mapped_to_infer_step1_no_candidate"],
                "has_candidates_but_no_pred": map_counts["has_candidates_but_no_pred"],
                "missing_infer_step1_artifact": map_counts["missing_infer_step1_artifact"],
                "top_samples": map_top_samples,
            },
            indent=2,
        )
        with fn_no_candidate_map_samples_path.open("w", encoding="utf-8") as fp_map:
            for row in fn_map_samples:
                write_jsonl_line(fp_map, row)
    except Exception as exc:
        raise ConfigError(f"fn_no_candidate_map write failed: {exc}")

    try:
        with fn_no_candidate_map_samples_path.open("r", encoding="utf-8") as fp_map:
            map_lines = sum(1 for _ in fp_map)
    except Exception as exc:
        raise ConfigError(f"fn_no_candidate_map samples read failed: {exc}")
    if map_lines != total_no_candidate_fn:
        raise ConfigError(
            "fn_no_candidate_map: samples lines mismatch "
            f"lines={map_lines} total={total_no_candidate_fn}"
        )
    logger.info(
        "[eval][fn_map] no_candidate_fn=%s mapped_infer_step1_no_candidate=%s has_candidates_but_no_pred=%s missing_infer_step1_artifact=%s",
        total_no_candidate_fn,
        map_counts["mapped_to_infer_step1_no_candidate"],
        map_counts["has_candidates_but_no_pred"],
        map_counts["missing_infer_step1_artifact"],
    )

    mapped_subbuckets_path = outputs_dir / "mapped_infer_step1_no_candidate_subbuckets.json"
    infer_step1_breakdown_path = None
    if infer_step1_run_dir_used is not None:
        infer_step1_breakdown_path = (
            Path(infer_step1_run_dir_used) / "outputs" / "no_candidate_breakdown.json"
        )
    exists = bool(infer_step1_breakdown_path and infer_step1_breakdown_path.exists())
    logger.info(
        "[eval][fn_map] infer_step1_no_candidate_breakdown_path=%s exists=%s",
        infer_step1_breakdown_path,
        exists,
    )
    subbuckets = {}
    infer_step1_total = None
    note = None
    if not exists:
        note = "missing infer_step1 no_candidate_breakdown.json"
        logger.warning("[eval][fn_map] missing infer_step1 no_candidate_breakdown.json")
    else:
        try:
            breakdown = json.loads(infer_step1_breakdown_path.read_text(encoding="utf-8"))
            if isinstance(breakdown, dict):
                infer_step1_total = breakdown.get("total_no_candidate")
                subbuckets = breakdown.get("subbuckets") or {}
        except Exception as exc:
            note = "missing infer_step1 no_candidate_breakdown.json"
            logger.warning(
                "[eval][fn_map] failed to read infer_step1 no_candidate_breakdown.json: %s",
                exc,
            )
    mapped_count = map_counts["mapped_to_infer_step1_no_candidate"]
    if infer_step1_total is not None and infer_step1_total != mapped_count:
        note = f"mismatch: mapped={mapped_count} infer_step1_total={infer_step1_total}"
        logger.warning("[eval][fn_map] %s", note)
    payload = {
        "infer_step1_run_dir_used": str(infer_step1_run_dir_used)
        if infer_step1_run_dir_used is not None
        else "",
        "mapped_infer_step1_no_candidate": mapped_count,
        "infer_step1_no_candidate_total": infer_step1_total,
        "subbuckets": subbuckets,
        "note": note,
    }
    try:
        write_json(mapped_subbuckets_path, payload, indent=2)
    except Exception as exc:
        raise ConfigError(
            f"mapped_infer_step1_no_candidate_subbuckets write failed: {exc}"
        ) from exc
    mapped_bytes = mapped_subbuckets_path.stat().st_size
    logger.info(
        "[eval][fn_map] wrote mapped_infer_step1_no_candidate_subbuckets.json bytes=%s",
        mapped_bytes,
    )

    audit_path = outputs_dir / "mapped_no_candidate_detect_no_hit_label_audit.json"
    infer_step1_no_candidate_samples_path = None
    if infer_step1_run_dir_used is not None:
        infer_step1_no_candidate_samples_path = (
            Path(infer_step1_run_dir_used) / "outputs" / "no_candidate_samples.jsonl"
        )
    detect_no_hit_example_ids: set[str] = set()
    if infer_step1_no_candidate_samples_path and infer_step1_no_candidate_samples_path.exists():
        try:
            with infer_step1_no_candidate_samples_path.open("r", encoding="utf-8") as fp:
                for line in fp:
                    line = line.strip()
                    if not line:
                        continue
                    row = json.loads(line)
                    if row.get("subbucket") != "detect_no_hit":
                        continue
                    example_id = row.get("example_id")
                    if example_id is None:
                        continue
                    detect_no_hit_example_ids.add(str(example_id))
        except Exception as exc:
            raise ConfigError(f"detect_no_hit audit read failed: {exc}") from exc

    role_by_example: dict[str, Any] = {}
    for row in gold_rows:
        example_id = row.get("example_id")
        if example_id is None:
            continue
        key = str(example_id)
        if key in role_by_example:
            continue
        role = row.get("gold_example_role")
        if role is None or (isinstance(role, str) and role.strip() == ""):
            role_by_example[key] = None
        else:
            role_by_example[key] = str(role)

    n_mapped = map_counts["mapped_to_infer_step1_no_candidate"]
    detect_no_hit_count = 0
    missing_role_count = 0
    role_counts: Counter[str] = Counter()
    top_samples: list[dict[str, Any]] = []
    for row in fn_map_samples:
        if row.get("mapped_bucket") != "mapped_to_infer_step1_no_candidate":
            continue
        example_id = row.get("example_id")
        if example_id is None:
            continue
        example_key = str(example_id)
        if example_key not in detect_no_hit_example_ids:
            continue
        detect_no_hit_count += 1
        role = role_by_example.get(example_key)
        if role is None:
            missing_role_count += 1
        else:
            role_counts[str(role)] += 1
        if len(top_samples) < 30:
            top_samples.append(
                {
                    "example_id": example_id,
                    "e_id": row.get("e_id"),
                    "gold_example_role": role,
                }
            )

    audit_note = None
    if detect_no_hit_count <= 0:
        audit_note = "n_detect_no_hit==0 (join mismatch or no samples)"
        logger.warning("[eval][audit] %s", audit_note)
    elif missing_role_count > 0:
        audit_note = f"missing gold_example_role for {missing_role_count} examples"
    try:
        write_json(
            audit_path,
            {
                "n_mapped_to_infer_step1_no_candidate": n_mapped,
                "n_detect_no_hit": detect_no_hit_count,
                "gold_example_role_counts": dict(role_counts),
                "top_samples": top_samples,
                "note": audit_note,
            },
            indent=2,
        )
    except Exception as exc:
        raise ConfigError(f"mapped_no_candidate_detect_no_hit_label_audit write failed: {exc}")
    audit_bytes = audit_path.stat().st_size
    logger.info(
        "[eval][audit] wrote mapped_no_candidate_detect_no_hit_label_audit.json bytes=%s",
        audit_bytes,
    )
    logger.info(
        "[eval][audit] detect_no_hit=%s role_counts=%s",
        detect_no_hit_count,
        json.dumps(dict(role_counts), ensure_ascii=False),
    )

    join_debug_path = outputs_dir / "mapped_no_candidate_detect_no_hit_join_debug.json"
    left_keys = {
        str(row.get("example_id"))
        for row in fn_map_samples
        if row.get("mapped_bucket") == "mapped_to_infer_step1_no_candidate"
        and row.get("example_id") is not None
    }
    right_keys = set(detect_no_hit_example_ids)
    intersection = left_keys.intersection(right_keys)
    left_only = sorted(left_keys - right_keys)
    right_only = sorted(right_keys - left_keys)
    notes: list[str] = []
    if not left_keys:
        notes.append("left_keys empty")
    if not right_keys:
        notes.append("right_keys empty")
    if left_keys and right_keys and left_keys.isdisjoint(right_keys):
        notes.append("example_id sets disjoint")
    join_debug = {
        "infer_step1_run_dir_used": str(infer_step1_run_dir_used)
        if infer_step1_run_dir_used is not None
        else "",
        "n_left_mapped": len(left_keys),
        "n_right_detect_no_hit": len(right_keys),
        "n_intersection": len(intersection),
        "key_format": "example_id",
        "left_only_samples": [{"example_id": key} for key in left_only[:10]],
        "right_only_samples": [{"example_id": key} for key in right_only[:10]],
        "notes": notes,
    }
    logger.info(
        "[eval][audit_join] left=%s right=%s intersection=%s left_only=%s right_only=%s",
        len(left_keys),
        len(right_keys),
        len(intersection),
        len(left_only),
        len(right_only),
    )
    try:
        write_json(join_debug_path, join_debug, indent=2)
    except Exception as exc:
        raise ConfigError(f"mapped_no_candidate_detect_no_hit_join_debug write failed: {exc}")
    join_bytes = join_debug_path.stat().st_size
    logger.info(
        "[eval][audit_join] wrote mapped_no_candidate_detect_no_hit_join_debug.json bytes=%s",
        join_bytes,
    )
    intersection_samples_path = (
        outputs_dir / "mapped_no_candidate_detect_no_hit_intersection_samples.jsonl"
    )
    role_by_example_id: dict[str, Any] = {}
    gold_xlsx_by_example_id: dict[str, dict[str, Any]] = {}
    n_role_non_null = 0
    n_role_missing = 0
    role_hits = 0
    role_misses = 0
    miss_example_ids: list[str] = []
    gold_rows_by_example: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in gold_rows:
        example_id = row.get("example_id")
        if example_id is None:
            continue
        example_key = str(example_id).strip()
        gold_rows_by_example[example_key].append(row)
    gold_xlsx_path = Path(cfg["paths"]["gold_xlsx"])
    if not gold_xlsx_path.exists():
        raise ConfigError(f"gold.xlsx not found: {gold_xlsx_path}")
    gold_xlsx_bytes = gold_xlsx_path.stat().st_size
    logger.info(
        "[eval][audit_intersection] gold_xlsx_path=%s exists=%s bytes=%s",
        gold_xlsx_path,
        gold_xlsx_path.exists(),
        gold_xlsx_bytes,
    )
    try:
        from openpyxl import load_workbook

        gold_sheet_name = cfg.get("paths", {}).get("gold_sheet_name")
        wb = load_workbook(gold_xlsx_path, read_only=True, data_only=True)
        sheet_name = (
            gold_sheet_name
            if (gold_sheet_name and gold_sheet_name in wb.sheetnames)
            else wb.sheetnames[0]
        )
        ws = wb[sheet_name]
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        logger.info(
            "[eval][audit_intersection] gold_xlsx_headers=%s",
            json.dumps(header, ensure_ascii=False),
        )
        if "example_id" not in header or "gold_example_role" not in header:
            raise ConfigError(
                "gold.xlsx missing required columns: example_id, gold_example_role"
            )
        idx_example = header.index("example_id")
        idx_role = header.index("gold_example_role")
        idx_eid = header.index("e_id") if "e_id" in header else None
        idx_span_key = header.index("span_key") if "span_key" in header else None
        idx_span_segments = header.index("span_segments") if "span_segments" in header else None
        idx_sentence = header.index("target_sentence") if "target_sentence" in header else None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            example_id = row[idx_example] if idx_example < len(row) else None
            if example_id is None:
                continue
            example_key = str(example_id).strip()
            if example_key in role_by_example_id:
                continue
            role = row[idx_role] if idx_role < len(row) else None
            role_by_example_id[example_key] = role
            gold_xlsx_by_example_id[example_key] = {
                "gold_eid": row[idx_eid] if idx_eid is not None and idx_eid < len(row) else None,
                "gold_span_key": row[idx_span_key]
                if idx_span_key is not None and idx_span_key < len(row)
                else None,
                "gold_span_segments": row[idx_span_segments]
                if idx_span_segments is not None and idx_span_segments < len(row)
                else None,
                "target_sentence": row[idx_sentence]
                if idx_sentence is not None and idx_sentence < len(row)
                else None,
            }
            if role is not None and not (isinstance(role, str) and role.strip() == ""):
                n_role_non_null += 1
    except Exception as exc:
        raise ConfigError(f"failed to read gold_example_role from gold.xlsx: {exc}") from exc
    intersection_example_ids = sorted({str(x).strip() for x in intersection})
    try:
        with intersection_samples_path.open("w", encoding="utf-8") as fp_intersection:
            for example_id in intersection_example_ids:
                example_key = str(example_id).strip()
                gold_role = role_by_example_id.get(example_key)
                gold_items = gold_rows_by_example.get(example_key, [])
                gold_item = gold_items[0] if gold_items else {}
                note = None
                if gold_role is None:
                    note = "missing_in_gold_xlsx"
                    n_role_missing += 1
                    role_misses += 1
                    if len(miss_example_ids) < 10:
                        miss_example_ids.append(example_key)
                else:
                    role_hits += 1
                write_jsonl_line(
                    fp_intersection,
                    {
                        "example_id": example_id,
                        "gold_example_role": gold_role,
                        "gold_eid": gold_item.get("e_id"),
                        "gold_span_key": gold_item.get("span_key"),
                        "gold_span_segments": gold_item.get("span_segments"),
                        "note": note,
                    },
                )
    except Exception as exc:
        raise ConfigError(
            f"mapped_no_candidate_detect_no_hit_intersection_samples write failed: {exc}"
        )
    logger.info(
        "[eval][audit_intersection] intersection_ids=%s role_hits=%s role_misses=%s miss_sample=%s",
        len(intersection_example_ids),
        role_hits,
        role_misses,
        json.dumps(miss_example_ids, ensure_ascii=False),
    )
    logger.info(
        "[eval][audit_intersection] role_source=gold_xlsx path=%s n_role_non_null=%s n_missing=%s",
        gold_xlsx_path,
        n_role_non_null,
        n_role_missing,
    )
    logger.info(
        "[eval][audit_intersection] wrote mapped_no_candidate_detect_no_hit_intersection_samples.jsonl lines=%s path=%s",
        len(intersection_example_ids),
        intersection_samples_path,
    )

    detect_no_hit_role_breakdown_path = (
        outputs_dir / "infer_step1_detect_no_hit_role_breakdown.json"
    )
    infer_step1_detect_no_hit_path = None
    if infer_step1_run_dir_used is not None:
        infer_step1_detect_no_hit_path = (
            Path(infer_step1_run_dir_used) / "outputs" / "no_candidate_samples.jsonl"
        )
    detect_no_hit_example_ids: set[str] = set()
    if infer_step1_detect_no_hit_path and infer_step1_detect_no_hit_path.exists():
        try:
            with infer_step1_detect_no_hit_path.open("r", encoding="utf-8") as fp:
                for line in fp:
                    line = line.strip()
                    if not line:
                        continue
                    row = json.loads(line)
                    if row.get("subbucket") != "detect_no_hit":
                        continue
                    example_id = row.get("example_id")
                    if example_id is None:
                        continue
                    detect_no_hit_example_ids.add(str(example_id).strip())
        except Exception as exc:
            raise ConfigError(
                f"infer_step1_detect_no_hit_role_breakdown read failed: {exc}"
            ) from exc
    n_detect_no_hit = len(detect_no_hit_example_ids)
    logger.info(
        "[eval][detect_no_hit_role] infer_step1_run_dir_used=%s n_detect_no_hit=%s",
        infer_step1_run_dir_used if infer_step1_run_dir_used is not None else "",
        n_detect_no_hit,
    )
    if n_detect_no_hit == 0:
        logger.warning("[eval][detect_no_hit_role] n_detect_no_hit=0 (no samples)")
    role_counts: Counter[str] = Counter()
    n_role_non_null = 0
    n_missing = 0
    top_examples_missing_role: list[str] = []
    for example_id in sorted(detect_no_hit_example_ids):
        role = role_by_example_id.get(example_id)
        if role is None or (isinstance(role, str) and role.strip() == ""):
            n_missing += 1
            if len(top_examples_missing_role) < 20:
                top_examples_missing_role.append(example_id)
            continue
        role_counts[str(role)] += 1
        n_role_non_null += 1
    breakdown_payload = {
        "infer_step1_run_dir_used": str(infer_step1_run_dir_used)
        if infer_step1_run_dir_used is not None
        else "",
        "n_detect_no_hit": n_detect_no_hit,
        "n_role_non_null": n_role_non_null,
        "n_missing_in_gold_xlsx": n_missing,
        "role_counts": dict(role_counts),
        "top_examples_missing_role": top_examples_missing_role,
    }
    try:
        write_json(detect_no_hit_role_breakdown_path, breakdown_payload, indent=2)
    except Exception as exc:
        raise ConfigError(
            f"infer_step1_detect_no_hit_role_breakdown write failed: {exc}"
        ) from exc
    detect_no_hit_bytes = detect_no_hit_role_breakdown_path.stat().st_size
    logger.info(
        "[eval][detect_no_hit_role] wrote infer_step1_detect_no_hit_role_breakdown.json bytes=%s n_role_non_null=%s n_missing=%s",
        detect_no_hit_bytes,
        n_role_non_null,
        n_missing,
    )

    locate_breakdown_path = outputs_dir / "infer_step1_locate_components_fail_role_breakdown.json"
    locate_pos_samples_path = (
        outputs_dir / "infer_step1_locate_components_fail_pos_samples.jsonl"
    )
    locate_pos_summary_path = (
        outputs_dir / "infer_step1_locate_components_fail_pos_summary.json"
    )
    locate_source_breakdown = None
    locate_source_samples = None
    if infer_step1_run_dir_used is not None:
        locate_source_breakdown = (
            Path(infer_step1_run_dir_used)
            / "outputs"
            / "infer_step1_locate_components_fail_breakdown.json"
        )
        locate_source_samples = (
            Path(infer_step1_run_dir_used)
            / "outputs"
            / "infer_step1_locate_components_fail_samples.jsonl"
        )
    locate_total = 0
    if locate_source_breakdown and locate_source_breakdown.exists():
        try:
            locate_payload = json.loads(locate_source_breakdown.read_text(encoding="utf-8"))
            locate_total = int(locate_payload.get("total_locate_components_fail") or 0)
        except Exception as exc:
            raise ConfigError(f"locate_components_fail breakdown read failed: {exc}") from exc

    role_counts: Counter[str] = Counter()
    n_role_non_null = 0
    n_missing_role = 0
    pos_roles = {"pos_conti", "pos_disconti"}
    pos_role_counts = {role: 0 for role in pos_roles}
    pos_samples: list[dict[str, Any]] = []
    if locate_source_samples and locate_source_samples.exists():
        try:
            with locate_source_samples.open("r", encoding="utf-8") as fp:
                for line in fp:
                    line = line.strip()
                    if not line:
                        continue
                    sample = json.loads(line)
                    example_id = sample.get("example_id")
                    example_key = str(example_id) if example_id is not None else ""
                    role = role_by_example_id.get(example_key)
                    if role is None or (isinstance(role, str) and role.strip() == ""):
                        n_missing_role += 1
                    else:
                        role_value = str(role)
                        role_counts[role_value] += 1
                        n_role_non_null += 1
                        if role_value in pos_roles:
                            pos_role_counts[role_value] = pos_role_counts.get(role_value, 0) + 1
                            pos_samples.append(
                                {
                                    "example_id": sample.get("example_id"),
                                    "instance_id": sample.get("instance_id"),
                                    "e_id": sample.get("e_id"),
                                    "reason_code": sample.get("reason_code"),
                                    "reason_detail": sample.get("reason_detail"),
                                    "gold_example_role": role_value,
                                }
                            )
        except Exception as exc:
            raise ConfigError(f"locate_components_fail samples read failed: {exc}") from exc

    locate_role_payload = {
        "infer_step1_run_dir_used": str(infer_step1_run_dir_used)
        if infer_step1_run_dir_used is not None
        else "",
        "n_locate_components_fail_total": locate_total,
        "role_counts": dict(role_counts),
        "n_role_non_null": n_role_non_null,
        "n_missing": n_missing_role,
    }
    try:
        write_json(locate_breakdown_path, locate_role_payload, indent=2)
    except Exception as exc:
        raise ConfigError(
            f"infer_step1_locate_components_fail_role_breakdown write failed: {exc}"
        ) from exc
    try:
        with locate_pos_samples_path.open("w", encoding="utf-8") as fp:
            if pos_samples:
                for row in pos_samples:
                    write_jsonl_line(fp, row)
            else:
                fp.write("\n")
    except Exception as exc:
        raise ConfigError(
            f"infer_step1_locate_components_fail_pos_samples write failed: {exc}"
        ) from exc
    locate_pos_summary_payload = {
        "n_locate_components_fail_total": locate_total,
        "n_pos_total": len(pos_samples),
        "pos_role_counts": pos_role_counts,
        "missing_role": n_missing_role,
    }
    try:
        write_json(locate_pos_summary_path, locate_pos_summary_payload, indent=2)
    except Exception as exc:
        raise ConfigError(
            f"infer_step1_locate_components_fail_pos_summary write failed: {exc}"
        ) from exc

    locate_breakdown_bytes = locate_breakdown_path.stat().st_size
    locate_pos_bytes = locate_pos_samples_path.stat().st_size
    logger.info(
        "[eval][locate_components_fail_role] infer_step1_run_dir_used=%s n_locate_components_fail=%s",
        infer_step1_run_dir_used if infer_step1_run_dir_used is not None else "",
        locate_total,
    )
    logger.info(
        "[eval][locate_components_fail_role] wrote infer_step1_locate_components_fail_role_breakdown.json bytes=%s n_role_non_null=%s n_missing=%s",
        locate_breakdown_bytes,
        n_role_non_null,
        n_missing_role,
    )
    logger.info(
        "[eval][locate_components_fail_pos] wrote pos_samples=%s lines=%s bytes=%s n_pos_total=%s",
        locate_pos_samples_path,
        len(pos_samples),
        locate_pos_bytes,
        len(pos_samples),
    )

    locate_pos_diag_path = (
        outputs_dir / "infer_step1_locate_components_fail_pos_eid_diagnosis.json"
    )
    locate_pos_enriched_path = (
        outputs_dir / "infer_step1_locate_components_fail_pos_samples_enriched.jsonl"
    )
    locate_samples_loaded = 0
    locate_pos_enriched: list[dict[str, Any]] = []
    locate_by_eid: dict[str, dict[str, Any]] = {}
    missing_eids: list[str] = []
    if locate_source_samples and locate_source_samples.exists():
        try:
            with locate_source_samples.open("r", encoding="utf-8") as fp:
                for line in fp:
                    line = line.strip()
                    if not line:
                        continue
                    sample = json.loads(line)
                    locate_samples_loaded += 1
                    example_id = sample.get("example_id")
                    example_key = str(example_id) if example_id is not None else ""
                    role = role_by_example_id.get(example_key)
                    if role is None or (isinstance(role, str) and role.strip() == ""):
                        continue
                    role_value = str(role)
                    if role_value not in pos_roles:
                        continue
                    gold_entry = gold_xlsx_by_example_id.get(example_key, {})
                    gold_eid = gold_entry.get("gold_eid")
                    reason_code = sample.get("reason_code")
                    reason_detail = sample.get("reason_detail")
                    debug_ptr = sample.get("debug_ptr")
                    locate_pos_enriched.append(
                        {
                            "example_id": sample.get("example_id"),
                            "instance_id": sample.get("instance_id"),
                            "gold_example_role": role_value,
                            "gold_eid": gold_eid,
                            "gold_span_segments": gold_entry.get("gold_span_segments"),
                            "gold_span_key": gold_entry.get("gold_span_key"),
                            "reason_code": reason_code,
                            "why": reason_detail,
                            "debug_ptr": debug_ptr,
                            "canonical_form": "",
                            "group": "",
                            "polyset_id": "",
                            "disconti_allowed": "",
                            "note": None,
                        }
                    )
        except Exception as exc:
            raise ConfigError(f"locate_components_fail pos read failed: {exc}") from exc

    logger.info(
        "[eval][locate_components_fail_pos_diag] infer_step1_run_dir_used=%s n_pos_samples=%s loaded_fail_samples=%s",
        infer_step1_run_dir_used if infer_step1_run_dir_used is not None else "",
        len(locate_pos_enriched),
        locate_samples_loaded,
    )

    expredict_by_eid: dict[str, dict[str, Any]] = {}
    components_by_eid: dict[str, list[dict[str, Any]]] = defaultdict(list)
    rules_by_ruleset_id: dict[str, list[dict[str, Any]]] = defaultdict(list)
    try:
        from openpyxl import load_workbook

        expredict_xlsx_path = Path(cfg["paths"]["dict_xlsx"])
        if not expredict_xlsx_path.exists():
            raise ConfigError(f"expredict.xlsx not found: {expredict_xlsx_path}")
        dict_cfg = cfg.get("dict", {}) or {}
        sheet_names = dict_cfg.get("sheet_names", {}) or {}
        expredict_sheet = sheet_names.get("expredict", "expredict")
        components_sheet = sheet_names.get("components", "components")
        rules_sheet = sheet_names.get("rules", "rules")

        wb = load_workbook(expredict_xlsx_path, read_only=True, data_only=True)
        if expredict_sheet in wb.sheetnames:
            ws = wb[expredict_sheet]
        else:
            ws = wb[wb.sheetnames[0]]
        expredict_header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        if "e_id" in expredict_header:
            idx_eid = expredict_header.index("e_id")
            idx_map = {name: expredict_header.index(name) for name in expredict_header if name}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                eid = row[idx_eid] if idx_eid < len(row) else None
                if eid is None:
                    continue
                eid_key = str(eid)
                if eid_key in expredict_by_eid:
                    continue
                expredict_by_eid[eid_key] = {
                    "canonical_form": row[idx_map["canonical_form"]]
                    if "canonical_form" in idx_map and idx_map["canonical_form"] < len(row)
                    else None,
                    "group": row[idx_map["group"]]
                    if "group" in idx_map and idx_map["group"] < len(row)
                    else None,
                    "polyset_id": row[idx_map["polyset_id"]]
                    if "polyset_id" in idx_map and idx_map["polyset_id"] < len(row)
                    else None,
                    "spacing_policy": row[idx_map["spacing_policy"]]
                    if "spacing_policy" in idx_map and idx_map["spacing_policy"] < len(row)
                    else None,
                    "disconti_allowed": row[idx_map["disconti_allowed"]]
                    if "disconti_allowed" in idx_map and idx_map["disconti_allowed"] < len(row)
                    else None,
                    "detect_ruleset_id": row[idx_map["detect_ruleset_id"]]
                    if "detect_ruleset_id" in idx_map and idx_map["detect_ruleset_id"] < len(row)
                    else None,
                }

        if components_sheet in wb.sheetnames:
            ws = wb[components_sheet]
            comp_header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            if "e_id" in comp_header:
                idx_map = {name: comp_header.index(name) for name in comp_header if name}
                idx_eid = idx_map["e_id"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row:
                        continue
                    eid = row[idx_eid] if idx_eid < len(row) else None
                    if eid is None:
                        continue
                    eid_key = str(eid)
                    components_by_eid[eid_key].append(
                        {
                            "comp_id": row[idx_map["comp_id"]]
                            if "comp_id" in idx_map and idx_map["comp_id"] < len(row)
                            else None,
                            "comp_surf": row[idx_map["comp_surf"]]
                            if "comp_surf" in idx_map and idx_map["comp_surf"] < len(row)
                            else None,
                            "anchor_rank": row[idx_map["anchor_rank"]]
                            if "anchor_rank" in idx_map and idx_map["anchor_rank"] < len(row)
                            else None,
                            "is_required": row[idx_map["is_required"]]
                            if "is_required" in idx_map and idx_map["is_required"] < len(row)
                            else None,
                            "min_gap_to_next": row[idx_map["min_gap_to_next"]]
                            if "min_gap_to_next" in idx_map and idx_map["min_gap_to_next"] < len(row)
                            else None,
                            "max_gap_to_next": row[idx_map["max_gap_to_next"]]
                            if "max_gap_to_next" in idx_map and idx_map["max_gap_to_next"] < len(row)
                            else None,
                        }
                    )

        if rules_sheet in wb.sheetnames:
            ws = wb[rules_sheet]
            rules_header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            if "ruleset_id" in rules_header:
                idx_map = {name: rules_header.index(name) for name in rules_header if name}
                idx_ruleset = idx_map["ruleset_id"]
                idx_stage = idx_map.get("stage")
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row:
                        continue
                    ruleset_id = row[idx_ruleset] if idx_ruleset < len(row) else None
                    if ruleset_id is None:
                        continue
                    if idx_stage is not None and idx_stage < len(row):
                        stage_val = row[idx_stage]
                        if stage_val is not None and str(stage_val).strip().lower() != "detect":
                            continue
                    rules_by_ruleset_id[str(ruleset_id)].append(
                        {
                            "rule_id": row[idx_map["rule_id"]]
                            if "rule_id" in idx_map and idx_map["rule_id"] < len(row)
                            else None,
                            "rule_type": row[idx_map["rule_type"]]
                            if "rule_type" in idx_map and idx_map["rule_type"] < len(row)
                            else None,
                            "target": row[idx_map["target"]]
                            if "target" in idx_map and idx_map["target"] < len(row)
                            else None,
                            "pattern": row[idx_map["pattern"]]
                            if "pattern" in idx_map and idx_map["pattern"] < len(row)
                            else None,
                            "priority": row[idx_map["priority"]]
                            if "priority" in idx_map and idx_map["priority"] < len(row)
                            else None,
                            "hard_fail": row[idx_map["hard_fail"]]
                            if "hard_fail" in idx_map and idx_map["hard_fail"] < len(row)
                            else None,
                        }
                    )
    except Exception as exc:
        raise ConfigError(f"locate_components_fail expredict read failed: {exc}") from exc

    for item in locate_pos_enriched:
        eid = item.get("gold_eid")
        if eid is None or str(eid).strip() == "":
            continue
        eid_key = str(eid)
        entry = locate_by_eid.get(eid_key)
        if entry is None:
            entry = {
                "n_samples": 0,
                "example_ids": [],
                "span_examples": [],
                "reason_code_counts": Counter(),
            }
            locate_by_eid[eid_key] = entry
        entry["n_samples"] += 1
        if len(entry["example_ids"]) < 50:
            entry["example_ids"].append(str(item.get("example_id")))
        reason_code = item.get("reason_code") or ""
        entry["reason_code_counts"][reason_code] += 1
        if len(entry["span_examples"]) < 5:
            sentence = gold_xlsx_by_example_id.get(str(item.get("example_id")), {}).get(
                "target_sentence"
            )
            short_target = ""
            if isinstance(sentence, str):
                short_target = sentence[:60]
            entry["span_examples"].append(
                {
                    "example_id": item.get("example_id"),
                    "instance_id": item.get("instance_id"),
                    "gold_span_key": item.get("gold_span_key"),
                    "gold_span_segments": item.get("gold_span_segments"),
                    "short_target": short_target,
                    "reason_code": reason_code,
                    "why": item.get("why"),
                }
            )

    for eid_key, entry in locate_by_eid.items():
        expredict = expredict_by_eid.get(eid_key)
        if expredict is None:
            missing_eids.append(eid_key)
            expredict = {
                "canonical_form": None,
                "group": None,
                "polyset_id": None,
                "spacing_policy": None,
                "disconti_allowed": None,
                "detect_ruleset_id": None,
            }
        entry["expredict_meta"] = expredict
        components = components_by_eid.get(eid_key, [])
        entry["components_preview"] = [
            {
                "comp_id": comp.get("comp_id"),
                "comp_surf": comp.get("comp_surf"),
                "is_required": comp.get("is_required"),
                "anchor_rank": comp.get("anchor_rank"),
                "min_gap": comp.get("min_gap_to_next"),
                "max_gap": comp.get("max_gap_to_next"),
            }
            for comp in components
            if comp.get("comp_id") is not None
        ]
        ruleset_id = expredict.get("detect_ruleset_id")
        rules = rules_by_ruleset_id.get(str(ruleset_id), []) if ruleset_id is not None else []
        if rules:
            rules = sorted(
                rules,
                key=lambda r: (r.get("priority") if r.get("priority") is not None else 0),
                reverse=True,
            )
        entry["detect_rules_preview"] = rules[:5]
        entry["reason_code_counts"] = dict(entry["reason_code_counts"])

    diagnosis_payload = {
        "n_pos_total": len(locate_pos_enriched),
        "by_eid": locate_by_eid,
        "missing_eids_in_expredict": missing_eids,
    }
    try:
        write_json(locate_pos_diag_path, diagnosis_payload, indent=2)
    except Exception as exc:
        raise ConfigError(
            f"infer_step1_locate_components_fail_pos_eid_diagnosis write failed: {exc}"
        ) from exc

    def _action_hint_from_reason(reason_code: Any) -> str | None:
        val = str(reason_code or "")
        if val in {"gap_out_of_bounds", "all_gap_failed"}:
            return "gap"
        if val == "required_comp_missing":
            return "required"
        if val == "anchor_not_found":
            return "anchor"
        if val == "special_generated_but_dropped":
            return "special_drop"
        return None

    def _span_key_from_segments(raw_segments: Any) -> str | None:
        if raw_segments is None:
            return None
        segments = raw_segments
        if isinstance(raw_segments, str):
            try:
                segments = ast.literal_eval(raw_segments)
            except Exception:
                return None
        if not isinstance(segments, list):
            return None
        parts: list[str] = []
        for seg in segments:
            if not isinstance(seg, (list, tuple)) or len(seg) != 2:
                return None
            try:
                s = int(seg[0])
                e = int(seg[1])
            except Exception:
                return None
            parts.append(f"{s}:{e}")
        if not parts:
            return None
        return "|".join(parts)

    n_span_key_filled = 0
    span_key_non_null = 0
    instance_id_non_null = 0
    try:
        with locate_pos_enriched_path.open("w", encoding="utf-8") as fp:
            for item in locate_pos_enriched:
                note_parts: list[str] = []
                span_key = item.get("gold_span_key")
                if not span_key:
                    span_key = _span_key_from_segments(item.get("gold_span_segments"))
                    if span_key is None and item.get("gold_span_segments") is not None:
                        note_parts.append("span_segments_parse_fail")
                if span_key:
                    n_span_key_filled += 1
                    span_key_non_null += 1

                instance_id = item.get("instance_id")
                if instance_id is None:
                    example_key = str(item.get("example_id") or "")
                    gold_eid = item.get("gold_eid")
                    candidates = gold_rows_by_example.get(example_key, [])
                    if gold_eid is not None:
                        candidates = [
                            r for r in candidates if str(r.get("e_id") or "") == str(gold_eid)
                        ]
                    inst_values: list[int] = []
                    for cand in candidates:
                        inst = cand.get("instance_id")
                        try:
                            if inst is not None:
                                inst_values.append(int(inst))
                        except Exception:
                            continue
                    if inst_values:
                        instance_id = min(inst_values)
                    else:
                        instance_id = 1
                        note_parts.append("instance_id_fallback=1")
                if instance_id is not None:
                    instance_id_non_null += 1
                item["canonical_form"] = expredict_by_eid.get(str(item.get("gold_eid") or ""), {}).get(
                    "canonical_form"
                )
                item["group"] = expredict_by_eid.get(str(item.get("gold_eid") or ""), {}).get("group")
                item["polyset_id"] = expredict_by_eid.get(str(item.get("gold_eid") or ""), {}).get(
                    "polyset_id"
                )
                item["disconti_allowed"] = expredict_by_eid.get(
                    str(item.get("gold_eid") or ""), {}
                ).get("disconti_allowed")
                action_hint = _action_hint_from_reason(item.get("reason_code"))
                if action_hint:
                    note_parts.append(action_hint)
                item["note"] = ";".join(note_parts) if note_parts else None
                item["instance_id"] = instance_id
                item["gold_span_key"] = span_key
                write_jsonl_line(fp, item)
    except Exception as exc:
        raise ConfigError(
            f"infer_step1_locate_components_fail_pos_samples_enriched write failed: {exc}"
        ) from exc

    diag_bytes = locate_pos_diag_path.stat().st_size
    enriched_bytes = locate_pos_enriched_path.stat().st_size
    logger.info(
        "[eval][locate_components_fail_pos_diag] wrote infer_step1_locate_components_fail_pos_eid_diagnosis.json bytes=%s n_eids=%s",
        diag_bytes,
        len(locate_by_eid),
    )
    logger.info(
        "[eval][locate_components_fail_pos_diag] wrote infer_step1_locate_components_fail_pos_samples_enriched.jsonl lines=%s bytes=%s n_span_key_filled=%s",
        len(locate_pos_enriched),
        enriched_bytes,
        n_span_key_filled,
    )
    logger.info(
        "[eval][locate_components_fail_pos_diag] post_fill instance_id_non_null=%s span_key_non_null=%s enriched_lines=%s",
        instance_id_non_null,
        span_key_non_null,
        len(locate_pos_enriched),
    )

    detect_no_hit_pos_samples_path = (
        outputs_dir / "infer_step1_detect_no_hit_pos_samples.jsonl"
    )
    detect_no_hit_pos_summary_path = outputs_dir / "infer_step1_detect_no_hit_pos_summary.json"
    n_detect_no_hit_total = len(detect_no_hit_example_ids)
    logger.info(
        "[eval][detect_no_hit_pos] infer_step1_run_dir_used=%s n_detect_no_hit_total=%s",
        infer_step1_run_dir_used if infer_step1_run_dir_used is not None else "",
        n_detect_no_hit_total,
    )
    pos_roles = {"pos_conti", "pos_disconti"}
    pos_role_counts = {role: 0 for role in pos_roles}
    n_pos_total = 0
    n_missing_in_gold_xlsx = 0
    top_examples_missing_in_gold_xlsx: list[str] = []
    try:
        with detect_no_hit_pos_samples_path.open("w", encoding="utf-8") as fp_pos:
            for example_id in sorted(detect_no_hit_example_ids):
                role = role_by_example_id.get(example_id)
                gold_entry = gold_xlsx_by_example_id.get(example_id)
                gold_eid = gold_entry.get("gold_eid") if gold_entry else None
                if role is None or gold_entry is None or gold_eid is None:
                    n_missing_in_gold_xlsx += 1
                    if len(top_examples_missing_in_gold_xlsx) < 20:
                        top_examples_missing_in_gold_xlsx.append(example_id)
                    continue
                role_value = str(role).strip() if isinstance(role, str) else role
                if role_value not in pos_roles:
                    continue
                write_jsonl_line(
                    fp_pos,
                    {
                        "example_id": example_id,
                        "gold_example_role": role_value,
                        "gold_eid": gold_eid,
                        "gold_span_segments": gold_entry.get("gold_span_segments"),
                        "gold_span_key": gold_entry.get("gold_span_key"),
                        "target_sentence": gold_entry.get("target_sentence"),
                        "note": None,
                    },
                )
                n_pos_total += 1
                pos_role_counts[role_value] = pos_role_counts.get(role_value, 0) + 1
    except Exception as exc:
        raise ConfigError(
            f"infer_step1_detect_no_hit_pos_samples write failed: {exc}"
        ) from exc
    if n_pos_total == 0:
        logger.warning("[eval][detect_no_hit_pos] n_pos_total=0 (no samples)")
    pos_samples_bytes = detect_no_hit_pos_samples_path.stat().st_size
    logger.info(
        "[eval][detect_no_hit_pos] wrote infer_step1_detect_no_hit_pos_samples.jsonl lines=%s bytes=%s",
        n_pos_total,
        pos_samples_bytes,
    )
    pos_summary_payload = {
        "infer_step1_run_dir_used": str(infer_step1_run_dir_used)
        if infer_step1_run_dir_used is not None
        else "",
        "n_detect_no_hit_total": n_detect_no_hit_total,
        "n_pos_total": n_pos_total,
        "pos_role_counts": pos_role_counts,
        "n_missing_in_gold_xlsx": n_missing_in_gold_xlsx,
        "top_examples_missing_in_gold_xlsx": top_examples_missing_in_gold_xlsx,
    }
    try:
        write_json(detect_no_hit_pos_summary_path, pos_summary_payload, indent=2)
    except Exception as exc:
        raise ConfigError(
            f"infer_step1_detect_no_hit_pos_summary write failed: {exc}"
        ) from exc
    pos_summary_bytes = detect_no_hit_pos_summary_path.stat().st_size
    logger.info(
        "[eval][detect_no_hit_pos] wrote infer_step1_detect_no_hit_pos_summary.json bytes=%s n_pos_total=%s n_missing=%s",
        pos_summary_bytes,
        n_pos_total,
        n_missing_in_gold_xlsx,
    )

    detect_no_hit_pos_diag_path = (
        outputs_dir / "infer_step1_detect_no_hit_pos_eid_diagnosis.json"
    )
    detect_no_hit_pos_enriched_path = (
        outputs_dir / "infer_step1_detect_no_hit_pos_samples_enriched.jsonl"
    )
    pos_samples_path = detect_no_hit_pos_samples_path
    pos_samples: list[dict[str, Any]] = []
    if pos_samples_path.exists():
        with pos_samples_path.open("r", encoding="utf-8") as fp_pos:
            for line in fp_pos:
                line = line.strip()
                if not line:
                    continue
                pos_samples.append(json.loads(line))
    logger.info(
        "[eval][detect_no_hit_pos_diag] pos_samples_path=%s n_pos_samples=%s",
        pos_samples_path,
        len(pos_samples),
    )

    expredict_xlsx_path = Path(cfg["paths"]["dict_xlsx"])
    if not expredict_xlsx_path.exists():
        raise ConfigError(f"expredict.xlsx not found: {expredict_xlsx_path}")
    expredict_xlsx_bytes = expredict_xlsx_path.stat().st_size
    logger.info(
        "[eval][detect_no_hit_pos_diag] loaded expredict_xlsx=%s exists=%s bytes=%s",
        expredict_xlsx_path,
        expredict_xlsx_path.exists(),
        expredict_xlsx_bytes,
    )
    expredict_by_eid: dict[str, dict[str, Any]] = {}
    components_by_eid: dict[str, list[dict[str, Any]]] = defaultdict(list)
    rules_by_ruleset_id: dict[str, list[dict[str, Any]]] = defaultdict(list)
    try:
        from openpyxl import load_workbook

        dict_cfg = cfg.get("dict", {}) or {}
        sheet_names = dict_cfg.get("sheet_names", {}) or {}
        expredict_sheet = sheet_names.get("expredict", "expredict")
        components_sheet = sheet_names.get("components", "components")
        rules_sheet = sheet_names.get("rules", "rules")

        wb = load_workbook(expredict_xlsx_path, read_only=True, data_only=True)
        if expredict_sheet in wb.sheetnames:
            ws = wb[expredict_sheet]
        else:
            ws = wb[wb.sheetnames[0]]
        expredict_header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        if "e_id" in expredict_header:
            idx_eid = expredict_header.index("e_id")
            idx_map = {name: expredict_header.index(name) for name in expredict_header if name}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                eid = row[idx_eid] if idx_eid < len(row) else None
                if eid is None:
                    continue
                eid_key = str(eid)
                if eid_key in expredict_by_eid:
                    continue
                expredict_by_eid[eid_key] = {
                    "canonical_form": row[idx_map["canonical_form"]]
                    if "canonical_form" in idx_map and idx_map["canonical_form"] < len(row)
                    else None,
                    "group": row[idx_map["group"]]
                    if "group" in idx_map and idx_map["group"] < len(row)
                    else None,
                    "polyset_id": row[idx_map["polyset_id"]]
                    if "polyset_id" in idx_map and idx_map["polyset_id"] < len(row)
                    else None,
                    "spacing_policy": row[idx_map["spacing_policy"]]
                    if "spacing_policy" in idx_map and idx_map["spacing_policy"] < len(row)
                    else None,
                    "disconti_allowed": row[idx_map["disconti_allowed"]]
                    if "disconti_allowed" in idx_map and idx_map["disconti_allowed"] < len(row)
                    else None,
                    "default_confidence": row[idx_map["default_confidence"]]
                    if "default_confidence" in idx_map and idx_map["default_confidence"] < len(row)
                    else None,
                    "detect_ruleset_id": row[idx_map["detect_ruleset_id"]]
                    if "detect_ruleset_id" in idx_map and idx_map["detect_ruleset_id"] < len(row)
                    else None,
                    "verify_ruleset_id": row[idx_map["verify_ruleset_id"]]
                    if "verify_ruleset_id" in idx_map and idx_map["verify_ruleset_id"] < len(row)
                    else None,
                    "context_positive_ruleset_id": row[idx_map["context_positive_ruleset_id"]]
                    if "context_positive_ruleset_id" in idx_map
                    and idx_map["context_positive_ruleset_id"] < len(row)
                    else None,
                    "context_negative_ruleset_id": row[idx_map["context_negative_ruleset_id"]]
                    if "context_negative_ruleset_id" in idx_map
                    and idx_map["context_negative_ruleset_id"] < len(row)
                    else None,
                    "gloss": row[idx_map["gloss"]]
                    if "gloss" in idx_map and idx_map["gloss"] < len(row)
                    else None,
                    "pragmatics": row[idx_map["pragmatics"]]
                    if "pragmatics" in idx_map and idx_map["pragmatics"] < len(row)
                    else None,
                    "disambiguation_hint": row[idx_map["disambiguation_hint"]]
                    if "disambiguation_hint" in idx_map and idx_map["disambiguation_hint"] < len(row)
                    else None,
                }

        if components_sheet in wb.sheetnames:
            ws = wb[components_sheet]
            comp_header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            if "e_id" in comp_header:
                idx_map = {name: comp_header.index(name) for name in comp_header if name}
                idx_eid = idx_map["e_id"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row:
                        continue
                    eid = row[idx_eid] if idx_eid < len(row) else None
                    if eid is None:
                        continue
                    eid_key = str(eid)
                    components_by_eid[eid_key].append(
                        {
                            "comp_id": row[idx_map["comp_id"]]
                            if "comp_id" in idx_map and idx_map["comp_id"] < len(row)
                            else None,
                            "comp_surf": row[idx_map["comp_surf"]]
                            if "comp_surf" in idx_map and idx_map["comp_surf"] < len(row)
                            else None,
                            "anchor_rank": row[idx_map["anchor_rank"]]
                            if "anchor_rank" in idx_map and idx_map["anchor_rank"] < len(row)
                            else None,
                            "is_required": row[idx_map["is_required"]]
                            if "is_required" in idx_map and idx_map["is_required"] < len(row)
                            else None,
                            "min_gap_to_next": row[idx_map["min_gap_to_next"]]
                            if "min_gap_to_next" in idx_map and idx_map["min_gap_to_next"] < len(row)
                            else None,
                            "max_gap_to_next": row[idx_map["max_gap_to_next"]]
                            if "max_gap_to_next" in idx_map and idx_map["max_gap_to_next"] < len(row)
                            else None,
                        }
                    )

        if rules_sheet in wb.sheetnames:
            ws = wb[rules_sheet]
            rules_header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            if "ruleset_id" in rules_header:
                idx_map = {name: rules_header.index(name) for name in rules_header if name}
                idx_ruleset = idx_map["ruleset_id"]
                idx_stage = idx_map.get("stage")
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row:
                        continue
                    ruleset_id = row[idx_ruleset] if idx_ruleset < len(row) else None
                    if ruleset_id is None:
                        continue
                    if idx_stage is not None and idx_stage < len(row):
                        stage_val = row[idx_stage]
                        if stage_val is not None and str(stage_val).strip().lower() != "detect":
                            continue
                    rules_by_ruleset_id[str(ruleset_id)].append(
                        {
                            "rule_id": row[idx_map["rule_id"]]
                            if "rule_id" in idx_map and idx_map["rule_id"] < len(row)
                            else None,
                            "rule_type": row[idx_map["rule_type"]]
                            if "rule_type" in idx_map and idx_map["rule_type"] < len(row)
                            else None,
                            "target": row[idx_map["target"]]
                            if "target" in idx_map and idx_map["target"] < len(row)
                            else None,
                            "pattern": row[idx_map["pattern"]]
                            if "pattern" in idx_map and idx_map["pattern"] < len(row)
                            else None,
                            "priority": row[idx_map["priority"]]
                            if "priority" in idx_map and idx_map["priority"] < len(row)
                            else None,
                            "hard_fail": row[idx_map["hard_fail"]]
                            if "hard_fail" in idx_map and idx_map["hard_fail"] < len(row)
                            else None,
                        }
                    )
    except Exception as exc:
        raise ConfigError(f"failed to read expredict.xlsx: {exc}") from exc

    by_eid: dict[str, Any] = {}
    missing_eids: list[str] = []
    for item in pos_samples:
        eid = item.get("gold_eid")
        if eid is None:
            continue
        eid_key = str(eid)
        entry = by_eid.get(eid_key)
        if entry is None:
            entry = {
                "n_samples": 0,
                "example_ids": [],
                "span_examples": [],
            }
            by_eid[eid_key] = entry
        entry["n_samples"] += 1
        if len(entry["example_ids"]) < 50:
            entry["example_ids"].append(str(item.get("example_id")))
        if len(entry["span_examples"]) < 10:
            entry["span_examples"].append(
                {
                    "example_id": str(item.get("example_id")),
                    "span_segments": item.get("gold_span_segments"),
                    "target_sentence": item.get("target_sentence"),
                }
            )

    for eid_key, entry in by_eid.items():
        expredict = expredict_by_eid.get(eid_key)
        if expredict is None:
            missing_eids.append(eid_key)
            expredict = {
                "canonical_form": None,
                "group": None,
                "polyset_id": None,
                "spacing_policy": None,
                "disconti_allowed": None,
                "default_confidence": None,
                "detect_ruleset_id": None,
                "verify_ruleset_id": None,
                "context_positive_ruleset_id": None,
                "context_negative_ruleset_id": None,
                "gloss": None,
                "pragmatics": None,
                "disambiguation_hint": None,
            }
        entry["expredict"] = expredict
        components = components_by_eid.get(eid_key, [])
        entry["components"] = {
            "n_components": len(components),
            "anchors": [
                {
                    "comp_id": comp.get("comp_id"),
                    "comp_surf": comp.get("comp_surf"),
                    "anchor_rank": comp.get("anchor_rank"),
                    "is_required": comp.get("is_required"),
                }
                for comp in components
                if comp.get("comp_id") is not None
            ],
            "gap_constraints": [
                {
                    "comp_id": comp.get("comp_id"),
                    "min_gap_to_next": comp.get("min_gap_to_next"),
                    "max_gap_to_next": comp.get("max_gap_to_next"),
                }
                for comp in components
                if comp.get("comp_id") is not None
            ],
        }
        ruleset_id = expredict.get("detect_ruleset_id")
        if ruleset_id is not None:
            rules = rules_by_ruleset_id.get(str(ruleset_id), [])
        else:
            rules = []
        entry["detect_rules_preview"] = rules

    diagnosis_payload = {
        "source_eval_run_dir": str(run_context.run_dir),
        "pos_samples_path": str(pos_samples_path),
        "n_pos_samples": len(pos_samples),
        "by_eid": by_eid,
        "missing_eids_in_expredict": missing_eids,
        "notes": [
            "This report is for rule coverage diagnosis only. Do not change rules based on neg_target_absent."
        ],
    }
    try:
        write_json(detect_no_hit_pos_diag_path, diagnosis_payload, indent=2)
    except Exception as exc:
        raise ConfigError(
            f"infer_step1_detect_no_hit_pos_eid_diagnosis write failed: {exc}"
        ) from exc
    diag_bytes = detect_no_hit_pos_diag_path.stat().st_size
    logger.info(
        "[eval][detect_no_hit_pos_diag] wrote infer_step1_detect_no_hit_pos_eid_diagnosis.json bytes=%s n_eids=%s missing_eids=%s",
        diag_bytes,
        len(by_eid),
        len(missing_eids),
    )

    def _span_key_from_segments(raw_segments: Any) -> str | None:
        if raw_segments is None:
            return None
        segments = raw_segments
        if isinstance(raw_segments, str):
            try:
                segments = json.loads(raw_segments)
            except Exception:
                return None
        if not isinstance(segments, list):
            return None
        pairs: list[str] = []
        for seg in segments:
            if not isinstance(seg, (list, tuple)) or len(seg) != 2:
                return None
            try:
                s = int(seg[0])
                e = int(seg[1])
            except Exception:
                return None
            pairs.append(f"{s}:{e}")
        if not pairs:
            return None
        return "|".join(pairs)

    n_span_key_filled = 0
    try:
        with detect_no_hit_pos_enriched_path.open("w", encoding="utf-8") as fp_enriched:
            for item in pos_samples:
                raw_segments = item.get("gold_span_segments")
                span_key = item.get("gold_span_key")
                if span_key is None or (isinstance(span_key, str) and span_key.strip() == ""):
                    span_key = _span_key_from_segments(raw_segments)
                if span_key is not None:
                    n_span_key_filled += 1
                eid = item.get("gold_eid")
                expredict = expredict_by_eid.get(str(eid)) if eid is not None else None
                write_jsonl_line(
                    fp_enriched,
                    {
                        **item,
                        "gold_span_key": span_key,
                        "expredict_detect_ruleset_id": expredict.get("detect_ruleset_id")
                        if expredict
                        else None,
                        "expredict_spacing_policy": expredict.get("spacing_policy")
                        if expredict
                        else None,
                        "expredict_disconti_allowed": expredict.get("disconti_allowed")
                        if expredict
                        else None,
                        "expredict_canonical_form": expredict.get("canonical_form")
                        if expredict
                        else None,
                    },
                )
    except Exception as exc:
        raise ConfigError(
            f"infer_step1_detect_no_hit_pos_samples_enriched write failed: {exc}"
        ) from exc
    enriched_bytes = detect_no_hit_pos_enriched_path.stat().st_size
    logger.info(
        "[eval][detect_no_hit_pos_diag] wrote infer_step1_detect_no_hit_pos_samples_enriched.jsonl lines=%s bytes=%s n_span_key_filled=%s",
        len(pos_samples),
        enriched_bytes,
        n_span_key_filled,
    )
    try:
        import hashlib
        size_bytes = min_csv_path.stat().st_size
        first_lines: list[str] = []
        with min_csv_path.open("r", encoding="utf-8", errors="replace") as fp_check:
            for _ in range(3):
                line = fp_check.readline()
                if not line:
                    break
                first_lines.append(line.rstrip("\n"))
        h = hashlib.sha256()
        with min_csv_path.open("rb") as fp_bin:
            for chunk in iter(lambda: fp_bin.read(65536), b""):
                h.update(chunk)
        logger.info("[for_users][SSOT] mincsv_path=%s", min_csv_path)
        logger.info("[for_users][SSOT] mincsv_size_bytes=%s", size_bytes)
        logger.info("[for_users][SSOT] mincsv_first3=%r", first_lines)
        logger.info("[for_users][SSOT] mincsv_sha256=%s", h.hexdigest())
    except Exception as exc:
        logger.warning("[for_users][SSOT] mincsv hash/peek failed: %s", exc)

    if write_readme:
        readme_path.write_text(
            _build_eval_export_readme()
            + "\n\n(min) eval_latest_min.csv columns:\n"
            + "- row_kind, view, match_key, example_id, instance_id, target_sentence\n"
            + "- gold_eid, gold_span_key, pred_eid, pred_span_key, pred_triage, pred_score\n"
            + "- pred_ignored_reason, neg_gold_conf_e_ids, neg_gold_span_keys, neg_gold_roles\n"
            + "- status (TP/FN/IGNORE), candidate_in_neg_gold (TRUE/FALSE)\n",
            encoding="utf-8",
        )
    return row_counters


def _build_eval_export_rows(
    *,
    cfg: dict[str, Any],
    pred_path: Path,
    gold_by_key: dict[str, list[dict[str, Any]]],
    neg_by_key: dict[str, list[dict[str, Any]]],
    views: dict[str, dict[str, Any]],
    report_views: list[str],
    match_key_policy: str,
    target_only: bool,
    only_keys_in_gold: bool,
    span_scope_policy: str,
    include_ignored: bool,
    max_json_chars: int,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, int]]:
    logger = logging.getLogger("kmwe")
    logger.info("[eval] pred_path=%s", pred_path)
    responses_rows_report = None
    decision_none_rows_report = None
    if pred_path.name == "infer_candidates.reranked.jsonl":
        report_path = pred_path.parent / "infer_step2_rerank_report.json"
        summary_path = pred_path.parent / "rerank_summary.json"
        for p in (report_path, summary_path):
            if not p.exists():
                continue
            try:
                obj = json.loads(p.read_text(encoding="utf-8"))
            except Exception:
                continue
            if responses_rows_report is None and isinstance(obj.get("responses_rows"), int):
                responses_rows_report = int(obj.get("responses_rows"))
            if decision_none_rows_report is None and isinstance(obj.get("decision_none_rows"), int):
                decision_none_rows_report = int(obj.get("decision_none_rows"))
    rows: list[dict[str, Any]] = []
    rows_jsonl: list[dict[str, Any]] = []
    row_id_counter = 0
    morph_cache: dict[str, tuple[str, str]] = {}
    n_gold_with_pred_joined = 0
    n_gold_tp_but_no_pred = 0
    pred_seen_keys: set[str] = set()
    appended_neg_confusable_eval_rows = 0
    row_counters = {
        "n_gold_occurrence_rows": 0,
        "n_pred_ignored_rows": 0,
        "pred_ignored_reason_counts": Counter(),
        "trap_fp_total": 0,
        "trap_fp_reason_counts": Counter(),
    }
    n_rerank_applied = 0
    n_confirm_after_rerank = 0
    n_tp_after_rerank = 0
    n_fn_after_rerank = 0
    expect_rerank = pred_path.name == "infer_candidates.reranked.jsonl"

    def next_row_id() -> str:
        nonlocal row_id_counter
        row_id_counter += 1
        return f"row_{row_id_counter:06d}"

    def json_cell(obj: Any, max_chars: int) -> str:
        text = json.dumps(obj, ensure_ascii=False)
        if max_chars > 0 and len(text) > max_chars:
            return text[: max_chars - 20] + "…(truncated)"
        return text

    def span_text(sentence: str, span_segments: Any) -> str:
        return _span_text_from_segments(sentence, span_segments)

    def raw_candidates(record: dict[str, Any]) -> list[dict[str, Any]]:
        rerank = record.get("rerank") or {}
        rerank_status = rerank.get("status")
        if rerank_status in ("applied", "fallback") and record.get("candidates"):
            candidates = list(record.get("candidates") or [])
        elif record.get("silver_labels"):
            candidates = list(record.get("silver_labels") or [])
        else:
            candidates = list(record.get("candidates") or [])
        return [_normalize_pred_candidate(c) for c in candidates]

    def triage_ok(cand: dict[str, Any], include_hold: bool) -> bool:
        triage = cand.get("triage")
        if triage is None:
            return True
        if triage == "confirm":
            return True
        return include_hold and triage == "hold"

    def morph_info(sentence: str) -> tuple[str, str]:
        if sentence in morph_cache:
            return morph_cache[sentence]
        try:
            tokens = analyze_with_kiwi(sentence, model="cong-global")
            if not tokens:
                morph_cache[sentence] = ("", "")
                return morph_cache[sentence]
            text = " ".join(
                f"{t.get('surface')}/{t.get('pos')}" for t in tokens if t.get("surface") and t.get("pos")
            )
            morph_cache[sentence] = (text, "")
        except Exception:
            morph_cache[sentence] = ("", "morph=unavailable")
        return morph_cache[sentence]

    def summarize_bridge(record: dict[str, Any]) -> str:
        debug_detect = (record.get("debug") or {}).get("detect") or {}
        summary = {
            "adnominal_candidates": debug_detect.get("adnominal_candidates", 0),
            "choose_adnominal_over_normal": debug_detect.get("choose_adnominal_over_normal", 0),
        }
        summary["bridge_applied"] = summary.get("choose_adnominal_over_normal", 0) > 0
        return json.dumps(summary, ensure_ascii=False)

    def summarize_thing_bridge(record: dict[str, Any]) -> str:
        debug_detect = (record.get("debug") or {}).get("detect") or {}
        summary = {
            "thing_bridge": debug_detect.get("n_component_match_thing_bridge", 0),
            "thing_bridge_fused": debug_detect.get("thing_bridge_fused", 0),
            "thing_bridge_form_counts": debug_detect.get("thing_bridge_form_counts", {}),
        }
        return json.dumps(summary, ensure_ascii=False)

    def summarize_detect_components(record: dict[str, Any]) -> str:
        debug_detect = (record.get("debug") or {}).get("detect") or {}
        summary = {
            "n_components_span_fail": debug_detect.get("n_components_span_fail", 0),
            "n_component_span_fail_required": debug_detect.get("n_component_span_fail_required", 0),
            "fail_samples": debug_detect.get("n_components_span_fail_samples", 0),
        }
        return json.dumps(summary, ensure_ascii=False)

    def _find_pred_for_gold(
        gold_item: dict[str, Any], preds: list[dict[str, Any]]
    ) -> dict[str, Any] | None:
        ge = str(gold_item.get("e_id") or "")
        gk = _canonical_span_key(gold_item.get("span_key"), gold_item.get("span_segments"))
        if not ge or not gk:
            return None
        for pred in preds:
            pk = _canonical_span_key(pred.get("span_key"), pred.get("span_segments"))
            if str(pred.get("e_id") or "") == ge and pk == gk:
                return pred
        same_e = [pred for pred in preds if str(pred.get("e_id") or "") == ge]
        if same_e:
            return sorted(same_e, key=lambda x: float(x.get("score") or 0.0), reverse=True)[0]
        return None

    uncertainty_cfg = cfg.get("infer_step1", {}).get("uncertainty", {}) or {}
    try:
        low_conf_threshold = float(uncertainty_cfg.get("low_conf_threshold", 0.55))
    except Exception:
        low_conf_threshold = 0.55
    try:
        margin_threshold = float(uncertainty_cfg.get("margin_threshold", 0.10))
    except Exception:
        margin_threshold = 0.10

    def _fn_reason_code(
        record: dict[str, Any],
        gold_item: dict[str, Any],
        status: str,
        gold_example_role: Any,
    ) -> str:
        if status != "FN":
            return ""
        if str(gold_example_role or "").strip() == "neg_target_absent":
            return ""
        rerank = record.get("rerank") or {}
        rerank_status = str(rerank.get("status") or "")
        decision_line = str(rerank.get("decision_line") or "")
        selected_eid = rerank.get("selected_eid")
        if expect_rerank and (
            not rerank or rerank_status in ("", "export_only", "no_candidates", "missing_match_key")
        ):
            return "RERANK_NOT_APPLIED"
        if rerank_status == "no_response" or (
            not decision_line and rerank_status in ("no_response", "parse_fail", "guard_reject")
        ):
            return "LLM_NO_RESPONSE"
        if decision_line.startswith("DECISION: NONE"):
            return "LLM_NONE"
        if rerank_status in ("applied", "fallback") and (selected_eid is None or str(selected_eid).strip() == ""):
            return "LLM_NONE"
        gold_eid = gold_item.get("e_id")
        candidates = record.get("candidates") or []
        if not isinstance(candidates, list):
            candidates = []
        if gold_eid:
            cand_eids = [
                str(c.get("e_id") or "")
                for c in candidates
                if isinstance(c, dict) and c.get("e_id") is not None
            ]
            if str(gold_eid) not in cand_eids:
                return "DETECT_MISS"
            gold_span_key = _canonical_span_key(gold_item.get("span_key"), gold_item.get("span_segments"))
            if gold_span_key:
                cand_span_keys = [
                    _canonical_span_key(c.get("span_key"), c.get("span_segments"))
                    for c in candidates
                    if isinstance(c, dict)
                ]
                if gold_span_key not in cand_span_keys:
                    return "BOUNDARY_ERROR"
        return "UNCLASSIFIED_FN"

    def _none_reason(record: dict[str, Any]) -> str:
        rerank = record.get("rerank") or {}
        rerank_status = str(rerank.get("status") or "")
        decision_line = str(rerank.get("decision_line") or "")
        if rerank_status in ("no_response", "missing"):
            return "NO_RESPONSE"
        if rerank_status.startswith("http") or rerank_status in ("http_fail", "http_401"):
            return "HTTP_FAIL"
        if rerank_status in ("empty_text",):
            return "EMPTY_TEXT"
        if rerank_status in ("parse_fail", "guard_reject"):
            return "PROTOCOL_FAIL"
        if rerank.get("protocol_ok") is False:
            return "PROTOCOL_FAIL"
        candidates = record.get("candidates") or []
        if not isinstance(candidates, list):
            candidates = []
        to_llm = [
            c for c in candidates if isinstance(c, dict) and c.get("to_llm") is True
        ]
        n_to_llm_candidates = rerank.get("n_to_llm_candidates")
        if not isinstance(n_to_llm_candidates, int):
            n_to_llm_candidates = len(to_llm)
        if n_to_llm_candidates == 0 or not to_llm:
            return "NO_CANDIDATES_TO_LLM"
        if any(isinstance(c, dict) and c.get("ambiguous") is True for c in to_llm):
            return "AMBIGUOUS_POLYSET"
        confidences = []
        for cand in to_llm:
            if not isinstance(cand, dict):
                continue
            conf = cand.get("confidence")
            try:
                conf_val = float(conf)
            except Exception:
                continue
            confidences.append(conf_val)
        if confidences and max(confidences) < low_conf_threshold:
            return "LOW_CONFIDENCE"
        if len(confidences) >= 2:
            top1, top2 = sorted(confidences, reverse=True)[:2]
            if (top1 - top2) < margin_threshold:
                return "LOW_MARGIN"
        return "UNKNOWN"

    for view_name in report_views:
        view = views.get(view_name) or {}
        include_hold = view_name == "lenient"
        policies = {
            "target_only": target_only,
            "only_keys_in_gold": only_keys_in_gold,
            "match_key_policy": match_key_policy,
            "include_hold": include_hold,
            "span_scope_policy": span_scope_policy,
        }
        overall = view.get("overall", {})
        row_id = next_row_id()
        rows.append(
            {
                "row_kind": "overall_summary",
                "view": view_name,
                "match_key": "",
                "example_id": "",
                "instance_id": "",
                "doc_id": "",
                "sent_index": "",
                "target_sentence": "",
                "gold_eid": "",
                "gold_span_key": "",
                "gold_span_segments": "",
                "gold_span_text": "",
                "pred_eid": "",
                "pred_span_key": "",
                "pred_span_segments": "",
                "pred_span_text": "",
                "pred_score": "",
                "pred_triage": "",
                "pred_hard_fail_triggered": "",
                "pred_hard_fail_reasons": "",
                "pred_stage_hits_json": "",
                "pred_target_candidates_json": "",
                "pred_target_span_keys": "",
                "pred_target_span_texts": "",
                "n_pred_all": "",
                "n_pred_target": "",
                "n_ignored_non_target": "",
                "n_ignored_out_of_scope": "",
                "n_ignored_by_triage": "",
                "n_tp": "",
                "n_fp": "",
                "n_fn": "",
                "status": "SUMMARY",
                "status_detail": f"row_id={row_id}",
                "debug_detect_json": "",
                "debug_verify_json": "",
                "debug_context_json": "",
                "overall_precision": overall.get("precision", ""),
                "overall_recall": overall.get("recall", ""),
                "overall_f1": overall.get("f1", ""),
                "overall_tp": overall.get("tp", ""),
                "overall_fp": overall.get("fp", ""),
                "overall_fn": overall.get("fn", ""),
                "morph_info": "",
                "bridge_info": "",
                "thing_bridge_info": "",
                "detect_components_info": "",
                "policies_json": json.dumps(policies, ensure_ascii=False),
            }
        )
        rows_jsonl.append(
            {
                "row_id": row_id,
                "row_kind": "overall_summary",
                "view": view_name,
                "full": {"overall": overall, "policies": policies},
            }
        )

    with pred_path.open("r", encoding="utf-8") as fp_pred:
        for line in fp_pred:
            line = line.strip()
            if not line:
                continue
            record = json.loads(line)
            sentence = record.get("target_sentence") or record.get("raw_sentence") or ""
            if not sentence:
                continue
            match_key = _match_key(record, match_key_policy, fallback_sentence=sentence)
            if match_key:
                pred_seen_keys.add(str(match_key))
            gold_items = list(gold_by_key.get(match_key, []))
            neg_items = list(neg_by_key.get(match_key, []))
            gold_item = gold_items[0] if gold_items else {}
            target_eid = None
            if target_only and gold_items:
                if len(gold_items) != 1:
                    raise ConfigError(
                        f"partial-label eval: match_key={match_key} 에 gold 행이 {len(gold_items)}개입니다. "
                        f"(example_id#instance_id 기준으로 1개여야 합니다)"
                    )
                target_eid = str(gold_item.get("e_id") or "")
            gold_keys = {(g["e_id"], g["span_key"]) for g in gold_items}
            gold_target_span_keys = {g["span_key"] for g in gold_items if g.get("span_key")}
            is_positive = bool(gold_target_span_keys)

            rerank = record.get("rerank") or {}
            rerank_status = rerank.get("status")
            if rerank_status == "applied":
                n_rerank_applied += 1
                confirm_candidates = [
                    c
                    for c in (record.get("candidates") or [])
                    if isinstance(c, dict) and c.get("triage") == "confirm"
                ]
                if confirm_candidates:
                    n_confirm_after_rerank += 1

            raw_preds = raw_candidates(record)
            join_info = _compute_joined(gold_items, raw_preds)
            n_gold_with_pred_joined += int(join_info.get("joined_count") or 0)
            neg_summary = _summarize_neg_items(neg_items)
            neg_pairs = neg_summary.get("neg_pairs_map", {})
            neg_roles_set = neg_summary.get("neg_roles_set", set())
            morph_text, morph_detail = morph_info(sentence)

            for view_name in report_views:
                include_hold = view_name == "lenient"
                preds = _extract_pred_candidates(record, include_hold=include_hold)
                policies = {
                    "target_only": target_only,
                    "only_keys_in_gold": only_keys_in_gold,
                    "match_key_policy": match_key_policy,
                    "include_hold": include_hold,
                    "span_scope_policy": span_scope_policy,
                }
                if only_keys_in_gold and not gold_items:
                    if include_ignored:
                        for cand in raw_preds:
                            cand_pair = (
                                str(cand.get("e_id") or ""),
                                _canonical_span_key(cand.get("span_key"), cand.get("span_segments")),
                            )
                            neg_role = neg_pairs.get(cand_pair)
                            if neg_role:
                                row_counters["trap_fp_total"] += 1
                                row_counters["trap_fp_reason_counts"][neg_role] += 1
                                rows_jsonl.append(
                                    {
                                        "row_kind": "trap_fp",
                                        "match_key": match_key,
                                        "example_id": record.get("example_id"),
                                        "instance_id": record.get("instance_id"),
                                        "candidate_e_id": cand_pair[0],
                                        "candidate_span_key": cand_pair[1],
                                        "neg_role": neg_role,
                                    }
                                )
                                ignored_reason = "neg_gold_only"
                                ignored_detail = {
                                    "triage": cand.get("triage"),
                                    "view": view_name,
                                    "include_hold": include_hold,
                                    "span_scope_policy": span_scope_policy,
                                    "target_only": target_only,
                                }
                                row_id = next_row_id()
                                row = _build_pred_row(
                                    row_kind="pred_ignored",
                                    view=view_name,
                                    match_key=match_key,
                                    record=record,
                                    sentence=sentence,
                                    gold_item={},
                                    cand=cand,
                                    status="IGNORE",
                                    status_detail=f"row_id={row_id};ignored_reason=neg_gold_only",
                                    ignored_reason=ignored_reason,
                                    ignored_reason_detail=ignored_detail,
                                    neg_summary=neg_summary,
                                    candidate_pair=f"{cand_pair[0]}@{cand_pair[1]}",
                                    gold_eid_override=neg_summary.get("neg_gold_conf_e_ids")
                                    or neg_summary.get("neg_gold_legacy_e_ids"),
                                    gold_span_key_override=neg_summary.get("neg_gold_span_keys"),
                                    gold_example_role_override=neg_role,
                                    candidate_in_neg_gold=_candidate_in_neg_gold(
                                        neg_role,
                                        neg_summary.get("neg_gold_conf_e_ids")
                                        or neg_summary.get("neg_gold_legacy_e_ids"),
                                        neg_summary.get("neg_gold_span_keys"),
                                        cand_pair[0],
                                        cand_pair[1],
                                        default_value="",
                                    ),
                                    n_pred_all=len(raw_preds),
                                    n_pred_target=0,
                                    n_ignored_non_target=0,
                                    n_ignored_out_of_scope=0,
                                    n_ignored_by_triage=0,
                                    n_tp=0,
                                    n_fp=0,
                                    n_fn=0,
                                    morph_text=morph_text,
                                    morph_detail=morph_detail,
                                    policies=policies,
                                    bridge_info=summarize_bridge(record),
                                    thing_bridge_info=summarize_thing_bridge(record),
                                    detect_components_info=summarize_detect_components(record),
                                )
                                rows.append(row)
                                row_counters["n_pred_ignored_rows"] += 1
                                row_counters["pred_ignored_reason_counts"]["neg_gold_only"] += 1
                                rows_jsonl.append(
                                    {
                                        "row_id": row_id,
                                        "row_kind": "pred_ignored",
                                        "view": view_name,
                                        "full": {
                                            "record": record,
                                            "candidate": cand,
                                            "ignored_reason": ignored_reason,
                                            "ignored_reason_detail": ignored_detail,
                                            "policies": policies,
                                        },
                                    }
                                )
                                continue
                            row_id = next_row_id()
                            ignored_reason = "key_not_in_gold"
                            ignored_detail = {
                                "triage": cand.get("triage"),
                                "view": view_name,
                                "include_hold": include_hold,
                                "span_scope_policy": span_scope_policy,
                                "target_only": target_only,
                            }
                            row = _build_pred_row(
                                row_kind="pred_ignored",
                                view=view_name,
                                match_key=match_key,
                                record=record,
                                sentence=sentence,
                                gold_item={},
                                cand=cand,
                                status="IGNORE",
                                status_detail=f"row_id={row_id};ignored_reason=key_not_in_gold",
                                ignored_reason=ignored_reason,
                                ignored_reason_detail=ignored_detail,
                                neg_summary=neg_summary,
                                candidate_pair=f"{cand_pair[0]}@{cand_pair[1]}",
                                gold_eid_override=neg_summary.get("neg_gold_conf_e_ids")
                                or neg_summary.get("neg_gold_legacy_e_ids"),
                                gold_span_key_override=neg_summary.get("neg_gold_span_keys"),
                                gold_example_role_override=(
                                    next(iter(neg_roles_set)) if len(neg_roles_set) == 1 else ""
                                ),
                                candidate_in_neg_gold=_candidate_in_neg_gold(
                                    next(iter(neg_roles_set)) if len(neg_roles_set) == 1 else "",
                                    neg_summary.get("neg_gold_conf_e_ids")
                                    or neg_summary.get("neg_gold_legacy_e_ids"),
                                    neg_summary.get("neg_gold_span_keys"),
                                    cand_pair[0],
                                    cand_pair[1],
                                    default_value="",
                                ),
                                n_pred_all=len(raw_preds),
                                n_pred_target=0,
                                n_ignored_non_target=0,
                                n_ignored_out_of_scope=0,
                                n_ignored_by_triage=0,
                                n_tp=0,
                                n_fp=0,
                                n_fn=0,
                                morph_text=morph_text,
                                morph_detail=morph_detail,
                                policies=policies,
                                bridge_info=summarize_bridge(record),
                                thing_bridge_info=summarize_thing_bridge(record),
                                detect_components_info=summarize_detect_components(record),
                            )
                            rows.append(row)
                            row_counters["n_pred_ignored_rows"] += 1
                            row_counters["pred_ignored_reason_counts"]["key_not_in_gold"] += 1
                            rows_jsonl.append(
                                {
                                    "row_id": row_id,
                                    "row_kind": "pred_ignored",
                                    "view": view_name,
                                    "full": {
                                        "record": record,
                                        "candidate": cand,
                                        "ignored_reason": ignored_reason,
                                        "ignored_reason_detail": ignored_detail,
                                        "policies": policies,
                                    },
                                }
                            )
                        continue
                    if include_ignored:
                        for cand in raw_preds:
                            row_id = next_row_id()
                            ignored_reason = "key_not_in_gold"
                            ignored_detail = {
                                "triage": cand.get("triage"),
                                "view": view_name,
                                "include_hold": include_hold,
                                "span_scope_policy": span_scope_policy,
                                "target_only": target_only,
                            }
                            row = _build_pred_row(
                                row_kind="pred_ignored",
                                view=view_name,
                                match_key=match_key,
                                record=record,
                                sentence=sentence,
                                gold_item={},
                                cand=cand,
                                status="IGNORE",
                                status_detail=f"row_id={row_id};ignored_reason=key_not_in_gold",
                                ignored_reason=ignored_reason,
                                ignored_reason_detail=ignored_detail,
                                neg_summary=neg_summary,
                                candidate_pair=f"{str(cand.get('e_id') or '')}@{_canonical_span_key(cand.get('span_key'), cand.get('span_segments'))}",
                                gold_eid_override=neg_summary.get("neg_gold_conf_e_ids")
                                or neg_summary.get("neg_gold_legacy_e_ids"),
                                gold_span_key_override=neg_summary.get("neg_gold_span_keys"),
                                gold_example_role_override=(
                                    next(iter(neg_roles_set)) if len(neg_roles_set) == 1 else ""
                                ),
                                candidate_in_neg_gold=_candidate_in_neg_gold(
                                    next(iter(neg_roles_set)) if len(neg_roles_set) == 1 else "",
                                    neg_summary.get("neg_gold_conf_e_ids")
                                    or neg_summary.get("neg_gold_legacy_e_ids"),
                                    neg_summary.get("neg_gold_span_keys"),
                                    str(cand.get("e_id") or ""),
                                    _canonical_span_key(cand.get("span_key"), cand.get("span_segments")),
                                    default_value="",
                                ),
                                n_pred_all=len(raw_preds),
                                n_pred_target=0,
                                n_ignored_non_target=0,
                                n_ignored_out_of_scope=0,
                                n_ignored_by_triage=0,
                                n_tp=0,
                                n_fp=0,
                                n_fn=0,
                                morph_text=morph_text,
                                morph_detail=morph_detail,
                                policies=policies,
                                bridge_info=summarize_bridge(record),
                                thing_bridge_info=summarize_thing_bridge(record),
                                detect_components_info=summarize_detect_components(record),
                            )
                            rows.append(row)
                            row_counters["n_pred_ignored_rows"] += 1
                            row_counters["pred_ignored_reason_counts"]["key_not_in_gold"] += 1
                            rows_jsonl.append(
                                {
                                    "row_id": row_id,
                                    "row_kind": "pred_ignored",
                                    "view": view_name,
                                    "full": {
                                        "record": record,
                                        "candidate": cand,
                                        "ignored_reason": ignored_reason,
                                        "ignored_reason_detail": ignored_detail,
                                        "policies": policies,
                                    },
                                }
                            )
                    continue

                ignored_non_target = 0
                ignored_out_of_scope = 0
                ignored_by_triage = 0
                in_scope: list[dict[str, Any]] = []
                ignored_rows: list[tuple[dict[str, Any], str]] = []
                for cand in raw_preds:
                    cand_eid = str(cand.get("e_id") or "")
                    span_key = cand.get("span_key", "")
                    if target_only and target_eid and cand_eid != target_eid:
                        ignored_non_target += 1
                        ignored_rows.append((cand, "non_target"))
                        continue
                    if (
                        span_scope_policy == "gold_only"
                        and target_eid
                        and cand_eid == target_eid
                        and span_key not in gold_target_span_keys
                        and is_positive
                    ):
                        ignored_out_of_scope += 1
                        ignored_rows.append((cand, "out_of_scope"))
                        continue
                    if not triage_ok(cand, include_hold):
                        ignored_by_triage += 1
                        ignored_rows.append((cand, "triage_filtered"))
                        continue
                    in_scope.append(cand)

                used_keys: set[tuple[str, str]] = set()
                fp_cands: list[dict[str, Any]] = []
                tp_i = fp_i = fn_i = 0
                for cand in in_scope:
                    cand_key = (str(cand.get("e_id") or ""), cand.get("span_key", ""))
                    if cand_key in gold_keys and cand_key not in used_keys:
                        tp_i += 1
                        used_keys.add(cand_key)
                    else:
                        fp_i += 1
                        fp_cands.append(cand)
                for g in gold_items:
                    gkey = (g["e_id"], g["span_key"])
                    if gkey not in used_keys:
                        fn_i += 1

                if gold_items:
                    target_candidates = [
                        c for c in in_scope if str(c.get("e_id") or "") == str(gold_item.get("e_id") or "")
                    ]
                    target_candidate_summaries = [_summarize_candidate(c, sentence) for c in target_candidates]
                    pred_target_span_keys = ";".join(
                        [c.get("span_key", "") for c in target_candidate_summaries if c.get("span_key")]
                    )
                    pred_target_span_texts = ";".join(
                        [c.get("span_text", "") for c in target_candidate_summaries if c.get("span_text")]
                    )
                    gold_span_segments = gold_item.get("span_segments")
                    gold_span_text = span_text(sentence, gold_span_segments)
                    row_id = next_row_id()
                    status = "TP" if tp_i > 0 else "FN"
                    rerank = record.get("rerank") or {}
                    rerank_status = rerank.get("status") or ""
                    rerank_decision_line = rerank.get("decision_line") or ""
                    rerank_selected_eid = rerank.get("selected_eid") or ""
                    protocol_ok = rerank.get("protocol_ok")
                    if rerank_status == "" and not rerank:
                        rerank_status = "not_run"
                    elif rerank_status == "":
                        rerank_status = "unknown"
                    has_candidates = bool(record.get("candidates"))
                    has_to_llm = False
                    candidates = record.get("candidates") or []
                    if isinstance(candidates, list):
                        for cand in candidates:
                            if isinstance(cand, dict) and cand.get("to_llm") is True:
                                has_to_llm = True
                                break
                    evidence_missing_fields = []
                    if not rerank_decision_line:
                        evidence_missing_fields.append("decision_line")
                    if not rerank_status:
                        evidence_missing_fields.append("rerank_status")
                    if protocol_ok is None:
                        evidence_missing_fields.append("protocol_ok")
                    if responses_rows_report is None:
                        evidence_missing_fields.append("responses_rows")
                    if decision_none_rows_report is None:
                        evidence_missing_fields.append("decision_none_rows")
                    fn_reason_code = _fn_reason_code(
                        record, gold_item, status, gold_item.get("gold_example_role") or ""
                    )
                    none_reason = _none_reason(record) if fn_reason_code == "LLM_NONE" else ""
                    reason = ""
                    if status == "FN":
                        reason = "reason=no_in_scope_target_pred"
                        if in_scope and not target_candidates:
                            reason = "reason=pred_exists_but_filtered"
                    else:
                        matched = next(
                            (
                                c.get("span_key", "")
                                for c in target_candidate_summaries
                                if (c.get("e_id"), c.get("span_key")) in used_keys
                            ),
                            "",
                        )
                        reason = f"matched_span_key={matched}"
                    status_detail = (
                        f"row_id={row_id};match_key={match_key};view_policy="
                        f"include_hold={include_hold},span_scope_policy={span_scope_policy};{reason}"
                    )
                    if morph_detail:
                        status_detail = f"{status_detail};{morph_detail}"
                    pred_for_gold = _find_pred_for_gold(gold_item, preds)
                    if status == "TP" and pred_for_gold is None:
                        n_gold_tp_but_no_pred += 1
                        logger.warning(
                            "[for_user] TP but no pred joined: match_key=%s gold_eid=%s gold_span_key=%s",
                            match_key,
                            str(gold_item.get("e_id") or ""),
                            str(gold_item.get("span_key") or ""),
                        )
                    row = {
                        "row_kind": "gold_occurrence",
                        "view": view_name,
                        "match_key": match_key,
                        "example_id": record.get("example_id"),
                        "instance_id": record.get("instance_id"),
                        "doc_id": record.get("doc_id"),
                        "sent_index": record.get("sent_index"),
                        "target_sentence": sentence,
                        "gold_example_role": gold_item.get("gold_example_role") or "",
                        "gold_eid": gold_item.get("e_id"),
                        "gold_span_key": gold_item.get("span_key"),
                        "gold_span_segments": json_cell(gold_span_segments, max_json_chars)
                        if gold_span_segments is not None
                        else "",
                        "gold_span_text": gold_span_text,
                        "pred_eid": pred_for_gold.get("e_id") if pred_for_gold else "",
                        "pred_span_key": pred_for_gold.get("span_key") if pred_for_gold else "",
                        "pred_span_segments": json_cell(
                            pred_for_gold.get("span_segments"), max_json_chars
                        )
                        if pred_for_gold and pred_for_gold.get("span_segments") is not None
                        else "",
                        "pred_span_text": pred_for_gold.get("span_text") if pred_for_gold else "",
                        "pred_score": pred_for_gold.get("score") if pred_for_gold else "",
                        "pred_triage": pred_for_gold.get("triage") if pred_for_gold else "",
                        "rerank_status": rerank_status,
                        "rerank_decision_line": rerank_decision_line,
                        "rerank_selected_eid": rerank_selected_eid,
                        "decision_line": rerank_decision_line,
                        "protocol_ok": protocol_ok if protocol_ok is not None else "",
                        "pred_ignored_reason": "",
                        "pred_ignored_reason_detail": "",
                        "neg_gold_roles": neg_summary.get("neg_gold_roles", ""),
                        "neg_gold_span_keys": neg_summary.get("neg_gold_span_keys", ""),
                        "neg_gold_conf_e_ids": neg_summary.get("neg_gold_conf_e_ids", ""),
                        "neg_gold_legacy_e_ids": neg_summary.get("neg_gold_legacy_e_ids", ""),
                        "neg_gold_pairs_n": neg_summary.get("neg_gold_pairs_n", ""),
                        "candidate_pair": "",
                        "candidate_in_neg_gold": "",
                        "pred_hard_fail_triggered": pred_for_gold.get("hard_fail_triggered")
                        if pred_for_gold
                        else "",
                        "pred_hard_fail_reasons": json_cell(
                            pred_for_gold.get("hard_fail_reasons") or [], max_json_chars
                        )
                        if pred_for_gold
                        else "",
                        "pred_stage_hits_json": json_cell(
                            pred_for_gold.get("stage_hits") or {}, max_json_chars
                        )
                        if pred_for_gold
                        else "",
                        "pred_target_candidates_json": json_cell(target_candidate_summaries, max_json_chars),
                        "pred_target_span_keys": pred_target_span_keys,
                        "pred_target_span_texts": pred_target_span_texts,
                        "n_pred_all": len(raw_preds),
                        "n_pred_target": len(
                            [c for c in raw_preds if str(c.get("e_id") or "") == str(gold_item.get("e_id") or "")]
                        ),
                        "n_ignored_non_target": ignored_non_target,
                        "n_ignored_out_of_scope": ignored_out_of_scope,
                        "n_ignored_by_triage": ignored_by_triage,
                        "n_tp": tp_i,
                        "n_fp": fp_i,
                        "n_fn": fn_i,
                        "status": status,
                        "status_detail": status_detail,
                        "debug_detect_json": json_cell((record.get("debug") or {}).get("detect") or {}, max_json_chars),
                        "debug_verify_json": json_cell((record.get("debug") or {}).get("verify") or {}, max_json_chars),
                        "debug_context_json": json_cell(
                            (record.get("debug") or {}).get("context") or {}, max_json_chars
                        ),
                        "overall_precision": "",
                        "overall_recall": "",
                        "overall_f1": "",
                        "overall_tp": "",
                        "overall_fp": "",
                        "overall_fn": "",
                        "morph_info": morph_text,
                        "bridge_info": summarize_bridge(record),
                        "thing_bridge_info": summarize_thing_bridge(record),
                        "detect_components_info": summarize_detect_components(record),
                        "policies_json": json.dumps(policies, ensure_ascii=False),
                        "fn_reason_code": fn_reason_code,
                        "none_reason": none_reason,
                    }
                    rows.append(row)
                    if rerank_status == "applied" and view_name == "strict":
                        if status == "TP":
                            n_tp_after_rerank += 1
                        else:
                            n_fn_after_rerank += 1
                    row_counters["n_gold_occurrence_rows"] += 1
                    rows_jsonl.append(
                        {
                            "row_id": row_id,
                            "row_kind": "gold_occurrence",
                            "view": view_name,
                            "fn_reason_code": fn_reason_code,
                            "none_reason": none_reason,
                            "decision_line": rerank_decision_line,
                            "protocol_ok": protocol_ok,
                            "evidence_missing_fields": evidence_missing_fields,
                            "has_candidates": has_candidates,
                            "has_to_llm": has_to_llm,
                            "rerank_status": rerank_status,
                            "responses_rows": responses_rows_report,
                            "decision_none_rows": decision_none_rows_report,
                            "full": {
                                "record": record,
                                "gold_item": gold_item,
                                "target_candidates": target_candidates,
                                "policies": policies,
                            },
                        }
                    )

    if gold_by_key:
        for match_key, gold_items in gold_by_key.items():
            if not match_key or str(match_key) in pred_seen_keys:
                continue
            for gold_item in gold_items:
                gold_role = str(gold_item.get("gold_example_role") or "").strip()
                if not gold_role.startswith("pos") and gold_role != "neg_confusable":
                    continue
                sentence = str(gold_item.get("target_sentence") or "")
                gold_span_segments = gold_item.get("span_segments")
                gold_span_text = span_text(sentence, gold_span_segments)
                for view_name in report_views:
                    include_hold = view_name == "lenient"
                    policies = {
                        "target_only": target_only,
                        "only_keys_in_gold": only_keys_in_gold,
                        "match_key_policy": match_key_policy,
                        "include_hold": include_hold,
                        "span_scope_policy": span_scope_policy,
                    }
                    row_id = next_row_id()
                    fn_reason_code = _fn_reason_code(
                        {
                            "example_id": gold_item.get("example_id"),
                            "instance_id": gold_item.get("instance_id"),
                        },
                        gold_item,
                        "FN",
                        gold_role,
                    )
                    status_detail = (
                        f"row_id={row_id};match_key={match_key};view_policy="
                        f"include_hold={include_hold},span_scope_policy={span_scope_policy};"
                        "reason=missing_pred_record"
                    )
                    row = {
                        "row_kind": "gold_occurrence",
                        "view": view_name,
                        "match_key": match_key,
                        "example_id": gold_item.get("example_id"),
                        "instance_id": gold_item.get("instance_id"),
                        "doc_id": gold_item.get("doc_id"),
                        "sent_index": gold_item.get("sent_index"),
                        "target_sentence": sentence,
                        "gold_example_role": gold_role,
                        "gold_eid": gold_item.get("e_id"),
                        "gold_span_key": gold_item.get("span_key"),
                        "gold_span_segments": json_cell(gold_span_segments, max_json_chars)
                        if gold_span_segments is not None
                        else "",
                        "gold_span_text": gold_span_text,
                        "pred_eid": "",
                        "pred_span_key": "",
                        "pred_span_segments": "",
                        "pred_span_text": "",
                        "pred_score": "",
                        "pred_triage": "",
                        "rerank_status": "",
                        "rerank_decision_line": "",
                        "rerank_selected_eid": "",
                        "pred_ignored_reason": "",
                        "pred_ignored_reason_detail": "",
                        "neg_gold_roles": "",
                        "neg_gold_span_keys": "",
                        "neg_gold_conf_e_ids": "",
                        "neg_gold_legacy_e_ids": "",
                        "neg_gold_pairs_n": "",
                        "candidate_pair": "",
                        "candidate_in_neg_gold": "",
                        "pred_hard_fail_triggered": "",
                        "pred_hard_fail_reasons": "",
                        "pred_stage_hits_json": "",
                        "pred_target_candidates_json": "[]",
                        "pred_target_span_keys": "",
                        "pred_target_span_texts": "",
                        "n_pred_all": 0,
                        "n_pred_target": 0,
                        "n_ignored_non_target": 0,
                        "n_ignored_out_of_scope": 0,
                        "n_ignored_by_triage": 0,
                        "n_tp": 0,
                        "n_fp": 0,
                        "n_fn": 1,
                        "status": "FN",
                        "status_detail": status_detail,
                        "debug_detect_json": "{}",
                        "debug_verify_json": "{}",
                        "debug_context_json": "{}",
                        "overall_precision": "",
                        "overall_recall": "",
                        "overall_f1": "",
                        "overall_tp": "",
                        "overall_fp": "",
                        "overall_fn": "",
                        "morph_info": "",
                        "bridge_info": "",
                        "thing_bridge_info": "",
                        "detect_components_info": "",
                        "policies_json": json.dumps(policies, ensure_ascii=False),
                        "fn_reason_code": fn_reason_code,
                        "none_reason": "",
                    }
                    rows.append(row)
                    row_counters["n_gold_occurrence_rows"] += 1
                    rows_jsonl.append(
                        {
                            "row_id": row_id,
                            "row_kind": "gold_occurrence",
                            "view": view_name,
                            "fn_reason_code": fn_reason_code,
                            "none_reason": "",
                            "full": {
                                "record": {
                                    "example_id": gold_item.get("example_id"),
                                    "instance_id": gold_item.get("instance_id"),
                                    "target_sentence": sentence,
                                },
                                "gold_item": gold_item,
                                "target_candidates": [],
                                "policies": policies,
                            },
                        }
                    )
                    if view_name == "strict" and gold_role == "neg_confusable":
                        appended_neg_confusable_eval_rows += 1

                for cand in fp_cands:
                    row_id = next_row_id()
                    cand_pair = (
                        str(cand.get("e_id") or ""),
                        _canonical_span_key(cand.get("span_key"), cand.get("span_segments")),
                    )
                    neg_role_for_row = neg_pairs.get(cand_pair)
                    if not neg_role_for_row and len(neg_roles_set) == 1:
                        neg_role_for_row = next(iter(neg_roles_set))
                    row = _build_pred_row(
                        row_kind="pred_only_fp",
                        view=view_name,
                        match_key=match_key,
                        record=record,
                        sentence=sentence,
                        gold_item=gold_item if gold_items else {},
                        cand=cand,
                        status="FP",
                        status_detail=f"row_id={row_id};reason=fp_in_scope",
                        neg_summary=neg_summary,
                        candidate_pair=f"{cand_pair[0]}@{cand_pair[1]}",
                        gold_eid_override=neg_summary.get("neg_gold_conf_e_ids")
                        or neg_summary.get("neg_gold_legacy_e_ids"),
                        gold_span_key_override=neg_summary.get("neg_gold_span_keys"),
                        gold_example_role_override=neg_role_for_row or "",
                        candidate_in_neg_gold=_candidate_in_neg_gold(
                            neg_role_for_row or "",
                            neg_summary.get("neg_gold_conf_e_ids")
                            or neg_summary.get("neg_gold_legacy_e_ids"),
                            neg_summary.get("neg_gold_span_keys"),
                            cand_pair[0],
                            cand_pair[1],
                            default_value="",
                        ),
                        n_pred_all=len(raw_preds),
                        n_pred_target=len(
                            [c for c in raw_preds if str(c.get("e_id") or "") == str(gold_item.get("e_id") or "")]
                        )
                        if gold_items
                        else 0,
                        n_ignored_non_target=ignored_non_target,
                        n_ignored_out_of_scope=ignored_out_of_scope,
                        n_ignored_by_triage=ignored_by_triage,
                        n_tp=0,
                        n_fp=1,
                        n_fn=0,
                        morph_text=morph_text,
                        morph_detail=morph_detail,
                        policies=policies,
                        bridge_info=summarize_bridge(record),
                        thing_bridge_info=summarize_thing_bridge(record),
                        detect_components_info=summarize_detect_components(record),
                    )
                    rows.append(row)
                    rows_jsonl.append(
                        {
                            "row_id": row_id,
                            "row_kind": "pred_only_fp",
                            "view": view_name,
                            "full": {
                                "record": record,
                                "gold_item": gold_item,
                                "candidate": cand,
                                "policies": policies,
                            },
                        }
                    )

                if include_ignored:
                    for cand, reason in ignored_rows:
                        row_id = next_row_id()
                        cand_pair = (
                            str(cand.get("e_id") or ""),
                            _canonical_span_key(cand.get("span_key"), cand.get("span_segments")),
                        )
                        neg_role_for_row = neg_pairs.get(cand_pair)
                        if not neg_role_for_row and len(neg_roles_set) == 1:
                            neg_role_for_row = next(iter(neg_roles_set))
                        ignored_detail = {
                            "triage": cand.get("triage"),
                            "view": view_name,
                            "include_hold": include_hold,
                            "span_scope_policy": span_scope_policy,
                            "target_only": target_only,
                        }
                        row = _build_pred_row(
                            row_kind="pred_ignored",
                            view=view_name,
                            match_key=match_key,
                            record=record,
                            sentence=sentence,
                            gold_item=gold_item if gold_items else {},
                            cand=cand,
                            status="IGNORE",
                            status_detail=f"row_id={row_id};ignored_reason={reason}",
                            ignored_reason=reason,
                            ignored_reason_detail=ignored_detail,
                            neg_summary=neg_summary,
                            candidate_pair=f"{cand_pair[0]}@{cand_pair[1]}",
                            gold_eid_override=neg_summary.get("neg_gold_conf_e_ids")
                            or neg_summary.get("neg_gold_legacy_e_ids"),
                            gold_span_key_override=neg_summary.get("neg_gold_span_keys"),
                            gold_example_role_override=neg_role_for_row or "",
                            candidate_in_neg_gold=_candidate_in_neg_gold(
                                neg_role_for_row or "",
                                neg_summary.get("neg_gold_conf_e_ids")
                                or neg_summary.get("neg_gold_legacy_e_ids"),
                                neg_summary.get("neg_gold_span_keys"),
                                cand_pair[0],
                                cand_pair[1],
                                default_value="",
                            ),
                            n_pred_all=len(raw_preds),
                            n_pred_target=len(
                                [c for c in raw_preds if str(c.get("e_id") or "") == str(gold_item.get("e_id") or "")]
                            )
                            if gold_items
                            else 0,
                            n_ignored_non_target=ignored_non_target,
                            n_ignored_out_of_scope=ignored_out_of_scope,
                            n_ignored_by_triage=ignored_by_triage,
                            n_tp=0,
                            n_fp=0,
                            n_fn=0,
                            morph_text=morph_text,
                            morph_detail=morph_detail,
                            policies=policies,
                            bridge_info=summarize_bridge(record),
                            thing_bridge_info=summarize_thing_bridge(record),
                            detect_components_info=summarize_detect_components(record),
                        )
                        rows.append(row)
                        row_counters["n_pred_ignored_rows"] += 1
                        row_counters["pred_ignored_reason_counts"][reason] += 1
                        rows_jsonl.append(
                            {
                                "row_id": row_id,
                                "row_kind": "pred_ignored",
                                "view": view_name,
                                "full": {
                                    "record": record,
                                    "gold_item": gold_item,
                                    "candidate": cand,
                                    "ignored_reason": reason,
                                    "ignored_reason_detail": ignored_detail,
                                    "policies": policies,
                                },
                            }
                        )

    logger.info(
        "[for_user] gold_occurrence join: joined=%s tp_without_pred=%s",
        n_gold_with_pred_joined,
        n_gold_tp_but_no_pred,
    )
    logger.info(
        "[eval][rerank] n_rerank_applied=%s n_confirm_after_rerank=%s",
        n_rerank_applied,
        n_confirm_after_rerank,
    )
    row_counters["neg_confusable_appended_eval_rows"] = appended_neg_confusable_eval_rows
    logger.info(
        "[eval][rerank] strict_tp_after_rerank=%s strict_fn_after_rerank=%s",
        n_tp_after_rerank,
        n_fn_after_rerank,
    )
    if n_gold_with_pred_joined == 0:
        pred_hint = str(pred_path)
        run_hint = ""
        parts = list(pred_path.parts)
        if "build_silver" in parts:
            idx = parts.index("build_silver")
            if idx + 1 < len(parts):
                run_hint = parts[idx + 1]
        logger.warning("HINT: pred was selected from %s", pred_hint)
        if run_hint:
            logger.warning("HINT: pred build_silver run_id=%s", run_hint)
        logger.warning(
            "HINT: if gold uses g#### ids, run ingest_corpus→build_silver on gold_eval_input.csv and re-run eval"
        )
        logger.warning("HINT: or set eval.pred_path in config (if you add this option)")
    return rows, rows_jsonl, row_counters


def _build_pred_row(
    *,
    row_kind: str,
    view: str,
    match_key: str,
    record: dict[str, Any],
    sentence: str,
    gold_item: dict[str, Any],
    cand: dict[str, Any],
    status: str,
    status_detail: str,
    ignored_reason: str | None = None,
    ignored_reason_detail: dict[str, Any] | None = None,
    neg_summary: dict[str, Any] | None = None,
    candidate_pair: str | None = None,
    candidate_in_neg_gold: bool | None = None,
    gold_eid_override: str | None = None,
    gold_span_key_override: str | None = None,
    gold_example_role_override: str | None = None,
    n_pred_all: int,
    n_pred_target: int,
    n_ignored_non_target: int,
    n_ignored_out_of_scope: int,
    n_ignored_by_triage: int,
    n_tp: int,
    n_fp: int,
    n_fn: int,
    morph_text: str,
    morph_detail: str,
    policies: dict[str, Any],
    bridge_info: str,
    thing_bridge_info: str,
    detect_components_info: str,
) -> dict[str, Any]:
    span_text = _span_text_from_segments(sentence, cand.get("span_segments"))
    status_detail_full = status_detail
    if morph_detail:
        status_detail_full = f"{status_detail_full};{morph_detail}"
    neg_summary = neg_summary or {}
    gold_eid_value = gold_eid_override if gold_eid_override is not None else gold_item.get("e_id")
    gold_span_key_value = (
        gold_span_key_override
        if gold_span_key_override is not None
        else gold_item.get("span_key")
    )
    return {
        "row_kind": row_kind,
        "view": view,
        "match_key": match_key,
        "example_id": record.get("example_id"),
        "instance_id": record.get("instance_id"),
        "doc_id": record.get("doc_id"),
        "sent_index": record.get("sent_index"),
        "target_sentence": sentence,
        "gold_example_role": gold_example_role_override or "",
        "gold_eid": gold_eid_value,
        "gold_span_key": gold_span_key_value,
        "gold_span_segments": json.dumps(gold_item.get("span_segments"), ensure_ascii=False)
        if gold_item.get("span_segments") is not None
        else "",
        "gold_span_text": _span_text_from_segments(sentence, gold_item.get("span_segments")),
        "pred_eid": cand.get("e_id"),
        "pred_span_key": cand.get("span_key", ""),
        "candidate_e_id": cand.get("e_id") or "",
        "candidate_span_key": cand.get("span_key", "") or "",
        "pred_span_segments": json.dumps(cand.get("span_segments"), ensure_ascii=False)
        if cand.get("span_segments") is not None
        else "",
        "pred_span_text": span_text,
        "pred_score": cand.get("score"),
        "pred_triage": cand.get("triage"),
        "pred_ignored_reason": ignored_reason or "",
        "pred_ignored_reason_detail": json.dumps(ignored_reason_detail or {}, ensure_ascii=False),
        "neg_gold_roles": neg_summary.get("neg_gold_roles", ""),
        "neg_gold_span_keys": neg_summary.get("neg_gold_span_keys", ""),
        "neg_gold_conf_e_ids": neg_summary.get("neg_gold_conf_e_ids", ""),
        "neg_gold_legacy_e_ids": neg_summary.get("neg_gold_legacy_e_ids", ""),
        "neg_gold_pairs_n": neg_summary.get("neg_gold_pairs_n", ""),
        "candidate_pair": candidate_pair or "",
        "candidate_in_neg_gold": candidate_in_neg_gold if candidate_in_neg_gold is not None else "",
        "pred_hard_fail_triggered": cand.get("hard_fail_triggered"),
        "pred_hard_fail_reasons": json.dumps(cand.get("hard_fail_reasons") or [], ensure_ascii=False),
        "pred_stage_hits_json": json.dumps(cand.get("stage_hits") or {}, ensure_ascii=False),
        "pred_target_candidates_json": json.dumps([_summarize_candidate(cand, sentence)], ensure_ascii=False),
        "pred_target_span_keys": cand.get("span_key", ""),
        "pred_target_span_texts": span_text,
        "n_pred_all": n_pred_all,
        "n_pred_target": n_pred_target,
        "n_ignored_non_target": n_ignored_non_target,
        "n_ignored_out_of_scope": n_ignored_out_of_scope,
        "n_ignored_by_triage": n_ignored_by_triage,
        "n_tp": n_tp,
        "n_fp": n_fp,
        "n_fn": n_fn,
        "status": status,
        "status_detail": status_detail_full,
        "eval_tag": "",
        "eval_tag_reason": "",
        "debug_detect_json": json.dumps((record.get("debug") or {}).get("detect") or {}, ensure_ascii=False),
        "debug_verify_json": json.dumps((record.get("debug") or {}).get("verify") or {}, ensure_ascii=False),
        "debug_context_json": json.dumps((record.get("debug") or {}).get("context") or {}, ensure_ascii=False),
        "overall_precision": "",
        "overall_recall": "",
        "overall_f1": "",
        "overall_tp": "",
        "overall_fp": "",
        "overall_fn": "",
        "morph_info": morph_text,
        "bridge_info": bridge_info,
        "thing_bridge_info": thing_bridge_info,
        "detect_components_info": detect_components_info,
        "policies_json": json.dumps(policies, ensure_ascii=False),
    }


def _build_eval_export_readme() -> str:
    return (
        "# eval_latest.csv 안내\n\n"
        "- row_kind: overall_summary/gold_occurrence/pred_only_fp/pred_ignored\n"
        "- view: strict/lenient\n"
        "- match_key: example_id#instance_id 또는 정책 기반 키\n"
        "- pred_only_fp는 실제 FP로 카운트된 후보 1행씩\n"
        "- pred_ignored는 정책으로 평가에서 제외된 후보\n"
        "- eval_latest_rows.jsonl에는 CSV 행의 원본(full) 데이터가 저장됩니다.\n"
    )


def _extract_pred_candidates(record: dict[str, Any], *, include_hold: bool) -> list[dict[str, Any]]:
    candidates = []
    rerank = record.get("rerank") or {}
    rerank_status = rerank.get("status")
    if rerank_status in ("applied", "fallback") and record.get("candidates"):
        candidates = list(record.get("candidates") or [])
    elif record.get("silver_labels"):
        candidates = list(record.get("silver_labels") or [])
    elif record.get("candidates"):
        candidates = list(record.get("candidates") or [])

    filtered: list[dict[str, Any]] = []
    for cand in candidates:
        triage = cand.get("triage")
        if triage is None:
            filtered.append(_normalize_pred_candidate(cand))
            continue
        if triage == "confirm" or (include_hold and triage == "hold"):
            filtered.append(_normalize_pred_candidate(cand))
    return filtered


def _normalize_pred_candidate(candidate: dict[str, Any]) -> dict[str, Any]:
    span_segments = candidate.get("span_segments") or []
    if isinstance(span_segments, str):
        try:
            span_segments = ast.literal_eval(span_segments)
        except Exception:
            span_segments = []
    span_key = silver_loader._span_key_from_segments(span_segments) if span_segments else ""
    return {**candidate, "span_key": span_key}


def _resolve_pred_path(
    cfg: dict[str, Any], run_context: RunContext, logger: logging.Logger
) -> tuple[Path, str]:
    eval_cfg = cfg.get("eval", {})
    pred_path = (
        eval_cfg.get("pred_path")
        or eval_cfg.get("input_pred_path")
        or eval_cfg.get("pred_jsonl")
        or eval_cfg.get("input_jsonl")
    )
    if pred_path:
        key = "eval.pred_path" if eval_cfg.get("pred_path") else "eval.input_pred_path"
        logger.info("[eval][pred_path_source] key=%s value=%s", key, pred_path)
        p = Path(pred_path)
        if not p.exists() or p.stat().st_size <= 0:
            raise ConfigError(f"eval input_pred_path invalid: {p}")
        return p, "explicit"

    artifacts_root = _artifacts_root_from_outputs_dir(run_context.outputs_dir, logger)
    rerank_candidate = _latest_stage_output(artifacts_root, run_context.exp_id, "infer_step2_rerank")
    if rerank_candidate:
        logger.info("[eval] pred auto-selected (infer_step2_rerank): %s", rerank_candidate)
        return rerank_candidate, "infer_step2_rerank"

    allow_fallback = bool(eval_cfg.get("allow_fallback_to_infer_step1", False))
    if not allow_fallback:
        raise ConfigError(
            "infer_step2_rerank 산출물이 없어 eval 기본 선택을 중단합니다. "
            "fallback이 필요하면 eval.allow_fallback_to_infer_step1=true로 명시하세요."
        )

    prefer = eval_cfg.get("pred_stage_preference") or eval_cfg.get("prefer_stage") or "infer_step1,build_silver"
    if isinstance(prefer, str):
        stages = [p.strip() for p in prefer.replace(";", ",").split(",") if p.strip()]
    elif isinstance(prefer, (list, tuple)):
        stages = [str(p).strip() for p in prefer if p is not None and str(p).strip()]
    else:
        stages = []
    stages = [s for s in stages if s in {"infer_step1", "build_silver"}]
    if not stages:
        stages = ["infer_step1", "build_silver"]
    for stage in stages:
        candidate = _latest_stage_output(artifacts_root, run_context.exp_id, stage)
        if candidate:
            logger.info("[eval] pred auto-selected (fallback=%s): %s", stage, candidate)
            return candidate, f"fallback_{stage}"
    raise ConfigError("eval pred_jsonl 경로를 찾지 못했습니다 (infer_step2_rerank 없음).")


def _latest_stage_output(artifacts_root: Path, exp_id: str, stage: str) -> Path | None:
    logger = logging.getLogger("kmwe")
    base_dir = artifacts_root / exp_id / stage
    reason = ""
    if stage == "infer_step1":
        logger.info("[eval][pred_guard_scan_enter] called=true")
        pattern = r"^\d{8}_\d{6}_[A-Za-z0-9]{6}$"
        logger.info(
            "[eval][pred_guard_scan] step1_base=%s artifacts_dir=%s exp_id=%s pattern=%s",
            base_dir,
            artifacts_root,
            exp_id,
            pattern,
        )
    if not base_dir.exists():
        if stage == "infer_step1":
            logger.info(
                "[eval][pred_guard_scan] n_step1_dirs=0 sample_step1_dirs=[]"
            )
            logger.info(
                "[eval][pred_guard_scan] n_match_pattern=0 n_has_report=0 n_has_pred=0"
            )
            logger.info(
                "[eval][pred_guard_scan] chosen_latest_infer_step1=None reason=base_missing"
            )
            logger.info(
                "[eval][pred_guard_hint] check infer_step1 base dir and required files under outputs/ (report+pred)"
            )
        return None
    run_dirs = sorted([p for p in base_dir.iterdir() if p.is_dir()])
    if stage == "infer_step1":
        sample_dirs = [p.name for p in run_dirs[:5]]
        logger.info(
            "[eval][pred_guard_scan] n_step1_dirs=%s sample_step1_dirs=%s",
            len(run_dirs),
            sample_dirs,
        )
        pattern_re = re.compile(r"^\d{8}_\d{6}_[A-Za-z0-9]{6}$")
        match_dirs = [d for d in run_dirs if pattern_re.match(d.name or "")]
        has_report = [
            d for d in match_dirs if (d / "outputs" / "infer_step1_report.json").exists()
        ]
        has_pred = [
            d for d in has_report if (d / "outputs" / "infer_candidates.jsonl").exists()
        ]
        logger.info(
            "[eval][pred_guard_scan] n_match_pattern=%s n_has_report=%s n_has_pred=%s",
            len(match_dirs),
            len(has_report),
            len(has_pred),
        )
        if not run_dirs:
            reason = "no_dirs"
        elif not match_dirs:
            reason = "no_match_pattern"
        elif not has_report:
            reason = "no_report"
        elif not has_pred:
            reason = "no_pred"
    if not run_dirs:
        if stage == "infer_step1":
            if not reason:
                reason = "no_dirs"
            logger.info(
                "[eval][pred_guard_scan] chosen_latest_infer_step1=None reason=%s",
                reason,
            )
            logger.info(
                "[eval][pred_guard_hint] check infer_step1 base dir and required files under outputs/ (report+pred)"
            )
        return None
    if stage == "build_silver":
        output_name = "silver.jsonl"
    elif stage == "infer_step2_rerank":
        output_name = "infer_candidates.reranked.jsonl"
    else:
        output_name = "infer_candidates.jsonl"
    if stage == "infer_step2_rerank":
        applied_runs: list[Path] = []
        for run_dir in reversed(run_dirs):
            log_path = run_dir / "logs" / "stage.log"
            if not log_path.exists():
                continue
            try:
                text = log_path.read_text(encoding="utf-8", errors="replace")
            except Exception:
                continue
            matches = re.findall(r"applied_selected=(\d+)", text)
            if matches and int(matches[-1]) > 0:
                applied_runs.append(run_dir)
        if applied_runs:
            candidate = applied_runs[0] / "outputs" / output_name
            if not candidate.exists() or candidate.stat().st_size <= 0:
                raise ConfigError(f"eval pred_path invalid (infer_step2_rerank): {candidate}")
            return candidate
        candidate = run_dirs[-1] / "outputs" / output_name
        if not candidate.exists() or candidate.stat().st_size <= 0:
            raise ConfigError(f"eval pred_path invalid (infer_step2_rerank): {candidate}")
        return candidate
    candidate = run_dirs[-1] / "outputs" / output_name
    if candidate.exists():
        return candidate
    if stage == "infer_step1":
        if not reason:
            reason = "all_filtered"
        logger.info(
            "[eval][pred_guard_scan] chosen_latest_infer_step1=None reason=%s",
            reason,
        )
        logger.info(
            "[eval][pred_guard_hint] check infer_step1 base dir and required files under outputs/ (report+pred)"
        )
    return None


def _artifacts_root_from_outputs_dir(outputs_dir: Path, logger: logging.Logger) -> Path:
    outputs_dir = Path(outputs_dir)
    if len(outputs_dir.parents) < 4:
        raise ValueError(f"outputs_dir 경로 깊이가 부족합니다: {outputs_dir}")
    artifacts_root = outputs_dir.parents[3]
    logger.info("eval artifacts_root(from outputs_dir): %s", artifacts_root)
    return artifacts_root
